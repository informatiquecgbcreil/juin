"""Générateur de budget prévisionnel éditable (modèle CERFA / CAF-FADO) :
structure éditable, calculs, export Excel, envoi vers subvention / budget."""
import json
import uuid
from io import BytesIO

from flask import url_for


def _sample_structure():
    """Structure de référence (mêmes ordres de grandeur que le modèle fourni)."""
    return {
        "charges": [
            {"code": "60", "libelle": "Achat", "lines": [
                {"libelle": "Prestations de services", "montant": 500},
                {"libelle": "Fournitures d'activité", "montant": 1200},
                {"libelle": "Ligne vide", "montant": 0},
            ]},
        ],
        "produits": [
            {"code": "74", "libelle": "Subventions d'exploitation", "lines": [
                {"libelle": "CREIL", "montant": 3000},
                {"libelle": "CAF OISE", "montant": 10000},
            ]},
        ],
        "contrib_emplois": [
            {"code": "86", "libelle": "Emplois", "lines": [{"libelle": "862 Prestations", "montant": 1500}]},
        ],
        "contrib_ressources": [
            {"code": "87", "libelle": "Contributions", "lines": [{"libelle": "870 Bénévolat", "montant": 1500}]},
        ],
    }


def test_parse_amount():
    from app.previsionnel.cerfa import parse_amount
    assert parse_amount("500") == 500.0
    assert parse_amount("1 000,50") == 1000.5
    assert parse_amount("220+144+270") == 634.0
    assert parse_amount("") == 0.0
    assert parse_amount("=DANGER()") == 0.0
    assert parse_amount("100abc") == 0.0
    assert parse_amount("1,000.50") == 0.0
    assert parse_amount(500) == 500.0
    assert parse_amount(None) == 0.0


def test_default_structure_est_le_modele_cerfa():
    from app.previsionnel.cerfa import default_structure
    s = default_structure()
    assert {"charges", "produits", "contrib_emplois", "contrib_ressources"} <= set(s)
    codes = [c["code"] for c in s["charges"]]
    assert "60" in codes and "69" in codes
    prod_codes = [c["code"] for c in s["produits"]]
    assert "74" in prod_codes and "RP" in prod_codes


def test_parse_structure_json_et_defaut():
    from app.previsionnel.cerfa import parse_structure, default_structure
    parsed = parse_structure(json.dumps(_sample_structure()))
    assert parsed["charges"][0]["code"] == "60"
    assert parsed["charges"][0]["lines"][0]["montant"] == 500.0
    # ligne totalement vide filtrée
    assert all(l["libelle"] or l["montant"] for l in parsed["charges"][0]["lines"])
    # entrée vide -> modèle par défaut
    assert parse_structure("") == default_structure()
    assert parse_structure("{}") == default_structure()


def test_parse_structure_cape_les_tailles():
    from app.previsionnel.cerfa import parse_structure, MAX_COMPTES, MAX_LINES
    big = {"charges": [{"code": "6", "libelle": "x", "lines": [{"libelle": "l", "montant": 1}]}
                       for _ in range(MAX_COMPTES + 50)],
           "produits": [], "contrib_emplois": [], "contrib_ressources": []}
    parsed = parse_structure(json.dumps(big))
    assert len(parsed["charges"]) <= MAX_COMPTES


def test_compute_totals():
    from app.previsionnel.cerfa import compute_totals
    t = compute_totals(_sample_structure())
    assert t["total_charges"] == 1700.0
    assert t["total_produits"] == 13000.0
    assert t["total_emplois"] == 1500.0
    assert t["total_ressources"] == 1500.0
    assert t["total_general_charges"] == 3200.0
    assert t["total_general_produits"] == 14500.0
    assert t["equilibre"] == 11300.0


def test_lignes_budget_exclut_zero_et_contributions():
    from app.previsionnel.cerfa import lignes_budget
    lignes = lignes_budget(_sample_structure())
    assert len(lignes) == 4      # 2 charges + 2 produits, ligne vide + contributions exclues
    creil = next(l for l in lignes if l["libelle"] == "CREIL")
    assert creil["nature"] == "produit"
    assert creil["compte"].startswith("74 - ")
    assert creil["montant"] == 3000.0
    assert all("86" not in l["compte"] and "87" not in l["compte"] for l in lignes)


def test_safe_text_neutralise_formule():
    from app.previsionnel.cerfa import _safe_text
    assert _safe_text("=1+1").startswith("'")
    assert _safe_text("Normal") == "Normal"


def test_page_generateur_rendue(app, admin_client):
    with app.app_context():
        with app.test_request_context():
            url = url_for("previsionnel.generateur")
    r = admin_client.get(url)
    assert r.status_code == 200
    body = r.get_data(as_text=True)
    assert "Générateur de budget prévisionnel" in body
    assert 'id="genRoot"' in body
    assert 'id="cerfaDefault"' in body and "Achat" in body      # structure par défaut embarquée
    assert "budget-generateur.js" in body
    # l'action reste portée par un champ caché (anti-régression HTTP 400)
    assert 'name="action"' in body and 'id="cerfaAction"' in body
    assert 'name="structure_json"' in body


def _post_data(action, **extra):
    d = {"action": action, "organisme": "Centre Test", "exercice": "2026",
         "secteur": "Numérique", "structure_json": json.dumps(_sample_structure())}
    d.update(extra)
    return d


def test_export_xlsx(app, admin_client):
    from openpyxl import load_workbook
    with app.app_context():
        with app.test_request_context():
            url = url_for("previsionnel.generateur")
    r = admin_client.post(url, data=_post_data("export_xlsx"))
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("Content-Type", "")
    wb = load_workbook(BytesIO(r.data))
    ws = wb.active
    vals = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
    assert 500.0 in vals and 3000.0 in vals
    assert any(isinstance(v, str) and v.startswith("=SUM(") for v in vals)
    assert any(isinstance(v, str) and "TOTAL DES CHARGES" in v for v in vals)
    assert any(isinstance(v, str) and "équilibre" in v.lower() for v in vals)


def test_export_xlsx_neutralise_injection(app, admin_client):
    from openpyxl import load_workbook
    with app.app_context():
        with app.test_request_context():
            url = url_for("previsionnel.generateur")
    r = admin_client.post(url, data=_post_data("export_xlsx", organisme="=cmd()"))
    wb = load_workbook(BytesIO(r.data))
    assert wb.active["B1"].value == "'=cmd()"


def test_export_xlsx_secteur_caractere_interdit(app, admin_client):
    from openpyxl import load_workbook
    with app.app_context():
        with app.test_request_context():
            url = url_for("previsionnel.generateur")
    r = admin_client.post(url, data=_post_data("export_xlsx", secteur="A/B:C*?"))
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.data))
    assert not any(ch in wb.active.title for ch in r"[]:*?/\\")


def test_envoi_vers_subvention(app, admin_client):
    from app.extensions import db
    from app.models import Subvention, LigneBudget

    nom = f"BudgetGen{uuid.uuid4().hex[:6]}"
    with app.app_context():
        with app.test_request_context():
            url = url_for("previsionnel.generateur")
    admin_client.post(url, data=_post_data("to_subvention", exercice="2035", subvention_nom=nom),
                      follow_redirects=True)
    with app.app_context():
        sub = Subvention.query.filter_by(nom=nom).first()
        assert sub is not None and sub.annee_exercice == 2035
        lignes = LigneBudget.query.filter_by(subvention_id=sub.id).all()
        assert len(lignes) == 4
        assert sum(l.montant_base for l in lignes if l.nature == "charge") == 1700.0
        assert sum(l.montant_base for l in lignes if l.nature == "produit") == 13000.0
        assert all(len(l.compte) <= 20 for l in lignes)   # tient dans VARCHAR(20)
        assert any(l.compte == "74" for l in lignes)


def test_envoi_vers_previsionnel_avec_projet(app, admin_client):
    from app.extensions import db
    from app.models import Projet, BudgetPrevisionnel, BudgetPrevisionnelLigne

    with app.app_context():
        p = Projet(nom=f"ProjGen{uuid.uuid4().hex[:6]}", secteur="Numérique")
        db.session.add(p)
        db.session.commit()
        pid = p.id
        with app.test_request_context():
            url = url_for("previsionnel.generateur")
    admin_client.post(url, data=_post_data("to_previsionnel", exercice="2036", projet_id=str(pid)),
                      follow_redirects=True)
    with app.app_context():
        budget = BudgetPrevisionnel.query.filter_by(annee=2036, secteur="Numérique").order_by(
            BudgetPrevisionnel.id.desc()).first()
        assert budget is not None and budget.statut == "brouillon"
        lignes = BudgetPrevisionnelLigne.query.filter_by(budget_id=budget.id).all()
        assert len(lignes) == 4
        assert all(l.projet_id == pid for l in lignes)
        assert all(len(l.compte) <= 20 for l in lignes)


def test_envoi_previsionnel_refuse_projet_autre_secteur(app, admin_client):
    from app.extensions import db
    from app.models import Projet

    with app.app_context():
        p = Projet(nom=f"ProjX{uuid.uuid4().hex[:6]}", secteur="Familles")
        db.session.add(p)
        db.session.commit()
        pid = p.id
        with app.test_request_context():
            url = url_for("previsionnel.generateur")
    r = admin_client.post(url, data=_post_data("to_previsionnel", exercice="2037", projet_id=str(pid)))
    assert r.status_code == 400
