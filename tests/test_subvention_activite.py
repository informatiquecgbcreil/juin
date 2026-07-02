"""Lien subvention <-> ateliers financés : rattachement, justificatif chiffré, export."""
import datetime as dt
import uuid

from flask import url_for


def _contexte(app, annee=2042):
    """Subvention (10 000 € attribués) + atelier du même secteur avec activité réelle."""
    from app.extensions import db
    from app.models import (Subvention, AtelierActivite, SessionActivite,
                            PresenceActivite, Participant)
    suf = uuid.uuid4().hex[:6]
    sub = Subvention(nom=f"Just{suf}", secteur="Numérique", annee_exercice=annee,
                     financeur="CAF", statut_cycle="versee",
                     montant_attribue=10000, montant_recu=10000)
    at = AtelierActivite(nom=f"AtJ{suf}", secteur="Numérique", type_atelier="COLLECTIF")
    db.session.add_all([sub, at])
    db.session.flush()
    s1 = SessionActivite(atelier_id=at.id, secteur="Numérique", session_type="COLLECTIF",
                         date_session=dt.date(annee, 3, 1), heure_debut="14:00", heure_fin="16:00")
    s2 = SessionActivite(atelier_id=at.id, secteur="Numérique", session_type="COLLECTIF",
                         date_session=dt.date(annee, 3, 8), heure_debut="14:00", heure_fin="16:00")
    db.session.add_all([s1, s2])
    db.session.flush()
    for i in range(4):
        p = Participant(nom=f"PJ{i}{suf}", prenom="T")
        db.session.add(p)
        db.session.flush()
        db.session.add(PresenceActivite(session_id=s1.id, participant_id=p.id))
        if i < 2:
            db.session.add(PresenceActivite(session_id=s2.id, participant_id=p.id))
    db.session.commit()
    return sub.id, at.id


def test_rattacher_et_detacher_atelier(app, admin_client):
    from app.extensions import db
    from app.models import SubventionAtelier, AuditLog

    with app.app_context():
        sid, aid = _contexte(app)
        with app.test_request_context():
            url = url_for("main.subvention_pilotage", subvention_id=sid)

    admin_client.post(url, data={"action": "add_atelier", "atelier_id": aid,
                                 "poids_pct": "60", "justification": "40 jeunes QPV"},
                      follow_redirects=True)
    with app.app_context():
        lien = SubventionAtelier.query.filter_by(subvention_id=sid, atelier_id=aid).first()
        assert lien is not None
        assert lien.poids_pct == 60.0
        assert lien.justification == "40 jeunes QPV"
        assert AuditLog.query.filter_by(action="subvention.atelier_lien").count() >= 1
        lid = lien.id

    # doublon refusé silencieusement
    admin_client.post(url, data={"action": "add_atelier", "atelier_id": aid, "poids_pct": "50"},
                      follow_redirects=True)
    with app.app_context():
        assert SubventionAtelier.query.filter_by(subvention_id=sid, atelier_id=aid).count() == 1

    admin_client.post(url, data={"action": "del_atelier", "lien_id": lid}, follow_redirects=True)
    with app.app_context():
        assert SubventionAtelier.query.filter_by(subvention_id=sid).count() == 0


def test_atelier_autre_secteur_refuse(app, admin_client):
    from app.extensions import db
    from app.models import AtelierActivite, SubventionAtelier

    with app.app_context():
        sid, _aid = _contexte(app)
        autre = AtelierActivite(nom=f"Autre{uuid.uuid4().hex[:5]}", secteur="Familles",
                                type_atelier="COLLECTIF")
        db.session.add(autre)
        db.session.commit()
        autre_id = autre.id
        with app.test_request_context():
            url = url_for("main.subvention_pilotage", subvention_id=sid)

    r = admin_client.post(url, data={"action": "add_atelier", "atelier_id": autre_id})
    assert r.status_code == 400
    with app.app_context():
        assert SubventionAtelier.query.filter_by(subvention_id=sid, atelier_id=autre_id).count() == 0


def test_justificatif_chiffre(app, admin_client):
    from app.extensions import db
    from app.models import SubventionAtelier

    with app.app_context():
        sid, aid = _contexte(app, annee=2043)
        db.session.add(SubventionAtelier(subvention_id=sid, atelier_id=aid,
                                         poids_pct=100, justification="Public jeunes"))
        db.session.commit()
        with app.test_request_context():
            url = url_for("main.subvention_justificatif", subvention_id=sid)

    body = admin_client.get(url).get_data(as_text=True)
    # activité : 2 séances × 2 h = 4 h, 4 uniques, 6 présences
    assert "Justificatif" in body
    assert "Public jeunes" in body
    assert "2500.00 €" in body or "2500,00" in body or "2500.00" in body  # 10 000 / 4 participants


def test_justificatif_export_xlsx(app, admin_client):
    from app.extensions import db
    from app.models import SubventionAtelier
    from openpyxl import load_workbook
    from io import BytesIO

    with app.app_context():
        sid, aid = _contexte(app, annee=2044)
        db.session.add(SubventionAtelier(subvention_id=sid, atelier_id=aid, poids_pct=100))
        db.session.commit()
        with app.test_request_context():
            url = url_for("main.subvention_justificatif_xlsx", subvention_id=sid)

    r = admin_client.get(url)
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.data))
    vals = [c.value for row in wb.active.iter_rows() for c in row if c.value is not None]
    assert "CAF" in vals
    assert 2500.0 in vals          # coût / participant : 10 000 € / 4 uniques
    assert 4 in vals               # participants uniques ou heures
