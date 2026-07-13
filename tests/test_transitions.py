"""Module Transitions : étiquetage, tableau de bord, défis, mesures, export.

- le référentiel se seede tout seul (6 thématiques standard) ;
- un atelier étiqueté fait remonter automatiquement séances réalisées,
  participations, publics uniques, heures et éco-conso (kWh/CO₂) ;
- les séances annulées ne comptent pas ;
- croisement thématique × objectif sectoriel ;
- défis : engagement → réalisé (date posée) ;
- mesures : saisie, cumul par (libellé, unité), suppression ;
- export XLSX annuel ; pages protégées ; permissions RBAC attribuées.
"""
import uuid
from datetime import date

import pytest

ANNEE = 2002  # exercice « bac à sable », loin des autres jeux de tests


@pytest.fixture()
def contexte(app):
    """Un atelier étiqueté mobilité + OS, 1 séance réalisée (2 présents,
    conso figée), 1 séance annulée, 1 atelier non étiqueté (témoin)."""
    suf = uuid.uuid4().hex[:6]
    with app.app_context():
        from app.extensions import db
        from app.models import (
            AtelierActivite,
            ObjectifSectoriel,
            Participant,
            PresenceActivite,
            PresenceMaterielConsommation,
            SessionActivite,
        )
        from app.services.transitions import etiqueter_atelier, seed_thematiques

        seed_thematiques()
        from app.models import TransitionThematique
        mobilite = TransitionThematique.query.filter_by(code="mobilite").first()

        objectif = ObjectifSectoriel(secteur="Transitions", code=f"OS-{suf}",
                                     libelle="Expérimenter", ordre=1)
        atelier = AtelierActivite(nom=f"Vélo-école {suf}", secteur="Transitions")
        temoin = AtelierActivite(nom=f"Témoin {suf}", secteur="Transitions")
        p1 = Participant(nom=f"Trans{suf}", prenom="Ada")
        p2 = Participant(nom=f"Trans{suf}", prenom="Bob")
        db.session.add_all([objectif, atelier, temoin, p1, p2])
        db.session.flush()

        s_ok = SessionActivite(atelier_id=atelier.id, secteur="Transitions",
                               session_type="COLLECTIF", date_session=date(ANNEE, 5, 10),
                               heure_debut="14:00", heure_fin="16:00", statut="realisee")
        s_no = SessionActivite(atelier_id=atelier.id, secteur="Transitions",
                               session_type="COLLECTIF", date_session=date(ANNEE, 5, 17),
                               heure_debut="14:00", heure_fin="16:00", statut="annulee")
        db.session.add_all([s_ok, s_no])
        db.session.flush()

        pr1 = PresenceActivite(session_id=s_ok.id, participant_id=p1.id)
        pr2 = PresenceActivite(session_id=s_ok.id, participant_id=p2.id)
        db.session.add_all([pr1, pr2])
        db.session.flush()
        # Conso individuelle figée (comme le ferait l'émargement).
        db.session.add(PresenceMaterielConsommation(
            presence_id=pr1.id, session_id=s_ok.id, participant_id=p1.id,
            quantite=1, kwh_snapshot=0.5, co2_kg_snapshot=0.03,
        ))
        db.session.commit()

        etiqueter_atelier(atelier, [mobilite.id], [objectif.id])

        return {"suf": suf, "atelier_id": atelier.id, "temoin_id": temoin.id,
                "mobilite_id": mobilite.id, "objectif_id": objectif.id,
                "participant_ids": [p1.id, p2.id]}


def test_seed_thematiques(app):
    with app.app_context():
        from app.models import TransitionThematique
        from app.services.transitions import seed_thematiques

        seed_thematiques()
        codes = {t.code for t in TransitionThematique.query.all()}
        assert {"alimentation", "mobilite", "energie", "reemploi", "nature", "numerique"} <= codes
        seed_thematiques()  # idempotent
        assert TransitionThematique.query.filter_by(code="mobilite").count() == 1


def test_tableau_de_bord_agrege_automatiquement(app, contexte):
    with app.app_context():
        from app.services.transitions import tableau_de_bord

        data = tableau_de_bord(ANNEE)
        ligne = next(l for l in data["thematiques"] if l["thematique"].id == contexte["mobilite_id"])

        assert ligne["nb_ateliers"] >= 1
        assert ligne["seances"] == 1, "la séance annulée ne compte pas"
        assert ligne["participations"] == 2
        assert ligne["uniques"] == 2
        assert ligne["heures"] == 2.0
        assert ligne["kwh"] == 0.5
        assert ligne["co2_kg"] == 0.03

        assert data["totaux"]["participations"] >= 2
        # L'atelier témoin non étiqueté n'apparaît pas.
        assert all(a.id != contexte["temoin_id"] for a in data["ateliers"])


def test_croisement_thematique_objectif(app, contexte):
    with app.app_context():
        from app.services.transitions import tableau_de_bord

        data = tableau_de_bord(ANNEE)
        assert any(o.id == contexte["objectif_id"] for o in data["objectifs"])
        idx = [o.id for o in data["objectifs"]].index(contexte["objectif_id"])
        ligne = next(c for c in data["croisement"] if c["thematique"].id == contexte["mobilite_id"])
        assert ligne["cellules"][idx] == 2, "2 participations sur la cellule mobilité × OS"


def test_defi_cycle_de_vie(admin_client, app, contexte):
    r = admin_client.post("/transitions/defis/creer", data={
        "titre": f"Un mois à vélo {contexte['suf']}",
        "thematique_id": contexte["mobilite_id"],
        "participant_id": contexte["participant_ids"][0],
        "date_engagement": f"{ANNEE}-06-01",
    }, follow_redirects=True)
    assert r.status_code == 200
    assert "engagé ✅" in r.get_data(as_text=True)

    with app.app_context():
        from app.models import DefiTransition
        defi = DefiTransition.query.filter(DefiTransition.titre.contains(contexte["suf"])).first()
        assert defi is not None and defi.statut == "en_cours"
        defi_id = defi.id

    r = admin_client.post(f"/transitions/defis/{defi_id}/statut",
                          data={"statut": "realise"}, follow_redirects=True)
    assert "réalisé 🎉" in r.get_data(as_text=True)

    with app.app_context():
        from app.extensions import db
        from app.models import DefiTransition
        from app.services.transitions import tableau_de_bord
        defi = db.session.get(DefiTransition, defi_id)
        assert defi.statut == "realise"
        assert defi.date_realisation is not None
        data = tableau_de_bord(ANNEE)
        ligne = next(l for l in data["thematiques"] if l["thematique"].id == contexte["mobilite_id"])
        assert ligne["defis_realises"] >= 1


def test_mesures_cumul_et_suppression(admin_client, app, contexte):
    for valeur in ("120", "80,5"):
        r = admin_client.post("/transitions/mesures/creer", data={
            "atelier_id": contexte["atelier_id"],
            "libelle": f"km évités {contexte['suf']}",
            "valeur": valeur,
            "unite": "km",
            "date_mesure": f"{ANNEE}-07-01",
        }, follow_redirects=True)
        assert "enregistrée ✅" in r.get_data(as_text=True)

    with app.app_context():
        from app.models import TransitionMesure
        from app.services.transitions import tableau_de_bord

        data = tableau_de_bord(ANNEE)
        mesure = next(m for m in data["mesures"] if contexte["suf"] in m["libelle"])
        assert mesure["total"] == 200.5
        assert mesure["nb_saisies"] == 2
        # Thématique déduite de l'atelier étiqueté.
        assert mesure["thematique"].id == contexte["mobilite_id"]
        mesure_id = TransitionMesure.query.filter(
            TransitionMesure.libelle.contains(contexte["suf"])).first().id

    r = admin_client.post(f"/transitions/mesures/{mesure_id}/supprimer", follow_redirects=True)
    assert "supprimée" in r.get_data(as_text=True)


def test_referentiel_thematiques(admin_client, app):
    suf = uuid.uuid4().hex[:6]
    r = admin_client.post("/transitions/thematiques", data={
        "action": "ajouter", "libelle": f"Eau {suf}",
    }, follow_redirects=True)
    assert "ajoutée ✅" in r.get_data(as_text=True)

    with app.app_context():
        from app.models import TransitionThematique
        them = TransitionThematique.query.filter(
            TransitionThematique.libelle == f"Eau {suf}").first()
        assert them is not None and them.actif
        them_id = them.id

    r = admin_client.post("/transitions/thematiques", data={
        "action": "basculer", "thematique_id": them_id,
    }, follow_redirects=True)
    assert "désactivée" in r.get_data(as_text=True)


def test_etiquetage_depuis_la_page(admin_client, app, contexte):
    r = admin_client.get("/transitions/ateliers")
    assert r.status_code == 200
    assert f"Témoin {contexte['suf']}" in r.get_data(as_text=True)

    r = admin_client.post("/transitions/ateliers", data={
        "atelier_id": contexte["temoin_id"],
        "thematique_ids": [contexte["mobilite_id"]],
    }, follow_redirects=True)
    assert "compte maintenant dans le suivi transitions ✅" in r.get_data(as_text=True)

    with app.app_context():
        from app.extensions import db
        from app.models import AtelierActivite
        temoin = db.session.get(AtelierActivite, contexte["temoin_id"])
        assert [t.id for t in temoin.transition_thematiques] == [contexte["mobilite_id"]]


def test_dashboard_page_et_export(admin_client, contexte):
    r = admin_client.get(f"/transitions/?annee={ANNEE}")
    assert r.status_code == 200
    page = r.get_data(as_text=True)
    assert "Transitions ·" in page
    assert f"Vélo-école {contexte['suf']}" in page or "Par thématique" in page

    r = admin_client.get(f"/transitions/export.xlsx?annee={ANNEE}")
    assert r.status_code == 200
    assert "spreadsheet" in r.content_type

    from io import BytesIO
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(r.data))
    assert {"Synthèse", "Par thématique", "Défis", "Mesures", "Détail ateliers", "Mode d'emploi"} <= set(wb.sheetnames)
    synthese = {row[0].value: row[1].value for row in wb["Synthèse"].iter_rows(min_row=2)}
    assert synthese["Séances réalisées"] >= 1
    assert synthese["Participations"] >= 2


def test_refus_anonymes(client):
    assert client.get("/transitions/").status_code in (302, 401, 403)
    assert client.get("/transitions/export.xlsx").status_code in (302, 401, 403)
    assert client.post("/transitions/defis/creer", data={}).status_code in (302, 401, 403)


def test_permissions_rbac(app):
    with app.app_context():
        from app.models import Permission, Role

        assert Permission.query.filter_by(code="transitions:view").first() is not None
        assert Permission.query.filter_by(code="transitions:edit").first() is not None
        direction = Role.query.filter_by(code="direction").first()
        codes = {p.code for p in direction.permissions}
        assert {"transitions:view", "transitions:edit"} <= codes
