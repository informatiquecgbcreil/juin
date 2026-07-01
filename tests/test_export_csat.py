"""Export des participants et des séances au format d'import du portail CSAT
(Centres Sociaux Acteurs des Transitions), pour éviter la ressaisie manuelle."""
import datetime as dt
import uuid
from io import BytesIO

from flask import url_for


def test_export_participants_csat_format(app, admin_client):
    from app.extensions import db
    from app.models import Participant

    tag = uuid.uuid4().hex[:6]
    with app.app_context():
        db.session.add_all([
            Participant(nom=f"Dupont{tag}", prenom="Benjamin", ville="Creil",
                       genre="Homme", date_naissance=dt.date(1995, 8, 15)),
            Participant(nom=f"Dutoit{tag}", prenom="Elise", ville="Creil",
                       genre="Femme", date_naissance=dt.date(2012, 1, 27)),
            Participant(nom=f"Martin{tag}", prenom="Alex", ville="Creil",
                       genre="Autre", date_naissance=dt.date(2000, 1, 1)),
            Participant(nom=f"Sansdate{tag}", prenom="Nul", ville=None, genre=None, date_naissance=None),
        ])
        db.session.commit()
        with app.test_request_context():
            url = url_for("participants.export_csat_csv")

    r = admin_client.get(url)
    assert r.status_code == 200
    assert "text/csv" in r.headers.get("Content-Type", "")
    body = r.data.decode("utf-8-sig")
    lines = body.strip().split("\r\n") if "\r\n" in body else body.strip().split("\n")
    assert lines[0] == "last_name;first_name;location;gender;birthdate"

    rows = {l.split(";")[0]: l.split(";") for l in lines[1:]}
    assert rows[f"Dupont{tag}"] == [f"Dupont{tag}", "Benjamin", "Creil", "M", "15/08/1995"]
    assert rows[f"Dutoit{tag}"] == [f"Dutoit{tag}", "Elise", "Creil", "F", "27/01/2012"]
    assert rows[f"Martin{tag}"] == [f"Martin{tag}", "Alex", "Creil", "N", "01/01/2000"]
    assert rows[f"Sansdate{tag}"] == [f"Sansdate{tag}", "Nul", "", "N", ""]


def test_export_participants_csat_exclut_anonymises(app, admin_client):
    from app.extensions import db
    from app.models import Participant
    from app.services.purge_rgpd import NOM_ANONYME

    with app.app_context():
        db.session.add(Participant(nom=NOM_ANONYME, prenom="P", ville="Creil", genre="Homme"))
        db.session.commit()
        with app.test_request_context():
            url = url_for("participants.export_csat_csv")

    body = admin_client.get(url).data.decode("utf-8-sig")
    assert NOM_ANONYME not in body


def test_export_sessions_csat_format(app, admin_client):
    from app.extensions import db
    from app.models import AtelierActivite, SessionActivite
    from openpyxl import load_workbook

    with app.app_context():
        at = AtelierActivite(nom=f"At{uuid.uuid4().hex[:6]}", secteur="Numérique", type_atelier="COLLECTIF")
        db.session.add(at)
        db.session.flush()
        db.session.add_all([
            SessionActivite(atelier_id=at.id, secteur="Numérique", session_type="COLLECTIF",
                            date_session=dt.date(2026, 4, 30), heure_debut="14:00", heure_fin="15:30"),
            SessionActivite(atelier_id=at.id, secteur="Numérique", session_type="COLLECTIF",
                            date_session=dt.date(2026, 3, 31), heure_debut="14:00", heure_fin="15:30"),
            # séance en corbeille : ne doit pas apparaître
            SessionActivite(atelier_id=at.id, secteur="Numérique", session_type="COLLECTIF",
                            date_session=dt.date(2026, 1, 1), heure_debut="10:00", heure_fin="11:00",
                            is_deleted=True),
        ])
        db.session.commit()
        atid = at.id
        with app.test_request_context():
            url = url_for("activite.export_csat_sessions", atelier_id=atid)

    r = admin_client.get(url)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("Content-Type", "")

    wb = load_workbook(BytesIO(r.data))
    ws = wb["Sessions"]
    header = [c.value for c in ws[1]]
    assert header == ["session_date", "session_debut", "session_fin"]

    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == 2   # la séance en corbeille est exclue
    # trié par date croissante
    assert data_rows[0][0] == dt.datetime(2026, 3, 31, 0, 0)
    assert data_rows[0][1] == "14:00"
    assert data_rows[0][2] == dt.time(15, 30)


def test_export_sessions_csat_ne_reexporte_pas_deux_fois(app, admin_client):
    """Une séance déjà exportée ne doit plus jamais réapparaître par défaut."""
    from app.extensions import db
    from app.models import AtelierActivite, SessionActivite

    with app.app_context():
        at = AtelierActivite(nom=f"AtOnce{uuid.uuid4().hex[:6]}", secteur="Numérique", type_atelier="COLLECTIF")
        db.session.add(at)
        db.session.flush()
        s = SessionActivite(atelier_id=at.id, secteur="Numérique", session_type="COLLECTIF",
                            date_session=dt.date(2026, 5, 1), heure_debut="09:00", heure_fin="10:00")
        db.session.add(s)
        db.session.commit()
        atid, sid = at.id, s.id
        with app.test_request_context():
            url = url_for("activite.export_csat_sessions", atelier_id=atid)

    # Premier export : la séance est incluse.
    r1 = admin_client.get(url)
    assert r1.status_code == 200
    assert "spreadsheetml" in r1.headers.get("Content-Type", "")

    with app.app_context():
        assert db.session.get(SessionActivite, sid).exported_csat_at is not None

    # Deuxième export : rien de nouveau -> redirection avec message, pas de fichier.
    r2 = admin_client.get(url, follow_redirects=True)
    assert r2.status_code == 200
    assert "Aucune nouvelle séance" in r2.get_data(as_text=True)


def test_export_sessions_csat_tout_force_reexport(app, admin_client):
    """Le paramètre tout=1 permet de forcer un export complet en cas d'échec côté CSAT."""
    from app.extensions import db
    from app.models import AtelierActivite, SessionActivite
    from openpyxl import load_workbook

    with app.app_context():
        at = AtelierActivite(nom=f"AtForce{uuid.uuid4().hex[:6]}", secteur="Numérique", type_atelier="COLLECTIF")
        db.session.add(at)
        db.session.flush()
        s = SessionActivite(atelier_id=at.id, secteur="Numérique", session_type="COLLECTIF",
                            date_session=dt.date(2026, 6, 1), heure_debut="09:00", heure_fin="10:00")
        db.session.add(s)
        db.session.commit()
        atid = at.id
        with app.test_request_context():
            url = url_for("activite.export_csat_sessions", atelier_id=atid)
            url_tout = url_for("activite.export_csat_sessions", atelier_id=atid, tout=1)

    admin_client.get(url)  # marque la séance comme exportée
    r = admin_client.get(url_tout)
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.data))
    assert len(list(wb["Sessions"].iter_rows(min_row=2, values_only=True))) == 1


def test_nouvelle_seance_apparait_apres_export_precedent(app, admin_client):
    """Une séance créée après un premier export doit apparaître au prochain export."""
    from app.extensions import db
    from app.models import AtelierActivite, SessionActivite
    from openpyxl import load_workbook

    with app.app_context():
        at = AtelierActivite(nom=f"AtNew{uuid.uuid4().hex[:6]}", secteur="Numérique", type_atelier="COLLECTIF")
        db.session.add(at)
        db.session.flush()
        s1 = SessionActivite(atelier_id=at.id, secteur="Numérique", session_type="COLLECTIF",
                             date_session=dt.date(2026, 5, 10), heure_debut="09:00", heure_fin="10:00")
        db.session.add(s1)
        db.session.commit()
        atid = at.id
        with app.test_request_context():
            url = url_for("activite.export_csat_sessions", atelier_id=atid)

    admin_client.get(url)  # s1 exportée

    with app.app_context():
        s2 = SessionActivite(atelier_id=atid, secteur="Numérique", session_type="COLLECTIF",
                             date_session=dt.date(2026, 5, 17), heure_debut="09:00", heure_fin="10:00")
        db.session.add(s2)
        db.session.commit()

    r = admin_client.get(url)
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.data))
    rows = list(wb["Sessions"].iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    assert rows[0][0] == dt.datetime(2026, 5, 17, 0, 0)
