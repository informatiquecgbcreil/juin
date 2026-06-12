"""Tests du bilan SENACS (publics dédoublonnés, tranches d'âge, actions)."""
import uuid
from datetime import date
from io import BytesIO

import pytest

ANNEE = 2025


@pytest.fixture(scope="module")
def donnees_senacs(app):
    """Jeu de données 2025 : 3 participants, 2 ateliers de secteurs différents.

    - Amina (née 2000, Femme, Rouher) : présente aux DEUX ateliers
      -> doit compter pour 1 participant unique et 2 participations.
    - Bruno (né 1950, Homme, sans quartier) : présent à un atelier.
    - Chloé (née 2020) : présente à un atelier.
    - Une présence en 2024 qui ne doit PAS compter pour 2025.
    """
    suffixe = uuid.uuid4().hex[:6]
    with app.app_context():
        from app.extensions import db
        from app.models import (
            AtelierActivite,
            Participant,
            Partenaire,
            PresenceActivite,
            Quartier,
            SessionActivite,
        )

        rouher = Quartier.query.filter(Quartier.nom.ilike("%rouher%")).first()
        if rouher is None:
            rouher = Quartier(nom="Rouher", ville="Creil", is_qpv=True)
            db.session.add(rouher)
            db.session.flush()

        amina = Participant(nom=f"Senacs-A-{suffixe}", prenom="Amina", genre="Femme",
                            date_naissance=date(2000, 5, 1), quartier_id=rouher.id, ville="Creil")
        bruno = Participant(nom=f"Senacs-B-{suffixe}", prenom="Bruno", genre="Homme",
                            date_naissance=date(1950, 1, 15))
        chloe = Participant(nom=f"Senacs-C-{suffixe}", prenom="Chloé",
                            date_naissance=date(2020, 9, 9))
        db.session.add_all([amina, bruno, chloe])

        at_num = AtelierActivite(nom=f"Atelier num {suffixe}", secteur="Numérique")
        at_fam = AtelierActivite(nom=f"Atelier fam {suffixe}", secteur="Familles")
        db.session.add_all([at_num, at_fam])
        db.session.flush()

        s_num = SessionActivite(atelier_id=at_num.id, secteur="Numérique",
                                date_session=date(ANNEE, 3, 10), duree_minutes=120)
        s_fam = SessionActivite(atelier_id=at_fam.id, secteur="Familles",
                                date_session=date(ANNEE, 4, 20), duree_minutes=90)
        s_2024 = SessionActivite(atelier_id=at_num.id, secteur="Numérique",
                                 date_session=date(ANNEE - 1, 6, 1), duree_minutes=60)
        db.session.add_all([s_num, s_fam, s_2024])
        db.session.flush()

        db.session.add_all([
            PresenceActivite(session_id=s_num.id, participant_id=amina.id),
            PresenceActivite(session_id=s_fam.id, participant_id=amina.id),
            PresenceActivite(session_id=s_num.id, participant_id=bruno.id),
            PresenceActivite(session_id=s_fam.id, participant_id=chloe.id),
            PresenceActivite(session_id=s_2024.id, participant_id=bruno.id),
        ])
        db.session.add(Partenaire(nom=f"CAF Oise {suffixe}"))
        db.session.commit()

        return {
            "suffixe": suffixe,
            "amina_id": amina.id,
            "atelier_num": at_num.nom,
            "atelier_fam": at_fam.nom,
        }


def test_dedoublonnage_inter_secteurs(app, donnees_senacs):
    with app.app_context():
        from app.services.senacs import publics_annee

        publics = publics_annee(ANNEE)
        # Amina (2 présences) ne compte qu'une fois : 3 uniques, 4 participations
        assert publics["participants_uniques"] == 3
        assert publics["participations"] == 4


def test_tranches_age_senacs(app, donnees_senacs):
    with app.app_context():
        from app.services.senacs import publics_annee

        ages = publics_annee(ANNEE)["ages"]
        assert ages["18-25 ans"] == 1, "Amina née en 2000 a 25 ans au 31/12/2025"
        assert ages["75 ans et plus"] == 1, "Bruno né en 1950 a 75 ans au 31/12/2025"
        assert ages["0-6 ans"] == 1, "Chloé née en 2020 a 5 ans au 31/12/2025"


def test_genres_et_quartiers(app, donnees_senacs):
    with app.app_context():
        from app.services.senacs import publics_annee

        publics = publics_annee(ANNEE)
        assert publics["genres"]["Femme"] == 1
        assert publics["genres"]["Homme"] == 1
        assert publics["genres"]["Non renseigné"] == 1
        assert publics["quartiers"]["Rouher"] == 1


def test_tableau_actions(app, donnees_senacs):
    with app.app_context():
        from app.services.senacs import tableau_actions

        lignes = {l["atelier"]: l for l in tableau_actions(ANNEE)}
        num = lignes[donnees_senacs["atelier_num"]]
        assert num["seances"] == 1, "la session 2024 ne doit pas compter en 2025"
        assert num["participants_uniques"] == 2
        assert num["participations"] == 2
        assert num["heures_face_public"] == 2.0

        fam = lignes[donnees_senacs["atelier_fam"]]
        assert fam["participations"] == 2
        assert fam["heures_face_public"] == 1.5


def test_page_senacs(admin_client, donnees_senacs):
    r = admin_client.get(f"/bilans/senacs?annee={ANNEE}")
    assert r.status_code == 200
    page = r.get_data(as_text=True)
    assert "Bilan SENACS" in page
    assert donnees_senacs["atelier_num"] in page
    assert "dédoublonnés" in page


def test_export_senacs_xlsx(admin_client, donnees_senacs):
    r = admin_client.get(f"/bilans/senacs/export.xlsx?annee={ANNEE}")
    assert r.status_code == 200
    assert "spreadsheet" in r.content_type

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(r.data))
    assert {"Public global", "Actions séances", "Partenariats", "Mode d'emploi"} <= set(wb.sheetnames)
    contenu = "\n".join(str(c.value) for row in wb["Actions séances"].iter_rows() for c in row)
    assert donnees_senacs["atelier_fam"] in contenu


def test_senacs_refuse_anonymes(client):
    assert client.get("/bilans/senacs").status_code == 302
    assert client.get("/bilans/senacs/export.xlsx").status_code == 302
