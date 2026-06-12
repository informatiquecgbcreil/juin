"""Tests RGPD : droit à l'image (consentement dématérialisé) et export droit d'accès."""
from io import BytesIO

import pytest


@pytest.fixture()
def participant_rgpd(app):
    """Participant dédié, rattaché au secteur Familles."""
    import uuid

    suffixe = uuid.uuid4().hex[:8]
    with app.app_context():
        from app.extensions import db
        from app.models import OrientationAccesDroit, Participant

        p = Participant(
            nom=f"Rgpd-{suffixe}",
            prenom="Test",
            ville="Creil",
            created_secteur="Familles",
        )
        db.session.add(p)
        db.session.flush()
        db.session.add(
            OrientationAccesDroit(
                participant_id=p.id,
                domaine="logement",
                demande="Demande de test pour export",
            )
        )
        db.session.commit()
        return p.id


def _modifier(client, pid, **champs):
    data = {"nom": "X", "prenom": "Y"}
    data.update(champs)
    return client.post(f"/participants/{pid}/edit", data=data)


def test_consentement_enregistre_avec_trace(app, admin_client, participant_rgpd):
    r = _modifier(admin_client, participant_rgpd, droit_image_statut="accepte")
    assert r.status_code == 302

    with app.app_context():
        from datetime import date

        from app.extensions import db
        from app.models import Participant

        p = db.session.get(Participant, participant_rgpd)
        assert p.droit_image_statut == "accepte"
        assert p.droit_image_date == date.today()
        assert p.droit_image_recueilli_par, "l'agent qui a recueilli doit être tracé"
        assert "admin@example.org" in p.droit_image_recueilli_par


def test_trace_conservee_si_statut_inchange(app, admin_client, participant_rgpd):
    _modifier(admin_client, participant_rgpd, droit_image_statut="refuse")
    with app.app_context():
        from app.extensions import db
        from app.models import Participant

        avant = db.session.get(Participant, participant_rgpd).droit_image_recueilli_par

    # Nouvelle sauvegarde sans changer le statut : la trace ne bouge pas
    _modifier(admin_client, participant_rgpd, droit_image_statut="refuse")
    with app.app_context():
        from app.extensions import db
        from app.models import Participant

        p = db.session.get(Participant, participant_rgpd)
        assert p.droit_image_recueilli_par == avant


def test_statut_invalide_ignore(app, admin_client, participant_rgpd):
    r = _modifier(admin_client, participant_rgpd, droit_image_statut="nimporte-quoi")
    assert r.status_code == 302
    with app.app_context():
        from app.extensions import db
        from app.models import Participant

        assert db.session.get(Participant, participant_rgpd).droit_image_statut == "non_renseigne"


def test_export_droit_acces_complet(app, admin_client, participant_rgpd):
    r = admin_client.get(f"/participants/{participant_rgpd}/export-rgpd.xlsx")
    assert r.status_code == 200
    assert "spreadsheet" in r.content_type

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(r.data))
    feuilles = set(wb.sheetnames)
    attendues = {
        "Identité",
        "Droit à l'image",
        "Orientations",
        "Présences activités",
        "Notes pédagogiques",
        "Pièces jointes",
        "Insertion",
    }
    manquantes = attendues - feuilles
    assert not manquantes, f"Feuilles manquantes : {manquantes}"

    # L'identité contient bien le nom, les orientations contiennent la demande
    identite = "\n".join(str(c.value) for row in wb["Identité"].iter_rows() for c in row)
    assert "Rgpd-" in identite
    orientations = "\n".join(str(c.value) for row in wb["Orientations"].iter_rows() for c in row)
    assert "Demande de test pour export" in orientations


def test_export_refuse_aux_anonymes_et_autres_secteurs(app, client, participant_rgpd):
    # Anonyme : redirection vers la connexion
    r = client.get(f"/participants/{participant_rgpd}/export-rgpd.xlsx")
    assert r.status_code == 302

    # Responsable d'un AUTRE secteur (Numérique) : refus
    import uuid

    email = f"resp-num-{uuid.uuid4().hex[:6]}@example.org"
    with app.app_context():
        from app.extensions import db
        from app.models import Role, User

        u = User(email=email, nom="Resp Numérique", secteur_assigne="Numérique")
        u.set_password("motdepasse-tests")
        u.roles.append(Role.query.filter_by(code="responsable_secteur").first())
        db.session.add(u)
        db.session.commit()

    c = app.test_client()
    assert c.post("/", data={"email": email, "password": "motdepasse-tests"}).status_code == 302
    r = c.get(f"/participants/{participant_rgpd}/export-rgpd.xlsx")
    assert r.status_code == 403, "un responsable d'un autre secteur ne doit pas pouvoir exporter"


def test_export_autorise_au_responsable_du_secteur(app, participant_rgpd):
    import uuid

    email = f"resp-fam-{uuid.uuid4().hex[:6]}@example.org"
    with app.app_context():
        from app.extensions import db
        from app.models import Role, User

        u = User(email=email, nom="Resp Familles", secteur_assigne="Familles")
        u.set_password("motdepasse-tests")
        u.roles.append(Role.query.filter_by(code="responsable_secteur").first())
        db.session.add(u)
        db.session.commit()

    c = app.test_client()
    assert c.post("/", data={"email": email, "password": "motdepasse-tests"}).status_code == 302
    r = c.get(f"/participants/{participant_rgpd}/export-rgpd.xlsx")
    assert r.status_code == 200, "le responsable du secteur de la personne doit pouvoir exporter"


def test_export_journalise(app, admin_client, participant_rgpd):
    import os

    admin_client.get(f"/participants/{participant_rgpd}/export-rgpd.xlsx")
    chemin = os.path.join(os.environ["ERP_LOG_DIR"], "erreurs.log")
    with open(chemin, encoding="utf-8") as fh:
        contenu = fh.read()
    assert f"#{participant_rgpd}" in contenu and "Export RGPD" in contenu


def test_anonymisation_efface_le_consentement(app, admin_client, participant_rgpd):
    _modifier(admin_client, participant_rgpd, droit_image_statut="accepte")
    r = admin_client.post(f"/participants/{participant_rgpd}/anonymize")
    assert r.status_code == 302

    with app.app_context():
        from app.extensions import db
        from app.models import Participant

        p = db.session.get(Participant, participant_rgpd)
        assert p.droit_image_statut == "non_renseigne"
        assert p.droit_image_date is None
        assert p.droit_image_recueilli_par is None
