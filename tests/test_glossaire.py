"""Glossaire métier (« dico du social ») + uniformisation du vocabulaire.

- la page rend tous les termes et la recherche est présente ;
- garde-fous de qualité : définitions non vides, sigles développés ;
- l'uniformisation « séance » (vs « session ») ne régresse pas sur les
  écrans principaux.
"""
import pytest


def test_glossaire_page_complete(admin_client):
    r = admin_client.get("/guides/glossaire")
    body = r.get_data(as_text=True)
    assert r.status_code == 200
    # Les termes clés demandés (bilan, subvention, présence…) sont là
    for terme in ["Bilan", "Subvention", "Présence", "Séance", "Atelier",
                  "SENACS", "QPV", "ETP", "Bénévole", "Passeport",
                  "Éducation populaire", "Développement du pouvoir", "DPA",
                  "Aller-vers", "CLAS", "REAAP", "Animation globale",
                  "Projet social",
                  # vocabulaire de l'évaluation (débutant absolu)
                  "Indicateur quantitatif", "Indicateur qualitatif",
                  "processus", "Indicateur de résultat", "Effet", "Impact",
                  "PAG", "SMART", "Participant unique", "File active",
                  # institutions / vie associative
                  "EVS", "ALSH", "LAEP", "CTG", "VACAF", "CCAS",
                  "Mission locale", "Bureau"]:
        assert terme in body, f"terme manquant : {terme}"
    # La recherche instantanée est en place
    assert "glossaire-filtre" in body
    # Le lien croisé vers les guides existe
    assert "/guides" in body


def test_glossaire_qualite_des_definitions(app):
    """Chaque terme a une définition substantielle ; les sigles sont développés."""
    from app.aide.glossaire import glossaire_termes_plats

    termes = glossaire_termes_plats()
    assert len(termes) >= 40, "le glossaire doit rester exhaustif"
    noms = [t["terme"] for t in termes]
    assert len(noms) == len(set(noms)), "pas de doublon de terme"
    for t in termes:
        assert len(t["definition"]) >= 40, f"définition trop courte : {t['terme']}"
    # Les sigles employés seuls doivent être développés dans leur définition
    corpus = {t["terme"]: t["definition"] for t in termes}
    assert "Caisse d'allocations familiales" in corpus["CAF (Caisse d'allocations familiales)"] or "CAF (" in str(corpus)
    assert any("Équivalent temps plein" in k or "Équivalent temps plein" in v for k, v in corpus.items())


def test_glossaire_accessible_a_tous_les_roles(app):
    """dashboard:view suffit : même un rôle minimal lit le glossaire."""
    from app.extensions import db
    from app.models import Role, User

    with app.app_context():
        u = User.query.filter_by(email="glo-tech@example.org").first()
        if u is None:
            u = User(email="glo-tech@example.org", nom="glo-tech")
            u.set_password("pw-test-123")
            u.roles.append(Role.query.filter_by(code="admin_tech").first())
            db.session.add(u)
            db.session.commit()
    c = app.test_client()
    c.post("/", data={"email": "glo-tech@example.org", "password": "pw-test-123"})
    assert c.get("/guides/glossaire").status_code == 200


def test_vocabulaire_seance_uniforme(admin_client):
    """Les écrans principaux disent « séance », plus « session » (hors code)."""
    for url in ("/dashboard", "/stats-impact/dashboard"):
        r = admin_client.get(url, follow_redirects=True)
        assert r.status_code == 200, url
        body = r.get_data(as_text=True)
        # Aucun libellé visible « >Session(s)< » ne doit rester
        assert ">Sessions<" not in body, url
        assert ">Session<" not in body, url


# ---------- Glossaire éditable (CRUD, masquage, import XLSX) ----------

def _login_role(app, email, role_code, secteur=None):
    from app.extensions import db
    from app.models import Role, User
    with app.app_context():
        u = User.query.filter_by(email=email).first()
        if u is None:
            u = User(email=email, nom=email.split("@")[0], secteur_assigne=secteur)
            u.set_password("pw-test-123")
            u.roles.append(Role.query.filter_by(code=role_code).first())
            db.session.add(u)
            db.session.commit()
    c = app.test_client()
    c.post("/", data={"email": email, "password": "pw-test-123"})
    return c


def _client_direction(app, email="glo-dir@example.org"):
    from app.extensions import db
    from app.models import Role, User
    with app.app_context():
        u = User.query.filter_by(email=email).first()
        if u is None:
            u = User(email=email, nom="glo-dir")
            u.set_password("pw-test-123")
            u.roles.append(Role.query.filter_by(code="direction").first())
            db.session.add(u)
            db.session.commit()
    c = app.test_client()
    c.post("/", data={"email": email, "password": "pw-test-123"})
    return c


def test_ajout_modification_suppression_dun_mot_perso(app):
    c = _client_direction(app)
    # Ajout
    c.post("/guides/glossaire/enregistrer", data={
        "terme": "Conseil citoyen", "definition": "Instance d'habitants tirée au sort dans les quartiers prioritaires pour participer aux décisions du contrat de ville.",
        "categorie": "Les institutions et les bilans"})
    body = c.get("/guides/glossaire").get_data(as_text=True)
    assert "Conseil citoyen" in body and "tirée au sort" in body
    # Modification (renommage compris)
    c.post("/guides/glossaire/enregistrer", data={
        "terme_original": "Conseil citoyen", "terme": "Conseil citoyen (quartiers)",
        "definition": "Instance d'habitants tirée au sort pour peser sur le contrat de ville.",
        "categorie": "Les institutions et les bilans"})
    body = c.get("/guides/glossaire").get_data(as_text=True)
    assert "Conseil citoyen (quartiers)" in body
    # Suppression (mot perso → disparition complète)
    c.post("/guides/glossaire/supprimer", data={"terme": "Conseil citoyen (quartiers)"})
    c.get("/guides/glossaire")  # consomme le message flash (qui cite le mot)
    body = c.get("/guides/glossaire").get_data(as_text=True)
    assert "Conseil citoyen (quartiers)" not in body


def test_modifier_puis_retablir_un_terme_de_base(app):
    c = _client_direction(app)
    c.post("/guides/glossaire/enregistrer", data={
        "terme_original": "Trésorerie", "terme": "Trésorerie",
        "definition": "Notre définition locale de la trésorerie, adaptée à la maison et suffisamment longue pour le test."})
    body = c.get("/guides/glossaire").get_data(as_text=True)
    assert "définition locale de la trésorerie" in body
    assert "modifié" in body  # badge
    # Rétablir l'origine
    c.post("/guides/glossaire/retablir", data={"terme": "Trésorerie"})
    body = c.get("/guides/glossaire").get_data(as_text=True)
    assert "définition locale de la trésorerie" not in body
    assert "instant donné" in body  # la définition d'origine est revenue


def test_masquer_puis_retablir_un_terme_de_base(app):
    c = _client_direction(app)
    c.post("/guides/glossaire/supprimer", data={"terme": "VACAF"})
    body = c.get("/guides/glossaire").get_data(as_text=True)
    assert "aide aux vacances de la CAF" not in body
    assert "Mots retirés" in body
    c.post("/guides/glossaire/retablir", data={"terme": "VACAF"})
    body = c.get("/guides/glossaire").get_data(as_text=True)
    assert "aide aux vacances de la CAF" in body


def test_import_et_export_xlsx(app):
    import io
    from openpyxl import Workbook, load_workbook

    c = _client_direction(app)
    wb = Workbook()
    ws = wb.active
    ws.append(["Terme", "Définition", "Catégorie", "Dans l'application"])
    ws.append(["Tiers-lieu", "Espace hybride ouvert à tous où l'on travaille, bricole et se rencontre — souvent partenaire du centre.", "Les mots de la structure", ""])
    ws.append(["Repair café", "Atelier où des bénévoles aident les habitants à réparer leurs objets au lieu de les jeter.", "", ""])
    ws.append(["", "ligne sans terme : ignorée", "", ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    r = c.post("/guides/glossaire/import",
               data={"fichier": (buf, "glossaire.xlsx")},
               content_type="multipart/form-data", follow_redirects=True)
    body = r.get_data(as_text=True)
    assert "2 mot(s) ajouté(s)" in body
    assert "Tiers-lieu" in body and "Repair café" in body

    # Export : le XLSX contient la base + les ajouts
    r = c.get("/guides/glossaire/export.xlsx")
    assert r.status_code == 200
    wb2 = load_workbook(io.BytesIO(r.data), read_only=True)
    termes = [row[0] for row in wb2.active.iter_rows(values_only=True)]
    assert "Tiers-lieu" in termes
    assert "Subvention" in termes  # la base est incluse

    # Nettoyage
    c.post("/guides/glossaire/supprimer", data={"terme": "Tiers-lieu"})
    c.post("/guides/glossaire/supprimer", data={"terme": "Repair café"})


def test_edition_reservee_a_glossaire_edit(app):
    """responsable_secteur n'a pas glossaire:edit : lecture oui, édition non."""
    c = _login_role(app, "glo-resp@example.org", "responsable_secteur", secteur="Numérique")
    assert c.get("/guides/glossaire").status_code == 200
    r = c.post("/guides/glossaire/enregistrer", data={
        "terme": "Interdit", "definition": "Ne doit pas passer, faute de permission glossaire:edit."})
    assert r.status_code == 403
    body = c.get("/guides/glossaire").get_data(as_text=True)
    assert "Ajouter un mot" not in body
