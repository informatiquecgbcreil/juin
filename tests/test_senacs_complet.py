"""SENACS complet : bénévolat, emplois/ETP, finances, événementiel."""
import datetime as dt
import uuid
from io import BytesIO

from flask import url_for


def test_evenementiel_annee(app):
    from app.extensions import db
    from app.models import AtelierActivite, SessionActivite, PresenceActivite, Participant
    from app.services.senacs import evenementiel_annee

    with app.app_context():
        suf = uuid.uuid4().hex[:6]
        at = AtelierActivite(nom=f"Ev{suf}", secteur="Numérique", type_atelier="COLLECTIF")
        db.session.add(at)
        db.session.flush()
        fete = SessionActivite(atelier_id=at.id, secteur="Numérique", session_type="COLLECTIF",
                               date_session=dt.date(2047, 6, 21), est_evenement=True)
        normale = SessionActivite(atelier_id=at.id, secteur="Numérique", session_type="COLLECTIF",
                                  date_session=dt.date(2047, 6, 22))
        db.session.add_all([fete, normale])
        db.session.flush()
        for i in range(3):
            p = Participant(nom=f"E{i}{suf}", prenom="T")
            db.session.add(p)
            db.session.flush()
            db.session.add(PresenceActivite(session_id=fete.id, participant_id=p.id))
        db.session.commit()

        ev = evenementiel_annee(2047)
        assert ev["nb_evenements"] == 1          # la séance normale ne compte pas
        assert ev["participations"] == 3
        assert ev["participants_uniques"] == 3


def test_emplois_crud_et_totaux(app, admin_client):
    from app.models import SenacsEmploi
    from app.services.senacs import emplois_annee

    with app.app_context():
        with app.test_request_context():
            url = url_for("bilans.bilan_senacs_emploi_create")

    admin_client.post(url, data={"annee": "2048", "intitule": "Directrice",
                                 "type_contrat": "cdi", "etp": "1"}, follow_redirects=True)
    admin_client.post(url, data={"annee": "2048", "intitule": "Animateur jeunesse",
                                 "type_contrat": "emploi_aide", "etp": "0,5"}, follow_redirects=True)
    with app.app_context():
        emplois = emplois_annee(2048)
        assert emplois["nb_postes"] == 2
        assert emplois["total_etp"] == 1.5
        assert emplois["par_contrat"]["CDI"]["nb"] == 1
        eid = emplois["postes"][0].id
        with app.test_request_context():
            url_del = url_for("bilans.bilan_senacs_emploi_supprimer", emploi_id=eid)

    admin_client.post(url_del, follow_redirects=True)
    with app.app_context():
        assert emplois_annee(2048)["nb_postes"] == 1


def test_finances_annee(app):
    from app.extensions import db
    from app.models import Subvention, LigneBudget
    from app.services.senacs import finances_annee

    with app.app_context():
        suf = uuid.uuid4().hex[:6]
        s = Subvention(nom=f"Fin{suf}", secteur="Numérique", annee_exercice=2049,
                       financeur="CAF", montant_attribue=8000, montant_recu=6000)
        db.session.add(s)
        db.session.flush()
        db.session.add(LigneBudget(subvention_id=s.id, nature="charge", compte="60",
                                   libelle="Achats", montant_base=5000, montant_reel=4500))
        db.session.commit()

        fin = finances_annee(2049)
        assert fin["par_financeur"]["CAF"]["attribue"] == 8000.0
        assert fin["par_financeur"]["CAF"]["recu"] == 6000.0
        assert fin["total_charges_reelles"] == 4500.0
        assert "valorisation_benevolat" in fin


def test_export_contient_les_nouveaux_onglets(app, admin_client):
    from openpyxl import load_workbook

    with app.app_context():
        with app.test_request_context():
            url = url_for("bilans.bilan_senacs_export", annee=2049)
    r = admin_client.get(url)
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.data))
    for onglet in ("Bénévolat", "Emplois", "Finances", "Événementiel"):
        assert onglet in wb.sheetnames, f"onglet {onglet} manquant : {wb.sheetnames}"


def test_page_senacs_affiche_les_volets(app, admin_client):
    with app.app_context():
        with app.test_request_context():
            url = url_for("bilans.bilan_senacs", annee=2049)
    body = admin_client.get(url).get_data(as_text=True)
    assert "Bénévolat" in body
    assert "Emplois / ETP" in body
    assert "Finances" in body
    assert "Événementiel" in body
    assert "Valorisation" in body


def test_creation_seance_evenement(app, admin_client):
    from app.extensions import db
    from app.models import AtelierActivite, SessionActivite

    with app.app_context():
        at = AtelierActivite(nom=f"AtEv{uuid.uuid4().hex[:6]}", secteur="Numérique",
                             type_atelier="COLLECTIF")
        db.session.add(at)
        db.session.commit()
        aid = at.id
        with app.test_request_context():
            url = url_for("activite.session_new", atelier_id=aid)

    admin_client.post(url, data={"date_session": "2050-07-14", "heure_debut": "14:00",
                                 "heure_fin": "18:00", "est_evenement": "1"},
                      follow_redirects=True)
    with app.app_context():
        s = SessionActivite.query.filter_by(atelier_id=aid).first()
        assert s is not None
        assert s.est_evenement is True
        sid = s.id
        with app.test_request_context():
            url_toggle = url_for("activite.session_toggle_evenement", session_id=sid)

    # bascule a posteriori depuis l'émargement
    admin_client.post(url_toggle, follow_redirects=True)
    with app.app_context():
        assert db.session.get(SessionActivite, sid).est_evenement is False
