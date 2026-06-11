from __future__ import annotations

from datetime import date
from io import BytesIO
import re
import unicodedata

from flask import Blueprint, abort, flash, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.extensions import db
from app.rbac import require_perm, can
from app.secteurs import get_secteur_labels
from app.models import (
    AppelProjetBudget,
    AppelProjetBudgetLigne,
    BudgetPrevisionnel,
    BudgetPrevisionnelLigne,
    Projet,
    Subvention,
    LigneBudget,
    SubventionProjet,
)
from app.previsionnel.referentiel import (
    BudgetCompteReferentiel,
    BudgetCategorieReferentiel,
    BudgetModeleReferentiel,
    BudgetModeleLigneReferentiel,
)

bp = Blueprint("previsionnel", __name__, url_prefix="/previsionnel")


def _is_all_scope() -> bool:
    return bool(can("scope:all_secteurs"))


def _allowed_secteurs() -> list[str]:
    if _is_all_scope():
        return get_secteur_labels()
    secteur = (getattr(current_user, "secteur_assigne", None) or "").strip()
    return [secteur] if secteur else []


def _check_budget_scope(budget: BudgetPrevisionnel):
    if _is_all_scope():
        return
    if budget.secteur != (getattr(current_user, "secteur_assigne", None) or ""):
        abort(403)


def _can_edit() -> bool:
    return can("subventions:edit") or can("subventions:create") or can("depenses:create")


def _parse_float(value, default=0.0) -> float:
    raw = str(value or "").replace(" ", "").replace(",", ".")
    try:
        return float(raw) if raw else float(default)
    except Exception:
        return float(default)


def _parse_int(value, default=0) -> int:
    try:
        return int(value or default)
    except Exception:
        return int(default)


def _money(value) -> float:
    return round(float(value or 0), 2)


EXCEL_FORBIDDEN_SHEET_CHARS = set(r'[]:*?/\\')


def _excel_ascii_identifier(raw: str, fallback: str = "Tableau") -> str:
    """Identifiant Excel sûr pour les noms de tableaux : ASCII, lettres/chiffres/_ uniquement."""
    text = unicodedata.normalize("NFKD", str(raw or "")).encode("ascii", "ignore").decode("ascii")
    text = "".join(ch if (ch.isalnum() or ch == "_") else "_" for ch in text)
    text = re.sub(r"_+", "_", text).strip("_") or fallback
    if not (text[0].isalpha() or text[0] == "_"):
        text = f"T_{text}"
    if re.match(r"^[A-Za-z]{1,3}[0-9]+$", text):
        text = f"T_{text}"
    return text[:180]


def _excel_safe_sheet_title(raw: str, used: set[str], fallback: str = "Feuille") -> str:
    """Titre d'onglet Excel sûr : max 31 caractères, caractères interdits retirés, unique."""
    title = "".join("_" if ch in EXCEL_FORBIDDEN_SHEET_CHARS else ch for ch in str(raw or "").strip())
    title = re.sub(r"\s+", " ", title).strip(" '") or fallback
    title = title[:31] or fallback[:31]

    base = title
    idx = 2
    while title in used:
        suffix = f" {idx}"
        title = f"{base[:31-len(suffix)]}{suffix}"
        idx += 1
    used.add(title)
    return title


def _excel_safe_table_name(raw: str, wb: Workbook, fallback: str = "Tableau") -> str:
    """Nom de tableau Excel sûr et unique dans tout le classeur."""
    name = _excel_ascii_identifier(raw, fallback=fallback)
    existing = set()
    for sh in wb.worksheets:
        try:
            existing.update(sh.tables.keys())
        except Exception:
            pass

    base = name
    idx = 2
    while name in existing:
        suffix = f"_{idx}"
        name = f"{base[:180-len(suffix)]}{suffix}"
        idx += 1
    return name


def _add_safe_table(ws, name: str, start_row: int, end_row: int, end_col: int, style="TableStyleMedium9"):
    """Ajoute un tableau structuré Excel sans déclencher de réparation à l'ouverture."""
    if end_row <= start_row or end_col <= 0:
        return None

    seen = set()
    for col in range(1, end_col + 1):
        cell = ws.cell(row=start_row, column=col)
        value = str(cell.value or f"Colonne {col}").strip() or f"Colonne {col}"
        base = value
        suffix = 2
        while value in seen:
            value = f"{base} {suffix}"
            suffix += 1
        seen.add(value)
        cell.value = value

    ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=_excel_safe_table_name(name, ws.parent), ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name=style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    return table


def _project_label_from_line(line):
    return line.projet.nom if getattr(line, "projet", None) else "Transversal / non affecté"


def _subvention_label_from_appel(appel):
    if getattr(appel, "subvention", None):
        return appel.subvention.nom
    return appel.financeur or "Sans financeur précisé"


def _subvention_demande_default(appel: AppelProjetBudget) -> float:
    """Montant demandé proposé par défaut.

    Si le dossier contient des produits, on prend les produits retenus.
    Sinon, on prend les charges retenues : cas fréquent où le budget financeur
    ne contient que le coût demandé.
    """
    return _money(appel.total_produits if appel.total_produits > 0 else appel.total_charges)


def _safe_ligne_budget_compte(value: str | None) -> str:
    """LigneBudget.compte est historiquement court dans certaines bases.

    On évite les plantages Postgres type VARCHAR trop court en privilégiant
    le code comptable s'il est présent.
    """
    raw = str(value or "").strip()
    m = re.match(r"^\s*(\d{2,3})", raw)
    if m:
        return m.group(1)
    return raw[:20] or "60"


def _sync_appel_lines_to_subvention(appel: AppelProjetBudget, subvention: Subvention, montant_attribue: float, montant_recu: float):
    """Crée/met à jour les lignes de budget de la subvention réelle depuis le dossier financeur.

    Principe :
    - les lignes de charges du dossier deviennent les lignes imputables de la subvention ;
    - leur montant_base = montant demandé/retenu ;
    - leur montant_reel = enveloppe attribuée répartie proportionnellement ;
    - les lignes produits éventuelles gardent une trace demandé/reçu ;
    - aucune ligne existante avec dépenses n'est supprimée.
    """
    rows = list(appel.lignes or [])
    charge_rows = [r for r in rows if r.ligne_budget and r.ligne_budget.nature == "charge"]
    produit_rows = [r for r in rows if r.ligne_budget and r.ligne_budget.nature == "produit"]

    total_charges = _money(sum(float(r.montant_retenu or 0) for r in charge_rows))
    total_produits = _money(sum(float(r.montant_retenu or 0) for r in produit_rows))

    # Si le financeur a accordé 0 pour l'instant, on laisse 0 en réel.
    # Si un montant est accordé, il devient l'enveloppe imputable sur les charges.
    charge_reel_base = _money(montant_attribue or 0)
    produit_reel_base = _money(montant_recu or 0)

    existing = {}
    for ligne in list(subvention.lignes or []):
        key = (ligne.nature, _safe_ligne_budget_compte(ligne.compte), (ligne.libelle or "").strip().lower())
        existing[key] = ligne

    def upsert_line(nature: str, compte: str, libelle: str, base: float, reel: float):
        compte_safe = _safe_ligne_budget_compte(compte)
        libelle = (libelle or "").strip() or "Ligne budgétaire"
        key = (nature, compte_safe, libelle.lower())
        ligne = existing.get(key)
        if not ligne:
            ligne = LigneBudget(
                subvention_id=subvention.id,
                nature=nature,
                compte=compte_safe,
                libelle=libelle,
                montant_base=0.0,
                montant_reel=0.0,
            )
            db.session.add(ligne)
            existing[key] = ligne
        ligne.montant_base = _money(base)
        ligne.montant_reel = _money(reel)
        return ligne

    for row in charge_rows:
        src = row.ligne_budget
        retained = _money(row.montant_retenu or 0)
        if retained <= 0:
            continue
        if total_charges > 0 and charge_reel_base > 0:
            reel = _money(charge_reel_base * retained / total_charges)
        else:
            reel = 0.0
        upsert_line("charge", src.compte, src.libelle, retained, reel)

    for row in produit_rows:
        src = row.ligne_budget
        retained = _money(row.montant_retenu or 0)
        if retained <= 0:
            continue
        if total_produits > 0 and produit_reel_base > 0:
            reel = _money(produit_reel_base * retained / total_produits)
        else:
            reel = 0.0
        upsert_line("produit", src.compte or "74", src.libelle, retained, reel)

    # Si aucun produit n'était présent dans le dossier, on crée une ligne produit lisible
    # pour garder une trace synthétique demandé/reçu côté subvention.
    if not produit_rows:
        upsert_line(
            "produit",
            "74",
            subvention.nom,
            _subvention_demande_default(appel),
            _money(montant_recu or 0),
        )


def _ensure_subvention_project_link(subvention: Subvention, projet_id: int | None):
    if not projet_id:
        return
    exists = SubventionProjet.query.filter_by(projet_id=projet_id, subvention_id=subvention.id).first()
    if not exists:
        db.session.add(SubventionProjet(projet_id=projet_id, subvention_id=subvention.id))


def _projects_for_budget(budget: BudgetPrevisionnel):
    return Projet.query.filter_by(secteur=budget.secteur).order_by(Projet.nom.asc()).all()


def _subventions_for_budget(budget: BudgetPrevisionnel):
    return (
        Subvention.query.filter_by(secteur=budget.secteur, annee_exercice=budget.annee)
        .order_by(Subvention.nom.asc())
        .all()
    )


def _ref_scope_filter(query, model, secteur: str | None):
    if secteur:
        return query.filter((model.secteur.is_(None)) | (model.secteur == secteur))
    return query


def _budget_comptes_ref(secteur: str | None = None, nature: str | None = None, active_only: bool = True):
    q = BudgetCompteReferentiel.query
    q = _ref_scope_filter(q, BudgetCompteReferentiel, secteur)
    if nature in ("charge", "produit"):
        q = q.filter(BudgetCompteReferentiel.nature == nature)
    if active_only:
        q = q.filter(BudgetCompteReferentiel.actif.is_(True))
    return q.order_by(BudgetCompteReferentiel.nature.asc(), BudgetCompteReferentiel.ordre.asc(), BudgetCompteReferentiel.code.asc()).all()


def _budget_categories_ref(secteur: str | None = None, nature: str | None = None, active_only: bool = True):
    q = BudgetCategorieReferentiel.query.join(BudgetCompteReferentiel)
    q = _ref_scope_filter(q, BudgetCategorieReferentiel, secteur)
    if secteur:
        q = q.filter((BudgetCompteReferentiel.secteur.is_(None)) | (BudgetCompteReferentiel.secteur == secteur))
    if nature in ("charge", "produit"):
        q = q.filter(BudgetCategorieReferentiel.nature == nature)
    if active_only:
        q = q.filter(BudgetCategorieReferentiel.actif.is_(True), BudgetCompteReferentiel.actif.is_(True))
    return q.order_by(
        BudgetCompteReferentiel.nature.asc(),
        BudgetCompteReferentiel.ordre.asc(),
        BudgetCompteReferentiel.code.asc(),
        BudgetCategorieReferentiel.ordre.asc(),
        BudgetCategorieReferentiel.libelle.asc(),
    ).all()


def _category_allowed_for_budget(cat: BudgetCategorieReferentiel, budget: BudgetPrevisionnel) -> bool:
    if not cat or not cat.compte:
        return False
    if cat.secteur and cat.secteur != budget.secteur:
        return False
    if cat.compte.secteur and cat.compte.secteur != budget.secteur:
        return False
    return bool(cat.actif and cat.compte.actif)


def _category_compte_label(cat: BudgetCategorieReferentiel) -> str:
    if not cat or not cat.compte:
        return ""
    return f"{cat.compte.code} - {cat.compte.libelle}".strip(" -")


def _budget_modeles_ref(secteur: str | None = None, active_only: bool = True):
    q = BudgetModeleReferentiel.query
    if secteur:
        q = q.filter((BudgetModeleReferentiel.secteur.is_(None)) | (BudgetModeleReferentiel.secteur == secteur))
    if active_only:
        q = q.filter(BudgetModeleReferentiel.actif.is_(True))
    return q.order_by(BudgetModeleReferentiel.ordre.asc(), BudgetModeleReferentiel.nom.asc()).all()


def _modeles_payload(secteur: str | None = None):
    payload = []
    for modele in _budget_modeles_ref(secteur):
        categorie_ids = []
        for row in getattr(modele, "lignes", []) or []:
            cat = row.categorie
            if not cat:
                continue
            if secteur and cat.secteur and cat.secteur != secteur:
                continue
            if secteur and cat.compte and cat.compte.secteur and cat.compte.secteur != secteur:
                continue
            if cat.actif and cat.compte and cat.compte.actif:
                categorie_ids.append(cat.id)
        payload.append({"id": modele.id, "nom": modele.nom, "categorie_ids": categorie_ids})
    return payload


def _line_category_ids(lines, categories_ref) -> dict[int, int]:
    """Rapproche une ligne prévisionnelle de sa catégorie source par nature/compte/libellé."""
    by_key = {}
    for cat in categories_ref:
        if not cat.compte:
            continue
        labels = {
            f"{cat.compte.code} - {cat.compte.libelle}".strip().lower(),
            str(cat.compte.code or "").strip().lower(),
            str(cat.compte.label or "").strip().lower(),
        }
        for compte_label in labels:
            by_key[(cat.nature, compte_label, str(cat.libelle or "").strip().lower())] = cat.id

    result = {}
    for line in lines:
        compte = str(line.compte or "").strip().lower()
        libelle = str(line.libelle or "").strip().lower()
        found = by_key.get((line.nature, compte, libelle))
        if not found and " - " in compte:
            code_only = compte.split(" - ", 1)[0].strip()
            found = by_key.get((line.nature, code_only, libelle))
        if found:
            result[line.id] = found
    return result


def _add_budget_line_from_category(budget: BudgetPrevisionnel, cat: BudgetCategorieReferentiel, ordre: int = 0):
    return BudgetPrevisionnelLigne(
        budget_id=budget.id,
        nature=cat.nature,
        compte=_category_compte_label(cat) or ("74" if cat.nature == "produit" else "60"),
        libelle=cat.libelle,
        montant=float(cat.montant_defaut or 0.0),
        commentaire=cat.description or None,
        ordre=ordre,
    )


def _normalize_budget_text(value: str | None) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _account_code(value: str | None) -> str:
    raw = str(value or "").strip()
    m = re.match(r"^\s*(\d{2,3})", raw)
    return m.group(1) if m else raw.lower()


def _subvention_project_ids(sub: Subvention) -> list[int | None]:
    ids = []
    for link in getattr(sub, "projets", []) or []:
        pid = getattr(link, "projet_id", None)
        if pid:
            ids.append(pid)
    return ids or [None]


def _subvention_project_labels(sub: Subvention) -> str:
    labels = []
    for link in getattr(sub, "projets", []) or []:
        proj = getattr(link, "projet", None)
        if proj and getattr(proj, "nom", None):
            labels.append(proj.nom)
    return ", ".join(labels) if labels else "Non rattaché"


def _budget_line_key(nature: str | None, compte: str | None, libelle: str | None):
    return ((nature or "charge").strip(), _account_code(compte), _normalize_budget_text(libelle))


def _actual_budget_lines_for_previsionnel(budget: BudgetPrevisionnel):
    subs = (
        Subvention.query
        .filter_by(secteur=budget.secteur, annee_exercice=budget.annee)
        .order_by(Subvention.nom.asc())
        .all()
    )
    lines = (
        LigneBudget.query
        .join(Subvention, LigneBudget.subvention_id == Subvention.id)
        .filter(Subvention.secteur == budget.secteur, Subvention.annee_exercice == budget.annee)
        .all()
    )
    return subs, lines


def _line_exec_amount(line: LigneBudget) -> float:
    if getattr(line, "nature", "charge") == "charge":
        return float(getattr(line, "engage", 0.0) or 0.0)
    return float(getattr(line, "montant_reel", 0.0) or 0.0)


def _realise_status(nature: str, planned: float, executed: float) -> str:
    if planned == 0 and executed > 0:
        return "Non prévu"
    if planned > 0 and executed == 0:
        return "Non réalisé"
    if planned <= 0:
        return "OK"

    taux = (executed / planned) * 100
    if nature == "charge":
        if taux > 110:
            return "Dépassement"
        if taux < 70:
            return "Sous-engagé"
        return "OK"
    else:
        if taux < 80:
            return "Produit à sécuriser"
        if taux > 110:
            return "Produit supérieur"
        return "OK"


def _compare_previsionnel_realise(budget: BudgetPrevisionnel, prev_lines):
    """Comparaison non destructive entre budget prévisionnel et données financières réelles.

    Important : aucune colonne de liaison stricte n'existe encore entre BudgetPrevisionnelLigne
    et LigneBudget. On rapproche donc par nature + code compte + libellé, puis par projet si possible.
    """
    subs, actual_lines = _actual_budget_lines_for_previsionnel(budget)

    by_key = {}
    by_project_key = {}
    actual_index = {}

    for actual in actual_lines:
        sub = actual.source_sub
        key = _budget_line_key(actual.nature, actual.compte, actual.libelle)
        project_ids = _subvention_project_ids(sub)
        actual_index[actual.id] = actual

        by_key.setdefault(key, []).append(actual)
        for pid in project_ids:
            by_project_key.setdefault((key, pid), []).append(actual)

    matched_ids = set()
    rows = []

    for line in prev_lines:
        key = _budget_line_key(line.nature, line.compte, line.libelle)
        matched = []
        if line.projet_id:
            matched = by_project_key.get((key, line.projet_id), [])
        if not matched:
            matched = by_key.get(key, [])

        for item in matched:
            matched_ids.add(item.id)

        planned = float(line.montant or 0.0)
        base = sum(float(item.montant_base or 0.0) for item in matched)
        reel = sum(float(item.montant_reel or 0.0) for item in matched)
        engage = sum(float(getattr(item, "engage", 0.0) or 0.0) for item in matched if getattr(item, "nature", "charge") == "charge")
        executed = engage if line.nature == "charge" else reel
        ecart = executed - planned
        taux = round((executed / planned) * 100, 1) if planned else None
        sub_names = sorted({item.source_sub.nom for item in matched if getattr(item, "source_sub", None)})

        rows.append({
            "source": "Prévisionnel",
            "nature": line.nature,
            "nature_label": "Produit" if line.nature == "produit" else "Charge",
            "compte": line.compte,
            "libelle": line.libelle,
            "projet": line.projet.nom if line.projet else "Transversal / non affecté",
            "prevu": _money(planned),
            "base": _money(base),
            "reel": _money(reel),
            "engage": _money(engage),
            "execute": _money(executed),
            "ecart": _money(ecart),
            "taux": taux,
            "statut": _realise_status(line.nature, planned, executed),
            "subventions": ", ".join(sub_names),
        })

    # Lignes réelles non retrouvées dans le prévisionnel : très utile pour repérer ce qui est parti hors budget.
    for actual in actual_lines:
        if actual.id in matched_ids:
            continue
        sub = actual.source_sub
        executed = _line_exec_amount(actual)
        rows.append({
            "source": "Réel seul",
            "nature": actual.nature,
            "nature_label": "Produit" if actual.nature == "produit" else "Charge",
            "compte": actual.compte,
            "libelle": actual.libelle,
            "projet": _subvention_project_labels(sub),
            "prevu": 0.0,
            "base": _money(float(actual.montant_base or 0.0)),
            "reel": _money(float(actual.montant_reel or 0.0)),
            "engage": _money(float(getattr(actual, "engage", 0.0) or 0.0)),
            "execute": _money(executed),
            "ecart": _money(executed),
            "taux": None,
            "statut": "Non prévu",
            "subventions": sub.nom if sub else "",
        })

    rows = sorted(rows, key=lambda r: (r["nature"], r["projet"].lower(), _account_code(r["compte"]), r["libelle"].lower()))

    subvention_rows = []
    for sub in subs:
        charge_lines = [l for l in sub.lignes if getattr(l, "nature", "charge") == "charge"]
        produit_lines = [l for l in sub.lignes if getattr(l, "nature", "charge") == "produit"]
        subvention_rows.append({
            "nom": sub.nom,
            "projets": _subvention_project_labels(sub),
            "demande": _money(sub.montant_demande or 0.0),
            "attribue": _money(sub.montant_attribue or 0.0),
            "recu": _money(sub.montant_recu or 0.0),
            "charges_base": _money(sum(float(l.montant_base or 0.0) for l in charge_lines)),
            "charges_reel": _money(sum(float(l.montant_reel or 0.0) for l in charge_lines)),
            "charges_engagees": _money(sum(float(getattr(l, "engage", 0.0) or 0.0) for l in charge_lines)),
            "produits_base": _money(sum(float(l.montant_base or 0.0) for l in produit_lines)),
            "produits_reel": _money(sum(float(l.montant_reel or 0.0) for l in produit_lines)),
        })

    charges_prevues = _money(sum(float(l.montant or 0.0) for l in prev_lines if l.nature == "charge"))
    produits_prevus = _money(sum(float(l.montant or 0.0) for l in prev_lines if l.nature == "produit"))

    charges_base = _money(sum(float(l.montant_base or 0.0) for l in actual_lines if getattr(l, "nature", "charge") == "charge"))
    charges_reel = _money(sum(float(l.montant_reel or 0.0) for l in actual_lines if getattr(l, "nature", "charge") == "charge"))
    charges_engagees = _money(sum(float(getattr(l, "engage", 0.0) or 0.0) for l in actual_lines if getattr(l, "nature", "charge") == "charge"))

    produits_base = _money(sum(float(l.montant_base or 0.0) for l in actual_lines if getattr(l, "nature", "charge") == "produit"))
    produits_reel = _money(sum(float(l.montant_reel or 0.0) for l in actual_lines if getattr(l, "nature", "charge") == "produit"))

    subventions_demande = _money(sum(float(s.montant_demande or 0.0) for s in subs))
    subventions_attribue = _money(sum(float(s.montant_attribue or 0.0) for s in subs))
    subventions_recu = _money(sum(float(s.montant_recu or 0.0) for s in subs))

    # Valeur de lecture : si des produits réels par ligne existent, on les prend ; sinon on utilise le reçu subvention.
    produits_execution = produits_reel if produits_reel else subventions_recu
    charges_execution = charges_engagees if charges_engagees else charges_reel

    kpis = {
        "charges_prevues": charges_prevues,
        "produits_prevus": produits_prevus,
        "solde_prevu": _money(produits_prevus - charges_prevues),
        "charges_base": charges_base,
        "charges_reel": charges_reel,
        "charges_engagees": charges_engagees,
        "charges_execution": _money(charges_execution),
        "produits_base": produits_base,
        "produits_reel": produits_reel,
        "produits_execution": _money(produits_execution),
        "subventions_demande": subventions_demande,
        "subventions_attribue": subventions_attribue,
        "subventions_recu": subventions_recu,
        "solde_execution": _money(produits_execution - charges_execution),
        "ecart_charges": _money(charges_execution - charges_prevues),
        "ecart_produits": _money(produits_execution - produits_prevus),
        "taux_charges": round((charges_execution / charges_prevues) * 100, 1) if charges_prevues else None,
        "taux_produits": round((produits_execution / produits_prevus) * 100, 1) if produits_prevus else None,
        "nb_subventions": len(subs),
        "nb_lignes_non_prevues": len([r for r in rows if r["source"] == "Réel seul"]),
    }

    return {"kpis": kpis, "rows": rows, "subventions": subvention_rows}


def _build_realise_workbook(budget: BudgetPrevisionnel, realise_data: dict):
    wb = Workbook()
    used_titles = set()

    ws = wb.active
    ws.title = _excel_safe_sheet_title("Synthèse réalisé", used_titles)
    k = realise_data["kpis"]
    ws.append([f"Prévisionnel vs réalisé — {budget.nom}"])
    ws["A1"].font = Font(bold=True, size=16)
    ws.append(["Année", budget.annee, "Secteur", budget.secteur])
    ws.append([])
    ws.append(["Indicateur", "Valeur"])
    header = ws.max_row
    for label, key in [
        ("Charges prévues", "charges_prevues"),
        ("Charges engagées / exécutées", "charges_execution"),
        ("Écart charges", "ecart_charges"),
        ("Taux charges (%)", "taux_charges"),
        ("Produits prévus", "produits_prevus"),
        ("Produits sécurisés / reçus", "produits_execution"),
        ("Écart produits", "ecart_produits"),
        ("Taux produits (%)", "taux_produits"),
        ("Solde prévu", "solde_prevu"),
        ("Solde exécuté", "solde_execution"),
        ("Subventions suivies", "nb_subventions"),
        ("Lignes réelles non prévues", "nb_lignes_non_prevues"),
    ]:
        ws.append([label, k.get(key)])
    _add_safe_table(ws, "SynthesePrevisionnelRealise", header, ws.max_row, 2, style="TableStyleMedium9")
    ws.freeze_panes = "A5"

    ws_rows = wb.create_sheet(_excel_safe_sheet_title("Écarts par ligne", used_titles))
    ws_rows.append(["Source", "Nature", "Compte", "Libellé", "Projet", "Prévu", "Base subvention", "Réel déclaré", "Engagé dépenses", "Exécuté", "Écart", "Taux %", "Statut", "Subventions"])
    for r in realise_data["rows"]:
        ws_rows.append([
            r["source"], r["nature_label"], r["compte"], r["libelle"], r["projet"],
            r["prevu"], r["base"], r["reel"], r["engage"], r["execute"], r["ecart"], r["taux"], r["statut"], r["subventions"],
        ])
    _add_safe_table(ws_rows, "EcartsPrevisionnelRealise", 1, ws_rows.max_row, 14, style="TableStyleMedium2")
    ws_rows.freeze_panes = "A2"

    ws_sub = wb.create_sheet(_excel_safe_sheet_title("Subventions réelles", used_titles))
    ws_sub.append(["Subvention", "Projets", "Demandé", "Attribué", "Reçu", "Charges base", "Charges réel", "Charges engagées", "Produits base", "Produits réel"])
    for s in realise_data["subventions"]:
        ws_sub.append([s["nom"], s["projets"], s["demande"], s["attribue"], s["recu"], s["charges_base"], s["charges_reel"], s["charges_engagees"], s["produits_base"], s["produits_reel"]])
    _add_safe_table(ws_sub, "SubventionsReelles", 1, ws_sub.max_row, 10, style="TableStyleMedium4")
    ws_sub.freeze_panes = "A2"

    for sheet in wb.worksheets:
        for col in range(1, sheet.max_column + 1):
            sheet.cell(1, col).font = Font(bold=True)
        _style_ws(sheet)
    return wb


def _project_totals(lines):
    totals = {}
    for line in lines:
        key = line.projet_id or 0
        label = line.projet.nom if line.projet else "Transversal / non affecté"
        bucket = totals.setdefault(key, {"label": label, "charges": 0.0, "produits": 0.0})
        if line.nature == "produit":
            bucket["produits"] += float(line.montant or 0)
        else:
            bucket["charges"] += float(line.montant or 0)
    rows = []
    for data in totals.values():
        rows.append({
            "label": data["label"],
            "charges": _money(data["charges"]),
            "produits": _money(data["produits"]),
            "solde": _money(data["produits"] - data["charges"]),
        })
    return sorted(rows, key=lambda r: r["label"].lower())


def _appel_usage_for_budget(budget_id: int, exclude_appel_id: int | None = None) -> dict[int, dict]:
    """Montants déjà retenus dans les budgets d'appel, par ligne prévisionnelle."""
    q = (
        db.session.query(AppelProjetBudgetLigne, AppelProjetBudget)
        .join(AppelProjetBudget, AppelProjetBudgetLigne.appel_id == AppelProjetBudget.id)
        .filter(AppelProjetBudget.budget_id == budget_id)
    )
    if exclude_appel_id:
        q = q.filter(AppelProjetBudget.id != exclude_appel_id)

    usage: dict[int, dict] = {}
    for row, appel in q.all():
        line_id = int(row.budget_ligne_id)
        bucket = usage.setdefault(line_id, {"montant": 0.0, "count": 0, "appels": []})
        bucket["montant"] += float(row.montant_retenu or 0.0)
        bucket["count"] += 1
        bucket["appels"].append(appel.nom)

    for bucket in usage.values():
        bucket["montant"] = _money(bucket["montant"])
    return usage


def _line_remaining_amount(line: BudgetPrevisionnelLigne, usage: dict[int, dict]) -> float:
    used = float((usage.get(line.id) or {}).get("montant") or 0.0)
    return _money(float(line.montant or 0.0) - used)


def _validate_retained_against_remaining(line: BudgetPrevisionnelLigne, retained: float, usage: dict[int, dict]) -> bool:
    return retained <= (_line_remaining_amount(line, usage) + 0.01)


def _add_table(ws, name: str, start_row: int, end_row: int, end_col: int, style="TableStyleMedium9"):
    if end_row <= start_row or end_col <= 0:
        return

    seen = set()
    for col in range(1, end_col + 1):
        cell = ws.cell(row=start_row, column=col)
        value = str(cell.value or f"Colonne {col}").strip() or f"Colonne {col}"
        base = value
        suffix = 2
        while value in seen:
            value = f"{base} {suffix}"
            suffix += 1
        seen.add(value)
        cell.value = value

    ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    safe = "".join(ch if ch.isalnum() else "_" for ch in name)
    if not safe or not safe[0].isalpha():
        safe = f"T_{safe or 'Table'}"
    existing = set()
    for sh in ws.parent.worksheets:
        existing.update(sh.tables.keys())
    base = safe[:180]
    idx = 1
    while safe in existing:
        idx += 1
        safe = f"{base}_{idx}"
    table = Table(displayName=safe, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name=style, showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)


def _style_ws(ws):
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if isinstance(cell.value, float):
                cell.number_format = '#,##0.00 €'
    ws.sheet_view.showGridLines = False
    for col in range(1, ws.max_column + 1):
        width = 12
        for cell in ws[get_column_letter(col)]:
            width = max(width, min(38, len(str(cell.value or "")) + 2))
        ws.column_dimensions[get_column_letter(col)].width = width


def _build_budget_workbook(budget: BudgetPrevisionnel, lines):
    """Export global prévisionnel : synthèse + charges + produits + détail par projet + DATA."""
    wb = Workbook()
    used_titles = set()

    ws = wb.active
    ws.title = _excel_safe_sheet_title("Synthèse financeur", used_titles)
    ws.append([f"Budget prévisionnel — {budget.nom}"])
    ws["A1"].font = Font(bold=True, size=16)
    ws.append(["Année", budget.annee, "Secteur", budget.secteur, "Statut", budget.statut])
    ws.append([])

    ws.append(["Indicateur", "Montant"])
    summary_header = ws.max_row
    ws.append(["Charges prévisionnelles", budget.total_charges])
    ws.append(["Produits prévisionnels", budget.total_produits])
    ws.append(["Solde produits - charges", budget.solde])
    _add_safe_table(ws, "SyntheseBudgetPrevisionnel", summary_header, ws.max_row, 2, style="TableStyleMedium9")
    ws.append([])

    ws.append(["Projet", "Charges", "Produits", "Solde"])
    project_header = ws.max_row
    for row in _project_totals(lines):
        ws.append([row["label"], row["charges"], row["produits"], row["solde"]])
    _add_safe_table(ws, "SyntheseBudgetParProjet", project_header, ws.max_row, 4, style="TableStyleMedium4")
    ws.freeze_panes = "A5"

    ws_ch = wb.create_sheet(_excel_safe_sheet_title("Charges", used_titles))
    ws_ch.append(["Nature", "Compte", "Libellé", "Projet", "Montant", "Commentaire"])
    for line in [l for l in lines if l.nature == "charge"]:
        ws_ch.append(["Charge", line.compte, line.libelle, _project_label_from_line(line), float(line.montant or 0), line.commentaire or ""])
    _add_safe_table(ws_ch, "ChargesPrevisionnelles", 1, ws_ch.max_row, 6, style="TableStyleMedium2")
    ws_ch.freeze_panes = "A2"

    ws_pr = wb.create_sheet(_excel_safe_sheet_title("Produits", used_titles))
    ws_pr.append(["Nature", "Compte", "Libellé", "Projet", "Montant", "Subvention liée", "Commentaire"])
    for line in [l for l in lines if l.nature == "produit"]:
        ws_pr.append([
            "Produit",
            line.compte,
            line.libelle,
            _project_label_from_line(line),
            float(line.montant or 0),
            getattr(getattr(line, "subvention", None), "nom", "") if getattr(line, "subvention", None) else "",
            line.commentaire or "",
        ])
    _add_safe_table(ws_pr, "ProduitsPrevisionnels", 1, ws_pr.max_row, 7, style="TableStyleMedium4")
    ws_pr.freeze_panes = "A2"

    grouped = {}
    for line in lines:
        grouped.setdefault(_project_label_from_line(line), []).append(line)

    for project_label in sorted(grouped.keys(), key=lambda x: x.lower()):
        project_lines = grouped[project_label]
        ws_p = wb.create_sheet(_excel_safe_sheet_title(f"Projet {project_label}", used_titles))
        charges = sum(float(l.montant or 0) for l in project_lines if l.nature == "charge")
        produits = sum(float(l.montant or 0) for l in project_lines if l.nature == "produit")
        ws_p.append([f"Projet — {project_label}"])
        ws_p["A1"].font = Font(bold=True, size=14)
        ws_p.append(["Charges", charges, "Produits", produits, "Solde", produits - charges])
        ws_p.append([])
        ws_p.append(["Nature", "Compte", "Libellé", "Montant", "Commentaire"])
        header = ws_p.max_row
        for line in project_lines:
            ws_p.append(["Produit" if line.nature == "produit" else "Charge", line.compte, line.libelle, float(line.montant or 0), line.commentaire or ""])
        _add_safe_table(ws_p, f"Projet_{project_label}", header, ws_p.max_row, 5, style="TableStyleMedium2")
        ws_p.freeze_panes = f"A{header + 1}"

    ws_data = wb.create_sheet(_excel_safe_sheet_title("DATA Budget", used_titles))
    ws_data.append(["Budget ID", "Année", "Secteur", "Nature", "Compte", "Libellé", "Projet", "Montant", "Commentaire", "Subvention ID"])
    for line in lines:
        ws_data.append([
            budget.id,
            budget.annee,
            budget.secteur,
            line.nature,
            line.compte,
            line.libelle,
            _project_label_from_line(line),
            float(line.montant or 0),
            line.commentaire or "",
            getattr(line, "subvention_id", None) or "",
        ])
    _add_safe_table(ws_data, "DataBudgetPrevisionnel", 1, ws_data.max_row, 10, style="TableStyleMedium7")
    ws_data.freeze_panes = "A2"

    for sheet in wb.worksheets:
        for col in range(1, sheet.max_column + 1):
            sheet.cell(1, col).font = Font(bold=True)
        _style_ws(sheet)
    return wb


def _build_appel_workbook(appel: AppelProjetBudget):
    """Export d'appel à projet : page 1 financeur + projets + financeur + data source."""
    wb = Workbook()
    used_titles = set()

    rows = []
    for row in appel.lignes:
        line = row.ligne_budget
        rows.append({
            "nature": line.nature,
            "nature_label": "Produit" if line.nature == "produit" else "Charge",
            "compte": line.compte or "",
            "libelle": line.libelle or "",
            "projet": _project_label_from_line(line),
            "montant_source": float(line.montant or 0),
            "montant_retenu": float(row.montant_retenu or 0),
            "pct": float(row.pourcentage_retenu or 0) if row.pourcentage_retenu is not None else "",
            "commentaire": row.commentaire or "",
            "subvention": getattr(getattr(appel, "subvention", None), "nom", None) or appel.financeur or "",
        })

    ws = wb.active
    ws.title = _excel_safe_sheet_title("Synthèse financeur", used_titles)
    ws.append([f"Budget appel à projet — {appel.nom}"])
    ws["A1"].font = Font(bold=True, size=16)
    ws.append(["Budget source", appel.budget.nom, "Projet", appel.projet.nom if appel.projet else "", "Financeur", appel.financeur or ""])
    ws.append([])
    ws.append(["Indicateur", "Montant"])
    summary_header = ws.max_row
    ws.append(["Charges retenues", appel.total_charges])
    ws.append(["Produits retenus", appel.total_produits])
    ws.append(["Solde produits - charges", appel.solde])
    _add_safe_table(ws, "SyntheseBudgetAppel", summary_header, ws.max_row, 2, style="TableStyleMedium9")
    ws.append([])

    ws.append(["Nature", "Compte", "Libellé", "Projet source", "Montant source", "Montant retenu", "% retenu", "Commentaire"])
    main_header = ws.max_row
    for item in rows:
        ws.append([item["nature_label"], item["compte"], item["libelle"], item["projet"], item["montant_source"], item["montant_retenu"], item["pct"], item["commentaire"]])
    _add_safe_table(ws, "BudgetAppelProjetSynthese", main_header, ws.max_row, 8, style="TableStyleMedium2")
    ws.freeze_panes = f"A{main_header + 1}"

    by_project = {}
    for item in rows:
        by_project.setdefault(item["projet"], []).append(item)

    ws_proj = wb.create_sheet(_excel_safe_sheet_title("Par projet", used_titles))
    ws_proj.append(["Projet", "Charges", "Produits", "Solde"])
    project_summary_header = ws_proj.max_row
    for project_label, items in sorted(by_project.items(), key=lambda kv: kv[0].lower()):
        charges = sum(i["montant_retenu"] for i in items if i["nature"] == "charge")
        produits = sum(i["montant_retenu"] for i in items if i["nature"] == "produit")
        ws_proj.append([project_label, charges, produits, produits - charges])
    _add_safe_table(ws_proj, "BudgetAppelParProjet", project_summary_header, ws_proj.max_row, 4, style="TableStyleMedium4")
    ws_proj.freeze_panes = "A2"

    for project_label, items in sorted(by_project.items(), key=lambda kv: kv[0].lower()):
        ws_p = wb.create_sheet(_excel_safe_sheet_title(f"Projet {project_label}", used_titles))
        charges = sum(i["montant_retenu"] for i in items if i["nature"] == "charge")
        produits = sum(i["montant_retenu"] for i in items if i["nature"] == "produit")
        ws_p.append([f"Projet — {project_label}"])
        ws_p["A1"].font = Font(bold=True, size=14)
        ws_p.append(["Charges retenues", charges, "Produits retenus", produits, "Solde", produits - charges])
        ws_p.append([])
        ws_p.append(["Nature", "Compte", "Libellé", "Montant source", "Montant retenu", "% retenu", "Commentaire"])
        header = ws_p.max_row
        for i in items:
            ws_p.append([i["nature_label"], i["compte"], i["libelle"], i["montant_source"], i["montant_retenu"], i["pct"], i["commentaire"]])
        _add_safe_table(ws_p, f"DetailProjet_{project_label}", header, ws_p.max_row, 7, style="TableStyleMedium2")
        ws_p.freeze_panes = f"A{header + 1}"

    ws_sub = wb.create_sheet(_excel_safe_sheet_title("Par financeur", used_titles))
    ws_sub.append(["Financeur / subvention", "Nature", "Compte", "Libellé", "Projet", "Montant retenu", "Commentaire"])
    header = ws_sub.max_row
    financeur = _subvention_label_from_appel(appel)
    for item in rows:
        ws_sub.append([financeur, item["nature_label"], item["compte"], item["libelle"], item["projet"], item["montant_retenu"], item["commentaire"]])
    _add_safe_table(ws_sub, "BudgetAppelParFinanceur", header, ws_sub.max_row, 7, style="TableStyleMedium4")
    ws_sub.freeze_panes = "A2"

    ws_data = wb.create_sheet(_excel_safe_sheet_title("DATA Appel", used_titles))
    ws_data.append(["Appel ID", "Budget ID", "Nature", "Compte", "Libellé", "Projet", "Montant source", "Montant retenu", "% retenu", "Commentaire", "Financeur"])
    for item in rows:
        ws_data.append([appel.id, appel.budget_id, item["nature"], item["compte"], item["libelle"], item["projet"], item["montant_source"], item["montant_retenu"], item["pct"], item["commentaire"], financeur])
    _add_safe_table(ws_data, "DataBudgetAppel", 1, ws_data.max_row, 11, style="TableStyleMedium7")
    ws_data.freeze_panes = "A2"

    for sheet in wb.worksheets:
        for col in range(1, sheet.max_column + 1):
            sheet.cell(1, col).font = Font(bold=True)
        _style_ws(sheet)
    return wb




@bp.route("/", methods=["GET", "POST"])
@login_required
@require_perm("subventions:view")
def index():
    secteurs = _allowed_secteurs()
    if request.method == "POST":
        if not _can_edit():
            abort(403)
        annee = _parse_int(request.form.get("annee"), date.today().year)
        secteur = (request.form.get("secteur") or "").strip()
        nom = (request.form.get("nom") or f"Budget prévisionnel {secteur} {annee}").strip()
        if secteur not in secteurs:
            abort(403)
        budget = BudgetPrevisionnel(nom=nom, annee=annee, secteur=secteur, statut="brouillon")
        db.session.add(budget)
        db.session.flush()

        selected_cat_ids = [int(x) for x in request.form.getlist("categorie_ref_id") if str(x).isdigit()]
        modele_id = _parse_int(request.form.get("modele_ref_id"), 0) or None

        cats_to_add = []
        seen_cat_ids = set()

        if modele_id:
            modele = db.get_or_404(BudgetModeleReferentiel, modele_id)
            if modele.secteur and modele.secteur != budget.secteur:
                abort(400)
            if not modele.actif:
                abort(400)
            for row in modele.lignes:
                cat = row.categorie
                if not cat or cat.id in seen_cat_ids or not _category_allowed_for_budget(cat, budget):
                    continue
                cats_to_add.append((cat, row.montant_defaut, row.commentaire))
                seen_cat_ids.add(cat.id)

        if selected_cat_ids:
            cats = BudgetCategorieReferentiel.query.filter(BudgetCategorieReferentiel.id.in_(selected_cat_ids)).all()
            for cat in cats:
                if cat.id in seen_cat_ids or not _category_allowed_for_budget(cat, budget):
                    continue
                cats_to_add.append((cat, None, None))
                seen_cat_ids.add(cat.id)

        ordre = 1
        for cat, montant_override, commentaire_override in cats_to_add:
            line = _add_budget_line_from_category(budget, cat, ordre=ordre)
            if montant_override is not None:
                line.montant = float(montant_override or 0.0)
            if commentaire_override:
                line.commentaire = commentaire_override
            db.session.add(line)
            ordre += 1

        db.session.commit()
        flash("Budget prévisionnel créé.", "success")
        return redirect(url_for("previsionnel.detail", budget_id=budget.id))

    q = BudgetPrevisionnel.query
    if not _is_all_scope():
        q = q.filter(BudgetPrevisionnel.secteur.in_(secteurs or ["__none__"]))
    budgets = q.order_by(BudgetPrevisionnel.annee.desc(), BudgetPrevisionnel.secteur.asc(), BudgetPrevisionnel.nom.asc()).all()
    default_secteur = secteurs[0] if secteurs else None
    categories_ref = _budget_categories_ref(default_secteur)
    modeles_ref = _budget_modeles_ref(default_secteur)
    return render_template(
        "previsionnel/index.html",
        budgets=budgets,
        secteurs=secteurs,
        categories_ref=categories_ref,
        modeles_ref=modeles_ref,
        current_year=date.today().year,
        can_edit_budget=_can_edit(),
    )


@bp.route("/<int:budget_id>", methods=["GET", "POST"])
@login_required
@require_perm("subventions:view")
def detail(budget_id: int):
    budget = db.get_or_404(BudgetPrevisionnel, budget_id)
    _check_budget_scope(budget)

    if request.method == "POST":
        if not _can_edit():
            abort(403)
        action = request.form.get("action") or ""

        if action == "update_budget":
            budget.nom = (request.form.get("nom") or budget.nom).strip()
            budget.statut = (request.form.get("statut") or budget.statut or "brouillon").strip()
            budget.notes = (request.form.get("notes") or "").strip() or None
            db.session.commit()
            flash("Budget mis à jour.", "success")
            return redirect(url_for("previsionnel.detail", budget_id=budget.id))

        if action == "add_line":
            nature = (request.form.get("nature") or "charge").strip()
            if nature not in ("charge", "produit"):
                nature = "charge"

            cat_id = _parse_int(request.form.get("categorie_ref_id"), 0) or None
            cat = None
            if cat_id:
                cat = db.get_or_404(BudgetCategorieReferentiel, cat_id)
                if not _category_allowed_for_budget(cat, budget):
                    abort(400)
                nature = cat.nature

            projet_id = _parse_int(request.form.get("projet_id"), 0) or None
            if projet_id:
                projet = db.get_or_404(Projet, projet_id)
                if projet.secteur != budget.secteur:
                    abort(400)

            compte_default = _category_compte_label(cat) if cat else ("74" if nature == "produit" else "60")
            libelle_default = cat.libelle if cat else ""
            montant_default = float(cat.montant_defaut or 0.0) if cat else 0.0

            line = BudgetPrevisionnelLigne(
                budget_id=budget.id,
                nature=nature,
                compte=(request.form.get("compte") or compte_default).strip(),
                libelle=(request.form.get("libelle") or libelle_default).strip(),
                montant=_parse_float(request.form.get("montant"), montant_default),
                projet_id=projet_id,
                commentaire=(request.form.get("commentaire") or (cat.description if cat else "") or "").strip() or None,
                ordre=len(budget.lignes) + 1,
            )
            if not line.libelle:
                flash("Le libellé de la ligne est obligatoire.", "danger")
                return redirect(url_for("previsionnel.detail", budget_id=budget.id, tab=("produits" if nature == "produit" else "charges")))
            db.session.add(line)
            db.session.commit()
            flash("Ligne ajoutée au prévisionnel.", "success")
            return redirect(url_for("previsionnel.detail", budget_id=budget.id, tab=("produits" if nature == "produit" else "charges")))

        if action == "create_subvention_from_line":
            line = db.get_or_404(BudgetPrevisionnelLigne, _parse_int(request.form.get("line_id")))
            if line.budget_id != budget.id:
                abort(400)
            if line.nature != "produit":
                flash("Seules les lignes de produit peuvent générer une subvention.", "danger")
                return redirect(url_for("previsionnel.detail", budget_id=budget.id, tab="produits"))
            existing_subvention_id = getattr(line, "subvention_id", None)
            if existing_subvention_id:
                flash("Cette ligne est déjà reliée à une subvention.", "info")
                return redirect(url_for("main.subvention_pilotage", subvention_id=existing_subvention_id))

            sub_name = (request.form.get("subvention_nom") or line.libelle or f"Subvention {budget.annee}").strip()
            montant_demande = _parse_float(request.form.get("montant_demande"), float(line.montant or 0))
            sub = Subvention(
                nom=sub_name,
                secteur=budget.secteur,
                annee_exercice=budget.annee,
                montant_demande=montant_demande,
                montant_attribue=0.0,
                montant_recu=0.0,
                est_archive=False,
            )
            db.session.add(sub)
            db.session.flush()

            db.session.add(LigneBudget(
                subvention_id=sub.id,
                nature="produit",
                compte=line.compte or "74",
                libelle=line.libelle,
                montant_base=montant_demande,
                montant_reel=0.0,
            ))

            if line.projet_id:
                already = SubventionProjet.query.filter_by(projet_id=line.projet_id, subvention_id=sub.id).first()
                if not already:
                    db.session.add(SubventionProjet(projet_id=line.projet_id, subvention_id=sub.id))

            if hasattr(line, "subvention_id"):
                line.subvention_id = sub.id
            db.session.commit()
            flash("Subvention créée depuis la ligne produit du prévisionnel.", "success")
            return redirect(url_for("main.subvention_pilotage", subvention_id=sub.id))

        if action == "delete_line":
            line = db.get_or_404(BudgetPrevisionnelLigne, _parse_int(request.form.get("line_id")))
            if line.budget_id != budget.id:
                abort(400)
            if line.lignes_appel:
                flash("Cette ligne est déjà utilisée dans un budget d'appel à projet : suppression bloquée.", "danger")
                return redirect(url_for("previsionnel.detail", budget_id=budget.id))
            tab = "produits" if line.nature == "produit" else "charges"
            db.session.delete(line)
            db.session.commit()
            flash("Ligne supprimée.", "warning")
            return redirect(url_for("previsionnel.detail", budget_id=budget.id, tab=tab))

        if action == "create_appel":
            nom = (request.form.get("appel_nom") or "").strip() or f"Budget appel à projet — {budget.nom}"
            projet_id = _parse_int(request.form.get("appel_projet_id"), 0) or None
            subvention_id = _parse_int(request.form.get("subvention_id"), 0) or None
            financeur = (request.form.get("financeur") or "").strip() or None
            if subvention_id:
                sub = db.get_or_404(Subvention, subvention_id)
                if sub.secteur != budget.secteur:
                    abort(400)
                financeur = financeur or sub.nom
            if projet_id:
                projet = db.get_or_404(Projet, projet_id)
                if projet.secteur != budget.secteur:
                    abort(400)
            selected_ids = [int(x) for x in request.form.getlist("include_line") if str(x).isdigit()]
            if not selected_ids:
                flash("Sélectionne au moins une ligne à inclure dans le budget d'appel à projet.", "danger")
                return redirect(url_for("previsionnel.detail", budget_id=budget.id))
            usage = _appel_usage_for_budget(budget.id)
            selected_lines = BudgetPrevisionnelLigne.query.filter(
                BudgetPrevisionnelLigne.id.in_(selected_ids),
                BudgetPrevisionnelLigne.budget_id == budget.id,
            ).all()

            prepared_rows = []
            for line in selected_lines:
                pct = _parse_float(request.form.get(f"pct_{line.id}"), 100.0)
                retained = round(float(line.montant or 0) * pct / 100.0, 2)
                manual = request.form.get(f"montant_{line.id}")
                if manual not in (None, ""):
                    retained = _parse_float(manual)
                retained = max(0.0, _money(retained))
                if retained <= 0:
                    continue
                if not _validate_retained_against_remaining(line, retained, usage):
                    flash(
                        f"Montant trop élevé pour '{line.libelle}' : reste disponible {_line_remaining_amount(line, usage):.2f} €.",
                        "danger",
                    )
                    return redirect(url_for("previsionnel.detail", budget_id=budget.id, tab="appels"))
                prepared_rows.append((line, pct, retained))

            if not prepared_rows:
                flash("Aucune ligne retenue avec un montant supérieur à 0 €.", "danger")
                return redirect(url_for("previsionnel.detail", budget_id=budget.id, tab="appels"))

            appel = AppelProjetBudget(budget_id=budget.id, projet_id=projet_id, subvention_id=subvention_id, nom=nom, financeur=financeur)
            db.session.add(appel)
            db.session.flush()
            for line, pct, retained in prepared_rows:
                db.session.add(AppelProjetBudgetLigne(
                    appel_id=appel.id,
                    budget_ligne_id=line.id,
                    montant_retenu=retained,
                    pourcentage_retenu=pct,
                ))
            db.session.commit()
            flash("Budget d'appel à projet généré depuis le prévisionnel.", "success")
            return redirect(url_for("previsionnel.appel_detail", appel_id=appel.id))

        abort(400)

    lines = BudgetPrevisionnelLigne.query.filter_by(budget_id=budget.id).order_by(BudgetPrevisionnelLigne.nature.asc(), BudgetPrevisionnelLigne.compte.asc(), BudgetPrevisionnelLigne.ordre.asc(), BudgetPrevisionnelLigne.id.asc()).all()
    charge_lines = [l for l in lines if l.nature == "charge"]
    produit_lines = [l for l in lines if l.nature == "produit"]
    projects = _projects_for_budget(budget)
    subventions = _subventions_for_budget(budget)
    appels = AppelProjetBudget.query.filter_by(budget_id=budget.id).order_by(AppelProjetBudget.created_at.desc()).all()
    project_rows = _project_totals(lines)
    categories_ref = _budget_categories_ref(budget.secteur)
    modeles_payload = _modeles_payload(budget.secteur)
    line_category_ids = _line_category_ids(lines, categories_ref)
    realise_data = _compare_previsionnel_realise(budget, lines)
    appel_usage = _appel_usage_for_budget(budget.id)
    line_usage_rows = {
        l.id: {
            "used": float((appel_usage.get(l.id) or {}).get("montant") or 0.0),
            "remaining": _line_remaining_amount(l, appel_usage),
            "count": int((appel_usage.get(l.id) or {}).get("count") or 0),
            "appels": ", ".join((appel_usage.get(l.id) or {}).get("appels") or []),
        }
        for l in lines
    }
    tab = (request.args.get("tab") or "synthese").strip().lower()
    if tab not in {"synthese", "charges", "produits", "ventilation", "realise", "appels"}:
        tab = "synthese"
    return render_template(
        "previsionnel/detail.html",
        budget=budget,
        lines=lines,
        charge_lines=charge_lines,
        produit_lines=produit_lines,
        projects=projects,
        subventions=subventions,
        appels=appels,
        project_rows=project_rows,
        categories_ref=categories_ref,
        modeles_payload=modeles_payload,
        line_category_ids=line_category_ids,
        realise_data=realise_data,
        line_usage_rows=line_usage_rows,
        tab=tab,
        can_edit_budget=_can_edit(),
    )


@bp.route("/referentiel", methods=["GET", "POST"])
@login_required
@require_perm("subventions:view")
def referentiel():
    secteurs = _allowed_secteurs()
    if request.method == "POST":
        if not _can_edit():
            abort(403)
        action = request.form.get("action") or ""

        def _posted_secteur():
            secteur = (request.form.get("secteur") or "").strip() or None
            if secteur and secteur not in secteurs:
                abort(403)
            return secteur

        if action == "create_compte":
            compte = BudgetCompteReferentiel(
                nature=(request.form.get("nature") or "charge").strip(),
                code=(request.form.get("code") or "").strip(),
                libelle=(request.form.get("libelle") or "").strip(),
                description=(request.form.get("description") or "").strip() or None,
                secteur=_posted_secteur(),
                ordre=_parse_int(request.form.get("ordre"), 0),
                actif=True,
            )
            if compte.nature not in ("charge", "produit") or not compte.code or not compte.libelle:
                flash("Nature, code et libellé du compte sont obligatoires.", "danger")
                return redirect(url_for("previsionnel.referentiel"))
            db.session.add(compte)
            db.session.commit()
            flash("Compte budgétaire ajouté.", "success")
            return redirect(url_for("previsionnel.referentiel"))

        if action == "toggle_compte":
            compte = db.get_or_404(BudgetCompteReferentiel, _parse_int(request.form.get("compte_id")))
            compte.actif = not bool(compte.actif)
            db.session.commit()
            flash("Statut du compte modifié.", "success")
            return redirect(url_for("previsionnel.referentiel"))

        if action == "create_categorie":
            compte = db.get_or_404(BudgetCompteReferentiel, _parse_int(request.form.get("compte_id")))
            cat = BudgetCategorieReferentiel(
                compte_id=compte.id,
                nature=compte.nature,
                code=(request.form.get("code") or "").strip() or None,
                libelle=(request.form.get("libelle") or "").strip(),
                description=(request.form.get("description") or "").strip() or None,
                montant_defaut=_parse_float(request.form.get("montant_defaut"), 0.0),
                secteur=_posted_secteur(),
                ordre=_parse_int(request.form.get("ordre"), 0),
                actif=True,
            )
            if not cat.libelle:
                flash("Le libellé de la catégorie est obligatoire.", "danger")
                return redirect(url_for("previsionnel.referentiel"))
            db.session.add(cat)
            db.session.commit()
            flash("Catégorie budgétaire ajoutée.", "success")
            return redirect(url_for("previsionnel.referentiel"))

        if action == "toggle_categorie":
            cat = db.get_or_404(BudgetCategorieReferentiel, _parse_int(request.form.get("categorie_id")))
            cat.actif = not bool(cat.actif)
            db.session.commit()
            flash("Statut de la catégorie modifié.", "success")
            return redirect(url_for("previsionnel.referentiel"))

        if action == "create_modele":
            modele = BudgetModeleReferentiel(
                nom=(request.form.get("nom") or "").strip(),
                description=(request.form.get("description") or "").strip() or None,
                secteur=_posted_secteur(),
                ordre=_parse_int(request.form.get("ordre"), 0),
                actif=True,
            )
            if not modele.nom:
                flash("Le nom du modèle est obligatoire.", "danger")
                return redirect(url_for("previsionnel.referentiel"))
            db.session.add(modele)
            db.session.commit()
            flash("Modèle budgétaire ajouté.", "success")
            return redirect(url_for("previsionnel.referentiel"))

        if action == "toggle_modele":
            modele = db.get_or_404(BudgetModeleReferentiel, _parse_int(request.form.get("modele_id")))
            modele.actif = not bool(modele.actif)
            db.session.commit()
            flash("Statut du modèle modifié.", "success")
            return redirect(url_for("previsionnel.referentiel"))

        if action == "add_modele_line":
            modele = db.get_or_404(BudgetModeleReferentiel, _parse_int(request.form.get("modele_id")))
            cat = db.get_or_404(BudgetCategorieReferentiel, _parse_int(request.form.get("categorie_id")))
            if modele.secteur and cat.secteur and modele.secteur != cat.secteur:
                abort(400)
            exists = BudgetModeleLigneReferentiel.query.filter_by(modele_id=modele.id, categorie_id=cat.id).first()
            if exists:
                flash("Cette catégorie est déjà présente dans ce modèle.", "warning")
                return redirect(url_for("previsionnel.referentiel"))
            row = BudgetModeleLigneReferentiel(
                modele_id=modele.id,
                categorie_id=cat.id,
                montant_defaut=_parse_float(request.form.get("montant_defaut"), 0.0),
                ordre=_parse_int(request.form.get("ordre"), len(modele.lignes) + 1),
                commentaire=(request.form.get("commentaire") or "").strip() or None,
            )
            db.session.add(row)
            db.session.commit()
            flash("Ligne ajoutée au modèle.", "success")
            return redirect(url_for("previsionnel.referentiel"))

        if action == "remove_modele_line":
            row = db.get_or_404(BudgetModeleLigneReferentiel, _parse_int(request.form.get("modele_ligne_id")))
            db.session.delete(row)
            db.session.commit()
            flash("Ligne retirée du modèle.", "warning")
            return redirect(url_for("previsionnel.referentiel"))

        abort(400)

    comptes = _budget_comptes_ref(None, active_only=False)
    categories = _budget_categories_ref(None, active_only=False)
    modeles = BudgetModeleReferentiel.query.order_by(BudgetModeleReferentiel.ordre.asc(), BudgetModeleReferentiel.nom.asc()).all()
    return render_template(
        "previsionnel/referentiel.html",
        comptes=comptes,
        categories=categories,
        modeles=modeles,
        secteurs=secteurs,
        can_edit_budget=_can_edit(),
    )


@bp.route("/<int:budget_id>/export-realise.xlsx")
@login_required
@require_perm("subventions:view")
def export_realise(budget_id: int):
    budget = db.get_or_404(BudgetPrevisionnel, budget_id)
    _check_budget_scope(budget)
    lines = BudgetPrevisionnelLigne.query.filter_by(budget_id=budget.id).order_by(
        BudgetPrevisionnelLigne.nature.asc(),
        BudgetPrevisionnelLigne.compte.asc(),
        BudgetPrevisionnelLigne.ordre.asc(),
        BudgetPrevisionnelLigne.id.asc(),
    ).all()
    realise_data = _compare_previsionnel_realise(budget, lines)
    wb = _build_realise_workbook(budget, realise_data)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"previsionnel_vs_realise_{budget.annee}_{budget.secteur}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/<int:budget_id>/export.xlsx")
@login_required
@require_perm("subventions:view")
def export_budget(budget_id: int):
    budget = db.get_or_404(BudgetPrevisionnel, budget_id)
    _check_budget_scope(budget)
    lines = BudgetPrevisionnelLigne.query.filter_by(budget_id=budget.id).order_by(BudgetPrevisionnelLigne.nature.asc(), BudgetPrevisionnelLigne.compte.asc(), BudgetPrevisionnelLigne.id.asc()).all()
    wb = _build_budget_workbook(budget, lines)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=f"budget_previsionnel_{budget.annee}_{budget.secteur}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/appel/<int:appel_id>", methods=["GET", "POST"])
@login_required
@require_perm("subventions:view")
def appel_detail(appel_id: int):
    appel = db.get_or_404(AppelProjetBudget, appel_id)
    _check_budget_scope(appel.budget)

    if request.method == "POST":
        if not _can_edit():
            abort(403)
        action = request.form.get("action") or ""

        if action == "update_appel":
            appel.nom = (request.form.get("nom") or appel.nom).strip()
            appel.financeur = (request.form.get("financeur") or "").strip() or None
            db.session.commit()
            flash("Budget d'appel à projet mis à jour.", "success")
            return redirect(url_for("previsionnel.appel_detail", appel_id=appel.id))

        if action == "convert_to_subvention":
            if not (can("subventions:create") or can("subventions:edit")):
                abort(403)

            nom = (request.form.get("subvention_nom") or appel.financeur or appel.nom).strip()
            montant_demande = _parse_float(request.form.get("montant_demande"), _subvention_demande_default(appel))
            montant_attribue = _parse_float(request.form.get("montant_attribue"), 0.0)
            montant_recu = _parse_float(request.form.get("montant_recu"), 0.0)
            statut = (request.form.get("statut") or "accorde").strip()
            sync_lines = (request.form.get("sync_lines") or "").strip() == "1"

            if not nom:
                flash("Nom de subvention obligatoire.", "danger")
                return redirect(url_for("previsionnel.appel_detail", appel_id=appel.id))

            if montant_recu > montant_attribue > 0:
                flash("Le montant reçu ne peut pas dépasser le montant attribué.", "danger")
                return redirect(url_for("previsionnel.appel_detail", appel_id=appel.id))

            if appel.subvention:
                sub = appel.subvention
                sub.nom = nom
                sub.secteur = appel.budget.secteur
                sub.annee_exercice = appel.budget.annee
                sub.montant_demande = _money(montant_demande)
                sub.montant_attribue = _money(montant_attribue)
                sub.montant_recu = _money(montant_recu)
            else:
                sub = Subvention(
                    nom=nom,
                    secteur=appel.budget.secteur,
                    annee_exercice=appel.budget.annee,
                    montant_demande=_money(montant_demande),
                    montant_attribue=_money(montant_attribue),
                    montant_recu=_money(montant_recu),
                )
                db.session.add(sub)
                db.session.flush()
                appel.subvention_id = sub.id

            appel.statut = statut
            if not appel.financeur:
                appel.financeur = nom

            _ensure_subvention_project_link(sub, appel.projet_id)

            if sync_lines:
                _sync_appel_lines_to_subvention(appel, sub, montant_attribue, montant_recu)

            db.session.commit()
            flash("Dossier financeur converti / synchronisé avec la subvention réelle.", "success")
            return redirect(url_for("previsionnel.appel_detail", appel_id=appel.id))

        if action == "update_lines":
            usage_without_current = _appel_usage_for_budget(appel.budget_id, exclude_appel_id=appel.id)
            for row in list(appel.lignes):
                retained = _parse_float(request.form.get(f"montant_{row.id}"), float(row.montant_retenu or 0))
                pct = _parse_float(request.form.get(f"pct_{row.id}"), float(row.pourcentage_retenu or 0))
                retained = max(0.0, _money(retained))
                if not _validate_retained_against_remaining(row.ligne_budget, retained, usage_without_current):
                    flash(
                        f"Montant trop élevé pour '{row.ligne_budget.libelle}' : reste disponible {_line_remaining_amount(row.ligne_budget, usage_without_current):.2f} € hors ce budget.",
                        "danger",
                    )
                    return redirect(url_for("previsionnel.appel_detail", appel_id=appel.id))
                row.montant_retenu = retained
                row.pourcentage_retenu = pct
                row.commentaire = (request.form.get(f"commentaire_{row.id}") or "").strip() or None
            db.session.commit()
            flash("Lignes retenues mises à jour.", "success")
            return redirect(url_for("previsionnel.appel_detail", appel_id=appel.id))

        if action == "remove_line":
            row = db.get_or_404(AppelProjetBudgetLigne, _parse_int(request.form.get("appel_ligne_id")))
            if row.appel_id != appel.id:
                abort(400)
            db.session.delete(row)
            db.session.commit()
            flash("Ligne retirée du budget d'appel.", "warning")
            return redirect(url_for("previsionnel.appel_detail", appel_id=appel.id))

        if action == "add_lines":
            existing_ids = {int(row.budget_ligne_id) for row in appel.lignes}
            selected_ids = [int(x) for x in request.form.getlist("include_line") if str(x).isdigit()]
            selected_ids = [x for x in selected_ids if x not in existing_ids]
            if not selected_ids:
                flash("Aucune nouvelle ligne sélectionnée.", "danger")
                return redirect(url_for("previsionnel.appel_detail", appel_id=appel.id))
            usage_without_current = _appel_usage_for_budget(appel.budget_id, exclude_appel_id=appel.id)
            added = 0
            for line in BudgetPrevisionnelLigne.query.filter(
                BudgetPrevisionnelLigne.id.in_(selected_ids),
                BudgetPrevisionnelLigne.budget_id == appel.budget_id,
            ).all():
                pct = _parse_float(request.form.get(f"pct_{line.id}"), 100.0)
                retained = round(float(line.montant or 0) * pct / 100.0, 2)
                manual = request.form.get(f"montant_{line.id}")
                if manual not in (None, ""):
                    retained = _parse_float(manual)
                retained = max(0.0, _money(retained))
                if retained <= 0:
                    continue
                if not _validate_retained_against_remaining(line, retained, usage_without_current):
                    flash(
                        f"Montant trop élevé pour '{line.libelle}' : reste disponible {_line_remaining_amount(line, usage_without_current):.2f} € hors ce budget.",
                        "danger",
                    )
                    return redirect(url_for("previsionnel.appel_detail", appel_id=appel.id))
                db.session.add(AppelProjetBudgetLigne(
                    appel_id=appel.id,
                    budget_ligne_id=line.id,
                    montant_retenu=retained,
                    pourcentage_retenu=pct,
                ))
                added += 1
            if not added:
                flash("Aucune ligne ajoutée : montants retenus à 0 €.", "danger")
                return redirect(url_for("previsionnel.appel_detail", appel_id=appel.id))
            db.session.commit()
            flash("Lignes ajoutées au budget d'appel.", "success")
            return redirect(url_for("previsionnel.appel_detail", appel_id=appel.id))

        abort(400)

    lines = BudgetPrevisionnelLigne.query.filter_by(budget_id=appel.budget_id).order_by(
        BudgetPrevisionnelLigne.nature.asc(),
        BudgetPrevisionnelLigne.compte.asc(),
        BudgetPrevisionnelLigne.ordre.asc(),
        BudgetPrevisionnelLigne.id.asc(),
    ).all()
    existing_ids = {int(row.budget_ligne_id) for row in appel.lignes}
    available_lines = [line for line in lines if line.id not in existing_ids]
    usage_without_current = _appel_usage_for_budget(appel.budget_id, exclude_appel_id=appel.id)
    line_usage_rows = {
        l.id: {
            "used_elsewhere": float((usage_without_current.get(l.id) or {}).get("montant") or 0.0),
            "remaining": _line_remaining_amount(l, usage_without_current),
            "count": int((usage_without_current.get(l.id) or {}).get("count") or 0),
            "appels": ", ".join((usage_without_current.get(l.id) or {}).get("appels") or []),
        }
        for l in lines
    }
    demande_default = _subvention_demande_default(appel)
    return render_template(
        "previsionnel/appel_detail.html",
        appel=appel,
        available_lines=available_lines,
        line_usage_rows=line_usage_rows,
        demande_default=demande_default,
        can_edit_budget=_can_edit(),
    )


@bp.route("/appel/<int:appel_id>/delete", methods=["POST"])
@login_required
@require_perm("subventions:view")
def appel_delete(appel_id: int):
    if not _can_edit():
        abort(403)
    appel = db.get_or_404(AppelProjetBudget, appel_id)
    _check_budget_scope(appel.budget)
    budget_id = appel.budget_id
    db.session.delete(appel)
    db.session.commit()
    flash("Budget d'appel à projet supprimé.", "warning")
    return redirect(url_for("previsionnel.detail", budget_id=budget_id))


@bp.route("/appel/<int:appel_id>/export.xlsx")
@login_required
@require_perm("subventions:view")
def export_appel(appel_id: int):
    appel = db.get_or_404(AppelProjetBudget, appel_id)
    _check_budget_scope(appel.budget)
    wb = _build_appel_workbook(appel)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=f"budget_appel_projet_{appel.id}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
