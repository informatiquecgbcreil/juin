"""Générateur de budget prévisionnel au format « compte de résultat prévisionnel »
associatif (type CERFA 12156 / modèle CAF-FADO).

Ce module définit la structure canonique du plan comptable (charges 60→69,
produits 70→79 + ressources propres, contributions volontaires 86/87), calcule
les totaux et l'équilibre, et produit :
- un classeur Excel fidèle au modèle (sous-totaux par compte via formules SUM,
  totaux généraux, vérification de l'équilibre) ;
- des lignes prêtes à alimenter une subvention (LigneBudget) ou un budget
  prévisionnel de projet (BudgetPrevisionnelLigne).

Aucune écriture en base ici : logique pure, testable isolément.
"""
from __future__ import annotations

import re
from io import BytesIO


# --- Structure canonique (dérivée du modèle CAF-FADO fourni) -----------------

def _cat(idx: int, code: str, libelle: str) -> dict:
    # clé de champ HTML stable et unique par catégorie
    return {"key": f"cat__{code}__{idx}", "libelle": libelle}


CERFA_CHARGES = [
    {"code": "60", "libelle": "Achat", "categories": [
        _cat(0, "60", "Prestations de services"),
        _cat(1, "60", "Achats matières et fournitures"),
        _cat(2, "60", "Fournitures d'activité et petit équipement"),
        _cat(3, "60", "Autres fournitures (alimentation)"),
    ]},
    {"code": "61", "libelle": "Services extérieurs", "categories": [
        _cat(0, "61", "Locations"),
        _cat(1, "61", "Entretien et réparation"),
        _cat(2, "61", "Assurance"),
        _cat(3, "61", "Documentation"),
        _cat(4, "61", "Logiciels"),
    ]},
    {"code": "62", "libelle": "Autres services extérieurs", "categories": [
        _cat(0, "62", "Rémunérations intermédiaires, honoraires"),
        _cat(1, "62", "Publicité, publications"),
        _cat(2, "62", "Déplacements, missions"),
        _cat(3, "62", "Frais postaux et internet"),
        _cat(4, "62", "Impressions tracts extérieur"),
    ]},
    {"code": "63", "libelle": "Impôts et taxes", "categories": [
        _cat(0, "63", "Impôts et taxes sur rémunérations"),
        _cat(1, "63", "Autres impôts et taxes"),
    ]},
    {"code": "64", "libelle": "Charges de personnel", "categories": [
        _cat(0, "64", "Rémunération des personnels"),
        _cat(1, "64", "Charges sociales"),
        _cat(2, "64", "Autres charges de personnel"),
    ]},
    {"code": "65", "libelle": "Autres charges de gestion courante", "categories": [
        _cat(0, "65", "Autres charges de gestion courante"),
    ]},
    {"code": "66", "libelle": "Charges financières", "categories": [
        _cat(0, "66", "Charges financières"),
    ]},
    {"code": "67", "libelle": "Charges exceptionnelles", "categories": [
        _cat(0, "67", "Charges exceptionnelles"),
    ]},
    {"code": "68", "libelle": "Dotations aux amortissements, provisions et engagements", "categories": [
        _cat(0, "68", "Dotations et provisions et engagements"),
    ]},
    {"code": "69", "libelle": "Charges indirectes liées à l'activité", "categories": [
        _cat(0, "69", "Charges indirectes"),
    ]},
]

CERFA_PRODUITS = [
    {"code": "70", "libelle": "Vente de produits finis, de marchandises, prestation de services", "categories": [
        _cat(0, "70", "Ventes"),
    ]},
    {"code": "73", "libelle": "Dotations et produits de tarification", "categories": [
        _cat(0, "73", "Tarifications"),
    ]},
    {"code": "74", "libelle": "Subventions d'exploitation", "categories": [
        _cat(0, "74", "Politique de la Ville / P147"),
        _cat(1, "74", "Fonjep"),
        _cat(2, "74", "DRAC"),
        _cat(3, "74", "ARS"),
        _cat(4, "74", "Région"),
        _cat(5, "74", "Hauts-de-France"),
        _cat(6, "74", "Département"),
        _cat(7, "74", "OISE"),
        _cat(8, "74", "Intercommunalité : EPCI"),
        _cat(9, "74", "DRAJES (Jeunesse Éducation Populaire)"),
        _cat(10, "74", "Fonds propres"),
        _cat(11, "74", "Commune"),
        _cat(12, "74", "CREIL"),
        _cat(13, "74", "Organismes sociaux : CAF OISE"),
        _cat(14, "74", "REAAP"),
        _cat(15, "74", "Fonds européens"),
        _cat(16, "74", "Recherche de financement"),
        _cat(17, "74", "ASP / emplois aidés"),
        _cat(18, "74", "Creil ACSO"),
        _cat(19, "74", "Aides privées (Caisse d'épargne)"),
    ]},
    {"code": "75", "libelle": "Autres produits de gestion courante", "categories": [
        _cat(0, "75", "Dont cotisations, dons manuels ou legs"),
    ]},
    {"code": "76", "libelle": "Produits financiers", "categories": [
        _cat(0, "76", "Produits financiers"),
    ]},
    {"code": "77", "libelle": "Produits exceptionnels", "categories": [
        _cat(0, "77", "Produits exceptionnels"),
    ]},
    {"code": "78", "libelle": "Report ressources non utilisées d'opérations antérieures", "categories": [
        _cat(0, "78", "Reports des années antérieures"),
    ]},
    {"code": "79", "libelle": "Transfert de charges", "categories": [
        _cat(0, "79", "Transfert de charges"),
    ]},
    {"code": "RP", "libelle": "Ressources propres", "categories": [
        _cat(0, "RP", "Ressources propres"),
    ]},
]

CERFA_CONTRIB_EMPLOIS = [
    {"code": "86", "libelle": "Emplois des contributions volontaires en nature", "categories": [
        _cat(0, "86", "860 - Secours en nature"),
        _cat(1, "86", "861 - Mise à disposition gratuite de biens et services"),
        _cat(2, "86", "862 - Prestations"),
        _cat(3, "86", "864 - Personnel bénévole"),
    ]},
]

CERFA_CONTRIB_RESSOURCES = [
    {"code": "87", "libelle": "Contributions volontaires en nature", "categories": [
        _cat(0, "87", "870 - Bénévolat"),
        _cat(1, "87", "871 - Prestations en nature"),
        _cat(2, "87", "875 - Dons en nature"),
    ]},
]


# --- Sections (une structure = 4 sections ordonnées) -------------------------

SECTIONS = [
    ("charges", "Charges", "charge"),
    ("produits", "Produits", "produit"),
    ("contrib_emplois", "Contributions volontaires — emplois", "charge"),
    ("contrib_ressources", "Contributions volontaires — ressources", "produit"),
]
SECTION_KEYS = [k for k, _, _ in SECTIONS]
SECTION_NATURE = {k: nat for k, _, nat in SECTIONS}
SECTION_LABEL = {k: lib for k, lib, _ in SECTIONS}

# Garde-fous anti-abus (structure envoyée par le client).
MAX_COMPTES = 120
MAX_LINES = 300


def default_structure() -> dict:
    """Structure de départ = modèle CERFA/CAF-FADO canonique, montants à 0."""
    def conv(block):
        return [
            {"code": c["code"], "libelle": c["libelle"],
             "lines": [{"libelle": cat["libelle"], "montant": 0.0} for cat in c["categories"]]}
            for c in block
        ]
    return {
        "charges": conv(CERFA_CHARGES),
        "produits": conv(CERFA_PRODUITS),
        "contrib_emplois": conv(CERFA_CONTRIB_EMPLOIS),
        "contrib_ressources": conv(CERFA_CONTRIB_RESSOURCES),
    }


def parse_amount(value) -> float:
    """Tolère nombre, espaces, virgule décimale et somme simple « a+b+c »."""
    raw = str(value if value is not None else "").replace(" ", "").replace(" ", "").replace(",", ".")
    if not raw:
        return 0.0
    try:
        return round(float(raw), 2)
    except Exception:
        pass
    if all(c in "0123456789.+" for c in raw):
        total = 0.0
        ok = False
        for part in raw.split("+"):
            if not part:
                continue
            try:
                total += float(part)
                ok = True
            except Exception:
                return 0.0
        if ok:
            return round(total, 2)
    return 0.0


def parse_structure(raw) -> dict:
    """Valide/normalise une structure (chaîne JSON ou dict). Retourne toujours
    une structure sûre ; revient au modèle par défaut si l'entrée est vide."""
    import json
    data = raw
    if isinstance(raw, str):
        try:
            data = json.loads(raw)
        except Exception:
            data = None
    if not isinstance(data, dict):
        return default_structure()

    out = {}
    for key in SECTION_KEYS:
        comptes = []
        block = data.get(key)
        if isinstance(block, list):
            for c in block[:MAX_COMPTES]:
                if not isinstance(c, dict):
                    continue
                code = str(c.get("code") or "").strip()[:20]
                libelle = str(c.get("libelle") or "").strip()[:200]
                lines = []
                raw_lines = c.get("lines")
                if isinstance(raw_lines, list):
                    for ln in raw_lines[:MAX_LINES]:
                        if not isinstance(ln, dict):
                            continue
                        llib = str(ln.get("libelle") or "").strip()[:255]
                        montant = parse_amount(ln.get("montant"))
                        if not llib and not montant:
                            continue
                        lines.append({"libelle": llib, "montant": montant})
                if code or libelle or lines:
                    comptes.append({"code": code, "libelle": libelle, "lines": lines})
        out[key] = comptes

    if not any(out[k] for k in SECTION_KEYS):
        return default_structure()
    return out


def _compte_total(compte: dict) -> float:
    return round(sum(float(l.get("montant") or 0) for l in compte.get("lines", [])), 2)


def _section_total(structure: dict, key: str) -> float:
    return round(sum(_compte_total(c) for c in structure.get(key, [])), 2)


def compute_totals(structure: dict) -> dict:
    total_charges = _section_total(structure, "charges")
    total_produits = _section_total(structure, "produits")
    total_emplois = _section_total(structure, "contrib_emplois")
    total_ressources = _section_total(structure, "contrib_ressources")
    tgc = round(total_charges + total_emplois, 2)
    tgp = round(total_produits + total_ressources, 2)
    return {
        "total_charges": total_charges,
        "total_produits": total_produits,
        "total_emplois": total_emplois,
        "total_ressources": total_ressources,
        "total_general_charges": tgc,
        "total_general_produits": tgp,
        "equilibre": round(tgp - tgc, 2),
    }


def _compte_label(compte: dict) -> str:
    code = (compte.get("code") or "").strip()
    lib = (compte.get("libelle") or "").strip()
    label = f"{code} - {lib}".strip(" -")
    return label or code or lib or "60"


def lignes_budget(structure: dict, *, inclure_zero: bool = False) -> list[dict]:
    """Aplati charges + produits en lignes {nature, compte, code, libelle, montant}
    pour alimenter une subvention ou un budget prévisionnel. Les contributions
    volontaires (86/87) sont exclues (ce ne sont pas des flux réels)."""
    out = []
    for key in ("charges", "produits"):
        nature = SECTION_NATURE[key]
        for compte in structure.get(key, []):
            label = _compte_label(compte)
            for ln in compte.get("lines", []):
                montant = round(float(ln.get("montant") or 0), 2)
                if montant or inclure_zero:
                    out.append({
                        "nature": nature,
                        "compte": label,
                        "code": (compte.get("code") or "").strip(),
                        "libelle": (ln.get("libelle") or "").strip() or label,
                        "montant": montant,
                    })
    return out


# --- Export Excel ------------------------------------------------------------

def _safe_text(value) -> str:
    """Neutralise l'injection de formule Excel dans un texte libre."""
    s = "" if value is None else str(value)
    if s and s[0] in ("=", "+", "-", "@"):
        return "'" + s
    return s


def build_workbook(structure: dict, *, organisme: str, exercice, secteur: str | None = None):
    """Classeur Excel fidèle au modèle : colonnes parallèles charges/produits,
    sous-totaux par compte en formules SUM, totaux généraux et équilibre.
    Piloté par la structure (éventuellement éditée par l'utilisateur)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    titre = re.sub(r"[\[\]:*?/\\]", " ", (secteur or "Budget")).strip()[:31]
    ws.title = titre or "Budget"

    bold = Font(bold=True)
    head_fill = PatternFill(fill_type="solid", fgColor="C6E0B4")
    prod_fill = PatternFill(fill_type="solid", fgColor="BDD7EE")
    total_fill = PatternFill(fill_type="solid", fgColor="FFE699")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "Nom de l'organisme"
    ws["A1"].font = bold
    ws["B1"] = _safe_text(organisme)
    ws.merge_cells("B1:D1")
    ws["A3"] = "EXERCICE"
    ws["A3"].font = bold
    ws["B3"] = exercice
    if secteur:
        ws["C3"] = "Secteur"
        ws["C3"].font = bold
        ws["D3"] = _safe_text(secteur)

    ws["A5"] = "CHARGES"
    ws["C5"] = "PRODUITS"
    for coord, fill in (("A5", head_fill), ("B5", head_fill), ("C5", prod_fill), ("D5", prod_fill)):
        ws[coord].font = bold
        ws[coord].fill = fill

    def write_block(comptes, col_lib, col_val, fill, cursor_start):
        row = cursor_start
        subtotal_cells = []
        for compte in comptes:
            first_cat_row = row + 1
            last_cat_row = row + len(compte.get("lines", []))
            lib_cell = ws.cell(row=row, column=col_lib, value=_compte_label(compte))
            lib_cell.font = bold
            lib_cell.fill = fill
            val_cell = ws.cell(row=row, column=col_val)
            col_letter = val_cell.column_letter
            if last_cat_row >= first_cat_row:
                val_cell.value = f"=SUM({col_letter}{first_cat_row}:{col_letter}{last_cat_row})"
            else:
                val_cell.value = 0
            val_cell.font = bold
            val_cell.fill = fill
            val_cell.number_format = "#,##0.00"
            subtotal_cells.append(f"{col_letter}{row}")
            row += 1
            for ln in compte.get("lines", []):
                ws.cell(row=row, column=col_lib, value=_safe_text(ln.get("libelle")))
                mc = ws.cell(row=row, column=col_val, value=round(float(ln.get("montant") or 0), 2))
                mc.number_format = "#,##0.00"
                row += 1
        return row, subtotal_cells

    end_charges, charge_subtotals = write_block(structure.get("charges", []), 1, 2, head_fill, 6)
    end_produits, produit_subtotals = write_block(structure.get("produits", []), 3, 4, prod_fill, 6)

    total_row = max(end_charges, end_produits) + 1
    ws.cell(row=total_row, column=1, value="TOTAL DES CHARGES").font = bold
    tc = ws.cell(row=total_row, column=2, value=("=" + "+".join(charge_subtotals)) if charge_subtotals else 0)
    ws.cell(row=total_row, column=3, value="TOTAL DES PRODUITS").font = bold
    tp = ws.cell(row=total_row, column=4, value=("=" + "+".join(produit_subtotals)) if produit_subtotals else 0)
    for cell in (tc, tp):
        cell.font = bold
        cell.number_format = "#,##0.00"
    for col in (1, 2, 3, 4):
        ws.cell(row=total_row, column=col).fill = total_fill

    contrib_title = total_row + 2
    ws.cell(row=contrib_title, column=1, value="CONTRIBUTIONS VOLONTAIRES EN NATURE").font = bold
    end_emplois, emploi_subtotals = write_block(structure.get("contrib_emplois", []), 1, 2, head_fill, contrib_title + 1)
    end_ressources, ressource_subtotals = write_block(structure.get("contrib_ressources", []), 3, 4, prod_fill, contrib_title + 1)

    tg_row = max(end_emplois, end_ressources) + 1
    ws.cell(row=tg_row, column=1, value="TOTAL GÉNÉRAL DES CHARGES").font = bold
    emplois_expr = ("+" + "+".join(emploi_subtotals)) if emploi_subtotals else ""
    ressources_expr = ("+" + "+".join(ressource_subtotals)) if ressource_subtotals else ""
    tgc = ws.cell(row=tg_row, column=2, value=f"=B{total_row}{emplois_expr}")
    ws.cell(row=tg_row, column=3, value="TOTAL GÉNÉRAL DES PRODUITS").font = bold
    tgp = ws.cell(row=tg_row, column=4, value=f"=D{total_row}{ressources_expr}")
    for cell in (tgc, tgp):
        cell.font = bold
        cell.number_format = "#,##0.00"
    for col in (1, 2, 3, 4):
        ws.cell(row=tg_row, column=col).fill = total_fill

    eq_row = tg_row + 2
    ws.cell(row=eq_row, column=3, value="Vérification de l'équilibre (produits - charges)").font = bold
    eq = ws.cell(row=eq_row, column=4, value=f"=D{tg_row}-B{tg_row}")
    eq.font = bold
    eq.number_format = "#,##0.00"

    for row in ws.iter_rows(min_row=5, max_row=eq_row, min_col=1, max_col=4):
        for cell in row:
            if cell.value is not None:
                cell.border = border
    for col_letter, width in (("A", 42), ("B", 15), ("C", 46), ("D", 15)):
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A6"
    return wb
