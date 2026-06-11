import os
import json
from io import BytesIO
from datetime import datetime, date
from flask import Blueprint, render_template, request, redirect, url_for, flash, abort, current_app, send_file
from flask_login import login_required, current_user
from app.rbac import require_perm, can, can_access_secteur
from werkzeug.utils import secure_filename

from app.extensions import db
from app.services.indicators import (
    INDICATOR_PACKS,
    INDICATOR_METRIC_GROUPS,
    INDICATOR_TEMPLATES,
    PERIOD_CHOICES,
    TARGET_OP_CHOICES,
    TARGET_OP_SYMBOLS,
    compute_project_indicators,
    indicator_counts,
    indicator_list_rows,
    indicator_metric_code,
    indicator_params,
    indicator_source,
    indicator_unique_code,
    indicator_value_type,
    parse_float_optional,
)
from app.utils.delete_guard import commit_delete

from app.models import (
    Projet,
    ChargeProjet,
    ProduitProjet,
    VentilationProjet,
    Depense,
    DepenseAffectation,
    Subvention,
    SubventionProjet,
    LigneBudget,
    BudgetPrevisionnel,
    BudgetPrevisionnelLigne,
    AppelProjetBudget,
    AtelierActivite,
    ProjetAtelier,
    ProjetIndicateur,
    ProjetJournalEntry,
    ProjetAction,
    ProjetActionAtelier,
    Competence,
    Objectif,
    objectif_competence,
    projet_competence,
    Referentiel,
)
from app.statsimpact.engine import (
    StatsFilters,
    compute_demography_stats,
    compute_participation_frequency_stats,
    compute_volume_activity_stats,
)

bp = Blueprint("projets", __name__)

ALLOWED_CR = {"pdf", "doc", "docx", "odt"}

PROJET_JOURNAL_CATEGORIES = {
    "fait_marquant": "Fait marquant",
    "reussite": "Réussite",
    "difficulte": "Difficulté",
    "partenariat": "Partenariat",
    "bilan": "Bilan",
}

PROJET_ACTION_CATEGORIES = {
    "atelier": "Atelier / cycle",
    "evenement": "Événement",
    "permanence": "Permanence",
    "sortie": "Sortie / temps fort",
    "partenariat": "Partenariat",
    "autre": "Autre",
}

PROJET_ACTION_STATUTS = {
    "prevue": "Prévue",
    "en_cours": "En cours",
    "realisee": "Réalisée",
    "annulee": "Annulée",
}


# ---------------------------------------------------------------------
# Helpers Budget AAP (UX)
# ---------------------------------------------------------------------

def _budget_stats(projet_id: int) -> dict:
    """Stats simples pour l'UX du budget AAP.

    Objectif : afficher en haut des pages un résumé (totaux + "reste à...")
    et permettre d'afficher des alertes rapides.
    """
    charges = ChargeProjet.query.filter_by(projet_id=projet_id).all()
    produits = ProduitProjet.query.filter_by(projet_id=projet_id).all()

    total_charges = float(sum((c.montant_previsionnel or 0) for c in charges))
    total_charges_ventile = float(sum((c.ventile or 0) for c in charges))
    total_charges_reste = max(0.0, total_charges - total_charges_ventile)

    total_demande = float(sum((p.montant_demande or 0) for p in produits))
    total_accorde = float(sum((p.montant_accorde or 0) for p in produits))
    total_recu = float(sum((p.montant_recu or 0) for p in produits))

    total_produits_ventile = float(sum((p.ventile or 0) for p in produits))
    total_produits_reste = float(sum((p.reste_a_ventiler or 0) for p in produits))

    # Base "produits" utilisée pour donner du contexte (reçu > accordé > demandé)
    base_produits = total_recu if total_recu > 0 else (total_accorde if total_accorde > 0 else total_demande)
    base_label = "reçu" if total_recu > 0 else ("accordé" if total_accorde > 0 else "demandé")

    def pct(a: float, b: float) -> int:
        if b <= 0:
            return 0
        v = int(round((a / b) * 100))
        return max(0, min(100, v))

    return {
        "total_charges": total_charges,
        "total_charges_ventile": total_charges_ventile,
        "total_charges_reste": total_charges_reste,
        "total_demande": total_demande,
        "total_accorde": total_accorde,
        "total_recu": total_recu,
        "base_produits": base_produits,
        "base_label": base_label,
        "total_produits_ventile": total_produits_ventile,
        "total_produits_reste": total_produits_reste,
        "pct_charges_financees": pct(total_charges_ventile, total_charges),
        "pct_produits_ventiles": pct(total_produits_ventile, base_produits if base_produits > 0 else total_charges),
    }


def can_see_secteur(secteur: str) -> bool:
    return can_access_secteur(secteur)


# ---------------------------------------------------------------------
# P0.2A — Dossier financier projet
# ---------------------------------------------------------------------

def _money(value) -> float:
    try:
        return round(float(value or 0), 2)
    except Exception:
        return 0.0


def _pct(part, total) -> int:
    try:
        total = float(total or 0)
        part = float(part or 0)
    except Exception:
        return 0
    if total <= 0:
        return 0
    return max(0, min(100, int(round((part / total) * 100))))


def _parse_float_fr(value, default=0.0) -> float:
    raw = str(value or '').strip()
    if not raw:
        return float(default)
    raw = raw.replace('\u00a0', '').replace(' ', '').replace('€', '').replace(',', '.')
    try:
        return round(float(raw), 2)
    except Exception:
        return float(default)


def _parse_date_or_none(value):
    raw = str(value or '').strip()
    if not raw:
        return None
    try:
        return date.fromisoformat(raw)
    except Exception:
        return None


def _clean_choice(value: str | None, choices: dict, default: str) -> str:
    raw = (value or default).strip()
    return raw if raw in choices else default


def _project_linked_atelier_ids(projet_id: int) -> set[int]:
    return {
        row.atelier_id for row in ProjetAtelier.query.filter_by(projet_id=projet_id).all()
    }


def _action_linked_atelier_ids(action: ProjetAction) -> list[int]:
    return [link.atelier_id for link in getattr(action, "atelier_links", []) or [] if link.atelier_id]


def _action_period(action: ProjetAction, year: int) -> tuple[date, date]:
    dmin = action.date_debut or date(year, 1, 1)
    dmax = action.date_fin or date(year, 12, 31)
    if dmax < dmin:
        dmin, dmax = dmax, dmin
    return dmin, dmax


def _empty_action_stats() -> dict:
    return {
        "kpi": {"sessions": 0, "presences": 0, "uniques": 0, "hours_people": 0},
        "freq": {"returning": 0, "returning_rate": 0, "freq_avg": 0},
        "demo": {"age_avg": None, "qpv": {"qpv": 0, "hors_qpv": 0, "inconnu": 0}},
    }


def _compute_action_stats(projet: Projet, action: ProjetAction, year: int) -> dict:
    atelier_ids = _action_linked_atelier_ids(action)
    if not atelier_ids:
        return _empty_action_stats()

    dmin, dmax = _action_period(action, year)
    flt = StatsFilters(
        secteur=projet.secteur,
        atelier_ids=atelier_ids,
        date_from=dmin,
        date_to=dmax,
    )
    return {
        "kpi": compute_volume_activity_stats(flt).get("kpi", {}),
        "freq": compute_participation_frequency_stats(flt),
        "demo": compute_demography_stats(flt),
    }


def _action_list_rows(projet: Projet, actions: list[ProjetAction], year: int) -> list[dict]:
    rows = []
    for action in actions:
        atelier_ids = _action_linked_atelier_ids(action)
        quality = _action_completion_payload(action, atelier_ids)
        rows.append({
            "action": action,
            "atelier_ids": atelier_ids,
            "stats": _compute_action_stats(projet, action, year),
            "statut_label": PROJET_ACTION_STATUTS.get(action.statut, action.statut),
            "categorie_label": PROJET_ACTION_CATEGORIES.get(action.categorie, action.categorie),
            "quality": quality,
        })
    return rows


def _quality_issue(severity: str, title: str, detail: str) -> dict:
    return {"severity": severity, "title": title, "detail": detail}


def _quality_counts(issues: list[dict]) -> dict:
    return {
        "bad": sum(1 for item in issues if item.get("severity") == "bad"),
        "warn": sum(1 for item in issues if item.get("severity") == "warn"),
        "info": sum(1 for item in issues if item.get("severity") == "info"),
    }


def _quality_score(issues: list[dict]) -> int:
    counts = _quality_counts(issues)
    score = 100 - (counts["bad"] * 18) - (counts["warn"] * 10) - (counts["info"] * 4)
    return max(0, min(100, score))


def _quality_payload(issues: list[dict]) -> dict:
    counts = _quality_counts(issues)
    score = _quality_score(issues)
    if counts["bad"]:
        status = "bad"
        label = "À sécuriser"
    elif counts["warn"]:
        status = "warn"
        label = "À compléter"
    else:
        status = "ok"
        label = "Prêt à partager"
    return {
        "issues": issues,
        "counts": counts,
        "score": score,
        "status": status,
        "label": label,
    }


def _action_completion_payload(action: ProjetAction, atelier_ids: list[int]) -> dict:
    issues = []
    if not atelier_ids:
        issues.append(_quality_issue(
            "bad",
            "Aucun atelier stats-impact",
            "Les chiffres de l'action resteront a zero tant qu'aucun atelier du projet n'est rattache.",
        ))
    if not getattr(action, "objectifs", None):
        issues.append(_quality_issue("warn", "Objectifs manquants", "Les objectifs de l'action ne sont pas encore renseignes."))
    if not getattr(action, "public_vise", None):
        issues.append(_quality_issue("info", "Public vise a preciser", "Le public vise aidera la lecture par l'equipe et les partenaires."))
    if not getattr(action, "referent", None):
        issues.append(_quality_issue("info", "Referent non renseigne", "La fiche ne precise pas encore qui pilote l'action."))
    if not getattr(action, "date_debut", None) and not getattr(action, "date_fin", None):
        issues.append(_quality_issue("warn", "Periode absente", "Sans periode, les stats utilisent l'annee selectionnee."))
    if getattr(action, "statut", None) == "realisee" and not getattr(action, "bilan_qualitatif", None):
        issues.append(_quality_issue("warn", "Bilan a completer", "L'action est realisee mais le bilan qualitatif n'est pas encore renseigne."))
    if not getattr(action, "description", None):
        issues.append(_quality_issue("info", "Description courte absente", "Une description courte facilite la lecture rapide de la fiche projet."))
    return _quality_payload(issues)


def _project_completion_payload(
    projet: Projet,
    linked_ateliers: list[AtelierActivite],
    action_rows: list[dict],
    indicators: list[dict],
    journal_entries: list[ProjetJournalEntry],
    finance: dict,
) -> dict:
    issues = []
    if not (projet.description or "").strip():
        issues.append(_quality_issue("warn", "Description projet manquante", "La fiche projet n'a pas encore de description lisible."))
    if not linked_ateliers:
        issues.append(_quality_issue("bad", "Aucun atelier stats-impact lie", "Les statistiques projet et indicateurs connectes resteront vides."))
    if not action_rows:
        issues.append(_quality_issue("warn", "Aucune fiche action", "Le projet n'a pas encore d'action detaillee partageable avec l'equipe."))
    else:
        action_without_stats = sum(1 for row in action_rows if not row.get("atelier_ids"))
        action_without_bilan = sum(
            1 for row in action_rows
            if getattr(row.get("action"), "statut", None) == "realisee"
            and not getattr(row.get("action"), "bilan_qualitatif", None)
        )
        if action_without_stats:
            issues.append(_quality_issue("warn", "Actions sans atelier", f"{action_without_stats} action(s) ne piochent pas encore dans stats-impact."))
        if action_without_bilan:
            issues.append(_quality_issue("warn", "Bilans d'action manquants", f"{action_without_bilan} action(s) realisee(s) n'ont pas encore de bilan qualitatif."))
    if not indicators:
        issues.append(_quality_issue("warn", "Aucun indicateur actif", "La page indicateurs ne contient pas encore d'indicateur actif pour ce projet."))
    else:
        pending = sum(1 for row in indicators if not row.get("status"))
        bad = sum(1 for row in indicators if row.get("status") == "bad")
        if pending:
            issues.append(_quality_issue("info", "Indicateurs a completer", f"{pending} indicateur(s) n'ont pas encore de cible ou de valeur exploitable."))
        if bad:
            issues.append(_quality_issue("warn", "Indicateurs non atteints", f"{bad} indicateur(s) sont sous l'objectif defini."))
    if not journal_entries:
        issues.append(_quality_issue("info", "Journal vide", "Aucun fait marquant n'est encore note pour faciliter le bilan."))
    finance_alerts = finance.get("alerts") or []
    finance_bad = sum(1 for alert in finance_alerts if alert.get("level") == "danger")
    finance_warn = sum(1 for alert in finance_alerts if alert.get("level") == "warn")
    if finance_bad:
        issues.append(_quality_issue("bad", "Alerte financiere critique", f"{finance_bad} point(s) financier(s) demandent une verification."))
    elif finance_warn:
        issues.append(_quality_issue("warn", "Points financiers a suivre", f"{finance_warn} alerte(s) financiere(s) sont a surveiller."))

    payload = _quality_payload(issues)
    payload["sections"] = {
        "description": bool((projet.description or "").strip()),
        "ateliers": bool(linked_ateliers),
        "actions": bool(action_rows),
        "indicateurs": bool(indicators),
        "journal": bool(journal_entries),
    }
    return payload


def _projet_finance_years(projet: Projet) -> list[int]:
    years = set()

    # Années des subventions rattachées au projet
    for link in getattr(projet, "subventions", []) or []:
        sub = getattr(link, "subvention", None)
        if sub and getattr(sub, "annee_exercice", None):
            years.add(int(sub.annee_exercice))

    # Années des budgets prévisionnels contenant des lignes pour ce projet
    rows = (
        db.session.query(BudgetPrevisionnel.annee)
        .join(BudgetPrevisionnelLigne, BudgetPrevisionnelLigne.budget_id == BudgetPrevisionnel.id)
        .filter(BudgetPrevisionnelLigne.projet_id == projet.id)
        .distinct()
        .all()
    )
    for (annee,) in rows:
        if annee:
            years.add(int(annee))

    # Années des dossiers/appels à projet liés
    rows = (
        db.session.query(BudgetPrevisionnel.annee)
        .join(AppelProjetBudget, AppelProjetBudget.budget_id == BudgetPrevisionnel.id)
        .filter(AppelProjetBudget.projet_id == projet.id)
        .distinct()
        .all()
    )
    for (annee,) in rows:
        if annee:
            years.add(int(annee))

    if not years:
        years.add(date.today().year)
    return sorted(years, reverse=True)


def _projet_selected_year(projet: Projet) -> int:
    years = _projet_finance_years(projet)
    raw = request.args.get("year")
    try:
        year = int(raw) if raw else years[0]
    except Exception:
        year = years[0]
    if year not in years:
        year = years[0]
    return year


def _depense_key(dep: Depense) -> str:
    return f"dep-{dep.id}"


def _collect_project_depenses(projet: Projet, subventions: list[Subvention]) -> list[Depense]:
    """Récupère les dépenses via deux chemins sans doublonner :
    - dépenses rattachées aux lignes budgétaires des subventions du projet ;
    - dépenses rattachées directement aux charges projet.
    """
    depenses_by_key = {}

    for sub in subventions:
        # Nouveau modèle : les imputations sont la source la plus fiable.
        for aff in getattr(sub, "depense_affectations", []) or []:
            dep = getattr(aff, "depense", None)
            if dep and not getattr(dep, "est_supprimee", False):
                depenses_by_key[_depense_key(dep)] = dep

        # Compatibilité ancien modèle : dépenses directement attachées aux lignes.
        for ligne in getattr(sub, "lignes", []) or []:
            if getattr(ligne, "nature", "charge") != "charge":
                continue
            for dep in getattr(ligne, "depenses", []) or []:
                if getattr(dep, "est_supprimee", False):
                    continue
                depenses_by_key[_depense_key(dep)] = dep

    for charge in getattr(projet, "charges_projet", []) or []:
        for dep in getattr(charge, "depenses", []) or []:
            if getattr(dep, "est_supprimee", False):
                continue
            depenses_by_key[_depense_key(dep)] = dep

    return sorted(
        depenses_by_key.values(),
        key=lambda d: (d.date_paiement or d.created_at.date() if getattr(d, "created_at", None) else date.min, d.id),
        reverse=True,
    )



# ---------------------------------------------------------------------
# P0.2D — Pilotage financier annuel du secteur
# ---------------------------------------------------------------------

def _finance_accessible_sectors() -> list[str]:
    sectors = set()

    for (s,) in db.session.query(Projet.secteur).distinct().all():
        if s and can_access_secteur(s):
            sectors.add(s)

    for (s,) in db.session.query(BudgetPrevisionnel.secteur).distinct().all():
        if s and can_access_secteur(s):
            sectors.add(s)

    for (s,) in db.session.query(Subvention.secteur).distinct().all():
        if s and can_access_secteur(s):
            sectors.add(s)

    return sorted(sectors)


def _finance_sector_years(secteur: str) -> list[int]:
    years = set()

    for (annee,) in db.session.query(BudgetPrevisionnel.annee).filter(BudgetPrevisionnel.secteur == secteur).distinct().all():
        if annee:
            years.add(int(annee))

    for (annee,) in db.session.query(Subvention.annee_exercice).filter(Subvention.secteur == secteur).distinct().all():
        if annee:
            years.add(int(annee))

    if not years:
        years.add(date.today().year)

    return sorted(years, reverse=True)


def _selected_sector_year():
    sectors = _finance_accessible_sectors()
    selected_sector = (request.args.get("secteur") or "").strip()
    if not selected_sector or selected_sector not in sectors:
        selected_sector = sectors[0] if sectors else None

    years = _finance_sector_years(selected_sector) if selected_sector else [date.today().year]
    try:
        year = int(request.args.get("year") or years[0])
    except Exception:
        year = years[0]

    if year not in years:
        year = years[0]

    return sectors, selected_sector, years, year


def _is_compte_fonds_propres(compte) -> bool:
    raw = str(compte or "").strip().lower().replace(" ", "")
    return raw.startswith("99")


def _label_is_fonds_propres(label) -> bool:
    raw = str(label or "").strip().lower()
    markers = (
        "fonds propre",
        "fonds propres",
        "autofinancement",
        "auto-financement",
        "reste à charge",
        "reste a charge",
        "interne",
    )
    return any(m in raw for m in markers)


def _is_fonds_propres_enveloppe(subvention=None, ligne=None) -> bool:
    """Convention métier P0.2G.1.

    On ne crée pas de table dédiée pour les fonds propres.
    Une enveloppe reste une Subvention côté code, mais elle est classée à part si :
    - une de ses lignes utilise un compte commençant par 99 ;
    - ou son nom contient fonds propres/autofinancement/reste à charge.
    """
    if ligne is not None and _is_compte_fonds_propres(getattr(ligne, "compte", None)):
        return True

    if subvention is not None:
        if _label_is_fonds_propres(getattr(subvention, "nom", "")):
            return True
        for l in getattr(subvention, "lignes", []) or []:
            if _is_compte_fonds_propres(getattr(l, "compte", None)):
                return True

    return False



def _collect_sector_depenses(secteur: str, year: int, subventions: list[Subvention]) -> list[Depense]:
    """Récupère les dépenses de l'exercice via les enveloppes réelles du secteur.

    On part des subventions réelles, parce que c'est le vrai point d'imputation.
    Les fonds propres remontent via les affectations de ces dépenses.
    """
    depenses = {}
    for sub in subventions:
        for ligne in getattr(sub, "lignes", []) or []:
            for dep in getattr(ligne, "depenses", []) or []:
                if dep and not dep.est_supprimee:
                    depenses[dep.id] = dep
        for aff in getattr(sub, "depense_affectations", []) or []:
            dep = getattr(aff, "depense", None)
            if dep and not dep.est_supprimee:
                depenses[dep.id] = dep

    return sorted(
        depenses.values(),
        key=lambda d: (d.date_paiement or d.created_at.date() if getattr(d, "created_at", None) else date.min, d.id),
        reverse=True,
    )


def _sector_finance_context(secteur: str, year: int) -> dict:
    budgets = (
        BudgetPrevisionnel.query
        .filter_by(secteur=secteur, annee=year)
        .order_by(BudgetPrevisionnel.created_at.desc(), BudgetPrevisionnel.id.desc())
        .all()
    )

    prev_lines = []
    for budget in budgets:
        prev_lines.extend(list(getattr(budget, "lignes", []) or []))

    subventions = (
        Subvention.query
        .filter(Subvention.secteur == secteur, Subvention.annee_exercice == year, Subvention.est_archive.is_(False))
        .order_by(Subvention.nom.asc())
        .all()
    )

    dossiers = (
        AppelProjetBudget.query
        .join(BudgetPrevisionnel, AppelProjetBudget.budget_id == BudgetPrevisionnel.id)
        .filter(BudgetPrevisionnel.secteur == secteur, BudgetPrevisionnel.annee == year)
        .order_by(AppelProjetBudget.created_at.desc(), AppelProjetBudget.id.desc())
        .all()
    )

    projets = Projet.query.filter_by(secteur=secteur).order_by(Projet.nom.asc()).all()
    depenses = _collect_sector_depenses(secteur, year, subventions)

    # Prévisionnel global
    charges_prevues = _money(sum(float(l.montant or 0) for l in prev_lines if l.nature == "charge"))
    produits_prevus = _money(sum(float(l.montant or 0) for l in prev_lines if l.nature == "produit"))

    # Enveloppes réelles : externes + fonds propres/autofinancement interne.
    # Techniquement, les fonds propres peuvent rester dans la table Subvention,
    # mais ils sont isolés par convention via compte 99 ou libellé.
    fonds_propres_enveloppes = [s for s in subventions if _is_fonds_propres_enveloppe(s)]
    subventions_externes = [s for s in subventions if not _is_fonds_propres_enveloppe(s)]

    demande_externe = _money(sum(float(s.montant_demande or 0) for s in subventions_externes))
    attribue_externe = _money(sum(float(s.montant_attribue or 0) for s in subventions_externes))
    recu_externe = _money(sum(float(s.montant_recu or 0) for s in subventions_externes))

    fonds_propres_demande = _money(sum(float(s.montant_demande or 0) for s in fonds_propres_enveloppes))
    fonds_propres_attribue = _money(sum(float(s.montant_attribue or 0) for s in fonds_propres_enveloppes))
    fonds_propres_recu = _money(sum(float(s.montant_recu or 0) for s in fonds_propres_enveloppes))

    demande = _money(demande_externe + fonds_propres_demande)
    attribue = _money(attribue_externe + fonds_propres_attribue)
    recu = _money(recu_externe + fonds_propres_recu)

    # Imputations réelles : subventions externes + fonds propres/autofinancement + autres
    impute_subventions = 0.0
    impute_fonds_propres = 0.0
    impute_autre = 0.0
    source_rows = {}

    for dep in depenses:
        for aff in getattr(dep, "affectations", []) or []:
            amount = float(aff.montant or 0)
            if aff.source_type == "subvention" and aff.subvention and aff.subvention.secteur == secteur and int(aff.subvention.annee_exercice or 0) == int(year):
                is_fp = _is_fonds_propres_enveloppe(aff.subvention, getattr(aff, "ligne_budget", None))
                if is_fp:
                    impute_fonds_propres += amount
                    key = f"fonds-sub-{aff.subvention_id}"
                    row = source_rows.setdefault(key, {
                        "type": "Fonds propres",
                        "label": aff.subvention.nom,
                        "demande": float(aff.subvention.montant_demande or 0),
                        "attribue": float(aff.subvention.montant_attribue or 0),
                        "recu": float(aff.subvention.montant_recu or 0),
                        "impute": 0.0,
                        "reste": 0.0,
                        "subvention": aff.subvention,
                        "is_enveloppe": True,
                    })
                else:
                    impute_subventions += amount
                    key = f"sub-{aff.subvention_id}"
                    row = source_rows.setdefault(key, {
                        "type": "Subvention",
                        "label": aff.subvention.nom,
                        "demande": float(aff.subvention.montant_demande or 0),
                        "attribue": float(aff.subvention.montant_attribue or 0),
                        "recu": float(aff.subvention.montant_recu or 0),
                        "impute": 0.0,
                        "reste": 0.0,
                        "subvention": aff.subvention,
                        "is_enveloppe": True,
                    })
                row["impute"] += amount
            elif aff.source_type == "fonds_propres":
                impute_fonds_propres += amount
                key = "fonds-propres"
                row = source_rows.setdefault(key, {
                    "type": "Fonds propres",
                    "label": aff.libelle_source or "Fonds propres",
                    "demande": 0.0,
                    "attribue": 0.0,
                    "recu": 0.0,
                    "impute": 0.0,
                    "reste": None,
                    "subvention": None,
                    "is_enveloppe": False,
                })
                row["impute"] += amount
            else:
                impute_autre += amount
                key = f"autre-{aff.libelle_source or 'autre'}"
                row = source_rows.setdefault(key, {
                    "type": "Autre",
                    "label": aff.libelle_source or "Autre source",
                    "demande": 0.0,
                    "attribue": 0.0,
                    "recu": 0.0,
                    "impute": 0.0,
                    "reste": None,
                    "subvention": None,
                    "is_enveloppe": False,
                })
                row["impute"] += amount

    for row in source_rows.values():
        if row["type"] in {"Subvention", "Fonds propres"} and row.get("is_enveloppe"):
            base = row["attribue"] if row["attribue"] > 0 else (row["recu"] if row["recu"] > 0 else row["demande"])
            row["reste"] = _money(base - row["impute"])
        row["impute"] = _money(row["impute"])

    impute_subventions = _money(impute_subventions)
    impute_fonds_propres = _money(impute_fonds_propres)
    impute_autre = _money(impute_autre)
    total_impute = _money(impute_subventions + impute_fonds_propres + impute_autre)
    depenses_uniques = _money(sum(float(d.montant or 0) for d in depenses))

    # Groupes budget par compte
    compte_rows = {}
    for line in prev_lines:
        key = (line.nature, line.compte or "")
        row = compte_rows.setdefault(key, {"nature": line.nature, "compte": line.compte or "", "prevu": 0.0, "nb": 0})
        row["prevu"] += float(line.montant or 0)
        row["nb"] += 1
    compte_rows = sorted(
        [
            {**r, "prevu": _money(r["prevu"]), "nature_label": "Produit" if r["nature"] == "produit" else "Charge"}
            for r in compte_rows.values()
        ],
        key=lambda r: (r["nature"], r["compte"])
    )

    # Projets : vue filtrée de l'année
    project_rows = []
    for projet in projets:
        pdata = _projet_finance_context(projet, year)
        pk = pdata["kpis"]
        has_data = any([
            pk.get("charges_prevues", 0),
            pk.get("produits_prevus", 0),
            pk.get("sub_attribue", 0),
            pk.get("sub_recu", 0),
            pk.get("depenses_total", 0),
            pk.get("nb_subventions", 0),
            pk.get("nb_dossiers", 0),
        ])
        project_rows.append({
            "projet": projet,
            "charges_prevues": pk.get("charges_prevues", 0),
            "produits_prevus": pk.get("produits_prevus", 0),
            "sub_attribue": pk.get("sub_attribue", 0),
            "sub_recu": pk.get("sub_recu", 0),
            "depenses_total": pk.get("depenses_total", 0),
            "reste_a_financer": pk.get("reste_a_financer", 0),
            "nb_subventions": pk.get("nb_subventions", 0),
            "nb_dossiers": pk.get("nb_dossiers", 0),
            "has_data": has_data,
        })

    project_rows = sorted(project_rows, key=lambda r: (not r["has_data"], r["projet"].nom.lower()))

    depense_rows = []
    for dep in depenses:
        depense_rows.append({
            "depense": dep,
            "date": dep.date_paiement or (dep.created_at.date() if dep.created_at else None),
            "montant": _money(dep.montant or 0),
            "affecte": _money(dep.total_affecte or 0),
            "reste": _money(dep.reste_a_affecter or 0),
            "statut": dep.statut_affectation,
        })

    sub_rows = []
    for sub in subventions:
        is_fp = _is_fonds_propres_enveloppe(sub)
        sub_rows.append({
            "subvention": sub,
            "type_enveloppe": "Fonds propres" if is_fp else "Subvention",
            "is_fonds_propres": is_fp,
            "demande": _money(sub.montant_demande or 0),
            "attribue": _money(sub.montant_attribue or 0),
            "recu": _money(sub.montant_recu or 0),
            "impute": _money(sub.total_impute_affectations or 0),
            "reste": _money(sub.reste_imputable or 0),
            "pct_impute": _pct(sub.total_impute_affectations or 0, sub.montant_attribue or sub.montant_recu or sub.montant_demande or 0),
        })

    controls = []
    partials = [r for r in depense_rows if r["statut"] != "ok"]

    if not budgets:
        _finance_add_control(controls, "warn", "Prévisionnel", f"Aucun budget prévisionnel global détecté pour {secteur} en {year}.", "Créer ou sélectionner un budget prévisionnel.")
    if charges_prevues > 0 and produits_prevus < charges_prevues:
        _finance_add_control(controls, "warn", "Prévisionnel", f"Déficit prévisionnel : {_money(charges_prevues - produits_prevus):.2f} € à couvrir.", "Vérifier les produits attendus ou les fonds propres.")
    if attribue > 0 and recu > attribue + 0.01:
        _finance_add_control(controls, "danger", "Enveloppes", f"Montant reçu supérieur au montant attribué de {_money(recu - attribue):.2f} €.", "Corriger les montants reçus/attribués.")
    if attribue > 0 and recu < attribue - 0.01:
        _finance_add_control(controls, "warn", "Encaissements", f"Subventions accordées non totalement reçues : {_money(attribue - recu):.2f} € restent à encaisser.", "Suivre les versements restants.")
    if partials:
        _finance_add_control(controls, "danger", "Dépenses", f"{len(partials)} dépense(s) ne sont pas totalement équilibrées en affectations.", "Ouvrir l'onglet Dépenses et corriger les répartitions.")
    if not subventions:
        _finance_add_control(controls, "info", "Enveloppes", "Aucune enveloppe/subvention réelle détectée pour cet exercice.", "Créer une subvention réelle ou une enveloppe fonds propres 99.")
    if not depenses:
        _finance_add_control(controls, "info", "Dépenses", "Aucune dépense imputée détectée pour cet exercice.", "Créer une dépense financée depuis la vue simple.")

    for r in sub_rows:
        sub = r["subvention"]
        base = float(r.get("attribue") or r.get("recu") or r.get("demande") or 0)
        if float(r.get("recu") or 0) > float(r.get("attribue") or 0) + 0.01 and float(r.get("attribue") or 0) > 0:
            _finance_add_control(controls, "danger", "Enveloppes", f"{sub.nom} : reçu supérieur à l'attribué.", "Corriger la fiche de l'enveloppe.")
        if float(r.get("impute") or 0) > base + 0.01 and base > 0:
            _finance_add_control(controls, "danger", "Surconsommation", f"{sub.nom} : dépenses imputées supérieures à l'enveloppe de {_money(float(r.get('impute') or 0) - base):.2f} €.", "Réduire les imputations ou augmenter l'enveloppe réelle.")
        charge_lines = [l for l in getattr(sub, "lignes", []) or [] if getattr(l, "nature", "charge") == "charge"]
        if not charge_lines:
            _finance_add_control(controls, "warn", "Enveloppes", f"{sub.nom} n'a aucune ligne de charge imputable.", "Créer au moins une ligne de charge.")
        for ligne in charge_lines:
            if float(getattr(ligne, "reste", 0) or 0) < -0.01:
                _finance_add_control(controls, "danger", "Lignes", f"{sub.nom} — {ligne.compte} {ligne.libelle} est surconsommée de {_money(abs(float(ligne.reste or 0))):.2f} €.", "Corriger les affectations de dépenses sur cette ligne.")

    for dossier in dossiers:
        if not getattr(dossier, "subvention_id", None) and str(getattr(dossier, "statut", "") or "").lower() in {"accorde", "partiel", "solde"}:
            _finance_add_control(controls, "warn", "Dossiers financeurs", f"{dossier.nom} est marqué {dossier.statut} mais n'est pas lié à une subvention réelle.", "Ouvrir le dossier et créer/mettre à jour l'enveloppe réelle.")

    for s in source_rows.values():
        if not s.get("is_enveloppe") and s.get("type") != "Subvention":
            _finance_add_control(controls, "info", "Sources libres", f"{s.get('label')} est utilisé hors enveloppe formelle.", "Si c'est récurrent, créer une enveloppe Fonds propres 99.")

    alerts = []
    for c in controls:
        if c["level"] in {"danger", "warn", "info"}:
            alerts.append({"level": c["level"], "text": c["message"]})
    if not alerts:
        alerts.append({"level": "ok", "text": "Aucune alerte bloquante détectée sur le pilotage annuel."})
        _finance_add_control(controls, "ok", "Global", "Aucun contrôle bloquant détecté sur l'exercice.", "")

    kpis = {
        "charges_prevues": charges_prevues,
        "produits_prevus": produits_prevus,
        "solde_prevu": _money(produits_prevus - charges_prevues),
        "demande": demande,
        "attribue": attribue,
        "recu": recu,
        "demande_externe": demande_externe,
        "attribue_externe": attribue_externe,
        "recu_externe": recu_externe,
        "fonds_propres_demande": fonds_propres_demande,
        "fonds_propres_attribue": fonds_propres_attribue,
        "fonds_propres_recu": fonds_propres_recu,
        "impute_subventions": impute_subventions,
        "impute_fonds_propres": impute_fonds_propres,
        "impute_autre": impute_autre,
        "total_impute": total_impute,
        "depenses_uniques": depenses_uniques,
        "reste_a_financer": _money(charges_prevues - (attribue if attribue > 0 else produits_prevus)),
        "reste_a_encaisser": _money(attribue - recu),
        "reste_subventions": _money(sum(float(r["reste"] or 0) for r in sub_rows)),
        "pct_finance": _pct(attribue if attribue > 0 else produits_prevus, charges_prevues),
        "pct_recu": _pct(recu, attribue if attribue > 0 else demande),
        "pct_impute": _pct(total_impute, charges_prevues),
        "nb_budgets": len(budgets),
        "nb_subventions": len(subventions_externes),
        "nb_fonds_propres": len(fonds_propres_enveloppes),
        "nb_dossiers": len(dossiers),
        "nb_projets": len([r for r in project_rows if r["has_data"]]),
        "nb_depenses": len(depenses),
        "nb_depenses_a_corriger": len(partials),
        "nb_controles": len([c for c in controls if c.get("level") != "ok"]),
        "nb_controles_danger": len([c for c in controls if c.get("level") == "danger"]),
        "nb_controles_warn": len([c for c in controls if c.get("level") == "warn"]),
    }

    return {
        "kpis": kpis,
        "alerts": alerts,
        "controls": controls,
        "budgets": budgets,
        "prev_lines": prev_lines,
        "subventions": sub_rows,
        "dossiers": dossiers,
        "projets": project_rows,
        "depenses": depense_rows,
        "sources": sorted(source_rows.values(), key=lambda r: (r["type"], r["label"].lower())),
        "comptes": compte_rows,
    }



# ---------------------------------------------------------------------
# P0.2H — Exports financiers annuels
# ---------------------------------------------------------------------

def _xlsx_clean(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, (datetime, date)):
        return value
    return str(value)


def _xlsx_money(value) -> float:
    try:
        return round(float(value or 0), 2)
    except Exception:
        return 0.0


def _xlsx_sheet_title(title: str, existing: set[str]) -> str:
    forbidden = ["\\", "/", "*", "[", "]", ":", "?"]
    clean = str(title or "Feuille")
    for ch in forbidden:
        clean = clean.replace(ch, " ")
    clean = " ".join(clean.split()).strip() or "Feuille"
    clean = clean[:31]
    base = clean
    i = 2
    while clean in existing:
        suffix = f" {i}"
        clean = (base[:31 - len(suffix)] + suffix).strip()
        i += 1
    existing.add(clean)
    return clean


def _xlsx_table_name(prefix: str, index: int) -> str:
    raw = "".join(ch if ch.isalnum() else "_" for ch in str(prefix or "Table"))
    raw = raw.strip("_") or "Table"
    if raw[0].isdigit():
        raw = "T_" + raw
    return f"{raw}_{index}"[:240]


def _xlsx_add_table(ws, table_name: str, n_rows: int, n_cols: int):
    if n_rows < 2 or n_cols < 1:
        return
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    ref = f"A1:{get_column_letter(n_cols)}{n_rows}"
    tab = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)


def _xlsx_autowidth(ws, max_width=48):
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 10
        for cell in col:
            value = cell.value
            if value is None:
                continue
            width = max(width, min(max_width, len(str(value)) + 2))
        ws.column_dimensions[letter].width = width


def _xlsx_style_sheet(ws, money_cols=None, percent_cols=None):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    money_cols = set(money_cols or [])
    percent_cols = set(percent_cols or [])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A2"

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.column in money_cols and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00 €'
            if cell.column in percent_cols and isinstance(cell.value, (int, float)):
                cell.number_format = '0"%"'

    _xlsx_autowidth(ws)


def _xlsx_write_table(wb, title: str, headers: list[str], rows: list[list], table_prefix: str, money_headers=None, percent_headers=None):
    existing = {ws.title for ws in wb.worksheets}
    ws = wb.create_sheet(_xlsx_sheet_title(title, existing))
    ws.append(headers)
    for row in rows:
        ws.append([_xlsx_clean(v) for v in row])

    money_headers = set(money_headers or [])
    percent_headers = set(percent_headers or [])
    money_cols = [idx + 1 for idx, h in enumerate(headers) if h in money_headers]
    percent_cols = [idx + 1 for idx, h in enumerate(headers) if h in percent_headers]

    _xlsx_style_sheet(ws, money_cols=money_cols, percent_cols=percent_cols)

    # Sécurité Excel :
    # ne pas cumuler filtre de feuille + table structurée.
    # Excel répare parfois /xl/tables/table*.xml quand les deux filtres se chevauchent.
    if rows:
        _xlsx_add_table(ws, _xlsx_table_name(table_prefix, len(wb.worksheets)), len(rows) + 1, len(headers))
    else:
        # Sur une feuille vide, un simple gel de ligne suffit.
        # Pas de table header-only, pas d'autoFilter fragile.
        pass

    return ws


def _build_sector_finance_workbook(secteur: str, year: int, data: dict):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference, PieChart

    wb = Workbook()
    ws = wb.active
    ws.title = "Lecture rapide"

    k = data.get("kpis", {})
    alerts = data.get("alerts", [])

    # Feuille lecture rapide, volontairement pas en table pour garder une mise en page financeur.
    rows = [
        ["Secteur", secteur],
        ["Exercice", year],
        ["Charges prévues", _xlsx_money(k.get("charges_prevues"))],
        ["Produits prévus", _xlsx_money(k.get("produits_prevus"))],
        ["Solde prévu", _xlsx_money(k.get("solde_prevu"))],
        ["Demandé total", _xlsx_money(k.get("demande"))],
        ["Attribué total", _xlsx_money(k.get("attribue"))],
        ["Reçu total", _xlsx_money(k.get("recu"))],
        ["Subventions externes attribuées", _xlsx_money(k.get("attribue_externe"))],
        ["Fonds propres / autofinancement disponible", _xlsx_money(k.get("fonds_propres_attribue"))],
        ["Imputé sur subventions externes", _xlsx_money(k.get("impute_subventions"))],
        ["Imputé fonds propres / autofinancement", _xlsx_money(k.get("impute_fonds_propres"))],
        ["Imputé total", _xlsx_money(k.get("total_impute"))],
        ["Reste à encaisser", _xlsx_money(k.get("reste_a_encaisser"))],
        ["Dépenses à corriger", k.get("nb_depenses_a_corriger") or 0],
    ]
    ws.append(["Indicateur", "Valeur"])
    for r in rows:
        ws.append(r)

    # Style lecture rapide
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    ws["A1"].fill = header_fill
    ws["B1"].fill = header_fill
    ws["A1"].font = header_font
    ws["B1"].font = header_font
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 22
    for row in ws.iter_rows(min_row=2, max_col=2):
        row[0].font = Font(bold=True)
        if isinstance(row[1].value, (int, float)) and "Dépenses à corriger" not in str(row[0].value):
            row[1].number_format = '#,##0.00 €'
    ws.freeze_panes = "A2"

    # Bloc alertes
    start_alert = len(rows) + 4
    ws.cell(start_alert, 1, "Alertes / points de contrôle")
    ws.cell(start_alert, 1).font = Font(bold=True, size=13)
    ws.cell(start_alert + 1, 1, "Niveau")
    ws.cell(start_alert + 1, 2, "Message")
    for idx, alert in enumerate(alerts, start=start_alert + 2):
        ws.cell(idx, 1, alert.get("level", ""))
        ws.cell(idx, 2, alert.get("text", ""))
    _xlsx_autowidth(ws)

    # Petit graphique demandé/attribué/reçu/imputé
    chart_start = 20
    ws.cell(chart_start, 4, "Flux financier")
    ws.cell(chart_start, 5, "Montant")
    chart_rows = [
        ("Demandé", _xlsx_money(k.get("demande"))),
        ("Attribué", _xlsx_money(k.get("attribue"))),
        ("Reçu", _xlsx_money(k.get("recu"))),
        ("Imputé", _xlsx_money(k.get("total_impute"))),
    ]
    for i, (label, value) in enumerate(chart_rows, start=chart_start + 1):
        ws.cell(i, 4, label)
        ws.cell(i, 5, value)
    chart = BarChart()
    chart.title = "Demandé / attribué / reçu / imputé"
    chart.y_axis.title = "Montant"
    chart.x_axis.title = "Étape"
    chart.add_data(Reference(ws, min_col=5, min_row=chart_start, max_row=chart_start + len(chart_rows)), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=4, min_row=chart_start + 1, max_row=chart_start + len(chart_rows)))
    chart.height = 7
    chart.width = 14
    ws.add_chart(chart, "G3")

    # Budgets
    _xlsx_write_table(
        wb,
        "Budgets",
        ["Nom", "Statut", "Charges", "Produits", "Solde", "Lignes"],
        [
            [
                b.nom,
                b.statut,
                _xlsx_money(getattr(b, "total_charges", 0)),
                _xlsx_money(getattr(b, "total_produits", 0)),
                _xlsx_money(getattr(b, "solde", 0)),
                len(getattr(b, "lignes", []) or []),
            ]
            for b in data.get("budgets", [])
        ],
        "Budgets",
        money_headers={"Charges", "Produits", "Solde"},
    )

    # Répartition par compte
    _xlsx_write_table(
        wb,
        "Comptes prévisionnels",
        ["Nature", "Compte", "Montant prévu", "Nombre de lignes"],
        [
            [r.get("nature_label"), r.get("compte"), _xlsx_money(r.get("prevu")), r.get("nb")]
            for r in data.get("comptes", [])
        ],
        "ComptesPrev",
        money_headers={"Montant prévu"},
    )

    # Subventions externes
    _xlsx_write_table(
        wb,
        "Financeurs externes",
        ["Nom", "Demandé", "Attribué", "Reçu", "Imputé", "Reste imputable", "% imputé"],
        [
            [
                r["subvention"].nom,
                _xlsx_money(r.get("demande")),
                _xlsx_money(r.get("attribue")),
                _xlsx_money(r.get("recu")),
                _xlsx_money(r.get("impute")),
                _xlsx_money(r.get("reste")),
                r.get("pct_impute") or 0,
            ]
            for r in data.get("subventions", [])
            if not r.get("is_fonds_propres")
        ],
        "FinanceursExt",
        money_headers={"Demandé", "Attribué", "Reçu", "Imputé", "Reste imputable"},
        percent_headers={"% imputé"},
    )

    # Fonds propres compte 99
    _xlsx_write_table(
        wb,
        "Fonds propres 99",
        ["Nom", "Disponible", "Reçu / disponible", "Imputé", "Reste", "% imputé"],
        [
            [
                r["subvention"].nom,
                _xlsx_money(r.get("attribue")),
                _xlsx_money(r.get("recu")),
                _xlsx_money(r.get("impute")),
                _xlsx_money(r.get("reste")),
                r.get("pct_impute") or 0,
            ]
            for r in data.get("subventions", [])
            if r.get("is_fonds_propres")
        ],
        "FondsPropres99",
        money_headers={"Disponible", "Reçu / disponible", "Imputé", "Reste"},
        percent_headers={"% imputé"},
    )

    # Sources imputées
    _xlsx_write_table(
        wb,
        "Sources imputées",
        ["Type", "Source", "Demandé", "Attribué", "Reçu", "Imputé", "Reste", "Enveloppe formelle"],
        [
            [
                s.get("type"),
                s.get("label"),
                _xlsx_money(s.get("demande")),
                _xlsx_money(s.get("attribue")),
                _xlsx_money(s.get("recu")),
                _xlsx_money(s.get("impute")),
                "" if s.get("reste") is None else _xlsx_money(s.get("reste")),
                "Oui" if s.get("is_enveloppe") else "Non",
            ]
            for s in data.get("sources", [])
        ],
        "SourcesImputees",
        money_headers={"Demandé", "Attribué", "Reçu", "Imputé", "Reste"},
    )

    # Dossiers financeurs
    _xlsx_write_table(
        wb,
        "Dossiers financeurs",
        ["Dossier", "Statut", "Projet", "Financeur", "Charges retenues", "Produits retenus", "Solde", "Subvention liée"],
        [
            [
                a.nom,
                a.statut,
                a.projet.nom if getattr(a, "projet", None) else "Transversal",
                a.financeur or (a.subvention.nom if getattr(a, "subvention", None) else ""),
                _xlsx_money(getattr(a, "total_charges", 0)),
                _xlsx_money(getattr(a, "total_produits", 0)),
                _xlsx_money(getattr(a, "solde", 0)),
                a.subvention.nom if getattr(a, "subvention", None) else "",
            ]
            for a in data.get("dossiers", [])
        ],
        "DossiersFinanceurs",
        money_headers={"Charges retenues", "Produits retenus", "Solde"},
    )

    # Projets
    _xlsx_write_table(
        wb,
        "Projets",
        ["Projet", "A des données", "Charges prévues", "Produits prévus", "Attribué", "Reçu", "Dépenses", "Reste à financer", "Subventions", "Dossiers"],
        [
            [
                r["projet"].nom,
                "Oui" if r.get("has_data") else "Non",
                _xlsx_money(r.get("charges_prevues")),
                _xlsx_money(r.get("produits_prevus")),
                _xlsx_money(r.get("sub_attribue")),
                _xlsx_money(r.get("sub_recu")),
                _xlsx_money(r.get("depenses_total")),
                _xlsx_money(r.get("reste_a_financer")),
                r.get("nb_subventions"),
                r.get("nb_dossiers"),
            ]
            for r in data.get("projets", [])
        ],
        "Projets",
        money_headers={"Charges prévues", "Produits prévus", "Attribué", "Reçu", "Dépenses", "Reste à financer"},
    )

    # Dépenses
    _xlsx_write_table(
        wb,
        "Dépenses",
        ["Date", "Libellé", "Fournisseur", "Référence", "Montant", "Affecté", "Reste à affecter", "Statut", "Source initiale"],
        [
            [
                r.get("date"),
                r["depense"].libelle,
                r["depense"].fournisseur or "",
                r["depense"].reference_piece or "",
                _xlsx_money(r.get("montant")),
                _xlsx_money(r.get("affecte")),
                _xlsx_money(r.get("reste")),
                r.get("statut"),
                r["depense"].budget_source.source_sub.nom if getattr(r["depense"], "budget_source", None) and getattr(r["depense"].budget_source, "source_sub", None) else "",
            ]
            for r in data.get("depenses", [])
        ],
        "Depenses",
        money_headers={"Montant", "Affecté", "Reste à affecter"},
    )

    # Affectations détaillées
    aff_rows = []
    for r in data.get("depenses", []):
        dep = r["depense"]
        for aff in getattr(dep, "affectations", []) or []:
            if aff.source_type == "subvention" and getattr(aff, "subvention", None):
                source = aff.subvention.nom
                type_source = "Fonds propres" if _is_fonds_propres_enveloppe(aff.subvention, getattr(aff, "ligne_budget", None)) else "Subvention"
                ligne = f"{aff.ligne_budget.compte} - {aff.ligne_budget.libelle}" if getattr(aff, "ligne_budget", None) else ""
            else:
                source = aff.libelle_source or aff.source_type
                type_source = aff.source_type
                ligne = ""
            aff_rows.append([
                dep.date_paiement or (dep.created_at.date() if dep.created_at else ""),
                dep.libelle,
                _xlsx_money(dep.montant),
                type_source,
                source,
                ligne,
                _xlsx_money(aff.montant),
                aff.commentaire or "",
            ])

    _xlsx_write_table(
        wb,
        "Affectations dépenses",
        ["Date dépense", "Dépense", "Montant dépense", "Type source", "Source", "Ligne", "Montant affecté", "Commentaire"],
        aff_rows,
        "AffectationsDep",
        money_headers={"Montant dépense", "Montant affecté"},
    )

    # Contrôles consolidés
    controles = []
    for c in data.get("controls", []):
        controles.append([c.get("level", ""), c.get("zone", ""), c.get("message", ""), c.get("action", "")])
    if not controles:
        controles.append(["ok", "Global", "Aucun contrôle bloquant détecté.", ""])

    _xlsx_write_table(
        wb,
        "Contrôles",
        ["Niveau", "Zone", "Message", "Action conseillée"],
        controles,
        "Controles",
    )

    # Supprimer l'éventuelle feuille vide si créée en trop : non, Lecture rapide est active.
    wb.active = 0
    return wb



# ---------------------------------------------------------------------
# P0.2I — Vue simple + dépense financée
# ---------------------------------------------------------------------

def _line_available_amount(ligne) -> float:
    try:
        return _money(getattr(ligne, "reste", 0))
    except Exception:
        return 0.0


def _validate_line_capacity_for_new_affectations(line_amounts: dict[int, float]) -> tuple[bool, str]:
    """Empêche de créer/imputer au-delà du reste disponible d'une ligne.

    C'est un garde-fou métier : une dépense peut être répartie entre plusieurs
    enveloppes, mais chaque enveloppe/ligne ne doit pas être surconsommée.
    """
    for line_id, amount in line_amounts.items():
        ligne = LigneBudget.query.get(line_id)
        if not ligne:
            return False, "Une ligne de financement est introuvable."
        available = _line_available_amount(ligne)
        if amount > available + 0.01:
            sub_name = ligne.source_sub.nom if getattr(ligne, "source_sub", None) else "Enveloppe inconnue"
            return False, (
                f"Budget insuffisant sur {sub_name} — {ligne.compte} {ligne.libelle} : "
                f"tu essaies d'imputer {amount:.2f} €, il reste {available:.2f} €."
            )
    return True, ""


def _finance_add_control(controls: list[dict], level: str, zone: str, message: str, action: str = ""):
    controls.append({
        "level": level,
        "zone": zone,
        "message": message,
        "action": action,
    })


def _simple_finance_sources(data: dict) -> list[dict]:
    rows = []
    seen_line_ids = set()

    for r in data.get("subventions", []):
        sub = r.get("subvention")
        if not sub:
            continue

        type_label = "Fonds propres / autofinancement" if r.get("is_fonds_propres") else "Subvention externe"
        for ligne in getattr(sub, "lignes", []) or []:
            if getattr(ligne, "nature", "charge") != "charge":
                continue
            if ligne.id in seen_line_ids:
                continue
            seen_line_ids.add(ligne.id)
            rows.append({
                "ligne": ligne,
                "subvention": sub,
                "type": type_label,
                "label": f"{sub.nom} — {ligne.compte} · {ligne.libelle}",
                "reste": _money(getattr(ligne, "reste", 0)),
                "is_fonds_propres": bool(r.get("is_fonds_propres")),
            })

    return sorted(rows, key=lambda x: (x["is_fonds_propres"], x["subvention"].nom.lower(), x["ligne"].compte or ""))


def _create_depense_from_simple_form(secteur: str, year: int):
    if not can("depenses:create"):
        abort(403)

    libelle = (request.form.get("libelle") or "").strip()
    montant = _parse_float_fr(request.form.get("montant"), 0.0)
    if not libelle:
        flash("Libellé obligatoire pour créer une dépense.", "danger")
        return False, None
    if montant <= 0:
        flash("Le montant de la dépense doit être supérieur à 0 €.", "danger")
        return False, None

    line_ids = request.form.getlist("ligne_budget_id[]") or request.form.getlist("ligne_budget_id")
    amounts = request.form.getlist("montant_affecte[]") or request.form.getlist("montant_affecte")
    comments = request.form.getlist("commentaire_affectation[]") or request.form.getlist("commentaire_affectation")

    affectations = []
    total_affecte = 0.0
    first_line = None

    for idx, raw_line_id in enumerate(line_ids):
        raw_amount = amounts[idx] if idx < len(amounts) else ""
        if not str(raw_amount or "").strip():
            continue
        amount = _parse_float_fr(raw_amount, 0.0)
        if amount <= 0:
            continue

        try:
            line_id = int(raw_line_id or 0)
        except Exception:
            line_id = 0

        ligne = LigneBudget.query.get(line_id)
        if not ligne or getattr(ligne, "nature", "charge") != "charge":
            flash("Une ligne de répartition est invalide.", "danger")
            return False, None

        sub = getattr(ligne, "source_sub", None)
        if not sub or sub.secteur != secteur or int(sub.annee_exercice or 0) != int(year):
            flash("Une ligne de répartition ne correspond pas au secteur / exercice sélectionné.", "danger")
            return False, None

        if not can_access_secteur(sub.secteur):
            abort(403)

        if first_line is None:
            first_line = ligne

        comment = (comments[idx] if idx < len(comments) else "").strip() or None
        affectations.append((ligne, amount, comment))
        total_affecte = round(total_affecte + amount, 2)

    if not affectations:
        flash("Ajoute au moins une source de financement pour cette dépense.", "danger")
        return False, None

    if total_affecte > montant + 0.01:
        flash(f"La répartition dépasse la dépense : {total_affecte:.2f} € affectés pour {montant:.2f} €.", "danger")
        return False, None

    auto_complete = (request.form.get("auto_complete_reste") or "").strip() == "1"
    if total_affecte < montant - 0.01 and auto_complete:
        remaining = round(montant - total_affecte, 2)
        affectations.append((first_line, remaining, "Complément automatique sur la première source"))
        total_affecte = round(total_affecte + remaining, 2)

    line_amounts = {}
    for ligne, amount, _comment in affectations:
        line_amounts[ligne.id] = round(float(line_amounts.get(ligne.id, 0.0)) + float(amount or 0), 2)
    ok_capacity, capacity_msg = _validate_line_capacity_for_new_affectations(line_amounts)
    if not ok_capacity:
        flash(capacity_msg, "danger")
        return False, None

    dep = Depense(
        ligne_budget_id=first_line.id,
        libelle=libelle,
        montant=montant,
        date_paiement=_parse_date_or_none(request.form.get("date_paiement")),
        type_depense=(request.form.get("type_depense") or "Fonctionnement").strip(),
        fournisseur=(request.form.get("fournisseur") or "").strip() or None,
        reference_piece=(request.form.get("reference_piece") or "").strip() or None,
        mode_paiement=(request.form.get("mode_paiement") or "").strip() or None,
    )
    db.session.add(dep)
    db.session.flush()

    for ligne, amount, comment in affectations:
        sub = ligne.source_sub
        db.session.add(DepenseAffectation(
            depense_id=dep.id,
            source_type="subvention",
            subvention_id=sub.id,
            ligne_budget_id=ligne.id,
            montant=amount,
            commentaire=comment,
        ))

    db.session.commit()

    if abs(total_affecte - montant) <= 0.01:
        flash("Dépense financée créée et équilibrée.", "success")
    else:
        flash(f"Dépense créée, mais il reste {round(montant - total_affecte, 2):.2f} € à affecter.", "warning")

    return True, dep.id


@bp.route("/finance")
@login_required
@require_perm("projets:view")
def finance_home():
    sectors, selected_sector, years, year = _selected_sector_year()
    if not selected_sector:
        flash("Aucun secteur accessible pour les finances.", "warning")
        return redirect(url_for("projets.projets_list"))
    data = _sector_finance_context(selected_sector, year)
    return render_template(
        "finance_home.html",
        sectors=sectors,
        secteur=selected_sector,
        years=years,
        year=year,
        data=data,
    )


@bp.route("/finance/simple", methods=["GET", "POST"])
@login_required
@require_perm("projets:view")
def finance_simple():
    sectors, selected_sector, years, year = _selected_sector_year()
    if not selected_sector:
        flash("Aucun secteur accessible pour la finance simplifiée.", "warning")
        return redirect(url_for("projets.projets_list"))

    if request.method == "POST":
        ok, dep_id = _create_depense_from_simple_form(selected_sector, year)
        if ok and dep_id:
            return redirect(url_for("budget.depense_edit", depense_id=dep_id))
        return redirect(url_for("projets.finance_simple", secteur=selected_sector, year=year))

    data = _sector_finance_context(selected_sector, year)
    sources = _simple_finance_sources(data)

    return render_template(
        "finance_simple.html",
        sectors=sectors,
        secteur=selected_sector,
        years=years,
        year=year,
        data=data,
        sources=sources,
    )



@bp.route("/finance/secteur/export.xlsx")
@login_required
@require_perm("projets:view")
def finance_secteur_export():
    sectors, selected_sector, years, year = _selected_sector_year()
    if not selected_sector:
        abort(404)

    data = _sector_finance_context(selected_sector, year)
    wb = _build_sector_finance_workbook(selected_sector, year, data)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_secteur = "".join(ch if ch.isalnum() else "_" for ch in str(selected_sector)).strip("_") or "secteur"
    filename = f"pilotage_financier_{year}_{safe_secteur}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )



@bp.route("/finance/secteur")
@login_required
@require_perm("projets:view")
def finance_secteur():
    sectors, selected_sector, years, year = _selected_sector_year()
    if not selected_sector:
        flash("Aucun secteur accessible pour le pilotage financier.", "warning")
        return redirect(url_for("projets.projets_list"))

    tab = (request.args.get("tab") or "vue").strip().lower()
    if tab not in {"vue", "budget", "financeurs", "projets", "depenses", "controles"}:
        tab = "vue"

    data = _sector_finance_context(selected_sector, year)
    return render_template(
        "finance_secteur.html",
        sectors=sectors,
        secteur=selected_sector,
        years=years,
        year=year,
        tab=tab,
        data=data,
    )



def _projet_finance_context(projet: Projet, year: int) -> dict:
    # Budget prévisionnel global filtré sur ce projet
    prev_lines = (
        BudgetPrevisionnelLigne.query
        .join(BudgetPrevisionnel, BudgetPrevisionnelLigne.budget_id == BudgetPrevisionnel.id)
        .filter(
            BudgetPrevisionnelLigne.projet_id == projet.id,
            BudgetPrevisionnel.secteur == projet.secteur,
            BudgetPrevisionnel.annee == year,
        )
        .order_by(BudgetPrevisionnelLigne.nature.asc(), BudgetPrevisionnelLigne.compte.asc(), BudgetPrevisionnelLigne.ordre.asc(), BudgetPrevisionnelLigne.id.asc())
        .all()
    )

    prev_charges = _money(sum(float(l.montant or 0) for l in prev_lines if l.nature == "charge"))
    prev_produits = _money(sum(float(l.montant or 0) for l in prev_lines if l.nature == "produit"))

    # Ancien budget AAP projet, conservé pour compatibilité
    aap_charges = list(getattr(projet, "charges_projet", []) or [])
    aap_produits = list(getattr(projet, "produits_projet", []) or [])
    aap_charges_prevues = _money(sum(float(c.montant_previsionnel or 0) for c in aap_charges))
    aap_charges_reelles = _money(sum(float(c.montant_reel or 0) for c in aap_charges))
    aap_produits_demandes = _money(sum(float(p.montant_demande or 0) for p in aap_produits))
    aap_produits_accordes = _money(sum(float(p.montant_accorde or 0) for p in aap_produits))
    aap_produits_recus = _money(sum(float(p.montant_recu or 0) for p in aap_produits))
    aap_ventile = _money(sum(float(c.ventile or 0) for c in aap_charges))

    # Base principale : prévisionnel si renseigné, sinon ancien budget projet
    charges_prevues = prev_charges if prev_charges > 0 else aap_charges_prevues
    produits_prevus = prev_produits if prev_produits > 0 else aap_produits_demandes

    subventions = [
        link.subvention for link in getattr(projet, "subventions", []) or []
        if getattr(link, "subvention", None) and int(link.subvention.annee_exercice or 0) == int(year)
    ]
    subventions = sorted(subventions, key=lambda s: (s.nom or "").lower())

    sub_demande = _money(sum(float(s.montant_demande or 0) for s in subventions))
    sub_attribue = _money(sum(float(s.montant_attribue or 0) for s in subventions))
    sub_recu = _money(sum(float(s.montant_recu or 0) for s in subventions))

    dossiers = (
        AppelProjetBudget.query
        .join(BudgetPrevisionnel, AppelProjetBudget.budget_id == BudgetPrevisionnel.id)
        .filter(
            AppelProjetBudget.projet_id == projet.id,
            BudgetPrevisionnel.annee == year,
        )
        .order_by(AppelProjetBudget.created_at.desc(), AppelProjetBudget.id.desc())
        .all()
    )

    depenses = _collect_project_depenses(projet, subventions)
    depenses_total = _money(sum(float(d.montant or 0) for d in depenses))

    linked_subvention_ids = {s.id for s in subventions}
    available_subventions = (
        Subvention.query
        .filter(
            Subvention.secteur == projet.secteur,
            Subvention.annee_exercice == year,
            Subvention.est_archive.is_(False),
        )
        .order_by(Subvention.nom.asc())
        .all()
    )
    available_subventions = [s for s in available_subventions if s.id not in linked_subvention_ids]

    subvention_charge_lines = []
    for sub in subventions:
        for ligne in getattr(sub, "lignes", []) or []:
            if getattr(ligne, "nature", "charge") == "charge":
                subvention_charge_lines.append(ligne)

    # Dépenses saisies via subvention mais non rattachées à une charge projet : utile à repérer
    depenses_non_affectees = [
        d for d in depenses
        if not getattr(d, "charge_projet_id", None)
    ]

    financement_reference = sub_attribue if sub_attribue > 0 else (sub_demande if sub_demande > 0 else produits_prevus)
    encaisse_reference = sub_recu if sub_recu > 0 else aap_produits_recus

    reste_a_financer = _money(charges_prevues - financement_reference)
    reste_a_encaisser = _money(sub_attribue - sub_recu)
    reste_a_depenser = _money(charges_prevues - depenses_total)

    alerts = []
    if charges_prevues <= 0:
        alerts.append({"level": "warn", "text": "Aucun budget de charges prévu n'est détecté pour ce projet sur l'exercice sélectionné."})
    if produits_prevus > 0 and produits_prevus < charges_prevues:
        alerts.append({"level": "warn", "text": f"Produits prévus inférieurs aux charges prévues : reste théorique {_money(charges_prevues - produits_prevus):.2f} €."})
    if sub_attribue > 0 and sub_recu < sub_attribue:
        alerts.append({"level": "warn", "text": f"Subventions accordées non totalement reçues : reste à encaisser {reste_a_encaisser:.2f} €."})
    if depenses_total > charges_prevues and charges_prevues > 0:
        alerts.append({"level": "danger", "text": f"Dépenses engagées supérieures au budget prévu de {_money(depenses_total - charges_prevues):.2f} €."})
    if depenses_non_affectees:
        alerts.append({"level": "info", "text": f"{len(depenses_non_affectees)} dépense(s) sont rattachées à une subvention mais pas encore à une charge projet."})
    if not subventions and not dossiers:
        alerts.append({"level": "info", "text": "Aucun financement ou dossier financeur rattaché à ce projet pour l'exercice sélectionné."})

    if not alerts:
        alerts.append({"level": "ok", "text": "Aucune alerte bloquante détectée sur ce dossier financier."})

    kpis = {
        "charges_prevues": charges_prevues,
        "produits_prevus": produits_prevus,
        "financement_reference": _money(financement_reference),
        "sub_demande": sub_demande,
        "sub_attribue": sub_attribue,
        "sub_recu": sub_recu,
        "depenses_total": depenses_total,
        "reste_a_financer": reste_a_financer,
        "reste_a_encaisser": reste_a_encaisser,
        "reste_a_depenser": reste_a_depenser,
        "solde_previsionnel": _money(produits_prevus - charges_prevues),
        "solde_execution": _money(encaisse_reference - depenses_total),
        "pct_financement": _pct(financement_reference, charges_prevues),
        "pct_recu": _pct(sub_recu, sub_attribue if sub_attribue > 0 else financement_reference),
        "pct_depenses": _pct(depenses_total, charges_prevues),
        "nb_subventions": len(subventions),
        "nb_dossiers": len(dossiers),
        "nb_depenses": len(depenses),
        "nb_depenses_non_affectees": len(depenses_non_affectees),
        "aap_charges_prevues": aap_charges_prevues,
        "aap_charges_reelles": aap_charges_reelles,
        "aap_produits_demandes": aap_produits_demandes,
        "aap_produits_accordes": aap_produits_accordes,
        "aap_produits_recus": aap_produits_recus,
        "aap_ventile": aap_ventile,
    }

    return {
        "year": year,
        "years": _projet_finance_years(projet),
        "kpis": kpis,
        "alerts": alerts,
        "prev_lines": prev_lines,
        "aap_charges": aap_charges,
        "aap_produits": aap_produits,
        "subventions": subventions,
        "available_subventions": available_subventions,
        "subvention_charge_lines": subvention_charge_lines,
        "dossiers": dossiers,
        "depenses": depenses,
        "depenses_non_affectees": depenses_non_affectees,
    }


def ensure_projets_folder():
    return ensure_upload_subdir("projets")

def allowed_cr(filename: str) -> bool:
    if not filename or "." not in filename:
        return False
    ext = filename.rsplit(".", 1)[1].lower()
    return ext in ALLOWED_CR

@bp.route("/projets")
@login_required
@require_perm("projets:view")
def projets_list():
    q = Projet.query
    if not can("scope:all_secteurs"):
        q = q.filter(Projet.secteur == current_user.secteur_assigne)

    projets = q.order_by(Projet.created_at.desc()).all()
    secteurs = current_app.config.get("SECTEURS", [])
    return render_template("projets_list.html", projets=projets, secteurs=secteurs)

@bp.route("/projets/new", methods=["GET", "POST"])
@login_required
@require_perm("projets:edit")
def projets_new():
    secteurs = current_app.config.get("SECTEURS", [])

    if request.method == "POST":
        nom = (request.form.get("nom") or "").strip()
        secteur = (request.form.get("secteur") or "").strip()
        description = (request.form.get("description") or "").strip()

        if not can("scope:all_secteurs"):
            secteur = current_user.secteur_assigne

        if not nom or not secteur:
            flash("Le nom du projet et le secteur sont obligatoires.", "danger")
            return redirect(url_for("projets.projets_new"))

        if not can_see_secteur(secteur):
            abort(403)

        p = Projet(nom=nom, secteur=secteur, description=description)
        db.session.add(p)
        db.session.commit()

        flash("Le projet a bien été créé.", "success")
        return redirect(url_for("projets.projets_edit", projet_id=p.id))

    return render_template("projets_new.html", secteurs=secteurs)


@bp.route("/projets/<int:projet_id>", methods=["GET", "POST"])
@login_required
@require_perm("projets:view")
def projets_edit(projet_id):
    p = Projet.query.get_or_404(projet_id)
    if not can_see_secteur(p.secteur):
        abort(403)

    if request.method == "POST":
        if not can("projets:edit"):
            abort(403)
        action = request.form.get("action") or ""

        if action == "update":
            p.nom = (request.form.get("nom") or "").strip()
            p.description = (request.form.get("description") or "").strip()

            if not p.nom:
                flash("Le nom du projet est obligatoire.", "danger")
                return redirect(url_for("projets.projets_edit", projet_id=p.id))

            db.session.commit()
            flash("Les informations du projet ont bien été enregistrées.", "success")
            return redirect(url_for("projets.projets_edit", projet_id=p.id))

        if action == "update_competences":
            competence_ids = [int(cid) for cid in request.form.getlist("competence_ids") if cid.isdigit()]
            if competence_ids:
                p.competences = Competence.query.filter(Competence.id.in_(competence_ids)).all()
            else:
                p.competences = []
            db.session.commit()
            flash("Compétences du projet mises à jour.", "success")
            return redirect(url_for("projets.projets_edit", projet_id=p.id))

        if action == "upload_cr":
            if not can("projets:files"):
                abort(403)
            file = request.files.get("cr_file")
            if not file or not file.filename:
                flash("Aucun fichier.", "danger")
                return redirect(url_for("projets.projets_edit", projet_id=p.id))

            if not allowed_cr(file.filename):
                flash("Type autorisé : pdf/doc/docx/odt", "danger")
                return redirect(url_for("projets.projets_edit", projet_id=p.id))

            folder = ensure_projets_folder()
            safe_original = secure_filename(file.filename)
            ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            stored = secure_filename(f"P{p.id}_{ts}_{safe_original}")
            file.save(os.path.join(folder, stored))

            p.cr_filename = stored
            p.cr_original_name = safe_original
            db.session.commit()

            flash("Le compte-rendu a bien été importé.", "success")
            return redirect(url_for("projets.projets_edit", projet_id=p.id))

        if action == "toggle_subvention":
            if not can("subventions:link"):
                abort(403)
            sub_id = int(request.form.get("subvention_id") or 0)
            s = Subvention.query.get_or_404(sub_id)

            if s.secteur != p.secteur:
                abort(400)

            link = SubventionProjet.query.filter_by(projet_id=p.id, subvention_id=s.id).first()
            if link:
                db.session.delete(link)
                db.session.commit()
                flash("La subvention a été retirée du projet.", "warning")
            else:
                db.session.add(SubventionProjet(projet_id=p.id, subvention_id=s.id))
                db.session.commit()
                flash("La subvention a été rattachée au projet.", "success")

            return redirect(url_for("projets.projets_edit", projet_id=p.id))

        # ---- Liens projet <-> ateliers ----
        if action == "toggle_atelier":
            if not can("ateliers:edit"):
                abort(403)
            atelier_id = int(request.form.get("atelier_id") or 0)
            a = AtelierActivite.query.get_or_404(atelier_id)
            if a.secteur != p.secteur or a.is_deleted:
                abort(400)

            link = ProjetAtelier.query.filter_by(projet_id=p.id, atelier_id=a.id).first()
            if link:
                db.session.delete(link)
                db.session.commit()
                flash("L'atelier a été retiré du projet.", "warning")
            else:
                db.session.add(ProjetAtelier(projet_id=p.id, atelier_id=a.id))
                db.session.commit()
                flash("L'atelier a été rattaché au projet.", "success")

            return redirect(url_for("projets.projets_edit", projet_id=p.id))

        # ---- Indicateurs projet ----

        if action == "add_pack":
            pack = (request.form.get("pack") or "").strip()
            cfg = INDICATOR_PACKS.get(pack)
            if not cfg:
                flash("Le pack sélectionné est invalide.", "danger")
                return redirect(url_for("projets.projets_edit", projet_id=p.id))

            added = 0
            for code in cfg["codes"]:
                if code not in INDICATOR_TEMPLATES:
                    continue
                exists = ProjetIndicateur.query.filter_by(projet_id=p.id, code=code).first()
                if exists:
                    continue
                db.session.add(ProjetIndicateur(
                    projet_id=p.id,
                    code=code,
                    label=INDICATOR_TEMPLATES.get(code, code),
                    is_active=True,
                    params_json=None,
                ))
                added += 1
            db.session.commit()
            flash(f"Le pack a bien été ajouté ({added} indicateur(s)).", "success")
            return redirect(url_for("projets.projets_edit", projet_id=p.id))

        if action == "add_indicateur":
            code = (request.form.get("code") or "").strip()
            label = (request.form.get("label") or "").strip()
            if code not in INDICATOR_TEMPLATES:
                flash("L'indicateur sélectionné est invalide.", "danger")
                return redirect(url_for("projets.projets_edit", projet_id=p.id))
            if not label:
                label = INDICATOR_TEMPLATES[code]

            exists = ProjetIndicateur.query.filter_by(projet_id=p.id, code=code).first()
            if exists:
                flash("Cet indicateur est déjà présent sur ce projet.", "warning")
                return redirect(url_for("projets.projets_edit", projet_id=p.id))

            db.session.add(ProjetIndicateur(projet_id=p.id, code=code, label=label, is_active=True))
            db.session.commit()
            flash("L'indicateur a bien été ajouté.", "success")
            return redirect(url_for("projets.projets_edit", projet_id=p.id))

        if action == "toggle_indicateur":
            indic_id = int(request.form.get("indicateur_id") or 0)
            ind = ProjetIndicateur.query.get_or_404(indic_id)
            if ind.projet_id != p.id:
                abort(400)
            ind.is_active = not bool(ind.is_active)
            db.session.commit()
            flash("L'indicateur a bien été mis à jour.", "success")
            return redirect(url_for("projets.projets_edit", projet_id=p.id))

        if action == "save_indicateur":
            indic_id = int(request.form.get("indicateur_id") or 0)
            ind = ProjetIndicateur.query.get_or_404(indic_id)
            if ind.projet_id != p.id:
                abort(400)

            # label editable (optionnel)
            label = (request.form.get("label") or "").strip()
            if label:
                ind.label = label

            period = (request.form.get("period") or "context").strip()
            if period not in PERIOD_CHOICES:
                period = "context"

            target_raw = (request.form.get("target") or "").strip().replace(",", ".")
            target = None
            if target_raw:
                try:
                    target = float(target_raw)
                except ValueError:
                    target = None

            target_op = (request.form.get("target_op") or "ge").strip()
            if target_op not in TARGET_OP_CHOICES:
                target_op = "ge"

            atelier_id_raw = (request.form.get("atelier_id") or "").strip()
            atelier_id = None
            if atelier_id_raw:
                try:
                    atelier_id = int(atelier_id_raw)
                except ValueError:
                    atelier_id = None

            # bornes custom
            start = (request.form.get("start") or "").strip()
            end = (request.form.get("end") or "").strip()

            params = ind.params()
            params.update({
                "period": period,
                "target": target,
                "target_op": target_op,
                "atelier_id": atelier_id,
                "start": start if period == "custom" else None,
                "end": end if period == "custom" else None,
            })
            # nettoyage
            if params.get("atelier_id") is None:
                params.pop("atelier_id", None)
            if params.get("target") is None:
                params.pop("target", None)
            if period != "custom":
                params.pop("start", None)
                params.pop("end", None)

            ind.params_json = json.dumps(params, ensure_ascii=False)
            db.session.commit()
            flash("Les paramètres de l'indicateur ont bien été enregistrés.", "success")
            return redirect(url_for("projets.projets_edit", projet_id=p.id))


        if action == "delete_indicateur":
            indic_id = int(request.form.get("indicateur_id") or 0)
            ind = ProjetIndicateur.query.get_or_404(indic_id)
            if ind.projet_id != p.id:
                abort(400)
            db.session.delete(ind)
            db.session.commit()
            flash("L'indicateur a été supprimé.", "warning")
            return redirect(url_for("projets.projets_edit", projet_id=p.id))

        abort(400)

    # ----- GET (lists) -----
    subs_q = Subvention.query.filter_by(est_archive=False).filter(Subvention.secteur == p.secteur)
    subs = subs_q.order_by(Subvention.annee_exercice.desc(), Subvention.nom.asc()).all()
    linked_subs = set(sp.subvention_id for sp in p.subventions)

    ateliers = AtelierActivite.query.filter_by(secteur=p.secteur, is_deleted=False).order_by(AtelierActivite.nom.asc()).all()
    linked_ateliers = set(link.atelier_id for link in ProjetAtelier.query.filter_by(projet_id=p.id).all())

    indicateurs = ProjetIndicateur.query.filter_by(projet_id=p.id).order_by(ProjetIndicateur.created_at.asc()).all()
    referentiels = Referentiel.query.order_by(Referentiel.nom.asc()).all()
    selected_competences = {c.id for c in p.competences}

    return render_template(
        "projets_edit.html",
        projet=p,
        subs=subs,
        linked=linked_subs,
        ateliers=ateliers,
        linked_ateliers=linked_ateliers,
        indicateurs=indicateurs,
        indicator_templates=INDICATOR_TEMPLATES,
        indicator_metric_groups=INDICATOR_METRIC_GROUPS,
        indicator_packs=INDICATOR_PACKS,
        period_choices=PERIOD_CHOICES,
        target_op_choices=TARGET_OP_CHOICES,
        referentiels=referentiels,
        selected_competences=selected_competences,
    )


@bp.route("/projets/<int:projet_id>/indicateurs", methods=["GET", "POST"])
@login_required
@require_perm("projets:view")
def projet_indicateurs(projet_id):
    p = Projet.query.get_or_404(projet_id)
    if not can_see_secteur(p.secteur):
        abort(403)

    ateliers = AtelierActivite.query.filter_by(secteur=p.secteur, is_deleted=False).order_by(AtelierActivite.nom.asc()).all()
    linked_atelier_ids = {
        link.atelier_id for link in ProjetAtelier.query.filter_by(projet_id=p.id).all()
    }

    def clean_target_params(params: dict) -> dict:
        target = parse_float_optional(request.form.get("target"))
        target_op = (request.form.get("target_op") or "ge").strip()
        if target_op not in TARGET_OP_CHOICES:
            target_op = "ge"
        params["target_op"] = target_op
        if target is None:
            params.pop("target", None)
        else:
            params["target"] = target
        return params

    def clean_period_params(params: dict) -> dict:
        period = (request.form.get("period") or "context").strip()
        if period not in PERIOD_CHOICES:
            period = "context"
        params["period"] = period
        if period == "custom":
            params["start"] = (request.form.get("start") or "").strip() or None
            params["end"] = (request.form.get("end") or "").strip() or None
        else:
            params.pop("start", None)
            params.pop("end", None)

        atelier_id = None
        raw = (request.form.get("atelier_id") or "").strip()
        if raw:
            try:
                candidate = int(raw)
                if candidate in linked_atelier_ids:
                    atelier_id = candidate
            except Exception:
                atelier_id = None
        if atelier_id:
            params["atelier_id"] = atelier_id
        else:
            params.pop("atelier_id", None)
        return params

    if request.method == "POST":
        if not can("projets:edit"):
            abort(403)

        action = request.form.get("action") or ""

        if action == "add_manual_number":
            label = (request.form.get("label") or "").strip()
            if not label:
                flash("Le libellé de l'indicateur est obligatoire.", "danger")
                return redirect(url_for("projets.projet_indicateurs", projet_id=p.id))

            params = {
                "source": "manual",
                "value_type": "number",
                "manual_value": parse_float_optional(request.form.get("manual_value")),
                "unit": (request.form.get("unit") or "").strip(),
            }
            clean_target_params(params)
            db.session.add(ProjetIndicateur(
                projet_id=p.id,
                code=indicator_unique_code("manual_number"),
                label=label,
                is_active=True,
                params_json=json.dumps(params, ensure_ascii=False),
            ))
            db.session.commit()
            flash("Indicateur manuel ajouté.", "success")
            return redirect(url_for("projets.projet_indicateurs", projet_id=p.id))

        if action == "add_manual_check":
            label = (request.form.get("label") or "").strip()
            if not label:
                flash("Le libellé de la coche est obligatoire.", "danger")
                return redirect(url_for("projets.projet_indicateurs", projet_id=p.id))

            params = {
                "source": "manual",
                "value_type": "check",
                "checked": bool(request.form.get("checked")),
            }
            db.session.add(ProjetIndicateur(
                projet_id=p.id,
                code=indicator_unique_code("manual_check"),
                label=label,
                is_active=True,
                params_json=json.dumps(params, ensure_ascii=False),
            ))
            db.session.commit()
            flash("Coche de validation ajoutée.", "success")
            return redirect(url_for("projets.projet_indicateurs", projet_id=p.id))

        if action == "add_stats_indicator":
            metric_code = (request.form.get("metric_code") or request.form.get("code") or "").strip()
            if metric_code not in INDICATOR_TEMPLATES:
                flash("La statistique sélectionnée est invalide.", "danger")
                return redirect(url_for("projets.projet_indicateurs", projet_id=p.id))

            label = (request.form.get("label") or "").strip() or INDICATOR_TEMPLATES[metric_code]
            params = {
                "source": "stats",
                "value_type": "number",
                "metric_code": metric_code,
            }
            clean_target_params(params)
            clean_period_params(params)
            db.session.add(ProjetIndicateur(
                projet_id=p.id,
                code=indicator_unique_code(f"stat_{metric_code}"),
                label=label,
                is_active=True,
                params_json=json.dumps(params, ensure_ascii=False),
            ))
            db.session.commit()
            flash("Indicateur connecté aux stats-impact ajouté.", "success")
            return redirect(url_for("projets.projet_indicateurs", projet_id=p.id))

        if action == "add_pack":
            pack = (request.form.get("pack") or "").strip()
            cfg = INDICATOR_PACKS.get(pack)
            if not cfg:
                flash("Le pack sélectionné est invalide.", "danger")
                return redirect(url_for("projets.projet_indicateurs", projet_id=p.id))

            existing_metrics = set()
            for ind in ProjetIndicateur.query.filter_by(projet_id=p.id).all():
                params = indicator_params(ind)
                existing_metrics.add(indicator_metric_code(ind, params) or ind.code)

            added = 0
            for metric_code in cfg["codes"]:
                if metric_code not in INDICATOR_TEMPLATES or metric_code in existing_metrics:
                    continue
                params = {
                    "source": "stats",
                    "value_type": "number",
                    "metric_code": metric_code,
                    "period": "context",
                    "target_op": "ge",
                }
                db.session.add(ProjetIndicateur(
                    projet_id=p.id,
                    code=indicator_unique_code(f"stat_{metric_code}"),
                    label=INDICATOR_TEMPLATES[metric_code],
                    is_active=True,
                    params_json=json.dumps(params, ensure_ascii=False),
                ))
                added += 1
            db.session.commit()
            flash(f"Le pack a bien été ajouté ({added} indicateur(s)).", "success")
            return redirect(url_for("projets.projet_indicateurs", projet_id=p.id))

        if action == "update_indicateur":
            indic_id = int(request.form.get("indicateur_id") or 0)
            ind = ProjetIndicateur.query.get_or_404(indic_id)
            if ind.projet_id != p.id:
                abort(400)

            params = indicator_params(ind)
            source = indicator_source(ind, params)
            value_type = indicator_value_type(ind, params)
            label = (request.form.get("label") or "").strip()
            if label:
                ind.label = label
            ind.is_active = bool(request.form.get("is_active"))

            if source == "stats":
                metric_code = (request.form.get("metric_code") or indicator_metric_code(ind, params)).strip()
                if metric_code not in INDICATOR_TEMPLATES:
                    flash("La statistique sélectionnée est invalide.", "danger")
                    return redirect(url_for("projets.projet_indicateurs", projet_id=p.id))
                params.update({
                    "source": "stats",
                    "value_type": "number",
                    "metric_code": metric_code,
                })
                clean_target_params(params)
                clean_period_params(params)
            elif value_type == "check":
                params = {
                    "source": "manual",
                    "value_type": "check",
                    "checked": bool(request.form.get("checked")),
                }
            else:
                params.update({
                    "source": "manual",
                    "value_type": "number",
                    "manual_value": parse_float_optional(request.form.get("manual_value")),
                    "unit": (request.form.get("unit") or "").strip(),
                })
                clean_target_params(params)

            ind.params_json = json.dumps(params, ensure_ascii=False)
            db.session.commit()
            flash("Indicateur enregistré.", "success")
            return redirect(url_for("projets.projet_indicateurs", projet_id=p.id))

        if action == "delete_indicateur":
            indic_id = int(request.form.get("indicateur_id") or 0)
            ind = ProjetIndicateur.query.get_or_404(indic_id)
            if ind.projet_id != p.id:
                abort(400)
            db.session.delete(ind)
            db.session.commit()
            flash("Indicateur supprimé.", "warning")
            return redirect(url_for("projets.projet_indicateurs", projet_id=p.id))

        abort(400)

    indicateurs = ProjetIndicateur.query.filter_by(projet_id=p.id).order_by(ProjetIndicateur.created_at.asc()).all()
    indicator_rows = indicator_list_rows(indicateurs)

    return render_template(
        "projets_indicateurs.html",
        projet=p,
        ateliers=ateliers,
        linked_ateliers=linked_atelier_ids,
        indicator_rows=indicator_rows,
        indicator_counts=indicator_counts(indicator_rows),
        indicator_templates=INDICATOR_TEMPLATES,
        indicator_metric_groups=INDICATOR_METRIC_GROUPS,
        indicator_packs=INDICATOR_PACKS,
        period_choices=PERIOD_CHOICES,
        target_op_choices=TARGET_OP_CHOICES,
        target_op_symbols=TARGET_OP_SYMBOLS,
        can_edit_indicateurs=can("projets:edit"),
    )


@bp.route("/projets/<int:projet_id>/synthese", methods=["GET", "POST"])
@login_required
@require_perm("projets:view")
def projet_synthese(projet_id):
    p = Projet.query.get_or_404(projet_id)
    if not can_see_secteur(p.secteur):
        abort(403)

    year = _projet_selected_year(p)
    if request.method == "POST":
        if not can("projets:edit"):
            abort(403)

        action = request.form.get("action") or ""
        if action == "add_journal":
            entry_date = _parse_date_or_none(request.form.get("entry_date")) or date.today()
            categorie = (request.form.get("categorie") or "fait_marquant").strip()
            if categorie not in PROJET_JOURNAL_CATEGORIES:
                categorie = "fait_marquant"
            titre = (request.form.get("titre") or "").strip() or None
            contenu = (request.form.get("contenu") or "").strip()
            if not contenu:
                flash("Le contenu du journal est obligatoire.", "danger")
                return redirect(url_for("projets.projet_synthese", projet_id=p.id, year=year))

            db.session.add(ProjetJournalEntry(
                projet_id=p.id,
                entry_date=entry_date,
                categorie=categorie,
                titre=titre,
                contenu=contenu,
                created_by_user_id=current_user.id,
            ))
            db.session.commit()
            flash("Note ajoutée au journal de bord.", "success")
            return redirect(url_for("projets.projet_synthese", projet_id=p.id, year=year))

        if action == "delete_journal":
            entry_id = int(request.form.get("entry_id") or 0)
            entry = ProjetJournalEntry.query.get_or_404(entry_id)
            if entry.projet_id != p.id:
                abort(400)
            db.session.delete(entry)
            db.session.commit()
            flash("Note supprimée du journal de bord.", "warning")
            return redirect(url_for("projets.projet_synthese", projet_id=p.id, year=year))

        if action == "add_project_action":
            titre = (request.form.get("titre") or "").strip()
            if not titre:
                flash("Le titre de la fiche action est obligatoire.", "danger")
                return redirect(url_for("projets.projet_synthese", projet_id=p.id, year=year))

            item = ProjetAction(
                projet_id=p.id,
                titre=titre,
                categorie=_clean_choice(request.form.get("categorie"), PROJET_ACTION_CATEGORIES, "atelier"),
                statut=_clean_choice(request.form.get("statut"), PROJET_ACTION_STATUTS, "prevue"),
                date_debut=_parse_date_or_none(request.form.get("date_debut")),
                date_fin=_parse_date_or_none(request.form.get("date_fin")),
                description=(request.form.get("description") or "").strip() or None,
                created_by_user_id=current_user.id,
            )
            db.session.add(item)
            db.session.commit()
            flash("Fiche action créée.", "success")
            return redirect(url_for("projets.projet_action_detail", projet_id=p.id, action_id=item.id, year=year))

        abort(400)

    finance = _projet_finance_context(p, year)
    linked_atelier_ids = [
        row.atelier_id for row in ProjetAtelier.query.filter_by(projet_id=p.id).all()
    ]
    linked_ateliers = []
    if linked_atelier_ids:
        linked_ateliers = (
            AtelierActivite.query
            .filter(AtelierActivite.id.in_(linked_atelier_ids))
            .order_by(AtelierActivite.nom.asc())
            .all()
        )

    flt = StatsFilters(
        secteur=p.secteur,
        atelier_ids=linked_atelier_ids,
        date_from=date(year, 1, 1),
        date_to=date(year, 12, 31),
    )
    if linked_atelier_ids:
        activity_stats = compute_volume_activity_stats(flt)
        freq_stats = compute_participation_frequency_stats(flt)
        demo_stats = compute_demography_stats(flt)
    else:
        activity_stats = {"kpi": {"presences": 0, "uniques": 0, "sessions": 0, "hours_people": 0}}
        freq_stats = {"returning": 0, "returning_rate": 0, "freq_avg": 0}
        demo_stats = {
            "age_avg": None,
            "qpv": {"qpv": 0, "hors_qpv": 0, "inconnu": 0},
            "genre": {},
            "age_buckets": {},
        }

    indicators = compute_project_indicators(p, selected_annee=year, subventions=finance.get("subventions"))
    indicator_summary = {
        "total": len(indicators),
        "ok": sum(1 for row in indicators if row.get("status") == "ok"),
        "warn": sum(1 for row in indicators if row.get("status") == "warn"),
        "bad": sum(1 for row in indicators if row.get("status") == "bad"),
        "pending": sum(1 for row in indicators if not row.get("status")),
    }

    journal_entries = (
        ProjetJournalEntry.query
        .filter_by(projet_id=p.id)
        .order_by(ProjetJournalEntry.entry_date.desc(), ProjetJournalEntry.created_at.desc())
        .all()
    )
    actions = (
        ProjetAction.query
        .filter_by(projet_id=p.id)
        .order_by(ProjetAction.date_debut.desc(), ProjetAction.created_at.desc())
        .all()
    )
    action_rows = _action_list_rows(p, actions, year)
    project_quality = _project_completion_payload(
        p,
        linked_ateliers,
        action_rows,
        indicators,
        journal_entries,
        finance,
    )

    return render_template(
        "projets_synthese.html",
        projet=p,
        year=year,
        years=_projet_finance_years(p),
        finance=finance,
        linked_ateliers=linked_ateliers,
        activity_stats=activity_stats,
        freq_stats=freq_stats,
        demo_stats=demo_stats,
        indicators=indicators,
        indicator_summary=indicator_summary,
        action_rows=action_rows,
        project_quality=project_quality,
        action_categories=PROJET_ACTION_CATEGORIES,
        action_statuts=PROJET_ACTION_STATUTS,
        journal_entries=journal_entries,
        journal_categories=PROJET_JOURNAL_CATEGORIES,
        can_edit_journal=can("projets:edit"),
        can_edit_project=can("projets:edit"),
    )


@bp.route("/projets/<int:projet_id>/fiche")
@login_required
@require_perm("projets:view")
def projet_fiche_document(projet_id):
    p = Projet.query.get_or_404(projet_id)
    if not can_see_secteur(p.secteur):
        abort(403)

    year = _projet_selected_year(p)
    finance = _projet_finance_context(p, year)
    linked_atelier_ids = [
        row.atelier_id for row in ProjetAtelier.query.filter_by(projet_id=p.id).all()
    ]
    linked_ateliers = []
    if linked_atelier_ids:
        linked_ateliers = (
            AtelierActivite.query
            .filter(AtelierActivite.id.in_(linked_atelier_ids))
            .order_by(AtelierActivite.nom.asc())
            .all()
        )

    flt = StatsFilters(
        secteur=p.secteur,
        atelier_ids=linked_atelier_ids,
        date_from=date(year, 1, 1),
        date_to=date(year, 12, 31),
    )
    if linked_atelier_ids:
        activity_stats = compute_volume_activity_stats(flt)
        freq_stats = compute_participation_frequency_stats(flt)
        demo_stats = compute_demography_stats(flt)
    else:
        activity_stats = {"kpi": {"presences": 0, "uniques": 0, "sessions": 0, "hours_people": 0}}
        freq_stats = {"returning": 0, "returning_rate": 0, "freq_avg": 0}
        demo_stats = {
            "age_avg": None,
            "qpv": {"qpv": 0, "hors_qpv": 0, "inconnu": 0},
            "genre": {},
            "age_buckets": {},
        }

    indicators = compute_project_indicators(p, selected_annee=year, subventions=finance.get("subventions"))
    indicator_summary = {
        "total": len(indicators),
        "ok": sum(1 for row in indicators if row.get("status") == "ok"),
        "warn": sum(1 for row in indicators if row.get("status") == "warn"),
        "bad": sum(1 for row in indicators if row.get("status") == "bad"),
        "pending": sum(1 for row in indicators if not row.get("status")),
    }

    journal_entries = (
        ProjetJournalEntry.query
        .filter_by(projet_id=p.id)
        .order_by(ProjetJournalEntry.entry_date.desc(), ProjetJournalEntry.created_at.desc())
        .limit(8)
        .all()
    )
    actions = (
        ProjetAction.query
        .filter_by(projet_id=p.id)
        .order_by(ProjetAction.date_debut.desc(), ProjetAction.created_at.desc())
        .all()
    )
    action_rows = _action_list_rows(p, actions, year)
    project_quality = _project_completion_payload(
        p,
        linked_ateliers,
        action_rows,
        indicators,
        journal_entries,
        finance,
    )

    return render_template(
        "projets_fiche_document.html",
        projet=p,
        year=year,
        years=_projet_finance_years(p),
        finance=finance,
        linked_ateliers=linked_ateliers,
        activity_stats=activity_stats,
        freq_stats=freq_stats,
        demo_stats=demo_stats,
        indicators=indicators,
        indicator_summary=indicator_summary,
        action_rows=action_rows,
        project_quality=project_quality,
        journal_entries=journal_entries,
        journal_categories=PROJET_JOURNAL_CATEGORIES,
        generated_at=datetime.now(),
    )


@bp.route("/projets/<int:projet_id>/actions/<int:action_id>", methods=["GET", "POST"])
@login_required
@require_perm("projets:view")
def projet_action_detail(projet_id, action_id):
    p = Projet.query.get_or_404(projet_id)
    if not can_see_secteur(p.secteur):
        abort(403)
    action = ProjetAction.query.filter_by(id=action_id, projet_id=p.id).first_or_404()
    year = _projet_selected_year(p)

    linked_project_atelier_ids = _project_linked_atelier_ids(p.id)
    ateliers = []
    if linked_project_atelier_ids:
        ateliers = (
            AtelierActivite.query
            .filter(AtelierActivite.id.in_(linked_project_atelier_ids))
            .order_by(AtelierActivite.nom.asc())
            .all()
        )
    action_atelier_ids = set(_action_linked_atelier_ids(action))

    if request.method == "POST":
        if not can("projets:edit"):
            abort(403)

        form_action = request.form.get("action") or ""
        if form_action == "update_project_action":
            titre = (request.form.get("titre") or "").strip()
            if not titre:
                flash("Le titre de la fiche action est obligatoire.", "danger")
                return redirect(url_for("projets.projet_action_detail", projet_id=p.id, action_id=action.id, year=year))

            action.titre = titre
            action.categorie = _clean_choice(request.form.get("categorie"), PROJET_ACTION_CATEGORIES, "atelier")
            action.statut = _clean_choice(request.form.get("statut"), PROJET_ACTION_STATUTS, "prevue")
            action.referent = (request.form.get("referent") or "").strip() or None
            action.date_debut = _parse_date_or_none(request.form.get("date_debut"))
            action.date_fin = _parse_date_or_none(request.form.get("date_fin"))
            action.lieu = (request.form.get("lieu") or "").strip() or None
            action.public_vise = (request.form.get("public_vise") or "").strip() or None
            action.territoire = (request.form.get("territoire") or "").strip() or None
            action.objectifs = (request.form.get("objectifs") or "").strip() or None
            action.description = (request.form.get("description") or "").strip() or None
            action.partenaires_text = (request.form.get("partenaires_text") or "").strip() or None
            action.bilan_qualitatif = (request.form.get("bilan_qualitatif") or "").strip() or None

            selected_ateliers = set()
            for raw in request.form.getlist("atelier_ids"):
                try:
                    atelier_id = int(raw)
                except Exception:
                    continue
                if atelier_id in linked_project_atelier_ids:
                    selected_ateliers.add(atelier_id)

            ProjetActionAtelier.query.filter_by(action_id=action.id).delete(synchronize_session=False)
            for atelier_id in sorted(selected_ateliers):
                db.session.add(ProjetActionAtelier(action_id=action.id, atelier_id=atelier_id))

            db.session.commit()
            flash("Fiche action enregistrée.", "success")
            return redirect(url_for("projets.projet_action_detail", projet_id=p.id, action_id=action.id, year=year))

        if form_action == "delete_project_action":
            db.session.delete(action)
            db.session.commit()
            flash("Fiche action supprimée.", "warning")
            return redirect(url_for("projets.projet_synthese", projet_id=p.id, year=year))

        abort(400)

    stats = _compute_action_stats(p, action, year)
    dmin, dmax = _action_period(action, year)
    action_quality = _action_completion_payload(action, list(action_atelier_ids))

    return render_template(
        "projets_action.html",
        projet=p,
        action=action,
        year=year,
        years=_projet_finance_years(p),
        action_categories=PROJET_ACTION_CATEGORIES,
        action_statuts=PROJET_ACTION_STATUTS,
        ateliers=ateliers,
        action_atelier_ids=action_atelier_ids,
        stats=stats,
        action_quality=action_quality,
        period_start=dmin,
        period_end=dmax,
        can_edit_project=can("projets:edit"),
    )


@bp.route("/projets/<int:projet_id>/actions/<int:action_id>/fiche")
@login_required
@require_perm("projets:view")
def projet_action_fiche_document(projet_id, action_id):
    p = Projet.query.get_or_404(projet_id)
    if not can_see_secteur(p.secteur):
        abort(403)
    action = ProjetAction.query.filter_by(id=action_id, projet_id=p.id).first_or_404()
    year = _projet_selected_year(p)

    linked_project_atelier_ids = _project_linked_atelier_ids(p.id)
    ateliers = []
    if linked_project_atelier_ids:
        ateliers = (
            AtelierActivite.query
            .filter(AtelierActivite.id.in_(linked_project_atelier_ids))
            .order_by(AtelierActivite.nom.asc())
            .all()
        )
    action_atelier_ids = set(_action_linked_atelier_ids(action))
    linked_action_ateliers = [atelier for atelier in ateliers if atelier.id in action_atelier_ids]
    stats = _compute_action_stats(p, action, year)
    period_start, period_end = _action_period(action, year)
    action_quality = _action_completion_payload(action, list(action_atelier_ids))

    return render_template(
        "projets_action_fiche_document.html",
        projet=p,
        action=action,
        year=year,
        years=_projet_finance_years(p),
        action_categories=PROJET_ACTION_CATEGORIES,
        action_statuts=PROJET_ACTION_STATUTS,
        action_atelier_ids=action_atelier_ids,
        linked_action_ateliers=linked_action_ateliers,
        stats=stats,
        action_quality=action_quality,
        period_start=period_start,
        period_end=period_end,
        generated_at=datetime.now(),
    )


@bp.route("/projets/<int:projet_id>/delete", methods=["POST"])
@login_required
@require_perm('projets:delete')
def projets_delete(projet_id: int):
    """Suppression définitive d'un projet (et de ses liens).

    On fait du *delete applicatif* (compatible SQLite/Postgres) :
    - objectifs + table pivot objectif_competence
    - liens projet_atelier
    - liens projet_competence (table pivot)
    - indicateurs
    - liens projet<->subvention (SubventionProjet)
    Puis on supprime le projet.

    ⚠️ Non réversible. (On pourrait faire un soft-delete plus tard si tu veux.)
    """
    projet = Projet.query.get_or_404(projet_id)

    if not can_access_secteur(getattr(projet, "secteur", None)):
        flash("Accès refusé.", "danger")
        return redirect(url_for("projets.projets_list"))

    try:
        # 1) Objectifs liés au projet + pivot objectif_competence
        objectif_ids = [o.id for o in Objectif.query.filter_by(projet_id=projet.id).all()]
        if objectif_ids:
            db.session.execute(
                objectif_competence.delete().where(objectif_competence.c.objectif_id.in_(objectif_ids))
            )
            # suppression des objectifs (les enfants sont aussi dans la liste via projet_id)
            Objectif.query.filter(Objectif.id.in_(objectif_ids)).delete(synchronize_session=False)

        # 2) Liens projet/ateliers
        ProjetAtelier.query.filter_by(projet_id=projet.id).delete(synchronize_session=False)

        # 3) Liens projet/compétences (table pivot)
        db.session.execute(
            projet_competence.delete().where(projet_competence.c.projet_id == projet.id)
        )

        # 4) Indicateurs
        ProjetIndicateur.query.filter_by(projet_id=projet.id).delete(synchronize_session=False)
        ProjetJournalEntry.query.filter_by(projet_id=projet.id).delete(synchronize_session=False)
        action_ids = [a.id for a in ProjetAction.query.filter_by(projet_id=projet.id).all()]
        if action_ids:
            ProjetActionAtelier.query.filter(ProjetActionAtelier.action_id.in_(action_ids)).delete(synchronize_session=False)
            ProjetAction.query.filter(ProjetAction.id.in_(action_ids)).delete(synchronize_session=False)

        # 5) Liens projet/subventions
        SubventionProjet.query.filter_by(projet_id=projet.id).delete(synchronize_session=False)

        # 6) Fichier CR (si présent) — suppression physique seulement après succès BDD
        cr_relpath = None
        try:
            if projet.cr_filename:
                cr_relpath = os.path.join(current_app.instance_path, "cr_projets", projet.cr_filename)
        except Exception:
            cr_relpath = None

        db.session.delete(projet)
        if commit_delete(
            f"le projet « {projet.nom} »",
            "Projet supprimé définitivement.",
            success_category="warning",
            blocked_message=f"Impossible de supprimer le projet « {projet.nom} » : il est encore utilisé ailleurs.",
        ) and cr_relpath:
            try:
                if os.path.exists(cr_relpath):
                    os.remove(cr_relpath)
            except Exception:
                pass
    except Exception as e:
        db.session.rollback()
        flash(f"Suppression impossible : {e}", "danger")

    return redirect(url_for("projets.projets_list"))


@bp.route("/projets/cr/<int:projet_id>/download")
@login_required
@require_perm("projets:files")
def projets_cr_download(projet_id):
    p = Projet.query.get_or_404(projet_id)
    if not can_see_secteur(p.secteur):
        abort(403)

    if not p.cr_filename:
        abort(404)

    folder = ensure_projets_folder()
    return send_media_file(media_relpath("projets", p.cr_filename), as_attachment=True, download_name=(p.cr_original_name or p.cr_filename))

# ---------------------------------------------------------------------
# Budget AAP par projet : Charges / Produits / Ventilation / Synthèse
# ---------------------------------------------------------------------

@bp.route("/projets/<int:projet_id>/finance")
@login_required
@require_perm("projets:view")
def projet_finance(projet_id):
    projet = Projet.query.get_or_404(projet_id)
    if not can_access_secteur(projet.secteur):
        abort(403)

    year = _projet_selected_year(projet)
    tab = (request.args.get("tab") or "vue").strip().lower()
    if tab not in {"vue", "budget", "financements", "depenses", "exports"}:
        tab = "vue"

    finance = _projet_finance_context(projet, year)
    return render_template(
        "projet_finance.html",
        projet=projet,
        finance=finance,
        year=year,
        tab=tab,
        can_edit_budget=can("aap:charges_edit") or can("aap:produits_edit") or can("subventions:create") or can("depenses:create"),
    )



@bp.route("/projets/<int:projet_id>/finance/action", methods=["POST"])
@login_required
@require_perm("projets:view")
def projet_finance_action(projet_id):
    projet = Projet.query.get_or_404(projet_id)
    if not can_access_secteur(projet.secteur):
        abort(403)

    action = request.form.get("action") or ""
    year = request.form.get("year") or request.args.get("year") or date.today().year
    try:
        year = int(year)
    except Exception:
        year = date.today().year

    if action == "quick_charge":
        if not can("aap:charges_edit"):
            abort(403)
        libelle = (request.form.get("libelle") or "").strip()
        if not libelle:
            flash("Libellé de charge obligatoire.", "danger")
            return redirect(url_for("projets.projet_finance", projet_id=projet.id, year=year, tab="budget"))

        charge = ChargeProjet(
            projet_id=projet.id,
            bloc=(request.form.get("bloc") or "directe").strip(),
            code_plan=(request.form.get("code_plan") or "60").strip(),
            libelle=libelle,
            montant_previsionnel=_parse_float_fr(request.form.get("montant_previsionnel"), 0.0),
            montant_reel=_parse_float_fr(request.form.get("montant_reel"), 0.0),
            commentaire=(request.form.get("commentaire") or "").strip() or None,
        )
        db.session.add(charge)
        db.session.commit()
        flash("Charge projet ajoutée.", "success")
        return redirect(url_for("projets.projet_finance", projet_id=projet.id, year=year, tab="budget"))

    if action == "quick_produit":
        if not can("aap:produits_edit"):
            abort(403)
        financeur = (request.form.get("financeur") or "").strip()
        if not financeur:
            flash("Financeur obligatoire.", "danger")
            return redirect(url_for("projets.projet_finance", projet_id=projet.id, year=year, tab="budget"))

        produit = ProduitProjet(
            projet_id=projet.id,
            financeur=financeur,
            categorie=(request.form.get("categorie") or "autre").strip(),
            statut=(request.form.get("statut") or "prevu").strip(),
            montant_demande=_parse_float_fr(request.form.get("montant_demande"), 0.0),
            montant_accorde=_parse_float_fr(request.form.get("montant_accorde"), 0.0),
            montant_recu=_parse_float_fr(request.form.get("montant_recu"), 0.0),
            reference_dossier=(request.form.get("reference_dossier") or "").strip() or None,
            commentaire=(request.form.get("commentaire") or "").strip() or None,
        )
        db.session.add(produit)
        db.session.commit()
        flash("Produit / financement projet ajouté.", "success")
        return redirect(url_for("projets.projet_finance", projet_id=projet.id, year=year, tab="budget"))

    if action == "create_subvention_linked":
        if not can("subventions:edit"):
            abort(403)
        nom = (request.form.get("nom") or "").strip()
        if not nom:
            flash("Nom de la subvention obligatoire.", "danger")
            return redirect(url_for("projets.projet_finance", projet_id=projet.id, year=year, tab="financements"))

        sub = Subvention(
            nom=nom,
            secteur=projet.secteur,
            annee_exercice=year,
            montant_demande=_parse_float_fr(request.form.get("montant_demande"), 0.0),
            montant_attribue=_parse_float_fr(request.form.get("montant_attribue"), 0.0),
            montant_recu=_parse_float_fr(request.form.get("montant_recu"), 0.0),
        )
        db.session.add(sub)
        db.session.flush()
        db.session.add(SubventionProjet(projet_id=projet.id, subvention_id=sub.id))

        if (request.form.get("create_product_line") or "").strip() == "1":
            db.session.add(LigneBudget(
                subvention_id=sub.id,
                nature="produit",
                compte="74",
                libelle=nom,
                montant_base=float(sub.montant_demande or 0.0),
                montant_reel=float(sub.montant_recu or 0.0),
            ))

        db.session.commit()
        flash("Subvention créée et rattachée au projet.", "success")
        return redirect(url_for("projets.projet_finance", projet_id=projet.id, year=year, tab="financements"))

    if action == "link_subvention":
        if not can("subventions:link"):
            abort(403)
        sub_id = int(request.form.get("subvention_id") or 0)
        sub = Subvention.query.get_or_404(sub_id)
        if sub.secteur != projet.secteur:
            abort(400)

        exists = SubventionProjet.query.filter_by(projet_id=projet.id, subvention_id=sub.id).first()
        if exists:
            flash("Cette subvention est déjà rattachée au projet.", "info")
        else:
            db.session.add(SubventionProjet(projet_id=projet.id, subvention_id=sub.id))
            db.session.commit()
            flash("Subvention existante rattachée au projet.", "success")
        return redirect(url_for("projets.projet_finance", projet_id=projet.id, year=year, tab="financements"))

    if action == "quick_depense":
        if not can("depenses:create"):
            abort(403)

        ligne_id = int(request.form.get("ligne_budget_id") or 0)
        ligne = LigneBudget.query.get_or_404(ligne_id)
        sub = ligne.source_sub
        if not sub or sub.secteur != projet.secteur:
            abort(400)
        linked = SubventionProjet.query.filter_by(projet_id=projet.id, subvention_id=sub.id).first()
        if not linked:
            abort(400)
        if getattr(ligne, "nature", "charge") != "charge":
            flash("Une dépense doit être rattachée à une ligne de charge.", "danger")
            return redirect(url_for("projets.projet_finance", projet_id=projet.id, year=year, tab="depenses"))

        charge_id = int(request.form.get("charge_projet_id") or 0)
        charge = None
        if charge_id:
            charge = ChargeProjet.query.get_or_404(charge_id)
            if charge.projet_id != projet.id:
                abort(400)

        libelle = (request.form.get("libelle") or "").strip()
        if not libelle:
            flash("Libellé de dépense obligatoire.", "danger")
            return redirect(url_for("projets.projet_finance", projet_id=projet.id, year=year, tab="depenses"))

        dep = Depense(
            ligne_budget_id=ligne.id,
            charge_projet_id=charge.id if charge else None,
            libelle=libelle,
            montant=_parse_float_fr(request.form.get("montant"), 0.0),
            date_paiement=_parse_date_or_none(request.form.get("date_paiement")),
            type_depense=(request.form.get("type_depense") or "Fonctionnement").strip(),
            fournisseur=(request.form.get("fournisseur") or "").strip() or None,
            reference_piece=(request.form.get("reference_piece") or "").strip() or None,
            mode_paiement=(request.form.get("mode_paiement") or "").strip() or None,
        )
        db.session.add(dep)
        db.session.flush()
        db.session.add(DepenseAffectation(
            depense_id=dep.id,
            source_type="subvention",
            subvention_id=sub.id,
            ligne_budget_id=ligne.id,
            montant=float(dep.montant or 0.0),
        ))
        db.session.commit()
        flash("Dépense rapide créée et imputée à 100% sur la ligne choisie. Tu peux affiner la répartition depuis la fiche dépense.", "success")
        return redirect(url_for("projets.projet_finance", projet_id=projet.id, year=year, tab="depenses"))

    abort(400)


@bp.route("/projets/<int:projet_id>/budget")
@login_required
@require_perm("aap:view")
def projet_budget_home(projet_id):
    projet = Projet.query.get_or_404(projet_id)
    if not can_access_secteur(projet.secteur):
        abort(403)
    return redirect(url_for("projets.projet_budget_charges", projet_id=projet_id))


@bp.route("/projets/<int:projet_id>/budget/charges", methods=["GET", "POST"])
@login_required
@require_perm("aap:charges_view")
def projet_budget_charges(projet_id):
    projet = Projet.query.get_or_404(projet_id)
    if not can_access_secteur(projet.secteur):
        abort(403)

    if request.method == "POST":
        if not can("aap:charges_edit"):
            abort(403)
        libelle = (request.form.get("libelle") or "").strip()
        bloc = (request.form.get("bloc") or "directe").strip()
        code_plan = (request.form.get("code_plan") or "60").strip()
        montant = float(request.form.get("montant_previsionnel") or 0)

        if not libelle:
            flash("Le libellé est obligatoire.", "warning")
        else:
            c = ChargeProjet(
                projet_id=projet.id,
                libelle=libelle,
                bloc=bloc,
                code_plan=code_plan,
                montant_previsionnel=montant,
                montant_reel=float(request.form.get("montant_reel") or 0),
                commentaire=(request.form.get("commentaire") or "").strip() or None,
            )
            db.session.add(c)
            db.session.commit()
            flash("La charge a bien été ajoutée.", "success")
            return redirect(url_for("projets.projet_budget_charges", projet_id=projet.id))

    charges = ChargeProjet.query.filter_by(projet_id=projet.id).order_by(ChargeProjet.bloc.asc(), ChargeProjet.code_plan.asc(), ChargeProjet.id.asc()).all()
    return render_template(
        "projets_budget_charges.html",
        projet=projet,
        charges=charges,
        budget_stats=_budget_stats(projet.id),
        active_tab="charges",
    )


@bp.route("/projets/<int:projet_id>/budget/charges/<int:charge_id>/edit", methods=["GET", "POST"])
@login_required
@require_perm("aap:charges_edit")
def projet_budget_charge_edit(projet_id, charge_id):
    projet = Projet.query.get_or_404(projet_id)
    if not can_access_secteur(projet.secteur):
        abort(403)
    charge = ChargeProjet.query.filter_by(id=charge_id, projet_id=projet.id).first_or_404()

    if request.method == "POST":
        charge.libelle = (request.form.get("libelle") or "").strip()
        charge.bloc = (request.form.get("bloc") or "directe").strip()
        charge.code_plan = (request.form.get("code_plan") or "60").strip()
        charge.montant_previsionnel = float(request.form.get("montant_previsionnel") or 0)
        charge.montant_reel = float(request.form.get("montant_reel") or 0)
        charge.commentaire = (request.form.get("commentaire") or "").strip() or None
        db.session.commit()
        flash("La charge a bien été mise à jour.", "success")
        return redirect(url_for("projets.projet_budget_charges", projet_id=projet.id))

    return render_template(
        "projets_budget_charge_edit.html",
        projet=projet,
        charge=charge,
        budget_stats=_budget_stats(projet.id),
        active_tab="charges",
    )


@bp.route("/projets/<int:projet_id>/budget/charges/<int:charge_id>/delete", methods=["POST"])
@login_required
@require_perm("aap:charges_edit")
def projet_budget_charge_delete(projet_id, charge_id):
    projet = Projet.query.get_or_404(projet_id)
    if not can_access_secteur(projet.secteur):
        abort(403)
    charge = ChargeProjet.query.filter_by(id=charge_id, projet_id=projet.id).first_or_404()
    db.session.delete(charge)
    commit_delete(
        f"la charge « {charge.libelle} »",
        "Charge supprimée.",
        blocked_message=f"Impossible de supprimer la charge « {charge.libelle} » : elle est encore utilisée ailleurs.",
    )
    return redirect(url_for("projets.projet_budget_charges", projet_id=projet.id))


@bp.route("/projets/<int:projet_id>/budget/produits", methods=["GET", "POST"])
@login_required
@require_perm("aap:produits_view")
def projet_budget_produits(projet_id):
    projet = Projet.query.get_or_404(projet_id)
    if not can_access_secteur(projet.secteur):
        abort(403)

    if request.method == "POST":
        if not can("aap:produits_edit"):
            abort(403)
        financeur = (request.form.get("financeur") or "").strip()
        categorie = (request.form.get("categorie") or "autre").strip()
        statut = (request.form.get("statut") or "prevu").strip()
        demande = float(request.form.get("montant_demande") or 0)
        accorde = float(request.form.get("montant_accorde") or 0)
        recu = float(request.form.get("montant_recu") or 0)

        if not financeur:
            flash("Le nom du financeur est obligatoire.", "warning")
        else:
            p = ProduitProjet(
                projet_id=projet.id,
                financeur=financeur,
                categorie=categorie,
                statut=statut,
                montant_demande=demande,
                montant_accorde=accorde,
                montant_recu=recu,
                reference_dossier=(request.form.get("reference_dossier") or "").strip() or None,
                commentaire=(request.form.get("commentaire") or "").strip() or None,
            )
            db.session.add(p)
            db.session.commit()
            flash("Le financeur a bien été ajouté.", "success")
            return redirect(url_for("projets.projet_budget_produits", projet_id=projet.id))

    produits = ProduitProjet.query.filter_by(projet_id=projet.id).order_by(ProduitProjet.categorie.asc(), ProduitProjet.financeur.asc()).all()
    return render_template(
        "projets_budget_produits.html",
        projet=projet,
        produits=produits,
        budget_stats=_budget_stats(projet.id),
        active_tab="produits",
    )


@bp.route("/projets/<int:projet_id>/budget/produits/<int:produit_id>/edit", methods=["GET", "POST"])
@login_required
@require_perm("aap:produits_edit")
def projet_budget_produit_edit(projet_id, produit_id):
    projet = Projet.query.get_or_404(projet_id)
    if not can_access_secteur(projet.secteur):
        abort(403)
    produit = ProduitProjet.query.filter_by(id=produit_id, projet_id=projet.id).first_or_404()

    if request.method == "POST":
        produit.financeur = (request.form.get("financeur") or "").strip()
        produit.categorie = (request.form.get("categorie") or "autre").strip()
        produit.statut = (request.form.get("statut") or "prevu").strip()
        produit.montant_demande = float(request.form.get("montant_demande") or 0)
        produit.montant_accorde = float(request.form.get("montant_accorde") or 0)
        produit.montant_recu = float(request.form.get("montant_recu") or 0)
        produit.reference_dossier = (request.form.get("reference_dossier") or "").strip() or None
        produit.commentaire = (request.form.get("commentaire") or "").strip() or None
        db.session.commit()
        flash("Le financeur a bien été mis à jour.", "success")
        return redirect(url_for("projets.projet_budget_produits", projet_id=projet.id))

    return render_template(
        "projets_budget_produit_edit.html",
        projet=projet,
        produit=produit,
        budget_stats=_budget_stats(projet.id),
        active_tab="produits",
    )


@bp.route("/projets/<int:projet_id>/budget/produits/<int:produit_id>/delete", methods=["POST"])
@login_required
@require_perm("aap:produits_edit")
def projet_budget_produit_delete(projet_id, produit_id):
    projet = Projet.query.get_or_404(projet_id)
    if not can_access_secteur(projet.secteur):
        abort(403)
    produit = ProduitProjet.query.filter_by(id=produit_id, projet_id=projet.id).first_or_404()
    db.session.delete(produit)
    commit_delete(
        f"le produit / financeur « {produit.financeur} »",
        "Produit/financeur supprimé.",
        blocked_message=f"Impossible de supprimer le produit / financeur « {produit.financeur} » : il est encore utilisé ailleurs.",
    )
    return redirect(url_for("projets.projet_budget_produits", projet_id=projet.id))


@bp.route("/projets/<int:projet_id>/budget/ventilation", methods=["GET", "POST"])
@login_required
@require_perm("aap:ventilation_view")
def projet_budget_ventilation(projet_id):
    projet = Projet.query.get_or_404(projet_id)
    if not can_access_secteur(projet.secteur):
        abort(403)

    charges = ChargeProjet.query.filter_by(projet_id=projet.id).order_by(ChargeProjet.bloc.asc(), ChargeProjet.code_plan.asc(), ChargeProjet.id.asc()).all()
    produits = ProduitProjet.query.filter_by(projet_id=projet.id).order_by(ProduitProjet.categorie.asc(), ProduitProjet.financeur.asc()).all()

    # index existing ventilations
    existing = VentilationProjet.query.join(ChargeProjet).filter(ChargeProjet.projet_id == projet.id).all()
    vmap = {(v.charge_id, v.produit_id): v for v in existing}

    if request.method == "POST":
        if not can("aap:ventilation_edit"):
            abort(403)
        # matrice : v_<charge>_<produit> = montant
        # UX/sécurité : on refuse les ventilations incohérentes (somme > charge, somme > financeur)

        # 1) Lire les valeurs du formulaire en mémoire
        new_vals: dict[tuple[int, int], float] = {}
        for c in charges:
            for p in produits:
                key = f"v_{c.id}_{p.id}"
                if key not in request.form:
                    continue
                raw = (request.form.get(key) or "").strip().replace(",", ".")
                try:
                    val = float(raw) if raw else 0.0
                except ValueError:
                    val = 0.0
                if val < 0:
                    val = 0.0
                new_vals[(c.id, p.id)] = val

        # 2) Contrôles par charge
        errors = []
        for c in charges:
            s = sum((new_vals.get((c.id, p.id), float(vmap.get((c.id, p.id)).montant_ventile or 0)) if vmap.get((c.id, p.id)) else 0.0) for p in produits)
            max_c = float(c.montant_previsionnel or 0)
            # tolérance 1 centime
            if max_c > 0 and s - max_c > 0.01:
                errors.append(f"Charge '{c.libelle}' : {s:.2f}€ ventilés pour {max_c:.2f}€ prévus")

        # 3) Contrôles par produit/financeur
        for p in produits:
            s = sum((new_vals.get((c.id, p.id), float(vmap.get((c.id, p.id)).montant_ventile or 0)) if vmap.get((c.id, p.id)) else 0.0) for c in charges)
            # base de comparaison : reçu > accordé > demandé
            base = float(p.montant_recu or 0) if float(p.montant_recu or 0) > 0 else (float(p.montant_accorde or 0) if float(p.montant_accorde or 0) > 0 else float(p.montant_demande or 0))
            if base > 0 and s - base > 0.01:
                errors.append(f"Financeur '{p.financeur}' : {s:.2f}€ ventilés pour {base:.2f}€ ({'reçu' if float(p.montant_recu or 0) > 0 else ('accordé' if float(p.montant_accorde or 0) > 0 else 'demandé')})")

        if errors:
            flash("La ventilation ne peut pas être enregistrée : des incohérences ont été détectées.", "danger")
            for e in errors[:8]:
                flash("• " + e, "warning")
            if len(errors) > 8:
                flash(f"(+{len(errors) - 8} autres incohérences)", "warning")
            return redirect(url_for("projets.projet_budget_ventilation", projet_id=projet.id))

        # 4) Appliquer en base (diff minimal)
        changed = 0
        for c in charges:
            for p in produits:
                val = new_vals.get((c.id, p.id), None)
                if val is None:
                    continue
                cur = vmap.get((c.id, p.id))

                if val <= 0:
                    if cur:
                        db.session.delete(cur)
                        changed += 1
                    continue

                if cur:
                    if float(cur.montant_ventile or 0) != val:
                        cur.montant_ventile = val
                        changed += 1
                else:
                    nv = VentilationProjet(charge_id=c.id, produit_id=p.id, montant_ventile=val)
                    db.session.add(nv)
                    vmap[(c.id, p.id)] = nv
                    changed += 1

        db.session.commit()
        flash(f"La ventilation a bien été enregistrée ({changed} modification(s)).", "success")
        return redirect(url_for("projets.projet_budget_ventilation", projet_id=projet.id))

    # rebuild vmap with numbers for template
    vvals = {(cid, pid): float(v.montant_ventile or 0) for (cid, pid), v in vmap.items()}
    return render_template(
        "projets_budget_ventilation.html",
        projet=projet,
        charges=charges,
        produits=produits,
        vvals=vvals,
        budget_stats=_budget_stats(projet.id),
        active_tab="ventilation",
    )


@bp.route("/projets/<int:projet_id>/budget/synthese")
@login_required
@require_perm("aap:synthese_view")
def projet_budget_synthese(projet_id):
    projet = Projet.query.get_or_404(projet_id)
    if not can_access_secteur(projet.secteur):
        abort(403)
    charges = ChargeProjet.query.filter_by(projet_id=projet.id).all()
    produits = ProduitProjet.query.filter_by(projet_id=projet.id).all()

    alertes = []
    # charges non financées
    for c in charges:
        if float(c.reste_a_financer or 0) > 0.01:
            alertes.append(f"Charge non financée : {c.libelle} (reste {c.reste_a_financer:.2f}€)")
    # produits non ventilés
    for p in produits:
        if float(p.reste_a_ventiler or 0) > 0.01:
            alertes.append(f"Produit non ventilé : {p.financeur} (reste {p.reste_a_ventiler:.2f}€)")

    return render_template(
        "projets_budget_synthese.html",
        projet=projet,
        charges=charges,
        produits=produits,
        alertes=alertes,
        budget_stats=_budget_stats(projet.id),
        active_tab="synthese",
    )
