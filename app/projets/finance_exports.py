from datetime import datetime, date





from io import BytesIO

from flask import abort, send_file
from flask_login import login_required

from app.rbac import require_perm
from app.projets.common import bp
from app.projets.finance_secteur import (
    _is_fonds_propres_enveloppe,
    _sector_finance_context,
    _selected_sector_year,
)

# ---------------------------------------------------------------------
# P0.2H — Exports financiers annuels
# ---------------------------------------------------------------------

def _xlsx_clean(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, (datetime, date)):
        return value
    return str(value)


def _xlsx_money(value) -> float:
    try:
        return round(float(value or 0), 2)
    except Exception:
        return 0.0


def _xlsx_sheet_title(title: str, existing: set[str]) -> str:
    forbidden = ["\\", "/", "*", "[", "]", ":", "?"]
    clean = str(title or "Feuille")
    for ch in forbidden:
        clean = clean.replace(ch, " ")
    clean = " ".join(clean.split()).strip() or "Feuille"
    clean = clean[:31]
    base = clean
    i = 2
    while clean in existing:
        suffix = f" {i}"
        clean = (base[:31 - len(suffix)] + suffix).strip()
        i += 1
    existing.add(clean)
    return clean


def _xlsx_table_name(prefix: str, index: int) -> str:
    raw = "".join(ch if ch.isalnum() else "_" for ch in str(prefix or "Table"))
    raw = raw.strip("_") or "Table"
    if raw[0].isdigit():
        raw = "T_" + raw
    return f"{raw}_{index}"[:240]


def _xlsx_add_table(ws, table_name: str, n_rows: int, n_cols: int):
    if n_rows < 2 or n_cols < 1:
        return
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    ref = f"A1:{get_column_letter(n_cols)}{n_rows}"
    tab = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)


def _xlsx_autowidth(ws, max_width=48):
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 10
        for cell in col:
            value = cell.value
            if value is None:
                continue
            width = max(width, min(max_width, len(str(value)) + 2))
        ws.column_dimensions[letter].width = width


def _xlsx_style_sheet(ws, money_cols=None, percent_cols=None):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    money_cols = set(money_cols or [])
    percent_cols = set(percent_cols or [])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A2"

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.column in money_cols and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00 €'
            if cell.column in percent_cols and isinstance(cell.value, (int, float)):
                cell.number_format = '0"%"'

    _xlsx_autowidth(ws)


def _xlsx_write_table(wb, title: str, headers: list[str], rows: list[list], table_prefix: str, money_headers=None, percent_headers=None):
    existing = {ws.title for ws in wb.worksheets}
    ws = wb.create_sheet(_xlsx_sheet_title(title, existing))
    ws.append(headers)
    for row in rows:
        ws.append([_xlsx_clean(v) for v in row])

    money_headers = set(money_headers or [])
    percent_headers = set(percent_headers or [])
    money_cols = [idx + 1 for idx, h in enumerate(headers) if h in money_headers]
    percent_cols = [idx + 1 for idx, h in enumerate(headers) if h in percent_headers]

    _xlsx_style_sheet(ws, money_cols=money_cols, percent_cols=percent_cols)

    # Sécurité Excel :
    # ne pas cumuler filtre de feuille + table structurée.
    # Excel répare parfois /xl/tables/table*.xml quand les deux filtres se chevauchent.
    if rows:
        _xlsx_add_table(ws, _xlsx_table_name(table_prefix, len(wb.worksheets)), len(rows) + 1, len(headers))
    else:
        # Sur une feuille vide, un simple gel de ligne suffit.
        # Pas de table header-only, pas d'autoFilter fragile.
        pass

    return ws


def _build_sector_finance_workbook(secteur: str, year: int, data: dict):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.chart import BarChart, Reference

    wb = Workbook()
    ws = wb.active
    ws.title = "Lecture rapide"

    k = data.get("kpis", {})
    alerts = data.get("alerts", [])

    # Feuille lecture rapide, volontairement pas en table pour garder une mise en page financeur.
    rows = [
        ["Secteur", secteur],
        ["Exercice", year],
        ["Charges prévues", _xlsx_money(k.get("charges_prevues"))],
        ["Produits prévus", _xlsx_money(k.get("produits_prevus"))],
        ["Solde prévu", _xlsx_money(k.get("solde_prevu"))],
        ["Demandé total", _xlsx_money(k.get("demande"))],
        ["Attribué total", _xlsx_money(k.get("attribue"))],
        ["Reçu total", _xlsx_money(k.get("recu"))],
        ["Subventions externes attribuées", _xlsx_money(k.get("attribue_externe"))],
        ["Fonds propres / autofinancement disponible", _xlsx_money(k.get("fonds_propres_attribue"))],
        ["Imputé sur subventions externes", _xlsx_money(k.get("impute_subventions"))],
        ["Imputé fonds propres / autofinancement", _xlsx_money(k.get("impute_fonds_propres"))],
        ["Imputé total", _xlsx_money(k.get("total_impute"))],
        ["Reste à encaisser", _xlsx_money(k.get("reste_a_encaisser"))],
        ["Dépenses à corriger", k.get("nb_depenses_a_corriger") or 0],
    ]
    ws.append(["Indicateur", "Valeur"])
    for r in rows:
        ws.append(r)

    # Style lecture rapide
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    ws["A1"].fill = header_fill
    ws["B1"].fill = header_fill
    ws["A1"].font = header_font
    ws["B1"].font = header_font
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 22
    for row in ws.iter_rows(min_row=2, max_col=2):
        row[0].font = Font(bold=True)
        if isinstance(row[1].value, (int, float)) and "Dépenses à corriger" not in str(row[0].value):
            row[1].number_format = '#,##0.00 €'
    ws.freeze_panes = "A2"

    # Bloc alertes
    start_alert = len(rows) + 4
    ws.cell(start_alert, 1, "Alertes / points de contrôle")
    ws.cell(start_alert, 1).font = Font(bold=True, size=13)
    ws.cell(start_alert + 1, 1, "Niveau")
    ws.cell(start_alert + 1, 2, "Message")
    for idx, alert in enumerate(alerts, start=start_alert + 2):
        ws.cell(idx, 1, alert.get("level", ""))
        ws.cell(idx, 2, alert.get("text", ""))
    _xlsx_autowidth(ws)

    # Petit graphique demandé/attribué/reçu/imputé
    chart_start = 20
    ws.cell(chart_start, 4, "Flux financier")
    ws.cell(chart_start, 5, "Montant")
    chart_rows = [
        ("Demandé", _xlsx_money(k.get("demande"))),
        ("Attribué", _xlsx_money(k.get("attribue"))),
        ("Reçu", _xlsx_money(k.get("recu"))),
        ("Imputé", _xlsx_money(k.get("total_impute"))),
    ]
    for i, (label, value) in enumerate(chart_rows, start=chart_start + 1):
        ws.cell(i, 4, label)
        ws.cell(i, 5, value)
    chart = BarChart()
    chart.title = "Demandé / attribué / reçu / imputé"
    chart.y_axis.title = "Montant"
    chart.x_axis.title = "Étape"
    chart.add_data(Reference(ws, min_col=5, min_row=chart_start, max_row=chart_start + len(chart_rows)), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=4, min_row=chart_start + 1, max_row=chart_start + len(chart_rows)))
    chart.height = 7
    chart.width = 14
    ws.add_chart(chart, "G3")

    # Budgets
    _xlsx_write_table(
        wb,
        "Budgets",
        ["Nom", "Statut", "Charges", "Produits", "Solde", "Lignes"],
        [
            [
                b.nom,
                b.statut,
                _xlsx_money(getattr(b, "total_charges", 0)),
                _xlsx_money(getattr(b, "total_produits", 0)),
                _xlsx_money(getattr(b, "solde", 0)),
                len(getattr(b, "lignes", []) or []),
            ]
            for b in data.get("budgets", [])
        ],
        "Budgets",
        money_headers={"Charges", "Produits", "Solde"},
    )

    # Répartition par compte
    _xlsx_write_table(
        wb,
        "Comptes prévisionnels",
        ["Nature", "Compte", "Montant prévu", "Nombre de lignes"],
        [
            [r.get("nature_label"), r.get("compte"), _xlsx_money(r.get("prevu")), r.get("nb")]
            for r in data.get("comptes", [])
        ],
        "ComptesPrev",
        money_headers={"Montant prévu"},
    )

    # Subventions externes
    _xlsx_write_table(
        wb,
        "Financeurs externes",
        ["Nom", "Demandé", "Attribué", "Reçu", "Imputé", "Reste imputable", "% imputé"],
        [
            [
                r["subvention"].nom,
                _xlsx_money(r.get("demande")),
                _xlsx_money(r.get("attribue")),
                _xlsx_money(r.get("recu")),
                _xlsx_money(r.get("impute")),
                _xlsx_money(r.get("reste")),
                r.get("pct_impute") or 0,
            ]
            for r in data.get("subventions", [])
            if not r.get("is_fonds_propres")
        ],
        "FinanceursExt",
        money_headers={"Demandé", "Attribué", "Reçu", "Imputé", "Reste imputable"},
        percent_headers={"% imputé"},
    )

    # Fonds propres compte 99
    _xlsx_write_table(
        wb,
        "Fonds propres 99",
        ["Nom", "Disponible", "Reçu / disponible", "Imputé", "Reste", "% imputé"],
        [
            [
                r["subvention"].nom,
                _xlsx_money(r.get("attribue")),
                _xlsx_money(r.get("recu")),
                _xlsx_money(r.get("impute")),
                _xlsx_money(r.get("reste")),
                r.get("pct_impute") or 0,
            ]
            for r in data.get("subventions", [])
            if r.get("is_fonds_propres")
        ],
        "FondsPropres99",
        money_headers={"Disponible", "Reçu / disponible", "Imputé", "Reste"},
        percent_headers={"% imputé"},
    )

    # Sources imputées
    _xlsx_write_table(
        wb,
        "Sources imputées",
        ["Type", "Source", "Demandé", "Attribué", "Reçu", "Imputé", "Reste", "Enveloppe formelle"],
        [
            [
                s.get("type"),
                s.get("label"),
                _xlsx_money(s.get("demande")),
                _xlsx_money(s.get("attribue")),
                _xlsx_money(s.get("recu")),
                _xlsx_money(s.get("impute")),
                "" if s.get("reste") is None else _xlsx_money(s.get("reste")),
                "Oui" if s.get("is_enveloppe") else "Non",
            ]
            for s in data.get("sources", [])
        ],
        "SourcesImputees",
        money_headers={"Demandé", "Attribué", "Reçu", "Imputé", "Reste"},
    )

    # Dossiers financeurs
    _xlsx_write_table(
        wb,
        "Dossiers financeurs",
        ["Dossier", "Statut", "Projet", "Financeur", "Charges retenues", "Produits retenus", "Solde", "Subvention liée"],
        [
            [
                a.nom,
                a.statut,
                a.projet.nom if getattr(a, "projet", None) else "Transversal",
                a.financeur or (a.subvention.nom if getattr(a, "subvention", None) else ""),
                _xlsx_money(getattr(a, "total_charges", 0)),
                _xlsx_money(getattr(a, "total_produits", 0)),
                _xlsx_money(getattr(a, "solde", 0)),
                a.subvention.nom if getattr(a, "subvention", None) else "",
            ]
            for a in data.get("dossiers", [])
        ],
        "DossiersFinanceurs",
        money_headers={"Charges retenues", "Produits retenus", "Solde"},
    )

    # Projets
    _xlsx_write_table(
        wb,
        "Projets",
        ["Projet", "A des données", "Charges prévues", "Produits prévus", "Attribué", "Reçu", "Dépenses", "Reste à financer", "Subventions", "Dossiers"],
        [
            [
                r["projet"].nom,
                "Oui" if r.get("has_data") else "Non",
                _xlsx_money(r.get("charges_prevues")),
                _xlsx_money(r.get("produits_prevus")),
                _xlsx_money(r.get("sub_attribue")),
                _xlsx_money(r.get("sub_recu")),
                _xlsx_money(r.get("depenses_total")),
                _xlsx_money(r.get("reste_a_financer")),
                r.get("nb_subventions"),
                r.get("nb_dossiers"),
            ]
            for r in data.get("projets", [])
        ],
        "Projets",
        money_headers={"Charges prévues", "Produits prévus", "Attribué", "Reçu", "Dépenses", "Reste à financer"},
    )

    # Dépenses
    _xlsx_write_table(
        wb,
        "Dépenses",
        ["Date", "Libellé", "Fournisseur", "Référence", "Montant", "Affecté", "Reste à affecter", "Statut", "Source initiale"],
        [
            [
                r.get("date"),
                r["depense"].libelle,
                r["depense"].fournisseur or "",
                r["depense"].reference_piece or "",
                _xlsx_money(r.get("montant")),
                _xlsx_money(r.get("affecte")),
                _xlsx_money(r.get("reste")),
                r.get("statut"),
                r["depense"].budget_source.source_sub.nom if getattr(r["depense"], "budget_source", None) and getattr(r["depense"].budget_source, "source_sub", None) else "",
            ]
            for r in data.get("depenses", [])
        ],
        "Depenses",
        money_headers={"Montant", "Affecté", "Reste à affecter"},
    )

    # Affectations détaillées
    aff_rows = []
    for r in data.get("depenses", []):
        dep = r["depense"]
        for aff in getattr(dep, "affectations", []) or []:
            if aff.source_type == "subvention" and getattr(aff, "subvention", None):
                source = aff.subvention.nom
                type_source = "Fonds propres" if _is_fonds_propres_enveloppe(aff.subvention, getattr(aff, "ligne_budget", None)) else "Subvention"
                ligne = f"{aff.ligne_budget.compte} - {aff.ligne_budget.libelle}" if getattr(aff, "ligne_budget", None) else ""
            else:
                source = aff.libelle_source or aff.source_type
                type_source = aff.source_type
                ligne = ""
            aff_rows.append([
                dep.date_paiement or (dep.created_at.date() if dep.created_at else ""),
                dep.libelle,
                _xlsx_money(dep.montant),
                type_source,
                source,
                ligne,
                _xlsx_money(aff.montant),
                aff.commentaire or "",
            ])

    _xlsx_write_table(
        wb,
        "Affectations dépenses",
        ["Date dépense", "Dépense", "Montant dépense", "Type source", "Source", "Ligne", "Montant affecté", "Commentaire"],
        aff_rows,
        "AffectationsDep",
        money_headers={"Montant dépense", "Montant affecté"},
    )

    # Contrôles consolidés
    controles = []
    for c in data.get("controls", []):
        controles.append([c.get("level", ""), c.get("zone", ""), c.get("message", ""), c.get("action", "")])
    if not controles:
        controles.append(["ok", "Global", "Aucun contrôle bloquant détecté.", ""])

    _xlsx_write_table(
        wb,
        "Contrôles",
        ["Niveau", "Zone", "Message", "Action conseillée"],
        controles,
        "Controles",
    )

    # Supprimer l'éventuelle feuille vide si créée en trop : non, Lecture rapide est active.
    wb.active = 0
    return wb


@bp.route("/finance/secteur/export.xlsx")
@login_required
@require_perm("projets:view")
def finance_secteur_export():
    sectors, selected_sector, years, year = _selected_sector_year()
    if not selected_sector:
        abort(404)

    data = _sector_finance_context(selected_sector, year)
    wb = _build_sector_finance_workbook(selected_sector, year, data)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_secteur = "".join(ch if ch.isalnum() else "_" for ch in str(selected_sector)).strip("_") or "secteur"
    filename = f"pilotage_financier_{year}_{safe_secteur}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
