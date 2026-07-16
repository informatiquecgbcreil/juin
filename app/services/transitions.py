"""Module Transitions — agrégations pour le tableau de bord et le bilan CSAT.

Principe : ZÉRO double saisie. Les ateliers existants sont étiquetés par
thématique de transition (et rattachés aux objectifs sectoriels) ; séances,
participations, publics uniques, heures et éco-consommation (kWh/CO₂)
remontent automatiquement depuis l'émargement et le suivi de consommation
déjà en place. S'y ajoutent deux saisies volontaires : les défis des
habitants et les mesures chiffrées par action (km évités, kg détournés…).
"""
from __future__ import annotations

from datetime import date

from sqlalchemy import extract, func

from app.extensions import db
from app.models import (
    TRANSITION_THEMATIQUES_DEFAUT,
    AtelierActivite,
    DefiTransition,
    ObjectifSectoriel,
    PresenceActivite,
    SessionActivite,
    TransitionMesure,
    TransitionThematique,
    atelier_transition_objectif,
    atelier_transition_thematique,
)


def seed_thematiques() -> None:
    """Crée les thématiques standard si le référentiel est vide (idempotent)."""
    if TransitionThematique.query.first() is not None:
        return
    for ordre, (code, libelle) in enumerate(TRANSITION_THEMATIQUES_DEFAUT):
        db.session.add(TransitionThematique(code=code, libelle=libelle, ordre=ordre))
    db.session.commit()


def thematiques_actives() -> list[TransitionThematique]:
    return (
        TransitionThematique.query
        .filter_by(actif=True)
        .order_by(TransitionThematique.ordre.asc(), TransitionThematique.libelle.asc())
        .all()
    )


def ateliers_transitions(secteur: str | None = None) -> list[AtelierActivite]:
    """Ateliers étiquetés d'au moins une thématique (non supprimés)."""
    q = (
        AtelierActivite.query
        .join(atelier_transition_thematique,
              atelier_transition_thematique.c.atelier_id == AtelierActivite.id)
        .filter(AtelierActivite.is_deleted.is_(False))
        .distinct()
    )
    if secteur:
        q = q.filter(AtelierActivite.secteur == secteur)
    return q.order_by(AtelierActivite.secteur.asc(), AtelierActivite.nom.asc()).all()


def etiqueter_atelier(atelier: AtelierActivite,
                      thematique_ids: list[int],
                      objectif_ids: list[int]) -> None:
    """Remplace l'étiquetage transitions d'un atelier (thématiques + OS)."""
    thematiques = (
        TransitionThematique.query.filter(TransitionThematique.id.in_(thematique_ids)).all()
        if thematique_ids else []
    )
    objectifs = (
        ObjectifSectoriel.query.filter(ObjectifSectoriel.id.in_(objectif_ids)).all()
        if objectif_ids else []
    )
    atelier.transition_thematiques = thematiques
    atelier.transition_objectifs = objectifs
    db.session.commit()


def _duree_minutes_stats(session: SessionActivite) -> int:
    """Durée pour les statistiques : 0 si inconnue (jamais de valeur inventée)."""
    if session.duree_minutes:
        try:
            return max(0, int(session.duree_minutes))
        except (TypeError, ValueError):
            return 0
    debut = (session.heure_debut or session.rdv_debut or "").strip()
    fin = (session.heure_fin or session.rdv_fin or "").strip()
    try:
        h1, m1 = (int(x) for x in debut.split(":")[:2])
        h2, m2 = (int(x) for x in fin.split(":")[:2])
        minutes = (h2 * 60 + m2) - (h1 * 60 + m1)
        return minutes if minutes > 0 else 0
    except (ValueError, AttributeError):
        return 0


def _stats_ateliers(atelier_ids: list[int], debut: date, fin: date) -> dict:
    """Séances réalisées, participations, uniques, heures sur la période."""
    if not atelier_ids:
        return {"seances": 0, "participations": 0, "uniques": 0, "heures": 0.0}

    eff = func.coalesce(SessionActivite.rdv_date, SessionActivite.date_session)
    sessions = (
        SessionActivite.query
        .filter(
            SessionActivite.atelier_id.in_(atelier_ids),
            SessionActivite.is_deleted.is_(False),
            func.lower(func.coalesce(SessionActivite.statut, "")) != "annulee",
            eff >= debut,
            eff <= fin,
        )
        .all()
    )
    session_ids = [s.id for s in sessions]
    heures = round(sum(_duree_minutes_stats(s) for s in sessions) / 60.0, 1)

    participations = 0
    uniques = 0
    if session_ids:
        participations = (
            db.session.query(func.count(PresenceActivite.id))
            .filter(PresenceActivite.session_id.in_(session_ids))
            .scalar() or 0
        )
        uniques = (
            db.session.query(func.count(func.distinct(PresenceActivite.participant_id)))
            .filter(PresenceActivite.session_id.in_(session_ids))
            .scalar() or 0
        )
    return {
        "seances": len(sessions),
        "participations": int(participations),
        "uniques": int(uniques),
        "heures": heures,
    }


def _conso_ateliers(atelier_ids: list[int], debut: date, fin: date) -> dict:
    """kWh / CO₂ de l'éco-consommation individuelle existante."""
    if not atelier_ids:
        return {"kwh": 0.0, "co2_kg": 0.0}
    from app.services.consumption import aggregate_individual_consumption

    agg = aggregate_individual_consumption(
        date_from=debut, date_to=fin, atelier_ids=atelier_ids,
    )
    return {"kwh": agg["total_kwh"], "co2_kg": agg["total_co2"]}


def _mesures_agregees_query(debut: date, fin: date):
    """Cumul des mesures par (libellé, unité, thématique) sur la période.

    PostgreSQL refuse de grouper sur un COALESCE à valeur bindée quand le
    même COALESCE figure aussi dans le SELECT : SQLAlchemy génère deux
    paramètres différents et Postgres ne reconnaît plus l'expression comme
    identique (même piège que dans consumption.py). On groupe donc sur la
    colonne BRUTE — sans effet métier : l'unité vide est normalisée à NULL
    à la saisie — et on garde le COALESCE côté SELECT pour l'affichage.
    """
    return (
        db.session.query(
            TransitionMesure.libelle,
            func.coalesce(TransitionMesure.unite, ""),
            TransitionMesure.thematique_id,
            func.sum(TransitionMesure.valeur),
            func.count(TransitionMesure.id),
        )
        .filter(TransitionMesure.date_mesure >= debut,
                TransitionMesure.date_mesure <= fin)
        .group_by(TransitionMesure.libelle,
                  TransitionMesure.unite,
                  TransitionMesure.thematique_id)
    )


def tableau_de_bord(annee: int, secteur: str | None = None) -> dict:
    """Toutes les données du tableau de bord transitions pour un exercice."""
    seed_thematiques()
    debut, fin = date(annee, 1, 1), date(annee, 12, 31)

    ateliers = ateliers_transitions(secteur)
    atelier_ids_tous = [a.id for a in ateliers]

    # --- Par thématique -----------------------------------------------------
    lignes_thematiques = []
    for them in thematiques_actives():
        ids = [a.id for a in ateliers if them in a.transition_thematiques]
        stats = _stats_ateliers(ids, debut, fin)
        conso = _conso_ateliers(ids, debut, fin)
        defis_q = DefiTransition.query.filter(
            DefiTransition.thematique_id == them.id,
            DefiTransition.date_engagement >= debut,
            DefiTransition.date_engagement <= fin,
        )
        defis_total = defis_q.count()
        defis_realises = defis_q.filter(DefiTransition.statut == "realise").count()
        lignes_thematiques.append({
            "thematique": them,
            "nb_ateliers": len(ids),
            **stats,
            **conso,
            "defis_total": defis_total,
            "defis_realises": defis_realises,
        })

    # --- Croisement thématique × objectif sectoriel -------------------------
    objectifs = sorted(
        {os_ for a in ateliers for os_ in getattr(a, "transition_objectifs", [])},
        key=lambda o: (o.secteur, o.ordre, o.code),
    )
    croisement = []
    for them in thematiques_actives():
        cellules = []
        for objectif in objectifs:
            ids = [
                a.id for a in ateliers
                if them in a.transition_thematiques and objectif in a.transition_objectifs
            ]
            cellules.append(_stats_ateliers(ids, debut, fin)["participations"] if ids else 0)
        if any(cellules) or any(them in a.transition_thematiques for a in ateliers):
            croisement.append({"thematique": them, "cellules": cellules})

    # --- Défis (liste de l'exercice) ----------------------------------------
    defis = (
        DefiTransition.query
        .filter(DefiTransition.date_engagement >= debut,
                DefiTransition.date_engagement <= fin)
        .order_by(DefiTransition.date_engagement.desc(), DefiTransition.id.desc())
        .all()
    )

    # --- Mesures manuelles, additionnées par (libellé, unité) ---------------
    mesures_rows = _mesures_agregees_query(debut, fin).all()
    them_par_id = {t.id: t for t in TransitionThematique.query.all()}
    mesures = [
        {
            "libelle": libelle,
            "unite": unite,
            "thematique": them_par_id.get(them_id),
            "total": round(float(total or 0), 2),
            "nb_saisies": int(nb or 0),
        }
        for libelle, unite, them_id, total, nb in mesures_rows
    ]
    mesures.sort(key=lambda m: (m["thematique"].libelle if m["thematique"] else "~", m["libelle"]))

    mesures_detail = (
        TransitionMesure.query
        .filter(TransitionMesure.date_mesure >= debut,
                TransitionMesure.date_mesure <= fin)
        .order_by(TransitionMesure.date_mesure.desc(), TransitionMesure.id.desc())
        .limit(50)
        .all()
    )

    # --- Totaux --------------------------------------------------------------
    stats_globales = _stats_ateliers(atelier_ids_tous, debut, fin)
    conso_globale = _conso_ateliers(atelier_ids_tous, debut, fin)

    return {
        "annee": annee,
        "secteur": secteur,
        "ateliers": ateliers,
        "thematiques": lignes_thematiques,
        "objectifs": objectifs,
        "croisement": croisement,
        "defis": defis,
        "defis_en_cours": sum(1 for d in defis if d.statut == "en_cours"),
        "defis_realises": sum(1 for d in defis if d.statut == "realise"),
        "mesures": mesures,
        "mesures_detail": mesures_detail,
        "totaux": {
            "nb_ateliers": len(ateliers),
            **stats_globales,
            **conso_globale,
        },
    }


def annees_disponibles() -> list[int]:
    """Années avec de l'activité transitions (séances d'ateliers étiquetés,
    défis ou mesures), plus l'année courante."""
    annees = {date.today().year}
    ids = [a.id for a in ateliers_transitions()]
    if ids:
        eff = func.coalesce(SessionActivite.rdv_date, SessionActivite.date_session)
        for (a,) in (
            db.session.query(func.distinct(extract("year", eff)))
            .filter(SessionActivite.atelier_id.in_(ids), eff.isnot(None))
            .all()
        ):
            if a:
                annees.add(int(a))
    for (a,) in db.session.query(func.distinct(extract("year", DefiTransition.date_engagement))).all():
        if a:
            annees.add(int(a))
    for (a,) in db.session.query(func.distinct(extract("year", TransitionMesure.date_mesure))).all():
        if a:
            annees.add(int(a))
    return sorted(annees, reverse=True)


# ---------------------------------------------------------------------------
# Export bilan annuel (XLSX)
# ---------------------------------------------------------------------------

def construire_export(annee: int, secteur: str | None = None):
    """Classeur de synthèse annuelle, prêt à recopier dans le portail CSAT."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    data = tableau_de_bord(annee, secteur)
    wb = Workbook()
    wb.remove(wb.active)

    def feuille(titre, entetes, lignes):
        ws = wb.create_sheet(title=titre[:31])
        ws.append(entetes)
        for cellule in ws[1]:
            cellule.font = Font(bold=True)
        for ligne in lignes:
            ws.append(["" if v is None else v for v in ligne])
        for idx, entete in enumerate(entetes, start=1):
            largeur = max([len(str(entete))] + [len(str(l[idx - 1] or "")) for l in lignes] or [10])
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(max(largeur + 2, 12), 50)
        return ws

    t = data["totaux"]
    feuille(
        "Synthèse",
        ["Indicateur", "Valeur"],
        [
            ["Année", annee],
            ["Ateliers transitions", t["nb_ateliers"]],
            ["Séances réalisées", t["seances"]],
            ["Participations", t["participations"]],
            ["Participants uniques", t["uniques"]],
            ["Heures d'animation", t["heures"]],
            ["Consommation estimée (kWh)", t["kwh"]],
            ["CO₂ estimé (kg)", t["co2_kg"]],
            ["Défis engagés", len(data["defis"])],
            ["Défis réalisés", data["defis_realises"]],
        ],
    )

    feuille(
        "Par thématique",
        ["Thématique", "Ateliers", "Séances", "Participations", "Participants uniques",
         "Heures", "kWh", "CO₂ (kg)", "Défis engagés", "Défis réalisés"],
        [[l["thematique"].libelle, l["nb_ateliers"], l["seances"], l["participations"],
          l["uniques"], l["heures"], l["kwh"], l["co2_kg"], l["defis_total"], l["defis_realises"]]
         for l in data["thematiques"]],
    )

    if data["objectifs"]:
        feuille(
            "Croisement OS",
            ["Thématique"] + [f"{o.code} — {o.libelle}" for o in data["objectifs"]],
            [[c["thematique"].libelle] + c["cellules"] for c in data["croisement"]],
        )

    feuille(
        "Défis",
        ["Titre", "Thématique", "Porteur", "Statut", "Engagé le", "Réalisé le", "Commentaire"],
        [[d.titre, d.thematique.libelle if d.thematique else "",
          d.porteur_label, d.statut_label,
          d.date_engagement.strftime("%d/%m/%Y") if d.date_engagement else "",
          d.date_realisation.strftime("%d/%m/%Y") if d.date_realisation else "",
          d.commentaire or ""]
         for d in data["defis"]],
    )

    feuille(
        "Mesures",
        ["Mesure", "Unité", "Thématique", "Total", "Nb saisies"],
        [[m["libelle"], m["unite"], m["thematique"].libelle if m["thematique"] else "",
          m["total"], m["nb_saisies"]]
         for m in data["mesures"]],
    )

    feuille(
        "Détail ateliers",
        ["Atelier", "Secteur", "Thématiques", "Objectifs sectoriels"],
        [[a.nom, a.secteur,
          ", ".join(th.libelle for th in a.transition_thematiques),
          ", ".join(f"{o.code}" for o in getattr(a, "transition_objectifs", []))]
         for a in data["ateliers"]],
    )

    feuille(
        "Mode d'emploi",
        ["À savoir"],
        [
            ["Synthèse annuelle du module Transitions, à recopier dans le portail CSAT."],
            ["Séances, participations, publics et heures viennent de l'émargement : aucune double saisie."],
            ["kWh et CO₂ viennent du suivi d'éco-consommation par présence (valeurs indicatives, non certifiées)."],
            ["Les défis et mesures sont des saisies volontaires depuis la page Transitions."],
            ["Étiquetez les ateliers (page Transitions → Étiqueter les ateliers) pour alimenter ces chiffres."],
        ],
    )

    return wb
