"""Moteur d'import d'un annuaire de participants depuis un tableur.

Conçu pour être GÉNÉRIQUE (n'importe quelle structure, pas seulement le
fichier d'origine) et SÛR : il ne touche jamais la base. Il se contente
de lire des feuilles de présence, d'en extraire les personnes, de
normaliser, de dédoublonner, et de produire un RAPPORT DE SIMULATION
indiquant pour chaque personne ce qui se passerait : création, rapprochement
avec une fiche existante, doublon interne, ou cas ambigu à arbitrer.

Principe : aucune création silencieuse. Les cas certains sont classés
automatiquement ; les cas incertains (nom sans année de naissance,
quasi-homonymes) sont remontés pour décision humaine.

La persistance (création réelle en base) est volontairement hors de ce
module : elle consommera le rapport validé par l'utilisateur.
"""
from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field

# Synonymes d'en-têtes de colonnes (normalisés) -> champ logique.
ENTETES = {
    "nom": "nom", "noms": "nom",
    "prenom": "prenom", "prenoms": "prenom",
    "anneenaissance": "annee", "annee": "annee", "naissance": "annee",
    "datenaissance": "annee", "ddn": "annee",
    "sexe": "sexe", "genre": "sexe",
    "quartier": "quartier", "ville": "ville", "commune": "ville",
    "adresse": "adresse", "adresses": "adresse", "rue": "adresse",
}

# Valeurs d'année non exploitables (saisies à la place d'une vraie année).
_ANNEE_NON_RENSEIGNEE = {"adulte", "enfant", "inconnu", "na", "nc", ""}


def normaliser(valeur) -> str:
    """Minuscule, sans accents, sans ponctuation ni espaces : clé de comparaison."""
    s = unicodedata.normalize("NFKD", str(valeur or "")).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]", "", s.lower())


def nettoyer_affichage(valeur) -> str:
    """Texte propre pour stockage : espaces réduits, sans espaces de bord."""
    return re.sub(r"\s+", " ", str(valeur or "").strip())


def parse_annee(valeur) -> int | None:
    """Extrait une année de naissance plausible, sinon None."""
    brut = str(valeur or "").strip()
    if normaliser(brut) in {normaliser(x) for x in _ANNEE_NON_RENSEIGNEE}:
        return None
    m = re.search(r"(19|20)\d{2}", brut)
    if m:
        return int(m.group(0))
    return None


def normaliser_sexe(valeur) -> str | None:
    c = normaliser(valeur)[:1]
    if c == "f":
        return "Femme"
    if c in {"m", "h"}:  # M (masculin) ou H (homme) selon les feuilles
        return "Homme"
    return None


@dataclass
class PersonneImportee:
    nom: str
    prenom: str
    annee: int | None
    sexe: str | None
    quartier: str | None
    feuille: str
    ligne: int
    adresse: str | None = None
    ville: str | None = None

    @property
    def cle_forte(self) -> tuple | None:
        if self.annee is None:
            return None
        return (normaliser(self.nom), normaliser(self.prenom), self.annee)

    @property
    def cle_faible(self) -> tuple:
        return (normaliser(self.nom), normaliser(self.prenom))


def detecter_entetes(lignes, max_scan: int = 20) -> tuple[int, dict] | None:
    """Trouve la ligne d'en-tête et la position des colonnes.

    Retourne (index_ligne, {champ: index_colonne}) ou None si introuvable.
    Une ligne est un en-tête si elle contient au moins « nom » et « prenom ».
    """
    for i, ligne in enumerate(lignes[:max_scan]):
        mapping: dict[str, int] = {}
        for j, cellule in enumerate(ligne):
            champ = ENTETES.get(normaliser(cellule))
            if champ and champ not in mapping:
                mapping[champ] = j
        if "nom" in mapping and "prenom" in mapping:
            return i, mapping
    return None


def parser_feuille(nom_feuille: str, lignes: list) -> tuple[list[PersonneImportee], str | None]:
    """Extrait les personnes d'une feuille. Retourne (personnes, raison_ignoree)."""
    detection = detecter_entetes(lignes)
    if not detection:
        return [], "en-tête (NOM/PRÉNOM) introuvable"

    debut, cols = detection
    i_nom, i_prenom = cols["nom"], cols["prenom"]
    i_annee, i_sexe = cols.get("annee"), cols.get("sexe")
    i_quartier = cols.get("quartier", cols.get("ville"))
    i_ville = cols.get("ville")
    i_adresse = cols.get("adresse")

    def val(ligne, idx):
        if idx is None or idx >= len(ligne):
            return None
        return ligne[idx]

    personnes: list[PersonneImportee] = []
    for offset, ligne in enumerate(lignes[debut + 1:], start=debut + 2):
        nom = nettoyer_affichage(val(ligne, i_nom))
        prenom = nettoyer_affichage(val(ligne, i_prenom))
        if not nom and not prenom:
            continue
        if normaliser(nom) in {"total", "noms", "nom", "etatdepresences"}:
            continue
        if not nom:  # un prénom sans nom n'est pas exploitable comme identité
            continue
        personnes.append(PersonneImportee(
            nom=nom,
            prenom=prenom,
            annee=parse_annee(val(ligne, i_annee)),
            sexe=normaliser_sexe(val(ligne, i_sexe)),
            quartier=nettoyer_affichage(val(ligne, i_quartier)) or None,
            feuille=nom_feuille,
            ligne=offset,
            adresse=nettoyer_affichage(val(ligne, i_adresse)) or None,
            ville=nettoyer_affichage(val(ligne, i_ville)) or None,
        ))
    return personnes, None


@dataclass
class RapportImport:
    nouveaux: list[PersonneImportee] = field(default_factory=list)
    rapproches: list[PersonneImportee] = field(default_factory=list)
    doublons_internes: list[PersonneImportee] = field(default_factory=list)
    ambigus: list[PersonneImportee] = field(default_factory=list)
    feuilles_ignorees: list[tuple[str, str]] = field(default_factory=list)
    feuilles_traitees: list[str] = field(default_factory=list)

    @property
    def total_lignes(self) -> int:
        return (len(self.nouveaux) + len(self.rapproches)
                + len(self.doublons_internes) + len(self.ambigus))

    def resume(self) -> dict:
        return {
            "nouveaux": len(self.nouveaux),
            "rapproches": len(self.rapproches),
            "doublons_internes": len(self.doublons_internes),
            "ambigus": len(self.ambigus),
            "feuilles_traitees": len(self.feuilles_traitees),
            "feuilles_ignorees": len(self.feuilles_ignorees),
            "total_lignes": self.total_lignes,
        }


def analyser(feuilles: dict[str, list], cles_existantes: set | None = None,
             cles_faibles_existantes: set | None = None) -> RapportImport:
    """Simulation d'import : classe chaque personne sans rien écrire.

    - ``feuilles`` : {nom_feuille: lignes (liste de listes de cellules)}.
    - ``cles_existantes`` : clés fortes (nom_norm, prenom_norm, annee) déjà en base.
    - ``cles_faibles_existantes`` : clés (nom_norm, prenom_norm) déjà en base,
      pour repérer les homonymes potentiels quand l'année manque.

    Classement :
    - nouveau            : clé forte inédite (en base et dans le fichier) ;
    - rapproché          : clé forte déjà en base (la fiche existe) ;
    - doublon interne    : clé forte déjà vue plus haut dans le fichier ;
    - ambigu             : pas d'année exploitable -> décision humaine requise.
    """
    cles_existantes = set(cles_existantes or set())
    cles_faibles_existantes = set(cles_faibles_existantes or set())
    rapport = RapportImport()

    vues_fortes: set = set()
    for nom_feuille, lignes in feuilles.items():
        personnes, raison = parser_feuille(nom_feuille, lignes)
        if raison:
            rapport.feuilles_ignorees.append((nom_feuille, raison))
            continue
        rapport.feuilles_traitees.append(nom_feuille)
        for p in personnes:
            if p.annee is None:
                rapport.ambigus.append(p)
                continue
            cle = p.cle_forte
            if cle in cles_existantes:
                rapport.rapproches.append(p)
            elif cle in vues_fortes:
                rapport.doublons_internes.append(p)
            else:
                vues_fortes.add(cle)
                rapport.nouveaux.append(p)
    return rapport


def charger_classeur(chemin: str, max_col: int = 12) -> dict[str, list]:
    """Charge un .xlsx en {nom_feuille: lignes}.

    Ne lit que les premières colonnes (identité) : les colonnes de dates de
    présence ne servent pas à l'import de l'annuaire et leur nombre peut
    être énorme (centaines de colonnes). On borne donc à `max_col` pour
    rester rapide même sur des feuilles à dizaines de milliers de lignes.
    """
    import openpyxl

    wb = openpyxl.load_workbook(chemin, read_only=True, data_only=True)
    feuilles: dict[str, list] = {}
    try:
        for nom in wb.sheetnames:
            ws = wb[nom]
            lignes = [
                list(row) if row else []
                for row in ws.iter_rows(max_col=max_col, values_only=True)
            ]
            feuilles[nom] = lignes
    finally:
        wb.close()
    return feuilles


def dedup_ambigus(ambigus: list[PersonneImportee],
                  cles_faibles_existantes: set | None = None) -> list[PersonneImportee]:
    """Réduit les cas ambigus (sans année) à une fiche par nom+prénom distinct,
    en écartant ceux dont le nom+prénom existe déjà en base.

    ATTENTION : faute d'année, deux homonymes réels seraient ici fusionnés.
    À n'utiliser que si l'utilisateur a explicitement choisi d'inclure les
    ambigus, en connaissance de cause.
    """
    cles_faibles_existantes = set(cles_faibles_existantes or set())
    vus: set = set()
    retenus: list[PersonneImportee] = []
    for p in ambigus:
        cle = p.cle_faible
        if cle in cles_faibles_existantes or cle in vus:
            continue
        vus.add(cle)
        retenus.append(p)
    return retenus
