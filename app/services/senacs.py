"""Bilan SENACS : pré-calcul des indicateurs du bilan structure.

SENACS (Système d'Échanges National des Centres Sociaux) demande chaque
année une photographie chiffrée de la structure. Ce service calcule
automatiquement ce que les données de l'application permettent de
produire, en respectant les règles de l'enquête :

- participants UNIQUES sans double comptage entre secteurs (le piège
  n°1 du SENACS : l'annuaire unique de l'application dédoublonne
  nativement),
- participations cumulées (présences),
- répartition par les 8 tranches d'âge SENACS, âge au 31/12 de l'année,
- genre, quartiers (Rouher / Hauts de Creil / autre QPV / hors QPV),
- tableau actions/séances par atelier (séances réalisées, heures face
  public, participants uniques, participations),
- partenaires recensés.

Le bénévolat, les emplois et les finances ne sont pas gérés par
l'application : ils restent à renseigner à la main dans SENACS.
"""
from datetime import date

from sqlalchemy import extract, or_

from app.extensions import db
from app.models import (
    Participant,
    Partenaire,
    PresenceActivite,
    SessionActivite,
)

# Tranches d'âge officielles SENACS (âge au 31 décembre de l'année).
TRANCHES_AGE = [
    ("0-6 ans", 0, 6),
    ("7-10 ans", 7, 10),
    ("11-14 ans", 11, 14),
    ("15-17 ans", 15, 17),
    ("18-25 ans", 18, 25),
    ("26-59 ans", 26, 59),
    ("60-74 ans", 60, 74),
    ("75 ans et plus", 75, 999),
]
NON_RENSEIGNE = "Non renseigné"


def annees_disponibles() -> list[int]:
    """Années pour lesquelles il existe des sessions datées."""
    lignes = (
        db.session.query(extract("year", SessionActivite.date_session))
        .filter(SessionActivite.date_session.isnot(None))
        .distinct()
        .all()
    )
    annees = sorted({int(a) for (a,) in lignes if a}, reverse=True)
    return annees or [date.today().year - 1]


def _presences_annee(annee: int):
    """Présences de l'année: date de session, sinon date d'enregistrement."""
    return (
        db.session.query(PresenceActivite, SessionActivite)
        .join(SessionActivite, PresenceActivite.session_id == SessionActivite.id)
        .filter(SessionActivite.is_deleted.is_(False))
        .filter(
            or_(
                extract("year", SessionActivite.date_session) == annee,
                SessionActivite.date_session.is_(None)
                & (extract("year", PresenceActivite.created_at) == annee),
            )
        )
        .all()
    )


def _age_au_31_decembre(naissance: date | None, annee: int) -> int | None:
    if not naissance:
        return None
    return max(0, annee - naissance.year)


def _tranche_age(age: int | None) -> str:
    if age is None:
        return NON_RENSEIGNE
    for libelle, mini, maxi in TRANCHES_AGE:
        if mini <= age <= maxi:
            return libelle
    return NON_RENSEIGNE


def _bucket_quartier(p: Participant) -> str:
    quartier = p.quartier
    nom = (quartier.nom if quartier else "").strip().lower()
    ville = ((quartier.ville if quartier else None) or p.ville or "").strip().lower()
    if "rouher" in nom:
        return "Rouher"
    if "hauts de creil" in nom or "hauts-de-creil" in nom:
        return "Hauts de Creil"
    if quartier is not None and (getattr(quartier, "is_qpv", False) or "qpv" in nom):
        return "Autre QPV"
    if ville:
        return "Hors QPV"
    return NON_RENSEIGNE


def publics_annee(annee: int) -> dict:
    """Indicateurs « publics » SENACS pour l'année (tableau 6.1)."""
    presences = _presences_annee(annee)

    participations = len(presences)
    participants: dict[int, Participant] = {}
    for presence, _session in presences:
        if presence.participant_id and presence.participant is not None:
            participants[presence.participant_id] = presence.participant

    ages = {libelle: 0 for libelle, _, _ in TRANCHES_AGE}
    ages[NON_RENSEIGNE] = 0
    genres: dict[str, int] = {}
    quartiers: dict[str, int] = {}
    for p in participants.values():
        ages[_tranche_age(_age_au_31_decembre(p.date_naissance, annee))] += 1
        genre = (p.genre or NON_RENSEIGNE).strip() or NON_RENSEIGNE
        genres[genre] = genres.get(genre, 0) + 1
        bucket = _bucket_quartier(p)
        quartiers[bucket] = quartiers.get(bucket, 0) + 1

    return {
        "annee": annee,
        "participants_uniques": len(participants),
        "participations": participations,
        "ages": ages,
        "genres": dict(sorted(genres.items())),
        "quartiers": dict(sorted(quartiers.items())),
    }


def _duree_heures(session: SessionActivite) -> float | None:
    if session.duree_minutes:
        return round(session.duree_minutes / 60.0, 2)
    debut, fin = (session.heure_debut or "").strip(), (session.heure_fin or "").strip()
    try:
        h1, m1 = (int(x) for x in debut.split(":"))
        h2, m2 = (int(x) for x in fin.split(":"))
        minutes = (h2 * 60 + m2) - (h1 * 60 + m1)
        return round(minutes / 60.0, 2) if minutes > 0 else None
    except (ValueError, AttributeError):
        return None


def tableau_actions(annee: int) -> list[dict]:
    """Tableau actions/séances par atelier (tableau 6.2)."""
    presences = _presences_annee(annee)

    par_atelier: dict[int, dict] = {}
    sessions_vues: set[int] = set()
    for presence, session in presences:
        atelier = session.atelier
        if atelier is None:
            continue
        ligne = par_atelier.setdefault(
            atelier.id,
            {
                "atelier": atelier.nom,
                "secteur": atelier.secteur,
                "seances": 0,
                "heures_face_public": 0.0,
                "heures_inconnues": 0,
                "participants_uniques": set(),
                "participations": 0,
            },
        )
        ligne["participations"] += 1
        if presence.participant_id:
            ligne["participants_uniques"].add(presence.participant_id)
        if session.id not in sessions_vues:
            sessions_vues.add(session.id)
            ligne["seances"] += 1
            heures = _duree_heures(session)
            if heures is None:
                ligne["heures_inconnues"] += 1
            else:
                ligne["heures_face_public"] += heures

    lignes = []
    for ligne in par_atelier.values():
        ligne["participants_uniques"] = len(ligne["participants_uniques"])
        ligne["heures_face_public"] = round(ligne["heures_face_public"], 1)
        lignes.append(ligne)
    lignes.sort(key=lambda l: (l["secteur"] or "", l["atelier"] or ""))
    return lignes


def partenaires_recenses() -> list[dict]:
    """Base du tableau partenariats (6.4) : la nature et le rôle restent à qualifier."""
    partenaires = Partenaire.query.order_by(Partenaire.nom.asc()).all()
    return [
        {
            "nom": pa.nom,
            "territoire": (pa.territoire_couvert or "").strip(),
            "description": (pa.description or "").strip(),
        }
        for pa in partenaires
    ]


# ---------------------------------------------------------------------
# Export XLSX (tableaux de consolidation)
# ---------------------------------------------------------------------

def construire_export_senacs(annee: int):
    """Classeur Excel calqué sur les tableaux de consolidation du guide."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    wb.remove(wb.active)

    def feuille(titre, entetes, lignes):
        ws = wb.create_sheet(title=titre[:31])
        ws.append(entetes)
        for cellule in ws[1]:
            cellule.font = Font(bold=True)
        for ligne in lignes:
            ws.append(["" if v is None else v for v in ligne])
        for idx, entete in enumerate(entetes, start=1):
            largeur = max([len(str(entete))] + [len(str(l[idx - 1] or "")) for l in lignes] or [10])
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(max(largeur + 2, 12), 55)
        return ws

    publics = publics_annee(annee)

    lignes_global = [
        ["Année des données", annee],
        ["Participants uniques (actions/projets/services, sans double comptage)", publics["participants_uniques"]],
        ["Participations cumulées (présences)", publics["participations"]],
        ["", ""],
        ["Répartition par âge (au 31/12)", ""],
    ]
    for libelle, _, _ in TRANCHES_AGE:
        lignes_global.append([f"  {libelle}", publics["ages"][libelle]])
    lignes_global.append([f"  {NON_RENSEIGNE}", publics["ages"][NON_RENSEIGNE]])
    lignes_global.append(["", ""])
    lignes_global.append(["Répartition par genre", ""])
    for genre, nombre in publics["genres"].items():
        lignes_global.append([f"  {genre}", nombre])
    lignes_global.append(["", ""])
    lignes_global.append(["Répartition par quartier", ""])
    for quartier, nombre in publics["quartiers"].items():
        lignes_global.append([f"  {quartier}", nombre])
    feuille("Public global", ["Indicateur", "Valeur"], lignes_global)

    feuille(
        "Actions séances",
        ["Action / atelier", "Secteur porteur", "Séances réalisées", "Heures face public",
         "Séances sans durée connue", "Participants uniques", "Participations cumulées"],
        [[l["atelier"], l["secteur"], l["seances"], l["heures_face_public"],
          l["heures_inconnues"], l["participants_uniques"], l["participations"]]
         for l in tableau_actions(annee)],
    )

    feuille(
        "Partenariats",
        ["Partenaire", "Territoire couvert", "Description",
         "Nature (à qualifier : formalisé / non formalisé)", "Rôle (à qualifier)"],
        [[p["nom"], p["territoire"], p["description"], "", ""] for p in partenaires_recenses()],
    )

    benevolat = benevolat_annee(annee)
    lignes_benevolat = [
        ["Bénévoles actifs (avec heures saisies)", benevolat["nb_actifs"]],
        ["Bénévoles déclarés (fiches marquées)", benevolat["nb_declares"]],
        ["Heures de bénévolat", benevolat["total_heures"]],
        ["Taux horaire de valorisation (€)", benevolat["taux_horaire"]],
        ["Valorisation (contributions volontaires, €)", benevolat["valorisation"]],
        ["", ""],
        ["Heures par mission", ""],
    ] + [[f"  {mission}", heures] for mission, heures in benevolat["par_mission"].items()]
    feuille("Bénévolat", ["Indicateur", "Valeur"], lignes_benevolat)

    emplois = emplois_annee(annee)
    lignes_emplois = [
        [s.poste or "Salarié·e", s.contrat_label, s.etp, "Module RH"]
        for s in emplois["rh_salaries"]
    ] + [
        [p.intitule, p.contrat_label, p.etp, p.commentaire or "Saisie manuelle"]
        for p in emplois["postes"]
    ]
    lignes_emplois.append(["", "", "", ""])
    lignes_emplois.append(["TOTAL", f"{emplois['nb_rh'] + emplois['nb_postes']} poste(s)", emplois["etp_global"], ""])
    for contrat, d in emplois["par_contrat"].items():
        lignes_emplois.append([f"  dont {contrat}", d["nb"], d["etp"], ""])
    feuille("Emplois", ["Fonction", "Contrat", "ETP", "Source / commentaire"], lignes_emplois)

    finances = finances_annee(annee)
    lignes_finances = [
        ["Produits — subventions par financeur", "", ""],
    ] + [
        [f"  {financeur}", d["attribue"], d["recu"]]
        for financeur, d in finances["par_financeur"].items()
    ] + [
        ["TOTAL subventions", finances["total_attribue"], finances["total_recu"]],
        ["", "", ""],
        ["Charges réalisées (lignes budgétaires des subventions)", finances["total_charges_reelles"], ""],
        ["Masse salariale (module RH, brut chargé)", finances["masse_salariale"], ""],
        ["Valorisation du bénévolat (compte 87)", finances["valorisation_benevolat"], ""],
    ]
    feuille("Finances", ["Poste", "Attribué (€)", "Reçu (€)"], lignes_finances)

    evenementiel = evenementiel_annee(annee)
    feuille(
        "Événementiel",
        ["Indicateur", "Valeur"],
        [
            ["Événements réalisés (séances marquées « événement »)", evenementiel["nb_evenements"]],
            ["Participations aux événements", evenementiel["participations"]],
            ["Participants uniques aux événements", evenementiel["participants_uniques"]],
        ],
    )

    feuille(
        "Mode d'emploi",
        ["À savoir"],
        [
            ["Ce fichier pré-remplit les tableaux de consolidation SENACS à partir des données de l'application."],
            ["Les participants uniques sont dédoublonnés entre secteurs : une personne inscrite dans plusieurs ateliers compte une seule fois."],
            ["Bénévolat : onglet alimenté par la page Ressources -> Bénévolat (heures saisies, valorisation au taux configuré)."],
            ["Emplois : onglet alimenté par les postes déclarés sur la page Bilan SENACS (fonction, contrat, ETP)."],
            ["Finances : produits par financeur et charges réalisées issus des subventions de l'exercice ; à rapprocher de la comptabilité officielle."],
            ["Événementiel : séances cochées « événement » à la création ou depuis la correction de séance."],
            ["Les publics des partenaires hébergés restent à estimer à part, sans les mélanger."],
            ["La colonne 'Séances sans durée connue' signale les heures face public sous-évaluées : renseignez les horaires des séances pour fiabiliser le chiffre."],
            ["Le hors les murs (aller-vers) et les liens à distance relèvent d'un suivi spécifique non couvert ici."],
            ["La saisie officielle se fait sur senacs.fr ; en cas de doute, consultez vos référents CAF / fédération."],
        ],
    )

    return wb


# ---------------------------------------------------------------------------
# Volets complémentaires : bénévolat, emplois, finances, événementiel
# ---------------------------------------------------------------------------

def benevolat_annee(annee: int) -> dict:
    """Onglet « vitalité démocratique » : bénévoles, heures, valorisation."""
    from app.services.benevolat import stats_annee
    stats = stats_annee(annee)
    return {
        "nb_actifs": stats["nb_actifs"],
        "nb_declares": stats["nb_declares"],
        "total_heures": stats["total_heures"],
        "taux_horaire": stats["taux_horaire"],
        "valorisation": stats["valorisation"],
        "par_mission": stats["par_mission"],
    }


def emplois_annee(annee: int) -> dict:
    """Onglet « emplois » : salariés du module RH actifs sur l'exercice
    + postes complémentaires saisis à la main, avec totaux consolidés."""
    from app.models import SenacsEmploi, Salarie

    postes = (SenacsEmploi.query.filter_by(annee=annee)
              .order_by(SenacsEmploi.intitule.asc()).all())
    rh_salaries = [s for s in Salarie.query.order_by(Salarie.nom.asc()).all()
                   if s.actif_sur(annee)]

    par_contrat: dict[str, dict] = {}
    for p in postes:
        d = par_contrat.setdefault(p.contrat_label, {"nb": 0, "etp": 0.0})
        d["nb"] += 1
        d["etp"] = round(d["etp"] + float(p.etp or 0), 2)
    for s in rh_salaries:
        d = par_contrat.setdefault(s.contrat_label, {"nb": 0, "etp": 0.0})
        d["nb"] += 1
        d["etp"] = round(d["etp"] + float(s.etp or 0), 2)

    total_etp_manuel = round(sum(float(p.etp or 0) for p in postes), 2)
    rh_etp = round(sum(float(s.etp or 0) for s in rh_salaries), 2)
    return {
        "postes": postes,
        "nb_postes": len(postes),
        "total_etp": total_etp_manuel,
        "rh_salaries": rh_salaries,
        "nb_rh": len(rh_salaries),
        "rh_etp": rh_etp,
        "masse_salariale": round(sum(float(s.salaire_brut_charge or 0) for s in rh_salaries), 2),
        "etp_global": round(total_etp_manuel + rh_etp, 2),
        "par_contrat": par_contrat,
    }


def finances_annee(annee: int) -> dict:
    """Onglet « finances » : produits par financeur et charges réalisées,
    à partir des subventions de l'exercice + valorisation du bénévolat."""
    from app.models import Subvention

    subs = Subvention.query.filter_by(est_archive=False, annee_exercice=annee).all()
    par_financeur: dict[str, dict] = {}
    total_charges_reelles = 0.0
    for s in subs:
        fin = (s.financeur or "").strip() or "— Non renseigné —"
        d = par_financeur.setdefault(fin, {"attribue": 0.0, "recu": 0.0, "nb": 0})
        d["attribue"] = round(d["attribue"] + float(s.montant_attribue or 0), 2)
        d["recu"] = round(d["recu"] + float(s.montant_recu or 0), 2)
        d["nb"] += 1
        total_charges_reelles += float(s.total_reel_lignes or 0)

    benevolat = benevolat_annee(annee)
    emplois = emplois_annee(annee)
    return {
        "par_financeur": dict(sorted(par_financeur.items(), key=lambda kv: -kv[1]["attribue"])),
        "total_attribue": round(sum(d["attribue"] for d in par_financeur.values()), 2),
        "total_recu": round(sum(d["recu"] for d in par_financeur.values()), 2),
        "total_charges_reelles": round(total_charges_reelles, 2),
        "valorisation_benevolat": benevolat["valorisation"],
        "masse_salariale": emplois["masse_salariale"],
    }


def evenementiel_annee(annee: int) -> dict:
    """Volet événementiel : séances marquées « événement » sur l'année."""
    eff = db.func.coalesce(SessionActivite.date_session, SessionActivite.rdv_date)
    sessions = (SessionActivite.query
                .filter(SessionActivite.is_deleted.is_(False))
                .filter(SessionActivite.est_evenement.is_(True))
                .filter(eff >= date(annee, 1, 1), eff <= date(annee, 12, 31))
                .all())
    ids = [s.id for s in sessions]
    participations = 0
    uniques = 0
    if ids:
        participations = (db.session.query(db.func.count(PresenceActivite.id))
                          .filter(PresenceActivite.session_id.in_(ids)).scalar() or 0)
        uniques = (db.session.query(db.func.count(db.distinct(PresenceActivite.participant_id)))
                   .filter(PresenceActivite.session_id.in_(ids)).scalar() or 0)
    return {
        "nb_evenements": len(sessions),
        "participations": int(participations),
        "participants_uniques": int(uniques),
        "sessions": sessions,
    }
