"""Export « droit d'accès » RGPD (article 15).

Construit un classeur Excel rassemblant toutes les données qu'une
personne est en droit de réclamer : identité, droit à l'image,
orientations, présences aux activités, parcours d'insertion, notes
pédagogiques, pièces jointes et questionnaires.

La CNIL impose de fournir cette copie sous un mois : ce module permet
de répondre en un clic, sans compétence technique.
"""
from openpyxl import Workbook
from openpyxl.styles import Font

from app.models import (
    Evaluation,
    ObjectifSuivi,
    OrientationAccesDroit,
    Participant,
    PasseportNote,
    PasseportPieceJointe,
    PresenceActivite,
    QuestionnaireResponseGroup,
)

DROIT_IMAGE_LIBELLES = {
    "non_renseigne": "Non renseigné",
    "accepte": "Accepté",
    "refuse": "Refusé",
}


def _texte(value) -> str:
    if value is None:
        return ""
    return str(value)


def _ecrire_feuille(wb: Workbook, titre: str, entetes: list[str], lignes: list[list]):
    ws = wb.create_sheet(title=titre[:31])
    ws.append(entetes)
    for cellule in ws[1]:
        cellule.font = Font(bold=True)
    for ligne in lignes:
        ws.append([_texte(v) for v in ligne])
    # Largeurs lisibles
    for idx, entete in enumerate(entetes, start=1):
        largeur = max([len(entete)] + [len(_texte(l[idx - 1])) for l in lignes] or [10])
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(max(largeur + 2, 12), 60)
    return ws


def construire_export_rgpd(participant: Participant) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    # --- Identité ---
    identite = [
        ("Nom", participant.nom),
        ("Prénom", participant.prenom),
        ("Date de naissance", participant.date_naissance),
        ("Genre", participant.genre),
        ("Adresse", participant.adresse),
        ("Ville", participant.ville),
        ("Quartier", participant.quartier.nom if participant.quartier else ""),
        ("Email", participant.email),
        ("Téléphone", participant.telephone),
        ("Type de public", participant.type_public),
        ("Secteur de rattachement", participant.created_secteur),
        ("Fiche créée le", participant.created_at),
        ("Dernière modification", participant.updated_at),
        ("Pays d'origine", participant.pays_origine),
        ("Titre de séjour", participant.titre_sejour_type),
        ("Diplôme obtenu", participant.diplome_obtenu),
    ]
    _ecrire_feuille(wb, "Identité", ["Donnée", "Valeur"], [list(l) for l in identite])

    # --- Droit à l'image ---
    _ecrire_feuille(
        wb,
        "Droit à l'image",
        ["Statut", "Date du recueil", "Recueilli par"],
        [[
            DROIT_IMAGE_LIBELLES.get(participant.droit_image_statut, participant.droit_image_statut),
            participant.droit_image_date,
            participant.droit_image_recueilli_par,
        ]],
    )

    # --- Orientations / accès aux droits ---
    orientations = (
        OrientationAccesDroit.query.filter_by(participant_id=participant.id)
        .order_by(OrientationAccesDroit.date_orientation.desc())
        .all()
    )
    _ecrire_feuille(
        wb,
        "Orientations",
        ["Date", "Domaine", "Demande", "Statut", "Urgence", "Suite prévue", "Note"],
        [[o.date_orientation, o.domaine, o.demande, o.statut, o.urgence, o.suite_prevue, o.note] for o in orientations],
    )

    # --- Présences aux activités ---
    presences = (
        PresenceActivite.query.filter_by(participant_id=participant.id)
        .order_by(PresenceActivite.id.desc())
        .all()
    )
    lignes_presences = []
    for pr in presences:
        session = pr.session if hasattr(pr, "session") else None
        atelier = getattr(session, "atelier", None) if session else None
        lignes_presences.append([
            getattr(session, "date_session", None) or getattr(session, "date", None),
            getattr(atelier, "nom", ""),
            getattr(session, "secteur", ""),
            pr.motif or pr.motif_autre or "",
        ])
    _ecrire_feuille(wb, "Présences activités", ["Date", "Atelier", "Secteur", "Motif"], lignes_presences)

    # --- Notes pédagogiques (passeport) ---
    notes = (
        PasseportNote.query.filter_by(participant_id=participant.id)
        .order_by(PasseportNote.created_at.desc())
        .all()
    )
    _ecrire_feuille(
        wb,
        "Notes pédagogiques",
        ["Date", "Catégorie", "Secteur", "Contenu"],
        [[n.created_at, n.categorie, n.secteur, n.contenu] for n in notes],
    )

    # --- Pièces jointes (passeport) ---
    pieces = (
        PasseportPieceJointe.query.filter_by(participant_id=participant.id)
        .order_by(PasseportPieceJointe.created_at.desc())
        .all()
    )
    _ecrire_feuille(
        wb,
        "Pièces jointes",
        ["Date", "Titre", "Nom du fichier", "Catégorie"],
        [[pj.created_at, pj.titre, pj.original_name, pj.categorie] for pj in pieces],
    )

    # --- Évaluations et objectifs suivis ---
    evaluations = Evaluation.query.filter_by(participant_id=participant.id).all()
    _ecrire_feuille(
        wb,
        "Évaluations",
        ["Identifiant", "Détail"],
        [[e.id, _resume_objet(e)] for e in evaluations],
    )
    objectifs = ObjectifSuivi.query.filter_by(participant_id=participant.id).all()
    _ecrire_feuille(
        wb,
        "Objectifs suivis",
        ["Identifiant", "Détail"],
        [[o.id, _resume_objet(o)] for o in objectifs],
    )

    # --- Questionnaires ---
    groupes = QuestionnaireResponseGroup.query.filter_by(participant_id=participant.id).all()
    _ecrire_feuille(
        wb,
        "Questionnaires",
        ["Identifiant", "Détail"],
        [[g.id, _resume_objet(g)] for g in groupes],
    )

    # --- Parcours insertion (tables dédiées) ---
    lignes_insertion = []
    for attr in ("insertion_profile", "insertion_parcours", "insertion_positionnements", "insertion_certifications"):
        objets = getattr(participant, attr, None)
        if objets is None:
            continue
        if not isinstance(objets, (list, tuple)):
            try:
                objets = list(objets)
            except TypeError:
                objets = [objets]
        for obj in objets:
            if obj is not None:
                lignes_insertion.append([attr, _resume_objet(obj)])
    _ecrire_feuille(wb, "Insertion", ["Rubrique", "Détail"], lignes_insertion)

    return wb


def _resume_objet(obj) -> str:
    """Représentation lisible de toutes les colonnes simples d'une ligne."""
    morceaux = []
    for colonne in obj.__table__.columns:
        nom = colonne.name
        if nom in {"id", "participant_id"} or nom.endswith("_id"):
            continue
        valeur = getattr(obj, nom, None)
        if valeur not in (None, ""):
            morceaux.append(f"{nom}: {valeur}")
    return " | ".join(morceaux)
