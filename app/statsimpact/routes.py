from __future__ import annotations

from collections import Counter
from copy import deepcopy

from datetime import date, timedelta
from io import BytesIO, StringIO
import csv

import os
import re
from urllib.parse import urlencode

from flask import Blueprint, abort, render_template, request, redirect, url_for, flash, current_app, send_file, Response
from flask_login import login_required, current_user
from app.rbac import can

from sqlalchemy import func
from sqlalchemy.orm import joinedload

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.extensions import db

from app.models import (
    AtelierActivite,
    Participant,
    Quartier,
    PresenceActivite,
    SessionActivite,
    Projet,
    ProjetAtelier,
    Competence,
    Evaluation,
    Referentiel,
    Objectif,
)

from app.activite.services.docx_utils import generate_participant_bilan_pdf
from app.services.quartiers import normalize_quartier_for_ville

from .occupancy import compute_occupancy_stats
from app.services.consumption import aggregate_sessions_consumption, calculate_session_consumption, aggregate_individual_consumption

from .engine import (
    compute_volume_activity_stats,
    compute_participation_frequency_stats,
    compute_transversalite_stats,
    compute_demography_stats,
    compute_participants_stats,
    compute_magatomatique,
    normalize_filters,
    _has_atelier_filter,
    _apply_common_filters,
    _session_date_expr,
    _session_duration_minutes,
    _resolve_secteur_scope,
)

bp = Blueprint("statsimpact", __name__, url_prefix="")


CSV_FIELD_GROUPS = [
    {
        "label": "Participant",
        "fields": [
            {"key": "participant_id", "label": "ID participant"},
            {"key": "participant_nom", "label": "Nom"},
            {"key": "participant_prenom", "label": "Prénom"},
            {"key": "participant_email", "label": "Email"},
            {"key": "participant_telephone", "label": "Téléphone"},
            {"key": "participant_ville", "label": "Ville"},
            {"key": "participant_quartier", "label": "Quartier"},
            {"key": "participant_genre", "label": "Genre"},
            {"key": "participant_type_public", "label": "Type public"},
            {"key": "participant_date_naissance", "label": "Date naissance"},
            {"key": "participant_pays_origine", "label": "Pays d'origine"},
            {"key": "participant_titre_sejour_type", "label": "Titre de séjour (type)"},
            {"key": "participant_date_entree_dispositif", "label": "Date entrée dispositif"},
            {"key": "participant_date_sortie_dispositif", "label": "Date sortie dispositif"},
            {"key": "participant_diplome_obtenu", "label": "Diplôme obtenu"},
            {"key": "participant_cir_obtenu", "label": "CIR obtenu"},
        ],
    },
    {
        "label": "Session",
        "fields": [
            {"key": "session_id", "label": "ID session"},
            {"key": "session_date", "label": "Date session"},
            {"key": "session_type", "label": "Type session"},
            {"key": "session_statut", "label": "Statut session"},
            {"key": "session_heure_debut", "label": "Heure début"},
            {"key": "session_heure_fin", "label": "Heure fin"},
            {"key": "session_duree_minutes", "label": "Durée (minutes)"},
        ],
    },
    {
        "label": "Atelier",
        "fields": [
            {"key": "atelier_id", "label": "ID atelier"},
            {"key": "atelier_nom", "label": "Nom atelier"},
            {"key": "atelier_secteur", "label": "Secteur atelier"},
            {"key": "atelier_type", "label": "Type atelier"},
        ],
    },
    {
        "label": "Présence",
        "fields": [
            {"key": "presence_id", "label": "ID présence"},
            {"key": "presence_motif", "label": "Motif"},
            {"key": "presence_motif_autre", "label": "Motif autre"},
            {"key": "presence_created_at", "label": "Date d'émargement"},
        ],
    },
]

CSV_DEFAULT_FIELDS = [
    "participant_nom",
    "participant_prenom",
    "atelier_nom",
    "atelier_secteur",
    "session_date",
    "session_type",
]


def _fmt_date(value) -> str:
    return value.strftime("%Y-%m-%d") if value else ""


def _fmt_datetime(value) -> str:
    return value.strftime("%Y-%m-%d %H:%M") if value else ""


CSV_FIELD_MAP = {
    "participant_id": {"label": "ID participant", "getter": lambda ctx: ctx["participant"].id},
    "participant_nom": {"label": "Nom", "getter": lambda ctx: ctx["participant"].nom or ""},
    "participant_prenom": {"label": "Prénom", "getter": lambda ctx: ctx["participant"].prenom or ""},
    "participant_email": {"label": "Email", "getter": lambda ctx: ctx["participant"].email or ""},
    "participant_telephone": {"label": "Téléphone", "getter": lambda ctx: ctx["participant"].telephone or ""},
    "participant_ville": {"label": "Ville", "getter": lambda ctx: ctx["participant"].ville or ""},
    "participant_quartier": {
        "label": "Quartier",
        "getter": lambda ctx: ctx["quartier"].nom if ctx["quartier"] else "",
    },
    "participant_genre": {"label": "Genre", "getter": lambda ctx: ctx["participant"].genre or ""},
    "participant_type_public": {"label": "Type public", "getter": lambda ctx: ctx["participant"].type_public or ""},
    "participant_date_naissance": {
        "label": "Date naissance",
        "getter": lambda ctx: _fmt_date(ctx["participant"].date_naissance),
    },
    "participant_pays_origine": {"label": "Pays d'origine", "getter": lambda ctx: ctx["participant"].pays_origine or ""},
    "participant_titre_sejour_type": {"label": "Titre de séjour (type)", "getter": lambda ctx: ctx["participant"].titre_sejour_type or ""},
    "participant_date_entree_dispositif": {
        "label": "Date entrée dispositif",
        "getter": lambda ctx: _fmt_date(ctx["participant"].date_entree_dispositif),
    },
    "participant_date_sortie_dispositif": {
        "label": "Date sortie dispositif",
        "getter": lambda ctx: _fmt_date(ctx["participant"].date_sortie_dispositif),
    },
    "participant_diplome_obtenu": {"label": "Diplôme obtenu", "getter": lambda ctx: ctx["participant"].diplome_obtenu or ""},
    "participant_cir_obtenu": {
        "label": "CIR obtenu",
        "getter": lambda ctx: (
            "Oui"
            if ctx["participant"].cir_obtenu is True
            else ("Non" if ctx["participant"].cir_obtenu is False else "")
        ),
    },
    "session_id": {"label": "ID session", "getter": lambda ctx: ctx["session"].id},
    "session_date": {
        "label": "Date session",
        "getter": lambda ctx: _fmt_date(ctx["session"].rdv_date or ctx["session"].date_session),
    },
    "session_type": {"label": "Type session", "getter": lambda ctx: ctx["session"].session_type or ""},
    "session_statut": {"label": "Statut session", "getter": lambda ctx: ctx["session"].statut or ""},
    "session_heure_debut": {
        "label": "Heure début",
        "getter": lambda ctx: ctx["session"].rdv_debut or ctx["session"].heure_debut or "",
    },
    "session_heure_fin": {
        "label": "Heure fin",
        "getter": lambda ctx: ctx["session"].rdv_fin or ctx["session"].heure_fin or "",
    },
    "session_duree_minutes": {
        "label": "Durée (minutes)",
        "getter": lambda ctx: ctx["session"].duree_minutes or "",
    },
    "atelier_id": {"label": "ID atelier", "getter": lambda ctx: ctx["atelier"].id},
    "atelier_nom": {"label": "Nom atelier", "getter": lambda ctx: ctx["atelier"].nom or ""},
    "atelier_secteur": {"label": "Secteur atelier", "getter": lambda ctx: ctx["atelier"].secteur or ""},
    "atelier_type": {"label": "Type atelier", "getter": lambda ctx: ctx["atelier"].type_atelier or ""},
    "presence_id": {"label": "ID présence", "getter": lambda ctx: ctx["presence"].id},
    "presence_motif": {"label": "Motif", "getter": lambda ctx: ctx["presence"].motif or ""},
    "presence_motif_autre": {"label": "Motif autre", "getter": lambda ctx: ctx["presence"].motif_autre or ""},
    "presence_created_at": {
        "label": "Date d'émargement",
        "getter": lambda ctx: _fmt_datetime(ctx["presence"].created_at),
    },
}

SENSITIVE_INSERTION_FIELDS = {
    "participant_pays_origine",
    "participant_titre_sejour_type",
    "participant_date_entree_dispositif",
    "participant_date_sortie_dispositif",
    "participant_diplome_obtenu",
    "participant_cir_obtenu",
}


def _csv_fields_for_current_user() -> tuple[list[dict], list[str], set[str]]:
    can_sensitive = can("insertion:sensitive_view") and can("statsimpact:insertion_export")
    allowed_fields = set(CSV_FIELD_MAP.keys())
    if not can_sensitive:
        allowed_fields -= SENSITIVE_INSERTION_FIELDS

    groups: list[dict] = []
    for group in CSV_FIELD_GROUPS:
        filtered = [f for f in group["fields"] if f["key"] in allowed_fields]
        if filtered:
            groups.append({"label": group["label"], "fields": filtered})

    defaults = [f for f in CSV_DEFAULT_FIELDS if f in allowed_fields]
    return groups, defaults, allowed_fields


def _can_view() -> bool:
    return can("statsimpact:view") or can("statsimpact:view_all")


def _quality_participant_item(participant: Participant, detail: str) -> dict:
    label = f"{participant.nom or ''} {participant.prenom or ''}".strip() or f"Participant #{participant.id}"
    return {
        "label": label,
        "detail": detail,
        "url": url_for("participants.edit_participant", participant_id=participant.id),
    }


def _statsimpact_quality_payload(flt) -> dict:
    base = db.session.query(SessionActivite, AtelierActivite).join(
        AtelierActivite, AtelierActivite.id == SessionActivite.atelier_id
    )
    base = _apply_common_filters(base, flt)
    sessions_rows = base.order_by(_session_date_expr().desc()).all()
    session_ids = [session.id for session, _ in sessions_rows]

    presences = []
    if session_ids:
        presences = PresenceActivite.query.filter(PresenceActivite.session_id.in_(session_ids)).all()

    presence_counts = Counter(pr.session_id for pr in presences)
    participant_ids = sorted({pr.participant_id for pr in presences if pr.participant_id})
    participants = []
    if participant_ids:
        participants = (
            Participant.query
            .filter(Participant.id.in_(participant_ids))
            .order_by(Participant.nom.asc(), Participant.prenom.asc())
            .all()
        )

    def participant_issue(code: str, title: str, why: str, rows: list[Participant], severity: str) -> dict:
        return {
            "code": code,
            "title": title,
            "why": why,
            "severity": severity,
            "count": len(rows),
            "items": [_quality_participant_item(p, f"{p.ville or 'Ville inconnue'} · {p.quartier.nom if p.quartier else 'Quartier inconnu'}") for p in rows[:12]],
        }

    issues = [
        participant_issue(
            "participant_no_quartier",
            "Participants sans quartier",
            "Impact direct sur les statistiques QPV, hors QPV et quartier inconnu.",
            [p for p in participants if not getattr(p, "quartier_id", None)],
            "bad",
        ),
        participant_issue(
            "participant_no_genre",
            "Participants sans genre",
            "Fragilise les répartitions femmes / hommes / autre dans les bilans.",
            [p for p in participants if not (p.genre or "").strip()],
            "warn",
        ),
        participant_issue(
            "participant_no_birthdate",
            "Participants sans date de naissance",
            "Empêche le calcul fiable des tranches d'âge et de l'âge moyen.",
            [p for p in participants if not p.date_naissance],
            "warn",
        ),
        participant_issue(
            "participant_no_city",
            "Participants sans ville",
            "Rend moins lisibles l'origine géographique et les contrôles Creil / hors Creil.",
            [p for p in participants if not (p.ville or "").strip()],
            "warn",
        ),
    ]

    empty_sessions = [
        (session, atelier)
        for session, atelier in sessions_rows
        if presence_counts.get(session.id, 0) == 0
        and (session.statut or "").lower() != "annulee"
        and ((session.rdv_date or session.date_session) is None or (session.rdv_date or session.date_session) <= date.today())
    ]
    issues.append({
        "code": "session_no_presence",
        "title": "Séances réalisées sans présence",
        "why": "Une séance sans émargement peut faire baisser les moyennes et masquer une feuille non saisie.",
        "severity": "bad",
        "count": len(empty_sessions),
        "items": [
            {
                "label": f"{atelier.nom} · {(session.rdv_date or session.date_session).strftime('%d/%m/%Y') if (session.rdv_date or session.date_session) else 'Date inconnue'}",
                "detail": atelier.secteur,
                "url": url_for("activite.emargement", session_id=session.id),
            }
            for session, atelier in empty_sessions[:12]
        ],
    })

    eff_secteur = _resolve_secteur_scope(flt)
    ateliers_q = AtelierActivite.query.filter(
        AtelierActivite.is_deleted.is_(False),
        AtelierActivite.is_active.is_(True),
    )
    if eff_secteur == "__restricted__":
        ateliers_q = ateliers_q.filter(AtelierActivite.id == -1)
    elif eff_secteur:
        ateliers_q = ateliers_q.filter(AtelierActivite.secteur == eff_secteur)
    if flt.atelier_ids:
        ateliers_q = ateliers_q.filter(AtelierActivite.id.in_(flt.atelier_ids))
    ateliers = ateliers_q.order_by(AtelierActivite.secteur.asc(), AtelierActivite.nom.asc()).all()
    linked_atelier_ids = {
        row[0] for row in db.session.query(ProjetAtelier.atelier_id).distinct().all() if row and row[0]
    }
    ateliers_without_project = [a for a in ateliers if a.id not in linked_atelier_ids]
    issues.append({
        "code": "atelier_no_project",
        "title": "Ateliers actifs sans projet lié",
        "why": "Les indicateurs de projet ne peuvent pas piocher ces ateliers tant qu'ils ne sont pas rattachés.",
        "severity": "info",
        "count": len(ateliers_without_project),
        "items": [
            {
                "label": a.nom,
                "detail": a.secteur,
                "url": url_for("activite.sessions", atelier_id=a.id),
            }
            for a in ateliers_without_project[:12]
        ],
    })

    bad = sum(issue["count"] for issue in issues if issue["severity"] == "bad")
    warn = sum(issue["count"] for issue in issues if issue["severity"] == "warn")
    score = max(0, min(100, 100 - (bad * 8) - (warn * 3)))

    return {
        "score": score,
        "sessions_count": len(sessions_rows),
        "presences_count": len(presences),
        "participants_count": len(participants),
        "issues": issues,
        "bad_count": bad,
        "warn_count": warn,
        "info_count": sum(issue["count"] for issue in issues if issue["severity"] == "info"),
    }




def _query_args_preserve_lists(multidict) -> dict:
    if not hasattr(multidict, "lists"):
        return dict(multidict or {})
    payload = {}
    for key, values in multidict.lists():
        cleaned = [value for value in values if value not in (None, "")]
        if not cleaned:
            continue
        payload[key] = cleaned if len(cleaned) > 1 else cleaned[0]
    return payload


def _safe_sheet_title(name: str, fallback: str = "Atelier") -> str:
    """Openpyxl: max 31 chars, no [ ] * ? / \\ etc."""
    if not name:
        name = fallback
    bad = set('[]:*?/\\')
    cleaned = "".join(c for c in name if c not in bad).strip()
    cleaned = cleaned[:31] if cleaned else fallback
    return cleaned




def _auto_fit_statsimpact_ws(ws, min_width: int = 10, max_width: int = 42):
    for col_cells in ws.columns:
        try:
            idx = col_cells[0].column
        except Exception:
            continue
        max_len = 0
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[get_column_letter(idx)].width = max(min_width, min(max_len + 2, max_width))


def _autosize_columns(ws, min_width: int = 10, max_width: int = 42):
    """Alias robuste utilisé par les exports Magatomatique.

    Certains blocs historiques appelaient `_auto_fit_statsimpact_ws`, le patch
    Magatomatique période appelle `_autosize_columns`. On centralise le même
    comportement pour éviter un NameError en export XLSX.
    """
    return _auto_fit_statsimpact_ws(ws, min_width=min_width, max_width=max_width)


def _append_section(ws, title: str, headers: list[str], rows: list[list], *, start_row: int = 1) -> int:
    row = start_row
    ws.cell(row=row, column=1, value=title)
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
    for idx, header in enumerate(headers, start=1):
        c = ws.cell(row=row, column=idx, value=header)
        c.font = Font(bold=True)
        c.fill = fill
        c.alignment = Alignment(horizontal="center")
    row += 1
    for data in rows:
        for idx, value in enumerate(data, start=1):
            ws.cell(row=row, column=idx, value=value)
        row += 1
    return row


def _effective_statsimpact_secteur(flt):
    try:
        eff = _resolve_secteur_scope(flt)
        return eff
    except Exception:
        return getattr(flt, 'secteur', None)


def _ensure_default_export_period(flt):
    """Garantit un périmètre temporel explicite pour les exports Stats-Impact.

    La page /stats-impact affiche déjà l'année civile courante quand aucune date
    n'est saisie. Les routes d'export doivent faire pareil, sinon les boutons
    peuvent exporter "tout l'historique" pendant que l'écran affiche "année
    courante" : recette parfaite pour des chiffres qui ne causent pas la même
    langue.
    """
    if not getattr(flt, 'date_from', None) and not getattr(flt, 'date_to', None):
        today = date.today()
        flt.date_from = date(today.year, 1, 1)
        flt.date_to = date(today.year, 12, 31)
    return flt


def _export_period_label(flt) -> str:
    start = getattr(flt, 'date_from', None)
    end = getattr(flt, 'date_to', None)
    start_txt = start.isoformat() if start else 'début non limité'
    end_txt = end.isoformat() if end else 'fin non limitée'
    return f"{start_txt} → {end_txt}"


def _export_period_slug(flt) -> str:
    start = getattr(flt, 'date_from', None)
    end = getattr(flt, 'date_to', None)
    start_txt = start.isoformat() if start else 'debut'
    end_txt = end.isoformat() if end else 'fin'
    return f"{start_txt}_{end_txt}"


def _export_query_string_from_filter(flt, extra: dict | None = None) -> str:
    """Construit des liens d'export avec les dates effectivement appliquées."""
    params: list[tuple[str, str]] = []
    secteur = getattr(flt, 'secteur', None)
    if secteur:
        params.append(('secteur', str(secteur)))
    for aid in getattr(flt, 'atelier_ids', None) or []:
        params.append(('atelier_ids', str(aid)))
    if getattr(flt, 'date_from', None):
        params.append(('date_from', flt.date_from.isoformat()))
    if getattr(flt, 'date_to', None):
        params.append(('date_to', flt.date_to.isoformat()))
    if getattr(flt, 'group_by', None):
        params.append(('group_by', str(flt.group_by)))
    if getattr(flt, 'periode_id', None):
        params.append(('periode_id', str(flt.periode_id)))
    for key, value in (extra or {}).items():
        if value is None:
            continue
        params.append((str(key), str(value)))
    return urlencode(params, doseq=True)


def _magatomatique_filename(prefix: str, flt) -> str:
    return f"{prefix}_{_export_period_slug(flt)}.xlsx"


def _build_statsimpact_scope_workbook(flt) -> Workbook:
    flt = _ensure_default_export_period(flt)
    stats = compute_volume_activity_stats(flt)
    stats = _enrich_stats_with_consumption(stats, flt)
    consumption_stats = _consumption_stats_for_filter(flt)
    individual_consumption_stats = _individual_consumption_stats_for_filter(flt)
    freq = compute_participation_frequency_stats(flt)
    demo = compute_demography_stats(flt)
    occupancy = compute_occupancy_stats(flt)
    participants = compute_participants_stats(flt)

    wb = Workbook()

    ws_params = wb.active
    ws_params.title = 'Paramètres'
    atelier_ids = list(getattr(flt, 'atelier_ids', None) or [])
    param_rows = [
        ['Secteur demandé', getattr(flt, 'secteur', None) or 'Tous'],
        ['Secteur effectif', _effective_statsimpact_secteur(flt) or 'Tous'],
        ['Du', getattr(flt, 'date_from', None).isoformat() if getattr(flt, 'date_from', None) else ''],
        ['Au', getattr(flt, 'date_to', None).isoformat() if getattr(flt, 'date_to', None) else ''],
        ['Groupement', getattr(flt, 'group_by', None) or 'MONTH'],
        ['Ateliers sélectionnés', ', '.join(str(x) for x in atelier_ids) if atelier_ids else 'Tous'],
        ["Date d'export", date.today().isoformat()],
        ['Utilisateur', getattr(current_user, 'email', None) or getattr(current_user, 'username', None) or f"user#{getattr(current_user, 'id', '')}"],
    ]
    _append_section(ws_params, "Périmètre exact de l'export", ['Paramètre', 'Valeur'], param_rows, start_row=1)

    ws_syn = wb.create_sheet('Synthèse')
    kpi = stats.get('kpi', {}) or {}
    _append_section(ws_syn, 'Indicateurs clés', ['Indicateur', 'Valeur'], [
        ['Sessions', kpi.get('sessions', 0)],
        ['Présences', kpi.get('presences', 0)],
        ['Participants uniques', kpi.get('uniques', 0)],
        ['Nouveaux participants', kpi.get('new_participants', 0)],
        ['Heures animateur', kpi.get('hours_animator', 0)],
        ['Heures participants', kpi.get('hours_people', 0)],
        ['Moyenne / session', kpi.get('avg_per_session', 0)],
        ['Durée activité (jours)', kpi.get('activity_duration_days', '')],
        ['Fréquence moyenne', freq.get('freq_avg', 0)],
        ['Taux de retour (%)', freq.get('returning_rate', 0)],
        ['Participants réguliers (4+)', freq.get('regulars_4plus', 0)],
        ['Âge moyen', demo.get('age_avg', '')],
        ['Consommation estimée (kWh)', consumption_stats.get('total_kwh', 0)],
        ['CO₂ estimé (kg)', consumption_stats.get('total_co2', 0)],
        ['Sessions équipées', consumption_stats.get('sessions_count', 0)],
        ['Moyenne kWh / séance équipée', consumption_stats.get('avg_kwh_per_session', 0)],
        ['Conso individuelle tracée (kWh)', individual_consumption_stats.get('total_kwh', 0)],
        ['CO₂ individuel tracé (kg)', individual_consumption_stats.get('total_co2', 0)],
        ['Présences avec matériel individuel', individual_consumption_stats.get('presences_count', 0)],
        ['Habitants concernés', individual_consumption_stats.get('participants_count', 0)],
    ], start_row=1)

    ws_at = wb.create_sheet('Ateliers')
    atelier_rows = []
    for r in stats.get('table_ateliers', []) or []:
        atelier_rows.append([
            r.get('secteur', ''), r.get('nom', ''), r.get('type_atelier', ''),
            r.get('sessions_planned', 0), r.get('sessions_real', 0), r.get('presences', 0),
            r.get('uniques', 0), r.get('real_hours', 0), r.get('occupation_rate', 0),
            r.get('avg_per_session_real', 0), r.get('activity_duration_days', ''),
            r.get('conso_kwh', 0), r.get('co2_kg', 0), r.get('conso_sessions', 0),
            r.get('conso_indiv_kwh', 0), r.get('co2_indiv_kg', 0), r.get('conso_indiv_presences', 0),
        ])
    _append_section(ws_at, 'Vue ateliers', ['Secteur','Atelier','Type','Sessions prév.','Sessions réelles','Présences','Uniques','Heures réalisées',"Taux d'occupation (%)",'Moyenne / séance','Durée activité (j)','Conso séance kWh','CO₂ séance kg','Séances équipées','Conso ind. kWh','CO₂ ind. kg','Présences équipées'], atelier_rows, start_row=1)

    ws_pub = wb.create_sheet('Publics')
    row = _append_section(ws_pub, 'Répartition par genre', ['Genre', 'Total'], [[k, v] for k, v in (demo.get('genre') or {}).items()], start_row=1)
    row += 1
    row = _append_section(ws_pub, 'Répartition par âge', ['Tranche', 'Total'], [[k, v] for k, v in (demo.get('age_buckets') or {}).items()], start_row=row)
    row += 1
    row = _append_section(ws_pub, 'Type de public', ['Type', 'Total'], [[k, v] for k, v in (demo.get('type_public') or {}).items()], start_row=row)
    row += 1
    qpv_detail_rows = []
    for item in demo.get('qpv_details') or []:
        genre = item.get('genre') or {}
        ages = item.get('age_buckets') or {}
        qpv_detail_rows.append([
            item.get('label', ''),
            item.get('total', 0),
            genre.get('Femmes', 0),
            genre.get('Hommes', 0),
            genre.get('Autre / non renseigné', 0),
            ages.get('0-10', 0),
            ages.get('11-17', 0),
            ages.get('18-25', 0),
            ages.get('26-59', 0),
            ages.get('60+', 0),
            ages.get('Inconnu', 0),
        ])
    row = _append_section(
        ws_pub,
        'Détail QPV / genre / âge',
        ['Typologie', 'Total', 'Femmes', 'Hommes', 'Autre / NR', '0-10', '11-17', '18-25', '26-59', '60+', 'Âge inconnu'],
        qpv_detail_rows,
        start_row=row,
    )
    row += 1
    villes = [[r.get('ville', ''), r.get('count', 0)] for r in (demo.get('villes_top') or [])]
    _append_section(ws_pub, 'Top villes', ['Ville', 'Total'], villes, start_row=row)

    ws_fill = wb.create_sheet('Remplissage')
    fill_rows = [
        ['Sessions collectives', occupancy.get('collective_sessions', 0)],
        ['Présences collectives', occupancy.get('collective_presences', 0)],
        ['Remplissage moyen (%)', occupancy.get('avg_fill_rate_pct', '')],
        ['Capacité par défaut', occupancy.get('default_capacity', '')],
        ['Bucket <50%', (occupancy.get('buckets') or {}).get('<50%', 0)],
        ['Bucket 50-79%', (occupancy.get('buckets') or {}).get('50-79%', 0)],
        ['Bucket 80-99%', (occupancy.get('buckets') or {}).get('80-99%', 0)],
        ['Bucket 100%+', (occupancy.get('buckets') or {}).get('100%+', 0)],
    ]
    row = _append_section(ws_fill, 'Synthèse remplissage', ['Indicateur', 'Valeur'], fill_rows, start_row=1)
    row += 1
    occ_rows = []
    for r in occupancy.get('per_atelier', []) or []:
        occ_rows.append([r.get('secteur', ''), r.get('nom', ''), r.get('sessions', 0), r.get('presences', 0), r.get('capacity_total', 0), r.get('avg_fill_rate_pct', 0)])
    _append_section(ws_fill, 'Remplissage par atelier', ['Secteur','Atelier','Sessions','Présences','Capacité cumulée','Remplissage moyen (%)'], occ_rows, start_row=row)

    if can('participants:view') or can('participants:view_all'):
        ws_part = wb.create_sheet('Participants')
        part_rows = []
        for p in participants.get('participants', []) or []:
            ateliers_txt = ', '.join(sorted([a.get('atelier', '') for a in (p.get('ateliers') or {}).values() if a.get('atelier')]))
            part_rows.append([
                p.get('nom', ''), p.get('prenom', ''), p.get('ville', ''), p.get('quartier', ''),
                p.get('genre', ''), p.get('type_public', ''), p.get('visites', 0), ateliers_txt
            ])
        _append_section(ws_part, 'Participants du périmètre', ['Nom','Prénom','Ville','Quartier','Genre','Type public','Visites','Ateliers fréquentés'], part_rows, start_row=1)

    for ws in wb.worksheets:
        _auto_fit_statsimpact_ws(ws)
    return wb

def _pedago_scope_secteur() -> str | None:
    if can("scope:all_secteurs"):
        return None
    return (getattr(current_user, "secteur_assigne", None) or "").strip() or None


def _build_bilan_rows(participant: Participant) -> list[dict]:
    eval_rows = (
        db.session.query(
            Evaluation,
            Competence,
            Referentiel,
            SessionActivite,
            AtelierActivite,
        )
        .join(Competence, Evaluation.competence_id == Competence.id)
        .join(Referentiel, Competence.referentiel_id == Referentiel.id)
        .outerjoin(SessionActivite, Evaluation.session_id == SessionActivite.id)
        .outerjoin(AtelierActivite, SessionActivite.atelier_id == AtelierActivite.id)
        .filter(Evaluation.participant_id == participant.id, Evaluation.etat == 2)
        .order_by(Referentiel.nom.asc(), Competence.code.asc(), Evaluation.date_evaluation.asc())
        .all()
    )

    rows: list[dict] = []
    for eval_obj, comp, ref, session, atelier in eval_rows:
        date_label = eval_obj.date_evaluation.strftime("%d/%m/%Y") if eval_obj.date_evaluation else ""
        atelier_label = atelier.nom if atelier else ""
        rows.append(
            {
                "referentiel": ref.nom,
                "competence": f"{comp.code} · {comp.nom}",
                "date": date_label,
                "atelier": atelier_label,
            }
        )
    return rows


def _participants_success_rate(session_id: int, competences: list[Competence]) -> dict:
    if not competences:
        return {"total": 0, "success": 0, "ratio": 0}
    presences = PresenceActivite.query.filter_by(session_id=session_id).all()
    total = len(presences)
    if total == 0:
        return {"total": 0, "success": 0, "ratio": 0}
    comp_ids = [c.id for c in competences]
    success_count = 0
    for pr in presences:
        evals = (
            Evaluation.query.filter(
                Evaluation.session_id == session_id,
                Evaluation.participant_id == pr.participant_id,
                Evaluation.competence_id.in_(comp_ids),
                Evaluation.etat >= 2,
            )
            .distinct()
            .count()
        )
        if evals == len(comp_ids):
            success_count += 1
    ratio = (success_count / total * 100) if total else 0
    return {"total": total, "success": success_count, "ratio": ratio}


def _objective_success(obj: Objectif) -> dict:
    if obj.type == "operationnel" and obj.session_id:
        stats = _participants_success_rate(obj.session_id, obj.competences)
        validated = stats["ratio"] >= (obj.seuil_validation or 0)
        return {"ratio": stats["ratio"], "validated": validated, "total": stats["total"], "success": stats["success"]}

    enfants = obj.enfants or []
    if not enfants:
        return {"ratio": 0, "validated": False, "total": 0, "success": 0}
    results = [ _objective_success(child) for child in enfants ]
    total = len(results)
    success = sum(1 for r in results if r["validated"])
    ratio = (success / total * 100) if total else 0
    validated = ratio >= (obj.seuil_validation or 0)
    return {"ratio": ratio, "validated": validated, "total": total, "success": success}


def _scoped_participants_query_for_stats(flt):
    eff_secteur = _resolve_secteur_scope(flt)
    q = db.session.query(Participant).distinct()
    if eff_secteur == "__restricted__":
        return q.filter(Participant.id == -1)
    if eff_secteur:
        q = (
            q.join(PresenceActivite, PresenceActivite.participant_id == Participant.id)
            .join(SessionActivite, PresenceActivite.session_id == SessionActivite.id)
            .join(AtelierActivite, SessionActivite.atelier_id == AtelierActivite.id)
            .filter(SessionActivite.is_deleted.is_(False))
            .filter(AtelierActivite.is_deleted.is_(False))
            .filter(AtelierActivite.secteur == eff_secteur)
        )
    return q


def _query_presence_export(flt, participant_q: str | None = None):
    query = (
        db.session.query(
            PresenceActivite,
            Participant,
            SessionActivite,
            AtelierActivite,
            Quartier,
        )
        .join(Participant, PresenceActivite.participant_id == Participant.id)
        .join(SessionActivite, PresenceActivite.session_id == SessionActivite.id)
        .join(AtelierActivite, SessionActivite.atelier_id == AtelierActivite.id)
        .outerjoin(Quartier, Participant.quartier_id == Quartier.id)
    )
    query = _apply_common_filters(query, flt)

    if participant_q:
        like = f"%{participant_q.lower()}%"
        query = query.filter(
            func.lower(func.coalesce(Participant.nom, "")).like(like)
            | func.lower(func.coalesce(Participant.prenom, "")).like(like)
        )

    query = query.order_by(_session_date_expr().asc(), Participant.nom.asc(), Participant.prenom.asc())
    return query


@bp.route("/stats/pedagogie", methods=["GET"])
@login_required
def stats_pedagogie():
    if not _can_view():
        abort(403)

    secteur = _pedago_scope_secteur()
    
    flt = normalize_filters(request.args, user=current_user)

    projets_q = Projet.query
    ateliers_q = _apply_atelier_lifecycle_filters(AtelierActivite.query, flt)
    if secteur:
        projets_q = projets_q.filter(Projet.secteur == secteur)
        ateliers_q = ateliers_q.filter(AtelierActivite.secteur == secteur)

    projets = projets_q.order_by(Projet.secteur.asc(), Projet.nom.asc()).all()
    ateliers = ateliers_q.order_by(AtelierActivite.secteur.asc(), AtelierActivite.nom.asc()).all()
    participants = _scoped_participants_query_for_stats(flt).order_by(Participant.nom.asc(), Participant.prenom.asc()).all()

    projet_id = request.args.get("projet_id", type=int)
    atelier_id = request.args.get("atelier_id", type=int)
    participant_id = request.args.get("participant_id", type=int)
    participant_q = (request.args.get("participant_q") or "").strip()

    projet = Projet.query.get(projet_id) if projet_id else None
    if projet and secteur and projet.secteur != secteur:
        abort(403)

    atelier = AtelierActivite.query.get(atelier_id) if atelier_id else None
    if atelier and secteur and atelier.secteur != secteur:
        abort(403)

    participant = Participant.query.get(participant_id) if participant_id else None

    participants_for_passport_q = _scoped_participants_query_for_stats(flt)
    if participant_q:
        like = f"%{participant_q.lower()}%"
        participants_for_passport_q = participants_for_passport_q.filter(
            func.lower(func.coalesce(Participant.nom, "")).like(like)
            | func.lower(func.coalesce(Participant.prenom, "")).like(like)
        )
    participants_for_passport = participants_for_passport_q.order_by(Participant.nom.asc(), Participant.prenom.asc()).limit(200).all()

    projet_objectifs = []
    if projet:
        objectifs = Objectif.query.filter_by(projet_id=projet.id, type="general").order_by(Objectif.created_at.asc()).all()
        for obj in objectifs:
            stats = _objective_success(obj)
            projet_objectifs.append({"objectif": obj, **stats})

    atelier_stats = {}
    if atelier:
        objectifs = Objectif.query.filter_by(atelier_id=atelier.id, type="specifique").order_by(Objectif.created_at.asc()).all()
        objectifs_stats = []
        for obj in objectifs:
            stats = _objective_success(obj)
            objectifs_stats.append({"objectif": obj, **stats})
        atelier_stats = {"objectifs": objectifs_stats}

    participant_groups = []
    bilan_rows = []
    if participant:
        bilan_rows = _build_bilan_rows(participant)
        grouped: dict[str, list[dict]] = {}
        for row in bilan_rows:
            grouped.setdefault(row["referentiel"], []).append(row)
        participant_groups = [
            {"referentiel": ref, "rows": rows} for ref, rows in grouped.items()
        ]

    return render_template(
        "stats_pedagogie.html",
        projets=projets,
        ateliers=ateliers,
        participants=participants,
        projet=projet,
        atelier=atelier,
        participant=participant,
        participant_q=participant_q,
        participants_for_passport=participants_for_passport,
        projet_objectifs=projet_objectifs,
        atelier_stats=atelier_stats,
        participant_groups=participant_groups,
    )


@bp.route("/stats/pedagogie/participant/<int:participant_id>/bilan", methods=["GET"])
@login_required
def stats_pedagogie_bilan(participant_id: int):
    if not _can_view():
        abort(403)
    participant = Participant.query.get_or_404(participant_id)
    rows = _build_bilan_rows(participant)
    pdf_path = generate_participant_bilan_pdf(current_app, participant, rows)
    if pdf_path and os.path.exists(pdf_path):
        return send_file(pdf_path, as_attachment=True)
    flash("Impossible de générer le PDF.", "warning")
    return redirect(url_for("statsimpact.stats_pedagogie", participant_id=participant_id))






def _apply_atelier_lifecycle_filters(query, flt, include_inactive: bool = False):
    """Applique le filtrage métier atelier (hors corbeille + actifs par défaut)."""
    query = query.filter(AtelierActivite.is_deleted.is_(False))
    if not include_inactive and not _has_atelier_filter(flt):
        query = query.filter(AtelierActivite.is_active.is_(True))
    return query


def _sessions_for_consumption_stats(flt):
    """Sessions filtrées pour les KPI consommation.

    On ne dépend pas de compute_volume_activity_stats(), car ce service ne renvoie
    pas les objets SessionActivite complets. On reconstruit donc une requête dédiée,
    avec les mêmes filtres stats-impact.
    """
    query = (
        db.session.query(SessionActivite)
        .join(AtelierActivite, SessionActivite.atelier_id == AtelierActivite.id)
        .filter(SessionActivite.is_deleted.is_(False))
        .filter(AtelierActivite.is_deleted.is_(False))
    )

    query = _apply_common_filters(query, flt)
    query = query.filter(func.lower(func.coalesce(SessionActivite.statut, "")) != "annulee")
    query = query.order_by(_session_date_expr().asc(), SessionActivite.id.asc())
    return query.all()


def _consumption_stats_for_filter(flt):
    try:
        return aggregate_sessions_consumption(_sessions_for_consumption_stats(flt))
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        current_app.logger.exception("Impossible de calculer les statistiques de consommation stats-impact")
        return {
            "total_kwh": 0.0,
            "total_co2": 0.0,
            "sessions_count": 0,
            "avg_kwh_per_session": 0.0,
            "top_materiels": [],
        }




def _individual_consumption_stats_for_filter(flt):
    try:
        return aggregate_individual_consumption(
            date_from=getattr(flt, 'date_from', None),
            date_to=getattr(flt, 'date_to', None),
            secteur=_resolve_secteur_scope(flt) if _resolve_secteur_scope(flt) != "__restricted__" else "__NO_ACCESS__",
            atelier_ids=getattr(flt, 'atelier_ids', None) or ([getattr(flt, 'atelier_id', None)] if getattr(flt, 'atelier_id', None) else None),
        )
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        current_app.logger.exception("Impossible de calculer les statistiques de consommation individuelle stats-impact")
        return {
            "total_kwh": 0.0,
            "total_co2": 0.0,
            "sessions_count": 0,
            "presences_count": 0,
            "participants_count": 0,
            "materiel_lines_count": 0,
            "avg_kwh_per_presence": 0.0,
            "top_materiels": [],
            "by_atelier": [],
            "by_secteur": [],
        }


def _individual_consumption_by_atelier_for_filter(flt):
    stats = _individual_consumption_stats_for_filter(flt)
    return {int(row.get('atelier_id')): row for row in (stats.get('by_atelier') or []) if row.get('atelier_id')}

def _consumption_by_atelier_for_filter(flt):
    """Agrège la consommation par atelier, en respectant les filtres Stats-Impact."""
    out = {}
    for session_obj in _sessions_for_consumption_stats(flt):
        try:
            atelier = getattr(session_obj, "atelier", None) or AtelierActivite.query.get(session_obj.atelier_id)
            if not atelier:
                continue
            key = _continuity_group_key(atelier)
            payload = calculate_session_consumption(session_obj, atelier=atelier)
            bucket = out.setdefault(key, {"kwh": 0.0, "co2_kg": 0.0, "sessions_count": 0})
            bucket["kwh"] += float(payload.get("total_kwh") or 0.0)
            bucket["co2_kg"] += float(payload.get("co2_kg") or 0.0)
            if payload.get("details"):
                bucket["sessions_count"] += 1
        except Exception:
            current_app.logger.exception("Impossible d'agréger la consommation pour une séance stats-impact")
    for value in out.values():
        value["kwh"] = round(value.get("kwh", 0.0), 3)
        value["co2_kg"] = round(value.get("co2_kg", 0.0), 3)
    return out


def _enrich_stats_with_consumption(stats, flt):
    """Ajoute kWh/CO₂ aux lignes table_ateliers sans casser les stats existantes."""
    if not isinstance(stats, dict):
        return stats
    by_atelier = _consumption_by_atelier_for_filter(flt)
    by_atelier_indiv = _individual_consumption_by_atelier_for_filter(flt)
    for row in stats.get("table_ateliers", []) or []:
        payload = by_atelier.get(row.get("atelier_id"), {})
        row["conso_kwh"] = round(float(payload.get("kwh") or 0.0), 3)
        row["co2_kg"] = round(float(payload.get("co2_kg") or 0.0), 3)
        row["conso_sessions"] = int(payload.get("sessions_count") or 0)
        indiv_payload = by_atelier_indiv.get(row.get("atelier_id"), {})
        row["conso_indiv_kwh"] = round(float(indiv_payload.get("kwh") or 0.0), 3)
        row["co2_indiv_kg"] = round(float(indiv_payload.get("co2_kg") or 0.0), 3)
        row["conso_indiv_presences"] = int(indiv_payload.get("presences_count") or 0)
        row["conso_indiv_participants"] = int(indiv_payload.get("participants_count") or 0)
    return stats


def _continuity_group_key(at: AtelierActivite) -> int:
    return int(getattr(at, "continuity_parent_id", None) or at.id)

def _quartier_bucket(name: str | None) -> str:
    """Regroupe les quartiers en 4 grandes catégories + Inconnu."""
    if not name:
        return "Inconnu"
    s = str(name).strip().lower()
    if not s:
        return "Inconnu"
    # règles simples : on privilégie le 'contains' (les noms dans ta DB peuvent varier)
    if "hors rouher" in s:
        return "Hauts de Creil"
    if "rouher" in s:
        return "Rouher"
    if "bas" in s:
        return "Bas de Creil"
    if "haut" in s or "hauts" in s:
        return "Hauts de Creil"
    # tout le reste
    return "Autres"


def _xlsx_table_name(raw: str, fallback: str = "Tableau") -> str:
    """Nom Excel safe pour un tableau structuré.

    Contraintes Excel :
    - unique dans tout le classeur ;
    - commence par une lettre ou underscore ;
    - pas d'espace ni caractères spéciaux ;
    - ne ressemble pas à une référence de cellule.
    """
    cleaned = "".join(ch if (ch.isalnum() or ch == "_") else "_" for ch in str(raw or ""))
    cleaned = cleaned.strip("_") or fallback
    if not (cleaned[0].isalpha() or cleaned[0] == "_"):
        cleaned = f"T_{cleaned}"
    if re.match(r"^[A-Za-z]{1,3}[0-9]+$", cleaned):
        cleaned = f"T_{cleaned}"
    return cleaned[:200]


def _unique_excel_table_name(wb: Workbook, raw: str, fallback: str = "Tableau") -> str:
    base = _xlsx_table_name(raw, fallback=fallback)
    existing = set()
    for ws in wb.worksheets:
        try:
            existing.update(ws.tables.keys())
        except Exception:
            pass

    candidate = base
    idx = 1
    while candidate in existing:
        suffix = f"_{idx}"
        candidate = f"{base[:200-len(suffix)]}{suffix}"
        idx += 1
    return candidate


def _add_excel_table(ws, name: str, start_row: int, end_row: int, end_col: int, *, style: str = "TableStyleMedium2"):
    """Ajoute un vrai tableau Excel, sans créer de fichier corrompu.

    On refuse les plages vides, on nettoie les en-têtes et on garantit un nom
    unique au niveau du classeur.
    """
    if end_row <= start_row or end_col <= 0:
        return None
    if ws.parent is None:
        return None

    seen = set()
    for col in range(1, end_col + 1):
        cell = ws.cell(row=start_row, column=col)
        value = str(cell.value or f"Colonne {col}").strip() or f"Colonne {col}"
        # Les en-têtes de table doivent être des chaînes non vides et uniques.
        base = value
        idx = 2
        while value in seen:
            value = f"{base} {idx}"
            idx += 1
        seen.add(value)
        cell.value = value

    ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=_unique_excel_table_name(ws.parent, name), ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name=style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    return table


def _format_xlsx_header_row(ws, row_idx: int, end_col: int):
    fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
    for col in range(1, end_col + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _format_magato_workbook(wb: Workbook):
    for ws in wb.worksheets:
        try:
            _auto_fit_statsimpact_ws(ws)
        except Exception:
            pass
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.00"


def _apply_kpi_color(cell, value, mode: str = "positive"):
    """Coloration douce des KPI. mode='positive' : haut = bon, mode='warning' : haut = vigilance."""
    try:
        v = float(value)
    except Exception:
        return

    if mode == "warning":
        if v >= 100:
            cell.fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
        elif v >= 75:
            cell.fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
        elif v > 0:
            cell.fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
        return

    if v >= 80:
        cell.fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    elif v >= 50:
        cell.fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    elif v > 0:
        cell.fill = PatternFill(fill_type="solid", fgColor="FCE4D6")


def _style_financeur_sheet(ws):
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.sheet_view.showGridLines = False


def _add_financeur_quick_sheet(wb: Workbook, summary_rows: list[list], alert_rows: list[list], top_rows: list[list]):
    ws = wb.create_sheet("Lecture financeur", 0)
    ws.append(["Lecture rapide financeur"])
    ws["A1"].font = Font(bold=True, size=16)
    ws.append([])
    ws.append(["Indicateur", "Valeur", "Lecture"])
    start_summary = ws.max_row

    for row in summary_rows:
        ws.append(row)

    end_summary = ws.max_row
    _format_xlsx_header_row(ws, start_summary, 3)
    _add_excel_table(ws, "LectureFinanceurSynthese", start_summary, end_summary, 3, style="TableStyleMedium9")

    ws.append([])
    ws.append(["Points d'attention"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=12)
    ws.append(["Niveau", "Message"])
    start_alert = ws.max_row
    if alert_rows:
        for row in alert_rows:
            ws.append(row)
    else:
        ws.append(["OK", "Aucune alerte majeure détectée sur le périmètre."])
    end_alert = ws.max_row
    _format_xlsx_header_row(ws, start_alert, 2)
    _add_excel_table(ws, "LectureFinanceurAlertes", start_alert, end_alert, 2, style="TableStyleMedium4")

    ws.append([])
    ws.append(["Top ateliers par présence"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=12)
    ws.append(["Secteur", "Atelier", "Présences", "Participants uniques", "Conso kWh", "CO₂ kg"])
    start_top = ws.max_row
    if top_rows:
        for row in top_rows:
            ws.append(row)
    else:
        ws.append(["", "Aucun atelier dans le périmètre", 0, 0, 0, 0])
    end_top = ws.max_row
    _format_xlsx_header_row(ws, start_top, 6)
    _add_excel_table(ws, "LectureFinanceurTopAteliers", start_top, end_top, 6, style="TableStyleMedium2")

    for row in range(start_summary + 1, end_summary + 1):
        label = str(ws.cell(row=row, column=1).value or "")
        val_cell = ws.cell(row=row, column=2)
        if "Remplissage" in label:
            _apply_kpi_color(val_cell, val_cell.value, mode="positive")
        elif "CO₂" in label or "Consommation" in label:
            _apply_kpi_color(val_cell, val_cell.value, mode="warning")

    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 80
    for col in range(4, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18
    _style_financeur_sheet(ws)
    return ws


def _build_magato_per_atelier_workbook(flt) -> Workbook:
    """Magatomatique complet : 1 feuille par atelier + stats riches sur la période filtrée."""
    flt = _ensure_default_export_period(flt)

    # Cloisonnement : un responsable_secteur ne doit exporter que son secteur
    eff_secteur = flt.secteur
    if not can("scope:all_secteurs"):
        eff_secteur = (getattr(current_user, "secteur_assigne", None) or "").strip() or eff_secteur

    # Liste des ateliers dans le périmètre
    aq = _apply_atelier_lifecycle_filters(AtelierActivite.query, flt)
    if eff_secteur:
        aq = aq.filter(AtelierActivite.secteur == eff_secteur)
    ateliers = aq.order_by(AtelierActivite.secteur.asc(), AtelierActivite.nom.asc()).all()
    consumption_by_atelier = _consumption_by_atelier_for_filter(flt)

    grouped_ateliers = {}
    for at in ateliers:
        grouped_ateliers.setdefault(_continuity_group_key(at), []).append(at)

    wb = Workbook()

    # Synthèse globale
    ws0 = wb.active
    ws0.title = "Synthese"

    # Feuille de cadrage : impossible de confondre deux exports qui n'ont pas
    # le même périmètre temporel ou sectoriel.
    ws_params = wb.create_sheet("Paramètres", 0)
    atelier_ids_label = ', '.join(str(x) for x in (getattr(flt, 'atelier_ids', None) or [])) or 'Tous'
    param_rows = [
        ['Période appliquée', _export_period_label(flt)],
        ['Du', getattr(flt, 'date_from', None).isoformat() if getattr(flt, 'date_from', None) else ''],
        ['Au', getattr(flt, 'date_to', None).isoformat() if getattr(flt, 'date_to', None) else ''],
        ['Secteur demandé', getattr(flt, 'secteur', None) or 'Tous'],
        ['Secteur effectif', _effective_statsimpact_secteur(flt) or 'Tous'],
        ['Ateliers sélectionnés', atelier_ids_label],
        ["Date d'export", date.today().isoformat()],
        ['Utilisateur', getattr(current_user, 'email', None) or getattr(current_user, 'username', None) or f"user#{getattr(current_user, 'id', '')}"],
    ]
    _append_section(ws_params, "Périmètre exact du Magatomatique", ['Paramètre', 'Valeur'], param_rows, start_row=1)
    _autosize_columns(ws_params, max_width=50)
    ws_params.sheet_view.showGridLines = False

    ws0.append([f"Magatomatique — période {_export_period_label(flt)} : 1 feuille par atelier (matrice) + stats détaillées"])
    ws0.append([f"Périmètre : secteur {eff_secteur or 'Tous'} · ateliers {atelier_ids_label}"])

    ws0.append([
        "Secteur",
        "Atelier",
        "Nb séances prévisionnelles",
        "Nb séances réelles",
        "Nb présences totales",
        "Nb inscrits (uniques)",
        "Capacité d'accueil (réel)",
        "Nb places prévisionnelles",
        "Nb heures prévisionnelles",
        "Nb heures réalisées",
        "Taux d'occupation (%)",
        "Moyenne / séance",
        "Durée de l'activité (jours)",
        "Moyenne d'âge",
        "Conso kWh",
        "CO₂ kg",
        "Séances équipées",
        "Bas de Creil",
        "Hauts de Creil",
        "Rouher",
        "Autres",
        "Inconnu",
    ])

    # Agrégats globaux provenance + lecture financeur
    global_prov = {"Bas de Creil": 0, "Hauts de Creil": 0, "Rouher": 0, "Autres": 0, "Inconnu": 0}
    global_sessions_real = 0
    global_presences = 0
    global_uniques_sum = 0
    global_real_hours = 0.0
    global_conso_kwh = 0.0
    global_co2_kg = 0.0
    global_conso_sessions = 0
    top_ateliers_rows = []
    financeur_alert_rows = []

    for _gid, members in grouped_ateliers.items():
        at = sorted(members, key=lambda x: x.nom)[0]
        atelier_ids = [m.id for m in members]
        # Sessions de l'atelier dans la période
        sess_q = db.session.query(SessionActivite).filter(SessionActivite.atelier_id.in_(atelier_ids))
        if flt.date_from:
            sess_q = sess_q.filter(func.coalesce(SessionActivite.rdv_date, SessionActivite.date_session) >= flt.date_from)
        if flt.date_to:
            sess_q = sess_q.filter(func.coalesce(SessionActivite.rdv_date, SessionActivite.date_session) <= flt.date_to)
        sess_q = sess_q.order_by(
            func.coalesce(SessionActivite.rdv_date, SessionActivite.date_session).asc(),
            SessionActivite.id.asc(),
        )
        sessions = sess_q.all()

        # 1) KPIs par atelier
        sessions_planned = len(sessions)
        sessions_real = sum(1 for s in sessions if (s.statut or "").lower() != "annulee")
        dates = [d for d in [(s.rdv_date or s.date_session) for s in sessions] if d]
        duration_days = (max(dates) - min(dates)).days if dates else None

        # heures + capacités
        planned_hours = 0.0
        real_hours = 0.0
        planned_capacity = 0
        real_capacity = 0
        for s in sessions:
            mins = _session_duration_minutes(s, at) or 0
            h = float(mins) / 60.0 if mins else 0.0
            planned_hours += h
            cap = s.capacite if s.capacite is not None else (getattr(at, "capacite_defaut", 0) or 0)
            planned_capacity += int(cap or 0)
            if (s.statut or "").lower() != "annulee":
                real_hours += h
                real_capacity += int(cap or 0)

        if sessions_planned == 0:
            # atelier sans session dans le filtre -> ligne 0 + pas de feuille matrice
            ws0.append([at.secteur, at.nom, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, "", "", 0, 0, 0, 0, 0, 0, 0, 0])
            financeur_alert_rows.append(["Info", f"{at.secteur} — {at.nom} : aucune séance dans le périmètre filtré."])
            continue

        session_ids = [s.id for s in sessions]

        pres_rows = db.session.query(PresenceActivite.participant_id, PresenceActivite.session_id)             .filter(PresenceActivite.session_id.in_(session_ids)).all()

        presences_total = len(pres_rows)
        pid_set = sorted({int(pid) for (pid, _sid) in pres_rows if pid is not None})
        inscrits_uniques = len(pid_set)

        # participants (objets) pour âge + provenance
        participants = []
        if pid_set:
            participants = (
                db.session.query(Participant)
                .options(joinedload(Participant.quartier))
                .filter(Participant.id.in_(pid_set))
                .all()
            )

        ages = [p.age for p in participants if getattr(p, "age", None) is not None]
        age_avg = round(sum(ages) / len(ages), 1) if ages else None

        prov = {"Bas de Creil": 0, "Hauts de Creil": 0, "Rouher": 0, "Autres": 0, "Inconnu": 0}
        for p in participants:
            qname = p.quartier.nom if getattr(p, "quartier", None) is not None else None
            b = _quartier_bucket(qname)
            prov[b] = prov.get(b, 0) + 1
        for k in global_prov:
            global_prov[k] += prov.get(k, 0)

        occupation_rate = round((presences_total / real_capacity * 100.0), 1) if real_capacity > 0 else 0.0
        avg_per_session = round((presences_total / sessions_real), 2) if sessions_real > 0 else 0.0
        conso_payload = consumption_by_atelier.get(_continuity_group_key(at), {})
        conso_kwh = round(float(conso_payload.get("kwh") or 0.0), 3)
        conso_co2 = round(float(conso_payload.get("co2_kg") or 0.0), 3)
        conso_sessions = int(conso_payload.get("sessions_count") or 0)

        global_sessions_real += int(sessions_real or 0)
        global_presences += int(presences_total or 0)
        global_uniques_sum += int(inscrits_uniques or 0)
        global_real_hours += float(real_hours or 0.0)
        global_conso_kwh += float(conso_kwh or 0.0)
        global_co2_kg += float(conso_co2 or 0.0)
        global_conso_sessions += int(conso_sessions or 0)

        top_ateliers_rows.append([
            at.secteur,
            at.nom,
            presences_total,
            inscrits_uniques,
            conso_kwh,
            conso_co2,
        ])

        if sessions_real > 0 and presences_total == 0:
            financeur_alert_rows.append(["Attention", f"{at.secteur} — {at.nom} : séances réalisées mais aucune présence saisie."])
        if real_capacity > 0 and occupation_rate < 30:
            financeur_alert_rows.append(["Vigilance", f"{at.secteur} — {at.nom} : taux d'occupation faible ({occupation_rate}%)."])
        if sessions_real > 0 and conso_sessions == 0:
            financeur_alert_rows.append(["Info", f"{at.secteur} — {at.nom} : aucune consommation matériel renseignée."])

        ws0.append([
            at.secteur,
            at.nom,
            sessions_planned,
            sessions_real,
            presences_total,
            inscrits_uniques,
            real_capacity,
            planned_capacity,
            round(planned_hours, 2),
            round(real_hours, 2),
            occupation_rate,
            avg_per_session,
            duration_days if duration_days is not None else "",
            age_avg if age_avg is not None else "",
            conso_kwh,
            conso_co2,
            conso_sessions,
            prov.get("Bas de Creil", 0),
            prov.get("Hauts de Creil", 0),
            prov.get("Rouher", 0),
            prov.get("Autres", 0),
            prov.get("Inconnu", 0),
        ])

        # 2) Feuille atelier : bloc stats + matrice
        group_label = at.nom
        if len(members) > 1:
            aliases = sorted({m.nom for m in members if m.nom != at.nom})
            if aliases:
                group_label = f"{at.nom} ({', '.join(aliases)})"
        ws = wb.create_sheet(_safe_sheet_title(f"{at.nom}"))

        ws.append([f"{at.secteur} — {group_label}"])
        ws.append([f"Période : {_export_period_label(flt)}"])

        ws.append(["Statistiques", "Valeur"])
        ws.append(["Nb séances prévisionnelles", sessions_planned])
        ws.append(["Nb séances réelles", sessions_real])
        ws.append(["Nb présences totales", presences_total])
        ws.append(["Nb inscrits (uniques)", inscrits_uniques])
        ws.append(["Capacité d'accueil (réel)", real_capacity])
        ws.append(["Nb places prévisionnelles", planned_capacity])
        ws.append(["Nb heures prévisionnelles", round(planned_hours, 2)])
        ws.append(["Nb heures réalisées", round(real_hours, 2)])
        ws.append(["Taux d'occupation (%)", occupation_rate])
        ws.append(["Moyenne / séance", avg_per_session])
        ws.append(["Durée de l'activité (jours)", duration_days if duration_days is not None else ""])
        ws.append(["Moyenne d'âge", age_avg if age_avg is not None else ""])
        ws.append(["Consommation estimée (kWh)", conso_kwh])
        ws.append(["CO₂ estimé (kg)", conso_co2])
        ws.append(["Séances équipées", conso_sessions])

        ws.append([])
        ws.append(["Provenance (quartiers)", "Nb inscrits"])
        ws.append(["Bas de Creil", prov.get("Bas de Creil", 0)])
        ws.append(["Hauts de Creil", prov.get("Hauts de Creil", 0)])
        ws.append(["Rouher", prov.get("Rouher", 0)])
        ws.append(["Autres", prov.get("Autres", 0)])
        ws.append(["Inconnu", prov.get("Inconnu", 0)])

        ws.append([])

        # Matrice participants × sessions
        # (A) Antoine wanted a quick visual identification: âge + ville + quartier à côté du nom/prénom.
        headers = ["Nom", "Prénom", "Âge", "Ville", "Quartier"] + [
            ((d.strftime("%d/%m/%Y")) if (d := (s.rdv_date or s.date_session)) else "Sans date")
            for s in sessions
        ]
        ws.append(headers)

        matrix_header_row = ws.max_row
        if not pres_rows or not pid_set:
            # matrice vide structurée
            _format_xlsx_header_row(ws, matrix_header_row, len(headers))
            ws.auto_filter.ref = f"A{matrix_header_row}:{get_column_letter(len(headers))}{matrix_header_row}"
            # Feuilles atelier : pas de volet figé.
            # Le bloc stats au-dessus de la matrice rendait le défilement pénible
            # en scindant Excel en 4 zones. Les filtres du tableau restent actifs.
            ws.freeze_panes = None
            ws.column_dimensions[get_column_letter(1)].width = 20
            ws.column_dimensions[get_column_letter(2)].width = 18
            ws.column_dimensions[get_column_letter(3)].width = 8
            ws.column_dimensions[get_column_letter(4)].width = 18
            ws.column_dimensions[get_column_letter(5)].width = 22
            for col_idx in range(6, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 12
            continue

        # Participants (objets) pour trier + infos d'identification (âge/ville/quartier)
        parts = (
            db.session.query(Participant)
            .options(joinedload(Participant.quartier))
            .filter(Participant.id.in_(pid_set))
            .order_by(Participant.nom.asc(), Participant.prenom.asc())
            .all()
        )

        sid_index = {int(s.id): idx for idx, s in enumerate(sessions)}
        present = {(int(pid), int(sid)) for (pid, sid) in pres_rows if pid is not None and sid is not None}

        for p in parts:
            pid = int(p.id)
            nom = p.nom or ""
            prenom = p.prenom or ""
            age = getattr(p, "age", None)
            age = age if age is not None else ""
            ville = p.ville or ""
            quartier = p.quartier.nom if getattr(p, "quartier", None) is not None else ""

            row = [nom, prenom, age, ville, quartier] + [""] * len(sessions)
            for sid, idx in sid_index.items():
                if (int(pid), int(sid)) in present:
                    row[5 + idx] = "1"
            ws.append(row)

        matrix_last_row = ws.max_row
        _format_xlsx_header_row(ws, matrix_header_row, len(headers))
        _add_excel_table(ws, f"Matrice_{at.id}", matrix_header_row, matrix_last_row, len(headers), style="TableStyleMedium4")
        # Feuilles atelier : pas de volet figé.
        # Le bloc stats au-dessus de la matrice rendait le défilement pénible
        # en scindant Excel en 4 zones. Les filtres du tableau restent actifs.
        ws.freeze_panes = None
        ws.sheet_view.showGridLines = False

        # Largeurs
        ws.column_dimensions[get_column_letter(1)].width = 20
        ws.column_dimensions[get_column_letter(2)].width = 18
        ws.column_dimensions[get_column_letter(3)].width = 8
        ws.column_dimensions[get_column_letter(4)].width = 18
        ws.column_dimensions[get_column_letter(5)].width = 22
        for col_idx in range(6, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12

    synthese_last_row = ws0.max_row
    _format_xlsx_header_row(ws0, 3, 22)
    _add_excel_table(ws0, "SyntheseAteliers", 3, synthese_last_row, 22, style="TableStyleMedium9")
    ws0.freeze_panes = "C4"
    ws0.sheet_view.showGridLines = False

    # Couleurs douces sur KPI clés de la synthèse
    for row_idx in range(4, synthese_last_row + 1):
        _apply_kpi_color(ws0.cell(row=row_idx, column=11), ws0.cell(row=row_idx, column=11).value, mode="positive")
        _apply_kpi_color(ws0.cell(row=row_idx, column=15), ws0.cell(row=row_idx, column=15).value, mode="warning")
        _apply_kpi_color(ws0.cell(row=row_idx, column=16), ws0.cell(row=row_idx, column=16).value, mode="warning")

    top_ateliers_rows = sorted(top_ateliers_rows, key=lambda r: int(r[2] or 0), reverse=True)[:10]
    total_real_capacity = sum(int(ws0.cell(row=r, column=7).value or 0) for r in range(4, synthese_last_row + 1))
    avg_fill_global = round((global_presences / total_real_capacity * 100), 1) if total_real_capacity else 0

    summary_rows = [
        ["Période", _export_period_label(flt), "Périmètre temporel exact appliqué à cet export."],
        ["Séances réalisées", global_sessions_real, "Volume total de séances non annulées dans le périmètre."],
        ["Présences totales", global_presences, "Nombre total d'émargements saisis."],
        ["Participants uniques cumulés", global_uniques_sum, "Somme des participants uniques par atelier ; un participant présent sur plusieurs ateliers peut être compté plusieurs fois."],
        ["Heures réalisées", round(global_real_hours, 2), "Volume horaire animé sur les séances réalisées."],
        ["Remplissage moyen estimé (%)", avg_fill_global, "Indicateur de lecture rapide basé sur présences / capacité réelle cumulée."],
        ["Consommation estimée (kWh)", round(global_conso_kwh, 3), "Estimation issue du matériel renseigné par séance."],
        ["CO₂ estimé (kg)", round(global_co2_kg, 3), "Conversion kWh × facteur CO₂ du référentiel de consommation."],
        ["Séances équipées", global_conso_sessions, "Nombre de séances avec au moins une ligne matériel renseignée."],
    ]
    _add_financeur_quick_sheet(wb, summary_rows, financeur_alert_rows[:20], top_ateliers_rows)
    # Bloc "Provenance globale" en bas de la synthèse
    ws0.append([])
    ws0.append(["Provenance globale (inscrits uniques sur l'ensemble du périmètre)"])
    ws0.append(["Bas de Creil", global_prov.get("Bas de Creil", 0)])
    ws0.append(["Hauts de Creil", global_prov.get("Hauts de Creil", 0)])
    ws0.append(["Rouher", global_prov.get("Rouher", 0)])
    ws0.append(["Autres", global_prov.get("Autres", 0)])
    ws0.append(["Inconnu", global_prov.get("Inconnu", 0)])

    # Largeurs synthèse
    for col in range(1, 23):
        ws0.column_dimensions[get_column_letter(col)].width = 20 if col <= 2 else 18

    _format_magato_workbook(wb)
    return wb


@bp.route("/stats-impact", methods=["GET"])
@login_required
def exports():
    """Page légère dédiée aux exports (pas de calcul de stats lourdes)."""
    if not _can_view():
        abort(403)

    flt = _ensure_default_export_period(normalize_filters(request.args, user=current_user))

    secteurs = []
    if can("statsimpact:view_all") or can("scope:all_secteurs"):
        secteurs = [
            s[0]
            for s in (
                AtelierActivite.query.with_entities(AtelierActivite.secteur)
                .filter(AtelierActivite.is_deleted.is_(False))
                .filter(AtelierActivite.is_active.is_(True))
                .distinct()
                .order_by(AtelierActivite.secteur.asc())
                .all()
            )
            if s and s[0]
        ]

    q = _apply_atelier_lifecycle_filters(AtelierActivite.query, flt)
    eff_secteur = _resolve_secteur_scope(flt)
    if eff_secteur and eff_secteur != "__restricted__":
        q = q.filter(AtelierActivite.secteur == eff_secteur)
    ateliers = q.order_by(AtelierActivite.secteur.asc(), AtelierActivite.nom.asc()).all()

    csv_groups, csv_defaults, _ = _csv_fields_for_current_user()
    export_qs = _export_query_string_from_filter(flt)
    consumption_stats = _consumption_stats_for_filter(flt)
    individual_consumption_stats = _individual_consumption_stats_for_filter(flt)
    return render_template(
        "statsimpact/exports.html",
        flt=flt,
        secteurs=secteurs,
        ateliers=ateliers,
        csv_field_groups=csv_groups,
        csv_default_fields=csv_defaults,
        export_qs=export_qs,
        consumption_stats=consumption_stats,
        individual_consumption_stats=individual_consumption_stats,
    )

@bp.route("/stats-impact/", methods=["GET"])
@login_required
def exports_slash():
    """Compat: /stats-impact/ -> /stats-impact"""
    if not _can_view():
        abort(403)
    return redirect(url_for("statsimpact.exports", **_query_args_preserve_lists(request.args)))


@bp.route("/stats-impact/qualite-donnees", methods=["GET"])
@login_required
def qualite_donnees():
    if not _can_view():
        abort(403)

    flt = _ensure_default_export_period(normalize_filters(request.args, user=current_user))
    quality = _statsimpact_quality_payload(flt)

    secteurs = []
    if can("statsimpact:view_all") or can("scope:all_secteurs"):
        secteurs = [
            s[0]
            for s in (
                AtelierActivite.query.with_entities(AtelierActivite.secteur)
                .filter(AtelierActivite.is_deleted.is_(False))
                .filter(AtelierActivite.is_active.is_(True))
                .distinct()
                .order_by(AtelierActivite.secteur.asc())
                .all()
            )
            if s and s[0]
        ]

    q = AtelierActivite.query.filter(
        AtelierActivite.is_deleted.is_(False),
        AtelierActivite.is_active.is_(True),
    )
    eff_secteur = _resolve_secteur_scope(flt)
    if eff_secteur == "__restricted__":
        q = q.filter(AtelierActivite.id == -1)
    elif eff_secteur:
        q = q.filter(AtelierActivite.secteur == eff_secteur)
    ateliers = q.order_by(AtelierActivite.secteur.asc(), AtelierActivite.nom.asc()).all()

    return render_template(
        "statsimpact/qualite_donnees.html",
        flt=flt,
        quality=quality,
        secteurs=secteurs,
        ateliers=ateliers,
    )


import json
from sqlalchemy import case

def _dialect_name() -> str:
    try:
        return (db.engine.dialect.name or "").lower()
    except Exception:
        return ""

def _month_key_expr(date_expr):
    """
    Retourne une expression SQL qui donne une clé de mois 'YYYY-MM'
    compatible SQLite / Postgres.
    """
    d = _dialect_name()
    if "sqlite" in d:
        # SQLite: strftime
        return func.strftime("%Y-%m", date_expr)
    # Postgres / autres
    # to_char(date_trunc('month', ...), 'YYYY-MM')
    return func.to_char(func.date_trunc("month", date_expr), "YYYY-MM")



def _build_activity_charts(flt):
    """
    2 séries:
      - presences par mois
      - sessions réelles par mois (statut != 'annulee')
    Respecte les filtres communs via _apply_common_filters.
    """
    date_expr = _session_date_expr()
    month_key = _month_key_expr(date_expr)

    # ---------- Présences / mois ----------
    pres_q = (
        db.session.query(
            month_key.label("m"),
            func.count(PresenceActivite.id).label("nb"),
        )
        .select_from(PresenceActivite)
        .join(SessionActivite, PresenceActivite.session_id == SessionActivite.id)
        .join(AtelierActivite, SessionActivite.atelier_id == AtelierActivite.id)
    )
    pres_q = _apply_common_filters(pres_q, flt)
    pres_q = pres_q.group_by("m").order_by("m")
    pres_rows = pres_q.all()

    # ---------- Sessions réelles / mois ----------
    sess_q = (
        db.session.query(
            month_key.label("m"),
            func.count(SessionActivite.id).label("nb"),
        )
        .select_from(SessionActivite)
        .join(AtelierActivite, SessionActivite.atelier_id == AtelierActivite.id)
    )
    sess_q = _apply_common_filters(sess_q, flt)
    # "réelles" = pas annulée (attention: statut peut être None)
    sess_q = sess_q.filter(func.lower(func.coalesce(SessionActivite.statut, "")) != "annulee")
    sess_q = sess_q.group_by("m").order_by("m")
    sess_rows = sess_q.all()

    # ---------- Fusion mois (pour aligner labels) ----------
    months = sorted({r.m for r in pres_rows if r.m} | {r.m for r in sess_rows if r.m})

    pres_map = {r.m: int(r.nb or 0) for r in pres_rows if r.m}
    sess_map = {r.m: int(r.nb or 0) for r in sess_rows if r.m}

    pres_values = [pres_map.get(m, 0) for m in months]
    sess_values = [sess_map.get(m, 0) for m in months]

    return {
        "months": months,
        "presences": pres_values,
        "sessions_real": sess_values,
    }


def _compute_compare_payload(flt, stats, freq):
    """Optionnel : calcule la période précédente + deltas KPI, sans alourdir toute la page."""
    if not getattr(flt, "date_from", None) or not getattr(flt, "date_to", None):
        return {"enabled": False}

    try:
        span_days = (flt.date_to - flt.date_from).days
        prev_to = flt.date_from - timedelta(days=1)
        prev_from = prev_to - timedelta(days=span_days)
    except Exception:
        return {"enabled": False}

    flt_prev = deepcopy(flt)
    flt_prev.date_from = prev_from
    flt_prev.date_to = prev_to

    stats_prev = compute_volume_activity_stats(flt_prev)
    freq_prev = compute_participation_frequency_stats(flt_prev)

    def _safe_num(v):
        try:
            return float(v)
        except Exception:
            return 0.0

    cur = getattr(stats, "kpi", {}) or {}
    prv = getattr(stats_prev, "kpi", {}) or {}

    def _get(obj, key, default=0):
        if hasattr(obj, key):
            return getattr(obj, key)
        if isinstance(obj, dict):
            return obj.get(key, default)
        return default

    def _delta(key):
        c = _safe_num(_get(cur, key, 0))
        p = _safe_num(_get(prv, key, 0))
        d = c - p
        pct = (d / p * 100.0) if p else None
        return {"current": c, "previous": p, "delta": d, "delta_pct": pct}

    return {
        "enabled": True,
        "current_period": {"from": flt.date_from, "to": flt.date_to},
        "previous_period": {"from": prev_from, "to": prev_to},
        "kpis": {
            "sessions": _delta("sessions"),
            "presences": _delta("presences"),
            "uniques": _delta("uniques"),
            "avg_per_session": _delta("avg_per_session"),
            "new_participants": _delta("new_participants"),
        },
        "freq": {
            "freq_avg": {"current": _safe_num(_get(freq, "freq_avg", 0)), "previous": _safe_num(_get(freq_prev, "freq_avg", 0))},
            "returning_rate": {"current": _safe_num(_get(freq, "returning_rate", 0)), "previous": _safe_num(_get(freq_prev, "returning_rate", 0))},
            "regulars_4plus": {"current": _safe_num(_get(freq, "regulars_4plus", 0)), "previous": _safe_num(_get(freq_prev, "regulars_4plus", 0))},
        },
    }




@bp.route("/stats-impact/dashboard", methods=["GET", "POST"])
@login_required
def dashboard():
    """Dashboard complet Stats & Impacts (calculs + onglets)."""
    if not _can_view():
        abort(403)

    args = request.args

    # Robust: normalize_filters supports MultiDict (checklistes) and kwargs-style.
    flt = normalize_filters(args, user=current_user)

    # Default: current year if no dates
    if not flt.date_from and not flt.date_to:
        today = date.today()
        flt.date_from = date(today.year, 1, 1)
        flt.date_to = date(today.year, 12, 31)

    # Pre-compute participants for access control if we need to handle edits.
    participants = compute_participants_stats(flt)

    if request.method == "POST":
        action = request.form.get("action")
        if action == "update_participant":
            try:
                participant_id = int(request.form.get("participant_id", "0"))
            except Exception:
                participant_id = 0

            allowed_ids = {p["id"] for p in participants.get("participants", [])}
            if not participant_id or participant_id not in allowed_ids:
                abort(403)

            participant = Participant.query.get(participant_id)
            if not participant:
                abort(404)

            participant.nom = (request.form.get("nom") or participant.nom or "").strip() or participant.nom
            participant.prenom = (request.form.get("prenom") or participant.prenom or "").strip() or participant.prenom
            participant.ville = (request.form.get("ville") or "").strip() or None
            participant.email = (request.form.get("email") or "").strip() or None
            participant.telephone = (request.form.get("telephone") or "").strip() or None
            participant.genre = (request.form.get("genre") or "").strip() or None
            participant.type_public = (request.form.get("type_public") or participant.type_public or "H").strip().upper()

            dn_raw = request.form.get("date_naissance") or None
            dn = None
            if dn_raw:
                try:
                    dn = date.fromisoformat(dn_raw)
                except Exception:
                    dn = None
            participant.date_naissance = dn

            quartier_id = request.form.get("quartier_id") or None
            participant.quartier_id = normalize_quartier_for_ville(participant.ville, quartier_id)

            try:
                from app.extensions import db

                db.session.commit()
                flash("Participant mis à jour.", "success")
            except Exception:
                db.session.rollback()
                flash("Impossible de sauvegarder ce participant.", "danger")

            args_redirect = _query_args_preserve_lists(request.args)
            args_redirect["tab"] = "participants"
            return redirect(url_for("statsimpact.dashboard", **args_redirect))

        if action == "delete_participant":
            try:
                participant_id = int(request.form.get("participant_id", "0"))
            except Exception:
                participant_id = 0

            allowed_ids = {p["id"] for p in participants.get("participants", [])}
            if not participant_id or participant_id not in allowed_ids:
                abort(403)

            participant = Participant.query.get(participant_id)
            if not participant:
                abort(404)

            # Sécurité secteur: un responsable_secteur ne peut purger un participant
            # que si ce participant n'a des présences que dans SON secteur (ou aucune).
            user_secteur = (getattr(current_user, "secteur_assigne", None) or "").strip()
            if not can("participants:view_all"):
                sectors = (
                    PresenceActivite.query.join(SessionActivite, PresenceActivite.session_id == SessionActivite.id)
                    .with_entities(SessionActivite.secteur)
                    .filter(PresenceActivite.participant_id == participant_id)
                    .distinct()
                    .all()
                )
                sectors = {s[0] for s in sectors if s and s[0]}
                # s'il n'a jamais émargé: OK (secteurs = vide)
                if sectors and sectors != {user_secteur}:
                    flash(
                        "Suppression refusée : ce participant a des émargements dans d'autres secteurs.",
                        "danger",
                    )
                    args_redirect = _query_args_preserve_lists(request.args)
                    args_redirect["tab"] = "participants"
                    return redirect(url_for("statsimpact.dashboard", **args_redirect))

            try:
                from app.extensions import db

                # Supprime d'abord les signatures des présences
                presences = PresenceActivite.query.filter_by(participant_id=participant_id).all()
                for pr in presences:
                    if pr.signature_path:
                        try:
                            if os.path.exists(pr.signature_path):
                                os.remove(pr.signature_path)
                        except Exception:
                            pass
                    db.session.delete(pr)

                db.session.delete(participant)
                db.session.commit()
                flash("Participant supprimé définitivement.", "success")
            except Exception:
                db.session.rollback()
                flash("Impossible de supprimer ce participant.", "danger")

            args_redirect = _query_args_preserve_lists(request.args)
            args_redirect["tab"] = "participants"
            return redirect(url_for("statsimpact.dashboard", **args_redirect))

    # Refresh computed stats after any potential mutation
    participants = compute_participants_stats(flt)
    stats = compute_volume_activity_stats(flt)
    stats = _enrich_stats_with_consumption(stats, flt)
    freq = compute_participation_frequency_stats(flt)
    trans = compute_transversalite_stats(flt)
    demo = compute_demography_stats(flt)
    occupancy = compute_occupancy_stats(flt)

    # Le Magatomatique : calcul uniquement si l'onglet est affiché (sinon on garde la page légère)
    tab = (request.args.get("tab") or "base").strip().lower()
    magato = None
    if tab in ("magato", "magatomatique"):
        participant_q = (request.args.get("participant_q") or "").strip() or None
        view = (request.args.get("magato_view") or "macro").strip().lower()
        try:
            max_sessions = int(request.args.get("max_sessions") or 40)
        except Exception:
            max_sessions = 40
        try:
            max_participants = int(request.args.get("max_participants") or 250)
        except Exception:
            max_participants = 250

        # bornes de sécurité
        max_sessions = max(5, min(max_sessions, 200))
        max_participants = max(20, min(max_participants, 1000))

        magato = compute_magatomatique(
            flt,
            participant_q=participant_q,
            view=view,
            max_sessions=max_sessions,
            max_participants=max_participants,
        )

    participants = compute_participants_stats(flt)

    secteurs = []
    if can("statsimpact:view_all") or can("scope:all_secteurs"):
        secteurs = [
            s[0]
            for s in (
                AtelierActivite.query.with_entities(AtelierActivite.secteur)
                .filter(AtelierActivite.is_deleted.is_(False))
                .filter(AtelierActivite.is_active.is_(True))
                .distinct()
                .order_by(AtelierActivite.secteur.asc())
                .all()
            )
            if s and s[0]
        ]

    q = _apply_atelier_lifecycle_filters(AtelierActivite.query, flt)
    eff_secteur = _resolve_secteur_scope(flt)
    if eff_secteur and eff_secteur != "__restricted__":
        q = q.filter(AtelierActivite.secteur == eff_secteur)
    ateliers = q.order_by(AtelierActivite.secteur.asc(), AtelierActivite.nom.asc()).all()

    quartiers = Quartier.query.order_by(Quartier.ville.asc(), Quartier.nom.asc()).all()

    # Années disponibles (pour presets "année") dans le périmètre accessible
    try:
        eff_secteur = flt.secteur
        if not can("scope:all_secteurs"):
            eff_secteur = (getattr(current_user, "secteur_assigne", None) or "").strip() or eff_secteur

        year_expr = func.extract("year", func.coalesce(SessionActivite.rdv_date, SessionActivite.date_session))
        years_q = (
            db.session.query(year_expr.label("y"))
            .select_from(SessionActivite)
            .join(AtelierActivite, SessionActivite.atelier_id == AtelierActivite.id)
            .filter(AtelierActivite.is_deleted.is_(False))
            .filter(AtelierActivite.is_active.is_(True))
        )
        if eff_secteur:
            years_q = years_q.filter(AtelierActivite.secteur == eff_secteur)
        years = [int(r.y) for r in years_q.distinct().order_by(year_expr.desc()).all() if r and r.y]
    except Exception:
        years = []

    charts = {"months": [], "presences": [], "sessions_real": []}
    try:
        charts = _build_activity_charts(flt)
    except Exception:
        # On garde le fallback vide pour éviter de casser le dashboard
        pass

    # Option : comparaison période précédente (même durée)
    compare = {"enabled": False}
    compare_flag = (request.args.get("compare") or "").strip().lower()
    if compare_flag in ("1", "true", "yes", "on"):
        try:
            compare = _compute_compare_payload(flt, stats, freq)
        except Exception:
            compare = {"enabled": False}

    csv_groups, csv_defaults, _ = _csv_fields_for_current_user()
    export_qs = _export_query_string_from_filter(flt)
    consumption_stats = _consumption_stats_for_filter(flt)
    individual_consumption_stats = _individual_consumption_stats_for_filter(flt)
    return render_template(
        ["statsimpact/dashboard.html", "statsimpact_dashboard.html"],
        flt=flt,
        stats=stats,
        freq=freq,
        trans=trans,
        demo=demo,
        secteurs=secteurs,
        ateliers=ateliers,
        occupancy=occupancy,
        participants=participants,
        magato=magato,
        quartiers=quartiers,
        available_years=years,
        charts=charts,
        compare=compare,
        csv_field_groups=csv_groups,
        csv_default_fields=csv_defaults,
        consumption_stats=consumption_stats,
        individual_consumption_stats=individual_consumption_stats,
    )



@bp.route("/stats-impact/export-fidele.xlsx", methods=["GET"])
@login_required
def export_fidele_xlsx():
    if not _can_view():
        abort(403)

    flt = _ensure_default_export_period(normalize_filters(request.args, user=current_user))

    wb = _build_statsimpact_scope_workbook(flt)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name=_magatomatique_filename("statsimpact_export_fidele", flt),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/stats-impact/magatomatique.csv", methods=["GET"])
@login_required
def magatomatique_export_csv():
    if not _can_view():
        abort(403)

    flt = _ensure_default_export_period(normalize_filters(request.args, user=current_user))
    participant_q = (request.args.get("participant_q") or "").strip() or None

    _csv_groups, csv_defaults, allowed_fields = _csv_fields_for_current_user()

    fields = request.args.getlist("fields")
    if not fields:
        fields = list(csv_defaults)
    fields = [f for f in fields if f in CSV_FIELD_MAP and f in allowed_fields]
    if not fields:
        fields = list(csv_defaults) if csv_defaults else [f for f in CSV_DEFAULT_FIELDS if f in allowed_fields]

    query = _query_presence_export(flt, participant_q=participant_q)

    output = StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow([CSV_FIELD_MAP[f]["label"] for f in fields])

    for presence, participant, session, atelier, quartier in query.all():
        ctx = {
            "presence": presence,
            "participant": participant,
            "session": session,
            "atelier": atelier,
            "quartier": quartier,
        }
        writer.writerow([CSV_FIELD_MAP[f]["getter"](ctx) for f in fields])

    csv_name = f"magatomatique_export_{_export_period_slug(flt)}.csv"
    resp = Response(output.getvalue(), mimetype="text/csv")
    resp.headers["Content-Disposition"] = f"attachment; filename={csv_name}"
    return resp


@bp.route("/stats-impact/magatomatique.xlsx", methods=["GET"])
@login_required
def magatomatique_export():
    if not _can_view():
        abort(403)

    flt = _ensure_default_export_period(normalize_filters(request.args, user=current_user))

    export_mode = (request.args.get("export_mode") or "complete").strip().lower()
    participant_q = (request.args.get("participant_q") or "").strip() or None
    view = (request.args.get("magato_view") or "macro").strip().lower()
    try:
        max_sessions = int(request.args.get("max_sessions") or 40)
    except Exception:
        max_sessions = 40
    try:
        max_participants = int(request.args.get("max_participants") or 250)
    except Exception:
        max_participants = 250

    # bornes (export raisonnable)
    max_sessions = max(5, min(max_sessions, 400))
    max_participants = max(20, min(max_participants, 5000))

    # Mode complet : période explicite + lecture financeur + synthèse + 1 feuille par atelier.
    # C'est désormais le Magatomatique principal, pour éviter les écarts avec
    # l'ancien export résumé qui ne portait pas les mêmes conventions de calcul.
    if export_mode in ("complete", "complet", "full", "per_atelier", "per-atelier", "atelier"):
        wb = _build_magato_per_atelier_workbook(flt)
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        prefix = "magatomatique_par_atelier" if export_mode in ("per_atelier", "per-atelier", "atelier") else "magatomatique_complet"
        filename = _magatomatique_filename(prefix, flt)
        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    magato = compute_magatomatique(
        flt,
        participant_q=participant_q,
        view=view,
        max_sessions=max_sessions,
        max_participants=max_participants,
    )

    if magato.get("restricted"):
        abort(403)

    wb = Workbook()
    ws = wb.active
    ws.title = "Synthese"

    # En-têtes synthèse macro (secteurs)
    ws.append(["Synthèse par secteur"])
    ws.append(["Secteur", "Nb sessions", "Nb présences", "Participants uniques"])
    for r in (magato.get("macro") or {}).get("by_secteur", []):
        ws.append([r["secteur"], r["nb_sessions"], r["nb_presences"], r["nb_participants_uniques"]])

    ws.append([])
    ws.append(["Synthèse par atelier"])
    ws.append(["Secteur", "Atelier", "Nb sessions", "Nb présences", "Participants uniques"])
    for r in (magato.get("macro") or {}).get("by_atelier", []):
        ws.append([r["secteur"], r["atelier_nom"], r["nb_sessions"], r["nb_presences"], r["nb_participants_uniques"]])

    # Feuille participants (si dispo)
    if magato.get("participants"):
        ws2 = wb.create_sheet("Participants")
        ws2.append(["Participants (dans le périmètre filtré)"])
        ws2.append(["Nom", "Prénom", "Ville", "Quartier", "Nb présences", "1ère venue", "Dernière venue"])
        for p in magato["participants"]:
            fd = p.get("first_date")
            ld = p.get("last_date")
            ws2.append([
                p.get("nom",""),
                p.get("prenom",""),
                p.get("ville") or "",
                p.get("quartier") or "",
                int(p.get("nb_presences",0)),
                fd.strftime("%Y-%m-%d") if fd else "",
                ld.strftime("%Y-%m-%d") if ld else "",
            ])

    # Feuille matrice (si view=matrix)
    if magato.get("view") == "matrix" and magato.get("sessions") and magato.get("participants"):
        ws3 = wb.create_sheet("Matrice")
        sessions = magato["sessions"]
        participants = magato["participants"]
        matrix = magato.get("matrix") or {}

        header = ["Nom", "Prénom"] + [f'{s["atelier"]} · {s["label"]}' for s in sessions]
        ws3.append(header)

        for p in participants:
            row = [p.get("nom",""), p.get("prenom","")]
            pid = int(p["id"])
            for s in sessions:
                sid = int(s["id"])
                row.append("1" if matrix.get((pid, sid)) else "")
            ws3.append(row)

        # Ajuste largeur colonnes
        for col_idx in range(1, len(header) + 1):
            ws3.column_dimensions[get_column_letter(col_idx)].width = 16 if col_idx <= 2 else 12

        ws4 = wb.create_sheet("Participations")
        ws4.append(["Nom", "Prénom", "Atelier", "Secteur", "Date session", "ID session"])
        for p in participants:
            pid = int(p["id"])
            for s in sessions:
                sid = int(s["id"])
                if matrix.get((pid, sid)):
                    ws4.append(
                        [
                            p.get("nom", ""),
                            p.get("prenom", ""),
                            s.get("atelier", ""),
                            s.get("secteur", ""),
                            s.get("date").strftime("%Y-%m-%d") if s.get("date") else "",
                            sid,
                        ]
                    )

        for col_idx in range(1, 7):
            ws4.column_dimensions[get_column_letter(col_idx)].width = 18

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = _magatomatique_filename("magatomatique_simple", flt)
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
