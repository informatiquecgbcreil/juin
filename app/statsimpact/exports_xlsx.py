from __future__ import annotations


from datetime import date
from io import BytesIO, StringIO
import csv

import re

from flask import abort, request, send_file, Response
from flask_login import login_required, current_user
from app.rbac import can

from sqlalchemy import func
from sqlalchemy.orm import joinedload

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.extensions import db

from app.models import (
    AtelierActivite,
    Participant,
    PresenceActivite,
    SessionActivite,
)


from .occupancy import compute_occupancy_stats

from .engine import (
    compute_volume_activity_stats,
    compute_participation_frequency_stats,
    compute_demography_stats,
    compute_participants_stats,
    compute_magatomatique,
    normalize_filters,
    _session_duration_minutes,
)


from app.statsimpact.common import bp
from app.statsimpact.common import (
    CSV_DEFAULT_FIELDS,
    CSV_FIELD_MAP,
    _append_section,
    _apply_atelier_lifecycle_filters,
    _auto_fit_statsimpact_ws,
    _autosize_columns,
    _can_view,
    _consumption_by_atelier_for_filter,
    _consumption_stats_for_filter,
    _continuity_group_key,
    _csv_fields_for_current_user,
    _effective_statsimpact_secteur,
    _enrich_stats_with_consumption,
    _ensure_default_export_period,
    _export_period_label,
    _export_period_slug,
    _individual_consumption_stats_for_filter,
    _magatomatique_filename,
    _safe_sheet_title,
)
from app.statsimpact.pedagogie import (
    _query_presence_export,
)

def _build_statsimpact_scope_workbook(flt) -> Workbook:
    flt = _ensure_default_export_period(flt)
    stats = compute_volume_activity_stats(flt)
    stats = _enrich_stats_with_consumption(stats, flt)
    consumption_stats = _consumption_stats_for_filter(flt)
    individual_consumption_stats = _individual_consumption_stats_for_filter(flt)
    freq = compute_participation_frequency_stats(flt)
    demo = compute_demography_stats(flt)
    occupancy = compute_occupancy_stats(flt)
    participants = compute_participants_stats(flt)

    wb = Workbook()

    ws_params = wb.active
    ws_params.title = 'Paramètres'
    atelier_ids = list(getattr(flt, 'atelier_ids', None) or [])
    param_rows = [
        ['Secteur demandé', getattr(flt, 'secteur', None) or 'Tous'],
        ['Secteur effectif', _effective_statsimpact_secteur(flt) or 'Tous'],
        ['Du', getattr(flt, 'date_from', None).isoformat() if getattr(flt, 'date_from', None) else ''],
        ['Au', getattr(flt, 'date_to', None).isoformat() if getattr(flt, 'date_to', None) else ''],
        ['Groupement', getattr(flt, 'group_by', None) or 'MONTH'],
        ['Ateliers sélectionnés', ', '.join(str(x) for x in atelier_ids) if atelier_ids else 'Tous'],
        ["Date d'export", date.today().isoformat()],
        ['Utilisateur', getattr(current_user, 'email', None) or getattr(current_user, 'username', None) or f"user#{getattr(current_user, 'id', '')}"],
    ]
    _append_section(ws_params, "Périmètre exact de l'export", ['Paramètre', 'Valeur'], param_rows, start_row=1)

    ws_syn = wb.create_sheet('Synthèse')
    kpi = stats.get('kpi', {}) or {}
    _append_section(ws_syn, 'Indicateurs clés', ['Indicateur', 'Valeur'], [
        ['Sessions', kpi.get('sessions', 0)],
        ['Présences', kpi.get('presences', 0)],
        ['Participants uniques', kpi.get('uniques', 0)],
        ['Nouveaux participants', kpi.get('new_participants', 0)],
        ['Heures animateur', kpi.get('hours_animator', 0)],
        ['Heures participants', kpi.get('hours_people', 0)],
        ['Moyenne / session', kpi.get('avg_per_session', 0)],
        ['Durée activité (jours)', kpi.get('activity_duration_days', '')],
        ['Fréquence moyenne', freq.get('freq_avg', 0)],
        ['Taux de retour (%)', freq.get('returning_rate', 0)],
        ['Participants réguliers (4+)', freq.get('regulars_4plus', 0)],
        ['Âge moyen', demo.get('age_avg', '')],
        ['Consommation estimée (kWh)', consumption_stats.get('total_kwh', 0)],
        ['CO₂ estimé (kg)', consumption_stats.get('total_co2', 0)],
        ['Sessions équipées', consumption_stats.get('sessions_count', 0)],
        ['Moyenne kWh / séance équipée', consumption_stats.get('avg_kwh_per_session', 0)],
        ['Conso individuelle tracée (kWh)', individual_consumption_stats.get('total_kwh', 0)],
        ['CO₂ individuel tracé (kg)', individual_consumption_stats.get('total_co2', 0)],
        ['Présences avec matériel individuel', individual_consumption_stats.get('presences_count', 0)],
        ['Habitants concernés', individual_consumption_stats.get('participants_count', 0)],
    ], start_row=1)

    ws_at = wb.create_sheet('Ateliers')
    atelier_rows = []
    for r in stats.get('table_ateliers', []) or []:
        atelier_rows.append([
            r.get('secteur', ''), r.get('nom', ''), r.get('type_atelier', ''),
            r.get('sessions_planned', 0), r.get('sessions_real', 0), r.get('presences', 0),
            r.get('uniques', 0), r.get('real_hours', 0), r.get('occupation_rate', 0),
            r.get('avg_per_session_real', 0), r.get('activity_duration_days', ''),
            r.get('conso_kwh', 0), r.get('co2_kg', 0), r.get('conso_sessions', 0),
            r.get('conso_indiv_kwh', 0), r.get('co2_indiv_kg', 0), r.get('conso_indiv_presences', 0),
        ])
    _append_section(ws_at, 'Vue ateliers', ['Secteur','Atelier','Type','Sessions prév.','Sessions réelles','Présences','Uniques','Heures réalisées',"Taux d'occupation (%)",'Moyenne / séance','Durée activité (j)','Conso séance kWh','CO₂ séance kg','Séances équipées','Conso ind. kWh','CO₂ ind. kg','Présences équipées'], atelier_rows, start_row=1)

    ws_pub = wb.create_sheet('Publics')
    row = _append_section(ws_pub, 'Répartition par genre', ['Genre', 'Total'], [[k, v] for k, v in (demo.get('genre') or {}).items()], start_row=1)
    row += 1
    row = _append_section(ws_pub, 'Répartition par âge', ['Tranche', 'Total'], [[k, v] for k, v in (demo.get('age_buckets') or {}).items()], start_row=row)
    row += 1
    row = _append_section(ws_pub, 'Type de public', ['Type', 'Total'], [[k, v] for k, v in (demo.get('type_public') or {}).items()], start_row=row)
    row += 1
    qpv_detail_rows = []
    for item in demo.get('qpv_details') or []:
        genre = item.get('genre') or {}
        ages = item.get('age_buckets') or {}
        qpv_detail_rows.append([
            item.get('label', ''),
            item.get('total', 0),
            genre.get('Femmes', 0),
            genre.get('Hommes', 0),
            genre.get('Autre / non renseigné', 0),
            ages.get('0-10', 0),
            ages.get('11-17', 0),
            ages.get('18-25', 0),
            ages.get('26-59', 0),
            ages.get('60+', 0),
            ages.get('Inconnu', 0),
        ])
    row = _append_section(
        ws_pub,
        'Détail QPV / genre / âge',
        ['Typologie', 'Total', 'Femmes', 'Hommes', 'Autre / NR', '0-10', '11-17', '18-25', '26-59', '60+', 'Âge inconnu'],
        qpv_detail_rows,
        start_row=row,
    )
    row += 1
    villes = [[r.get('ville', ''), r.get('count', 0)] for r in (demo.get('villes_top') or [])]
    _append_section(ws_pub, 'Top villes', ['Ville', 'Total'], villes, start_row=row)

    ws_fill = wb.create_sheet('Remplissage')
    fill_rows = [
        ['Sessions collectives', occupancy.get('collective_sessions', 0)],
        ['Présences collectives', occupancy.get('collective_presences', 0)],
        ['Remplissage moyen (%)', occupancy.get('avg_fill_rate_pct', '')],
        ['Capacité par défaut', occupancy.get('default_capacity', '')],
        ['Bucket <50%', (occupancy.get('buckets') or {}).get('<50%', 0)],
        ['Bucket 50-79%', (occupancy.get('buckets') or {}).get('50-79%', 0)],
        ['Bucket 80-99%', (occupancy.get('buckets') or {}).get('80-99%', 0)],
        ['Bucket 100%+', (occupancy.get('buckets') or {}).get('100%+', 0)],
    ]
    row = _append_section(ws_fill, 'Synthèse remplissage', ['Indicateur', 'Valeur'], fill_rows, start_row=1)
    row += 1
    occ_rows = []
    for r in occupancy.get('per_atelier', []) or []:
        occ_rows.append([r.get('secteur', ''), r.get('nom', ''), r.get('sessions', 0), r.get('presences', 0), r.get('capacity_total', 0), r.get('avg_fill_rate_pct', 0)])
    _append_section(ws_fill, 'Remplissage par atelier', ['Secteur','Atelier','Sessions','Présences','Capacité cumulée','Remplissage moyen (%)'], occ_rows, start_row=row)

    if can('participants:view') or can('participants:view_all'):
        ws_part = wb.create_sheet('Participants')
        part_rows = []
        for p in participants.get('participants', []) or []:
            ateliers_txt = ', '.join(sorted([a.get('atelier', '') for a in (p.get('ateliers') or {}).values() if a.get('atelier')]))
            part_rows.append([
                p.get('nom', ''), p.get('prenom', ''), p.get('ville', ''), p.get('quartier', ''),
                p.get('genre', ''), p.get('type_public', ''), p.get('visites', 0), ateliers_txt
            ])
        _append_section(ws_part, 'Participants du périmètre', ['Nom','Prénom','Ville','Quartier','Genre','Type public','Visites','Ateliers fréquentés'], part_rows, start_row=1)

    for ws in wb.worksheets:
        _auto_fit_statsimpact_ws(ws)
    return wb

def _quartier_bucket(name: str | None) -> str:
    """Regroupe les quartiers en 4 grandes catégories + Inconnu."""
    if not name:
        return "Inconnu"
    s = str(name).strip().lower()
    if not s:
        return "Inconnu"
    # règles simples : on privilégie le 'contains' (les noms dans ta DB peuvent varier)
    if "hors rouher" in s:
        return "Hauts de Creil"
    if "rouher" in s:
        return "Rouher"
    if "bas" in s:
        return "Bas de Creil"
    if "haut" in s or "hauts" in s:
        return "Hauts de Creil"
    # tout le reste
    return "Autres"


def _xlsx_table_name(raw: str, fallback: str = "Tableau") -> str:
    """Nom Excel safe pour un tableau structuré.

    Contraintes Excel :
    - unique dans tout le classeur ;
    - commence par une lettre ou underscore ;
    - pas d'espace ni caractères spéciaux ;
    - ne ressemble pas à une référence de cellule.
    """
    cleaned = "".join(ch if (ch.isalnum() or ch == "_") else "_" for ch in str(raw or ""))
    cleaned = cleaned.strip("_") or fallback
    if not (cleaned[0].isalpha() or cleaned[0] == "_"):
        cleaned = f"T_{cleaned}"
    if re.match(r"^[A-Za-z]{1,3}[0-9]+$", cleaned):
        cleaned = f"T_{cleaned}"
    return cleaned[:200]


def _unique_excel_table_name(wb: Workbook, raw: str, fallback: str = "Tableau") -> str:
    base = _xlsx_table_name(raw, fallback=fallback)
    existing = set()
    for ws in wb.worksheets:
        try:
            existing.update(ws.tables.keys())
        except Exception:
            pass

    candidate = base
    idx = 1
    while candidate in existing:
        suffix = f"_{idx}"
        candidate = f"{base[:200-len(suffix)]}{suffix}"
        idx += 1
    return candidate


def _add_excel_table(ws, name: str, start_row: int, end_row: int, end_col: int, *, style: str = "TableStyleMedium2"):
    """Ajoute un vrai tableau Excel, sans créer de fichier corrompu.

    On refuse les plages vides, on nettoie les en-têtes et on garantit un nom
    unique au niveau du classeur.
    """
    if end_row <= start_row or end_col <= 0:
        return None
    if ws.parent is None:
        return None

    seen = set()
    for col in range(1, end_col + 1):
        cell = ws.cell(row=start_row, column=col)
        value = str(cell.value or f"Colonne {col}").strip() or f"Colonne {col}"
        # Les en-têtes de table doivent être des chaînes non vides et uniques.
        base = value
        idx = 2
        while value in seen:
            value = f"{base} {idx}"
            idx += 1
        seen.add(value)
        cell.value = value

    ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=_unique_excel_table_name(ws.parent, name), ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name=style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    return table


def _format_xlsx_header_row(ws, row_idx: int, end_col: int):
    fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
    for col in range(1, end_col + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _format_magato_workbook(wb: Workbook):
    for ws in wb.worksheets:
        try:
            _auto_fit_statsimpact_ws(ws)
        except Exception:
            pass
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.00"


def _apply_kpi_color(cell, value, mode: str = "positive"):
    """Coloration douce des KPI. mode='positive' : haut = bon, mode='warning' : haut = vigilance."""
    try:
        v = float(value)
    except Exception:
        return

    if mode == "warning":
        if v >= 100:
            cell.fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
        elif v >= 75:
            cell.fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
        elif v > 0:
            cell.fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
        return

    if v >= 80:
        cell.fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    elif v >= 50:
        cell.fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    elif v > 0:
        cell.fill = PatternFill(fill_type="solid", fgColor="FCE4D6")


def _style_financeur_sheet(ws):
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.sheet_view.showGridLines = False


def _add_financeur_quick_sheet(wb: Workbook, summary_rows: list[list], alert_rows: list[list], top_rows: list[list]):
    ws = wb.create_sheet("Lecture financeur", 0)
    ws.append(["Lecture rapide financeur"])
    ws["A1"].font = Font(bold=True, size=16)
    ws.append([])
    ws.append(["Indicateur", "Valeur", "Lecture"])
    start_summary = ws.max_row

    for row in summary_rows:
        ws.append(row)

    end_summary = ws.max_row
    _format_xlsx_header_row(ws, start_summary, 3)
    _add_excel_table(ws, "LectureFinanceurSynthese", start_summary, end_summary, 3, style="TableStyleMedium9")

    ws.append([])
    ws.append(["Points d'attention"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=12)
    ws.append(["Niveau", "Message"])
    start_alert = ws.max_row
    if alert_rows:
        for row in alert_rows:
            ws.append(row)
    else:
        ws.append(["OK", "Aucune alerte majeure détectée sur le périmètre."])
    end_alert = ws.max_row
    _format_xlsx_header_row(ws, start_alert, 2)
    _add_excel_table(ws, "LectureFinanceurAlertes", start_alert, end_alert, 2, style="TableStyleMedium4")

    ws.append([])
    ws.append(["Top ateliers par présence"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=12)
    ws.append(["Secteur", "Atelier", "Présences", "Participants uniques", "Conso kWh", "CO₂ kg"])
    start_top = ws.max_row
    if top_rows:
        for row in top_rows:
            ws.append(row)
    else:
        ws.append(["", "Aucun atelier dans le périmètre", 0, 0, 0, 0])
    end_top = ws.max_row
    _format_xlsx_header_row(ws, start_top, 6)
    _add_excel_table(ws, "LectureFinanceurTopAteliers", start_top, end_top, 6, style="TableStyleMedium2")

    for row in range(start_summary + 1, end_summary + 1):
        label = str(ws.cell(row=row, column=1).value or "")
        val_cell = ws.cell(row=row, column=2)
        if "Remplissage" in label:
            _apply_kpi_color(val_cell, val_cell.value, mode="positive")
        elif "CO₂" in label or "Consommation" in label:
            _apply_kpi_color(val_cell, val_cell.value, mode="warning")

    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 80
    for col in range(4, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18
    _style_financeur_sheet(ws)
    return ws


def _build_magato_per_atelier_workbook(flt) -> Workbook:
    """Magatomatique complet : 1 feuille par atelier + stats riches sur la période filtrée."""
    flt = _ensure_default_export_period(flt)

    # Cloisonnement : un responsable_secteur ne doit exporter que son secteur
    eff_secteur = flt.secteur
    if not can("scope:all_secteurs"):
        eff_secteur = (getattr(current_user, "secteur_assigne", None) or "").strip() or eff_secteur

    # Liste des ateliers dans le périmètre
    aq = _apply_atelier_lifecycle_filters(AtelierActivite.query, flt)
    if eff_secteur:
        aq = aq.filter(AtelierActivite.secteur == eff_secteur)
    ateliers = aq.order_by(AtelierActivite.secteur.asc(), AtelierActivite.nom.asc()).all()
    consumption_by_atelier = _consumption_by_atelier_for_filter(flt)

    grouped_ateliers = {}
    for at in ateliers:
        grouped_ateliers.setdefault(_continuity_group_key(at), []).append(at)

    wb = Workbook()

    # Synthèse globale
    ws0 = wb.active
    ws0.title = "Synthese"

    # Feuille de cadrage : impossible de confondre deux exports qui n'ont pas
    # le même périmètre temporel ou sectoriel.
    ws_params = wb.create_sheet("Paramètres", 0)
    atelier_ids_label = ', '.join(str(x) for x in (getattr(flt, 'atelier_ids', None) or [])) or 'Tous'
    param_rows = [
        ['Période appliquée', _export_period_label(flt)],
        ['Du', getattr(flt, 'date_from', None).isoformat() if getattr(flt, 'date_from', None) else ''],
        ['Au', getattr(flt, 'date_to', None).isoformat() if getattr(flt, 'date_to', None) else ''],
        ['Secteur demandé', getattr(flt, 'secteur', None) or 'Tous'],
        ['Secteur effectif', _effective_statsimpact_secteur(flt) or 'Tous'],
        ['Ateliers sélectionnés', atelier_ids_label],
        ["Date d'export", date.today().isoformat()],
        ['Utilisateur', getattr(current_user, 'email', None) or getattr(current_user, 'username', None) or f"user#{getattr(current_user, 'id', '')}"],
    ]
    _append_section(ws_params, "Périmètre exact du Magatomatique", ['Paramètre', 'Valeur'], param_rows, start_row=1)
    _autosize_columns(ws_params, max_width=50)
    ws_params.sheet_view.showGridLines = False

    ws0.append([f"Magatomatique — période {_export_period_label(flt)} : 1 feuille par atelier (matrice) + stats détaillées"])
    ws0.append([f"Périmètre : secteur {eff_secteur or 'Tous'} · ateliers {atelier_ids_label}"])

    ws0.append([
        "Secteur",
        "Atelier",
        "Nb séances prévisionnelles",
        "Nb séances réelles",
        "Nb présences totales",
        "Nb inscrits (uniques)",
        "Capacité d'accueil (réel)",
        "Nb places prévisionnelles",
        "Nb heures prévisionnelles",
        "Nb heures réalisées",
        "Taux d'occupation (%)",
        "Moyenne / séance",
        "Durée de l'activité (jours)",
        "Moyenne d'âge",
        "Conso kWh",
        "CO₂ kg",
        "Séances équipées",
        "Bas de Creil",
        "Hauts de Creil",
        "Rouher",
        "Autres",
        "Inconnu",
    ])

    # Agrégats globaux provenance + lecture financeur
    global_prov = {"Bas de Creil": 0, "Hauts de Creil": 0, "Rouher": 0, "Autres": 0, "Inconnu": 0}
    global_sessions_real = 0
    global_presences = 0
    global_uniques_sum = 0
    global_real_hours = 0.0
    global_conso_kwh = 0.0
    global_co2_kg = 0.0
    global_conso_sessions = 0
    top_ateliers_rows = []
    financeur_alert_rows = []

    for _gid, members in grouped_ateliers.items():
        at = sorted(members, key=lambda x: x.nom)[0]
        atelier_ids = [m.id for m in members]
        # Sessions de l'atelier dans la période
        sess_q = db.session.query(SessionActivite).filter(SessionActivite.atelier_id.in_(atelier_ids))
        if flt.date_from:
            sess_q = sess_q.filter(func.coalesce(SessionActivite.rdv_date, SessionActivite.date_session) >= flt.date_from)
        if flt.date_to:
            sess_q = sess_q.filter(func.coalesce(SessionActivite.rdv_date, SessionActivite.date_session) <= flt.date_to)
        sess_q = sess_q.order_by(
            func.coalesce(SessionActivite.rdv_date, SessionActivite.date_session).asc(),
            SessionActivite.id.asc(),
        )
        sessions = sess_q.all()

        # 1) KPIs par atelier
        sessions_planned = len(sessions)
        sessions_real = sum(1 for s in sessions if (s.statut or "").lower() != "annulee")
        dates = [d for d in [(s.rdv_date or s.date_session) for s in sessions] if d]
        duration_days = (max(dates) - min(dates)).days if dates else None

        # heures + capacités
        planned_hours = 0.0
        real_hours = 0.0
        planned_capacity = 0
        real_capacity = 0
        for s in sessions:
            mins = _session_duration_minutes(s, at) or 0
            h = float(mins) / 60.0 if mins else 0.0
            planned_hours += h
            cap = s.capacite if s.capacite is not None else (getattr(at, "capacite_defaut", 0) or 0)
            planned_capacity += int(cap or 0)
            if (s.statut or "").lower() != "annulee":
                real_hours += h
                real_capacity += int(cap or 0)

        if sessions_planned == 0:
            # atelier sans session dans le filtre -> ligne 0 + pas de feuille matrice
            ws0.append([at.secteur, at.nom, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, "", "", 0, 0, 0, 0, 0, 0, 0, 0])
            financeur_alert_rows.append(["Info", f"{at.secteur} — {at.nom} : aucune séance dans le périmètre filtré."])
            continue

        session_ids = [s.id for s in sessions]

        pres_rows = db.session.query(PresenceActivite.participant_id, PresenceActivite.session_id)             .filter(PresenceActivite.session_id.in_(session_ids)).all()

        presences_total = len(pres_rows)
        pid_set = sorted({int(pid) for (pid, _sid) in pres_rows if pid is not None})
        inscrits_uniques = len(pid_set)

        # participants (objets) pour âge + provenance
        participants = []
        if pid_set:
            participants = (
                db.session.query(Participant)
                .options(joinedload(Participant.quartier))
                .filter(Participant.id.in_(pid_set))
                .all()
            )

        ages = [p.age for p in participants if getattr(p, "age", None) is not None]
        age_avg = round(sum(ages) / len(ages), 1) if ages else None

        prov = {"Bas de Creil": 0, "Hauts de Creil": 0, "Rouher": 0, "Autres": 0, "Inconnu": 0}
        for p in participants:
            qname = p.quartier.nom if getattr(p, "quartier", None) is not None else None
            b = _quartier_bucket(qname)
            prov[b] = prov.get(b, 0) + 1
        for k in global_prov:
            global_prov[k] += prov.get(k, 0)

        occupation_rate = round((presences_total / real_capacity * 100.0), 1) if real_capacity > 0 else 0.0
        avg_per_session = round((presences_total / sessions_real), 2) if sessions_real > 0 else 0.0
        conso_payload = consumption_by_atelier.get(_continuity_group_key(at), {})
        conso_kwh = round(float(conso_payload.get("kwh") or 0.0), 3)
        conso_co2 = round(float(conso_payload.get("co2_kg") or 0.0), 3)
        conso_sessions = int(conso_payload.get("sessions_count") or 0)

        global_sessions_real += int(sessions_real or 0)
        global_presences += int(presences_total or 0)
        global_uniques_sum += int(inscrits_uniques or 0)
        global_real_hours += float(real_hours or 0.0)
        global_conso_kwh += float(conso_kwh or 0.0)
        global_co2_kg += float(conso_co2 or 0.0)
        global_conso_sessions += int(conso_sessions or 0)

        top_ateliers_rows.append([
            at.secteur,
            at.nom,
            presences_total,
            inscrits_uniques,
            conso_kwh,
            conso_co2,
        ])

        if sessions_real > 0 and presences_total == 0:
            financeur_alert_rows.append(["Attention", f"{at.secteur} — {at.nom} : séances réalisées mais aucune présence saisie."])
        if real_capacity > 0 and occupation_rate < 30:
            financeur_alert_rows.append(["Vigilance", f"{at.secteur} — {at.nom} : taux d'occupation faible ({occupation_rate}%)."])
        if sessions_real > 0 and conso_sessions == 0:
            financeur_alert_rows.append(["Info", f"{at.secteur} — {at.nom} : aucune consommation matériel renseignée."])

        ws0.append([
            at.secteur,
            at.nom,
            sessions_planned,
            sessions_real,
            presences_total,
            inscrits_uniques,
            real_capacity,
            planned_capacity,
            round(planned_hours, 2),
            round(real_hours, 2),
            occupation_rate,
            avg_per_session,
            duration_days if duration_days is not None else "",
            age_avg if age_avg is not None else "",
            conso_kwh,
            conso_co2,
            conso_sessions,
            prov.get("Bas de Creil", 0),
            prov.get("Hauts de Creil", 0),
            prov.get("Rouher", 0),
            prov.get("Autres", 0),
            prov.get("Inconnu", 0),
        ])

        # 2) Feuille atelier : bloc stats + matrice
        group_label = at.nom
        if len(members) > 1:
            aliases = sorted({m.nom for m in members if m.nom != at.nom})
            if aliases:
                group_label = f"{at.nom} ({', '.join(aliases)})"
        ws = wb.create_sheet(_safe_sheet_title(f"{at.nom}"))

        ws.append([f"{at.secteur} — {group_label}"])
        ws.append([f"Période : {_export_period_label(flt)}"])

        ws.append(["Statistiques", "Valeur"])
        ws.append(["Nb séances prévisionnelles", sessions_planned])
        ws.append(["Nb séances réelles", sessions_real])
        ws.append(["Nb présences totales", presences_total])
        ws.append(["Nb inscrits (uniques)", inscrits_uniques])
        ws.append(["Capacité d'accueil (réel)", real_capacity])
        ws.append(["Nb places prévisionnelles", planned_capacity])
        ws.append(["Nb heures prévisionnelles", round(planned_hours, 2)])
        ws.append(["Nb heures réalisées", round(real_hours, 2)])
        ws.append(["Taux d'occupation (%)", occupation_rate])
        ws.append(["Moyenne / séance", avg_per_session])
        ws.append(["Durée de l'activité (jours)", duration_days if duration_days is not None else ""])
        ws.append(["Moyenne d'âge", age_avg if age_avg is not None else ""])
        ws.append(["Consommation estimée (kWh)", conso_kwh])
        ws.append(["CO₂ estimé (kg)", conso_co2])
        ws.append(["Séances équipées", conso_sessions])

        ws.append([])
        ws.append(["Provenance (quartiers)", "Nb inscrits"])
        ws.append(["Bas de Creil", prov.get("Bas de Creil", 0)])
        ws.append(["Hauts de Creil", prov.get("Hauts de Creil", 0)])
        ws.append(["Rouher", prov.get("Rouher", 0)])
        ws.append(["Autres", prov.get("Autres", 0)])
        ws.append(["Inconnu", prov.get("Inconnu", 0)])

        ws.append([])

        # Matrice participants × sessions
        # (A) Antoine wanted a quick visual identification: âge + ville + quartier à côté du nom/prénom.
        headers = ["Nom", "Prénom", "Âge", "Ville", "Quartier"] + [
            ((d.strftime("%d/%m/%Y")) if (d := (s.rdv_date or s.date_session)) else "Sans date")
            for s in sessions
        ]
        ws.append(headers)

        matrix_header_row = ws.max_row
        if not pres_rows or not pid_set:
            # matrice vide structurée
            _format_xlsx_header_row(ws, matrix_header_row, len(headers))
            ws.auto_filter.ref = f"A{matrix_header_row}:{get_column_letter(len(headers))}{matrix_header_row}"
            # Feuilles atelier : pas de volet figé.
            # Le bloc stats au-dessus de la matrice rendait le défilement pénible
            # en scindant Excel en 4 zones. Les filtres du tableau restent actifs.
            ws.freeze_panes = None
            ws.column_dimensions[get_column_letter(1)].width = 20
            ws.column_dimensions[get_column_letter(2)].width = 18
            ws.column_dimensions[get_column_letter(3)].width = 8
            ws.column_dimensions[get_column_letter(4)].width = 18
            ws.column_dimensions[get_column_letter(5)].width = 22
            for col_idx in range(6, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 12
            continue

        # Participants (objets) pour trier + infos d'identification (âge/ville/quartier)
        parts = (
            db.session.query(Participant)
            .options(joinedload(Participant.quartier))
            .filter(Participant.id.in_(pid_set))
            .order_by(Participant.nom.asc(), Participant.prenom.asc())
            .all()
        )

        sid_index = {int(s.id): idx for idx, s in enumerate(sessions)}
        present = {(int(pid), int(sid)) for (pid, sid) in pres_rows if pid is not None and sid is not None}

        for p in parts:
            pid = int(p.id)
            nom = p.nom or ""
            prenom = p.prenom or ""
            age = getattr(p, "age", None)
            age = age if age is not None else ""
            ville = p.ville or ""
            quartier = p.quartier.nom if getattr(p, "quartier", None) is not None else ""

            row = [nom, prenom, age, ville, quartier] + [""] * len(sessions)
            for sid, idx in sid_index.items():
                if (int(pid), int(sid)) in present:
                    row[5 + idx] = "1"
            ws.append(row)

        matrix_last_row = ws.max_row
        _format_xlsx_header_row(ws, matrix_header_row, len(headers))
        _add_excel_table(ws, f"Matrice_{at.id}", matrix_header_row, matrix_last_row, len(headers), style="TableStyleMedium4")
        # Feuilles atelier : pas de volet figé.
        # Le bloc stats au-dessus de la matrice rendait le défilement pénible
        # en scindant Excel en 4 zones. Les filtres du tableau restent actifs.
        ws.freeze_panes = None
        ws.sheet_view.showGridLines = False

        # Largeurs
        ws.column_dimensions[get_column_letter(1)].width = 20
        ws.column_dimensions[get_column_letter(2)].width = 18
        ws.column_dimensions[get_column_letter(3)].width = 8
        ws.column_dimensions[get_column_letter(4)].width = 18
        ws.column_dimensions[get_column_letter(5)].width = 22
        for col_idx in range(6, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12

    synthese_last_row = ws0.max_row
    _format_xlsx_header_row(ws0, 3, 22)
    _add_excel_table(ws0, "SyntheseAteliers", 3, synthese_last_row, 22, style="TableStyleMedium9")
    ws0.freeze_panes = "C4"
    ws0.sheet_view.showGridLines = False

    # Couleurs douces sur KPI clés de la synthèse
    for row_idx in range(4, synthese_last_row + 1):
        _apply_kpi_color(ws0.cell(row=row_idx, column=11), ws0.cell(row=row_idx, column=11).value, mode="positive")
        _apply_kpi_color(ws0.cell(row=row_idx, column=15), ws0.cell(row=row_idx, column=15).value, mode="warning")
        _apply_kpi_color(ws0.cell(row=row_idx, column=16), ws0.cell(row=row_idx, column=16).value, mode="warning")

    top_ateliers_rows = sorted(top_ateliers_rows, key=lambda r: int(r[2] or 0), reverse=True)[:10]
    total_real_capacity = sum(int(ws0.cell(row=r, column=7).value or 0) for r in range(4, synthese_last_row + 1))
    avg_fill_global = round((global_presences / total_real_capacity * 100), 1) if total_real_capacity else 0

    summary_rows = [
        ["Période", _export_period_label(flt), "Périmètre temporel exact appliqué à cet export."],
        ["Séances réalisées", global_sessions_real, "Volume total de séances non annulées dans le périmètre."],
        ["Présences totales", global_presences, "Nombre total d'émargements saisis."],
        ["Participants uniques cumulés", global_uniques_sum, "Somme des participants uniques par atelier ; un participant présent sur plusieurs ateliers peut être compté plusieurs fois."],
        ["Heures réalisées", round(global_real_hours, 2), "Volume horaire animé sur les séances réalisées."],
        ["Remplissage moyen estimé (%)", avg_fill_global, "Indicateur de lecture rapide basé sur présences / capacité réelle cumulée."],
        ["Consommation estimée (kWh)", round(global_conso_kwh, 3), "Estimation issue du matériel renseigné par séance."],
        ["CO₂ estimé (kg)", round(global_co2_kg, 3), "Conversion kWh × facteur CO₂ du référentiel de consommation."],
        ["Séances équipées", global_conso_sessions, "Nombre de séances avec au moins une ligne matériel renseignée."],
    ]
    _add_financeur_quick_sheet(wb, summary_rows, financeur_alert_rows[:20], top_ateliers_rows)
    # Bloc "Provenance globale" en bas de la synthèse
    ws0.append([])
    ws0.append(["Provenance globale (inscrits uniques sur l'ensemble du périmètre)"])
    ws0.append(["Bas de Creil", global_prov.get("Bas de Creil", 0)])
    ws0.append(["Hauts de Creil", global_prov.get("Hauts de Creil", 0)])
    ws0.append(["Rouher", global_prov.get("Rouher", 0)])
    ws0.append(["Autres", global_prov.get("Autres", 0)])
    ws0.append(["Inconnu", global_prov.get("Inconnu", 0)])

    # Largeurs synthèse
    for col in range(1, 23):
        ws0.column_dimensions[get_column_letter(col)].width = 20 if col <= 2 else 18

    _format_magato_workbook(wb)
    return wb


@bp.route("/stats-impact/export-fidele.xlsx", methods=["GET"])
@login_required
def export_fidele_xlsx():
    if not _can_view():
        abort(403)

    flt = _ensure_default_export_period(normalize_filters(request.args, user=current_user))

    wb = _build_statsimpact_scope_workbook(flt)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name=_magatomatique_filename("statsimpact_export_fidele", flt),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/stats-impact/magatomatique.csv", methods=["GET"])
@login_required
def magatomatique_export_csv():
    if not _can_view():
        abort(403)

    flt = _ensure_default_export_period(normalize_filters(request.args, user=current_user))
    participant_q = (request.args.get("participant_q") or "").strip() or None

    _csv_groups, csv_defaults, allowed_fields = _csv_fields_for_current_user()

    fields = request.args.getlist("fields")
    if not fields:
        fields = list(csv_defaults)
    fields = [f for f in fields if f in CSV_FIELD_MAP and f in allowed_fields]
    if not fields:
        fields = list(csv_defaults) if csv_defaults else [f for f in CSV_DEFAULT_FIELDS if f in allowed_fields]

    query = _query_presence_export(flt, participant_q=participant_q)

    output = StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow([CSV_FIELD_MAP[f]["label"] for f in fields])

    for presence, participant, session, atelier, quartier in query.all():
        ctx = {
            "presence": presence,
            "participant": participant,
            "session": session,
            "atelier": atelier,
            "quartier": quartier,
        }
        writer.writerow([CSV_FIELD_MAP[f]["getter"](ctx) for f in fields])

    csv_name = f"magatomatique_export_{_export_period_slug(flt)}.csv"
    resp = Response(output.getvalue(), mimetype="text/csv")
    resp.headers["Content-Disposition"] = f"attachment; filename={csv_name}"
    return resp


@bp.route("/stats-impact/magatomatique.xlsx", methods=["GET"])
@login_required
def magatomatique_export():
    if not _can_view():
        abort(403)

    flt = _ensure_default_export_period(normalize_filters(request.args, user=current_user))

    export_mode = (request.args.get("export_mode") or "complete").strip().lower()
    participant_q = (request.args.get("participant_q") or "").strip() or None
    view = (request.args.get("magato_view") or "macro").strip().lower()
    try:
        max_sessions = int(request.args.get("max_sessions") or 40)
    except Exception:
        max_sessions = 40
    try:
        max_participants = int(request.args.get("max_participants") or 250)
    except Exception:
        max_participants = 250

    # bornes (export raisonnable)
    max_sessions = max(5, min(max_sessions, 400))
    max_participants = max(20, min(max_participants, 5000))

    # Mode complet : période explicite + lecture financeur + synthèse + 1 feuille par atelier.
    # C'est désormais le Magatomatique principal, pour éviter les écarts avec
    # l'ancien export résumé qui ne portait pas les mêmes conventions de calcul.
    if export_mode in ("complete", "complet", "full", "per_atelier", "per-atelier", "atelier"):
        wb = _build_magato_per_atelier_workbook(flt)
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        prefix = "magatomatique_par_atelier" if export_mode in ("per_atelier", "per-atelier", "atelier") else "magatomatique_complet"
        filename = _magatomatique_filename(prefix, flt)
        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    magato = compute_magatomatique(
        flt,
        participant_q=participant_q,
        view=view,
        max_sessions=max_sessions,
        max_participants=max_participants,
    )

    if magato.get("restricted"):
        abort(403)

    wb = Workbook()
    ws = wb.active
    ws.title = "Synthese"

    # En-têtes synthèse macro (secteurs)
    ws.append(["Synthèse par secteur"])
    ws.append(["Secteur", "Nb sessions", "Nb présences", "Participants uniques"])
    for r in (magato.get("macro") or {}).get("by_secteur", []):
        ws.append([r["secteur"], r["nb_sessions"], r["nb_presences"], r["nb_participants_uniques"]])

    ws.append([])
    ws.append(["Synthèse par atelier"])
    ws.append(["Secteur", "Atelier", "Nb sessions", "Nb présences", "Participants uniques"])
    for r in (magato.get("macro") or {}).get("by_atelier", []):
        ws.append([r["secteur"], r["atelier_nom"], r["nb_sessions"], r["nb_presences"], r["nb_participants_uniques"]])

    # Feuille participants (si dispo)
    if magato.get("participants"):
        ws2 = wb.create_sheet("Participants")
        ws2.append(["Participants (dans le périmètre filtré)"])
        ws2.append(["Nom", "Prénom", "Ville", "Quartier", "Nb présences", "1ère venue", "Dernière venue"])
        for p in magato["participants"]:
            fd = p.get("first_date")
            ld = p.get("last_date")
            ws2.append([
                p.get("nom",""),
                p.get("prenom",""),
                p.get("ville") or "",
                p.get("quartier") or "",
                int(p.get("nb_presences",0)),
                fd.strftime("%Y-%m-%d") if fd else "",
                ld.strftime("%Y-%m-%d") if ld else "",
            ])

    # Feuille matrice (si view=matrix)
    if magato.get("view") == "matrix" and magato.get("sessions") and magato.get("participants"):
        ws3 = wb.create_sheet("Matrice")
        sessions = magato["sessions"]
        participants = magato["participants"]
        matrix = magato.get("matrix") or {}

        header = ["Nom", "Prénom"] + [f'{s["atelier"]} · {s["label"]}' for s in sessions]
        ws3.append(header)

        for p in participants:
            row = [p.get("nom",""), p.get("prenom","")]
            pid = int(p["id"])
            for s in sessions:
                sid = int(s["id"])
                row.append("1" if matrix.get((pid, sid)) else "")
            ws3.append(row)

        # Ajuste largeur colonnes
        for col_idx in range(1, len(header) + 1):
            ws3.column_dimensions[get_column_letter(col_idx)].width = 16 if col_idx <= 2 else 12

        ws4 = wb.create_sheet("Participations")
        ws4.append(["Nom", "Prénom", "Atelier", "Secteur", "Date session", "ID session"])
        for p in participants:
            pid = int(p["id"])
            for s in sessions:
                sid = int(s["id"])
                if matrix.get((pid, sid)):
                    ws4.append(
                        [
                            p.get("nom", ""),
                            p.get("prenom", ""),
                            s.get("atelier", ""),
                            s.get("secteur", ""),
                            s.get("date").strftime("%Y-%m-%d") if s.get("date") else "",
                            sid,
                        ]
                    )

        for col_idx in range(1, 7):
            ws4.column_dimensions[get_column_letter(col_idx)].width = 18

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = _magatomatique_filename("magatomatique_simple", flt)
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
