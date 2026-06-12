from __future__ import annotations


from datetime import date

from urllib.parse import urlencode

from flask import Blueprint, current_app
from app.rbac import can

from sqlalchemy import func

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.extensions import db

from app.models import (
    AtelierActivite,
    SessionActivite,
)


from app.services.consumption import aggregate_sessions_consumption, calculate_session_consumption, aggregate_individual_consumption

from .engine import (
    _has_atelier_filter,
    _apply_common_filters,
    _session_date_expr,
    _resolve_secteur_scope,
)


bp = Blueprint("statsimpact", __name__, url_prefix="")



CSV_FIELD_GROUPS = [
    {
        "label": "Participant",
        "fields": [
            {"key": "participant_id", "label": "ID participant"},
            {"key": "participant_nom", "label": "Nom"},
            {"key": "participant_prenom", "label": "Prénom"},
            {"key": "participant_email", "label": "Email"},
            {"key": "participant_telephone", "label": "Téléphone"},
            {"key": "participant_ville", "label": "Ville"},
            {"key": "participant_quartier", "label": "Quartier"},
            {"key": "participant_genre", "label": "Genre"},
            {"key": "participant_type_public", "label": "Type public"},
            {"key": "participant_date_naissance", "label": "Date naissance"},
            {"key": "participant_pays_origine", "label": "Pays d'origine"},
            {"key": "participant_titre_sejour_type", "label": "Titre de séjour (type)"},
            {"key": "participant_date_entree_dispositif", "label": "Date entrée dispositif"},
            {"key": "participant_date_sortie_dispositif", "label": "Date sortie dispositif"},
            {"key": "participant_diplome_obtenu", "label": "Diplôme obtenu"},
            {"key": "participant_cir_obtenu", "label": "CIR obtenu"},
        ],
    },
    {
        "label": "Session",
        "fields": [
            {"key": "session_id", "label": "ID session"},
            {"key": "session_date", "label": "Date session"},
            {"key": "session_type", "label": "Type session"},
            {"key": "session_statut", "label": "Statut session"},
            {"key": "session_heure_debut", "label": "Heure début"},
            {"key": "session_heure_fin", "label": "Heure fin"},
            {"key": "session_duree_minutes", "label": "Durée (minutes)"},
        ],
    },
    {
        "label": "Atelier",
        "fields": [
            {"key": "atelier_id", "label": "ID atelier"},
            {"key": "atelier_nom", "label": "Nom atelier"},
            {"key": "atelier_secteur", "label": "Secteur atelier"},
            {"key": "atelier_type", "label": "Type atelier"},
        ],
    },
    {
        "label": "Présence",
        "fields": [
            {"key": "presence_id", "label": "ID présence"},
            {"key": "presence_motif", "label": "Motif"},
            {"key": "presence_motif_autre", "label": "Motif autre"},
            {"key": "presence_created_at", "label": "Date d'émargement"},
        ],
    },
]

CSV_DEFAULT_FIELDS = [
    "participant_nom",
    "participant_prenom",
    "atelier_nom",
    "atelier_secteur",
    "session_date",
    "session_type",
]


def _fmt_date(value) -> str:
    return value.strftime("%Y-%m-%d") if value else ""


def _fmt_datetime(value) -> str:
    return value.strftime("%Y-%m-%d %H:%M") if value else ""


CSV_FIELD_MAP = {
    "participant_id": {"label": "ID participant", "getter": lambda ctx: ctx["participant"].id},
    "participant_nom": {"label": "Nom", "getter": lambda ctx: ctx["participant"].nom or ""},
    "participant_prenom": {"label": "Prénom", "getter": lambda ctx: ctx["participant"].prenom or ""},
    "participant_email": {"label": "Email", "getter": lambda ctx: ctx["participant"].email or ""},
    "participant_telephone": {"label": "Téléphone", "getter": lambda ctx: ctx["participant"].telephone or ""},
    "participant_ville": {"label": "Ville", "getter": lambda ctx: ctx["participant"].ville or ""},
    "participant_quartier": {
        "label": "Quartier",
        "getter": lambda ctx: ctx["quartier"].nom if ctx["quartier"] else "",
    },
    "participant_genre": {"label": "Genre", "getter": lambda ctx: ctx["participant"].genre or ""},
    "participant_type_public": {"label": "Type public", "getter": lambda ctx: ctx["participant"].type_public or ""},
    "participant_date_naissance": {
        "label": "Date naissance",
        "getter": lambda ctx: _fmt_date(ctx["participant"].date_naissance),
    },
    "participant_pays_origine": {"label": "Pays d'origine", "getter": lambda ctx: ctx["participant"].pays_origine or ""},
    "participant_titre_sejour_type": {"label": "Titre de séjour (type)", "getter": lambda ctx: ctx["participant"].titre_sejour_type or ""},
    "participant_date_entree_dispositif": {
        "label": "Date entrée dispositif",
        "getter": lambda ctx: _fmt_date(ctx["participant"].date_entree_dispositif),
    },
    "participant_date_sortie_dispositif": {
        "label": "Date sortie dispositif",
        "getter": lambda ctx: _fmt_date(ctx["participant"].date_sortie_dispositif),
    },
    "participant_diplome_obtenu": {"label": "Diplôme obtenu", "getter": lambda ctx: ctx["participant"].diplome_obtenu or ""},
    "participant_cir_obtenu": {
        "label": "CIR obtenu",
        "getter": lambda ctx: (
            "Oui"
            if ctx["participant"].cir_obtenu is True
            else ("Non" if ctx["participant"].cir_obtenu is False else "")
        ),
    },
    "session_id": {"label": "ID session", "getter": lambda ctx: ctx["session"].id},
    "session_date": {
        "label": "Date session",
        "getter": lambda ctx: _fmt_date(ctx["session"].rdv_date or ctx["session"].date_session),
    },
    "session_type": {"label": "Type session", "getter": lambda ctx: ctx["session"].session_type or ""},
    "session_statut": {"label": "Statut session", "getter": lambda ctx: ctx["session"].statut or ""},
    "session_heure_debut": {
        "label": "Heure début",
        "getter": lambda ctx: ctx["session"].rdv_debut or ctx["session"].heure_debut or "",
    },
    "session_heure_fin": {
        "label": "Heure fin",
        "getter": lambda ctx: ctx["session"].rdv_fin or ctx["session"].heure_fin or "",
    },
    "session_duree_minutes": {
        "label": "Durée (minutes)",
        "getter": lambda ctx: ctx["session"].duree_minutes or "",
    },
    "atelier_id": {"label": "ID atelier", "getter": lambda ctx: ctx["atelier"].id},
    "atelier_nom": {"label": "Nom atelier", "getter": lambda ctx: ctx["atelier"].nom or ""},
    "atelier_secteur": {"label": "Secteur atelier", "getter": lambda ctx: ctx["atelier"].secteur or ""},
    "atelier_type": {"label": "Type atelier", "getter": lambda ctx: ctx["atelier"].type_atelier or ""},
    "presence_id": {"label": "ID présence", "getter": lambda ctx: ctx["presence"].id},
    "presence_motif": {"label": "Motif", "getter": lambda ctx: ctx["presence"].motif or ""},
    "presence_motif_autre": {"label": "Motif autre", "getter": lambda ctx: ctx["presence"].motif_autre or ""},
    "presence_created_at": {
        "label": "Date d'émargement",
        "getter": lambda ctx: _fmt_datetime(ctx["presence"].created_at),
    },
}

SENSITIVE_INSERTION_FIELDS = {
    "participant_pays_origine",
    "participant_titre_sejour_type",
    "participant_date_entree_dispositif",
    "participant_date_sortie_dispositif",
    "participant_diplome_obtenu",
    "participant_cir_obtenu",
}


def _csv_fields_for_current_user() -> tuple[list[dict], list[str], set[str]]:
    can_sensitive = can("insertion:sensitive_view") and can("statsimpact:insertion_export")
    allowed_fields = set(CSV_FIELD_MAP.keys())
    if not can_sensitive:
        allowed_fields -= SENSITIVE_INSERTION_FIELDS

    groups: list[dict] = []
    for group in CSV_FIELD_GROUPS:
        filtered = [f for f in group["fields"] if f["key"] in allowed_fields]
        if filtered:
            groups.append({"label": group["label"], "fields": filtered})

    defaults = [f for f in CSV_DEFAULT_FIELDS if f in allowed_fields]
    return groups, defaults, allowed_fields


def _can_view() -> bool:
    return can("statsimpact:view") or can("statsimpact:view_all")


def _query_args_preserve_lists(multidict) -> dict:
    if not hasattr(multidict, "lists"):
        return dict(multidict or {})
    payload = {}
    for key, values in multidict.lists():
        cleaned = [value for value in values if value not in (None, "")]
        if not cleaned:
            continue
        payload[key] = cleaned if len(cleaned) > 1 else cleaned[0]
    return payload


def _safe_sheet_title(name: str, fallback: str = "Atelier") -> str:
    """Openpyxl: max 31 chars, no [ ] * ? / \\ etc."""
    if not name:
        name = fallback
    bad = set('[]:*?/\\')
    cleaned = "".join(c for c in name if c not in bad).strip()
    cleaned = cleaned[:31] if cleaned else fallback
    return cleaned




def _auto_fit_statsimpact_ws(ws, min_width: int = 10, max_width: int = 42):
    for col_cells in ws.columns:
        try:
            idx = col_cells[0].column
        except Exception:
            continue
        max_len = 0
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[get_column_letter(idx)].width = max(min_width, min(max_len + 2, max_width))


def _autosize_columns(ws, min_width: int = 10, max_width: int = 42):
    """Alias robuste utilisé par les exports Magatomatique.

    Certains blocs historiques appelaient `_auto_fit_statsimpact_ws`, le patch
    Magatomatique période appelle `_autosize_columns`. On centralise le même
    comportement pour éviter un NameError en export XLSX.
    """
    return _auto_fit_statsimpact_ws(ws, min_width=min_width, max_width=max_width)


def _append_section(ws, title: str, headers: list[str], rows: list[list], *, start_row: int = 1) -> int:
    row = start_row
    ws.cell(row=row, column=1, value=title)
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
    for idx, header in enumerate(headers, start=1):
        c = ws.cell(row=row, column=idx, value=header)
        c.font = Font(bold=True)
        c.fill = fill
        c.alignment = Alignment(horizontal="center")
    row += 1
    for data in rows:
        for idx, value in enumerate(data, start=1):
            ws.cell(row=row, column=idx, value=value)
        row += 1
    return row


def _effective_statsimpact_secteur(flt):
    try:
        eff = _resolve_secteur_scope(flt)
        return eff
    except Exception:
        return getattr(flt, 'secteur', None)


def _ensure_default_export_period(flt):
    """Garantit un périmètre temporel explicite pour les exports Stats-Impact.

    La page /stats-impact affiche déjà l'année civile courante quand aucune date
    n'est saisie. Les routes d'export doivent faire pareil, sinon les boutons
    peuvent exporter "tout l'historique" pendant que l'écran affiche "année
    courante" : recette parfaite pour des chiffres qui ne causent pas la même
    langue.
    """
    if not getattr(flt, 'date_from', None) and not getattr(flt, 'date_to', None):
        today = date.today()
        flt.date_from = date(today.year, 1, 1)
        flt.date_to = date(today.year, 12, 31)
    return flt


def _export_period_label(flt) -> str:
    start = getattr(flt, 'date_from', None)
    end = getattr(flt, 'date_to', None)
    start_txt = start.isoformat() if start else 'début non limité'
    end_txt = end.isoformat() if end else 'fin non limitée'
    return f"{start_txt} → {end_txt}"


def _export_period_slug(flt) -> str:
    start = getattr(flt, 'date_from', None)
    end = getattr(flt, 'date_to', None)
    start_txt = start.isoformat() if start else 'debut'
    end_txt = end.isoformat() if end else 'fin'
    return f"{start_txt}_{end_txt}"


def _export_query_string_from_filter(flt, extra: dict | None = None) -> str:
    """Construit des liens d'export avec les dates effectivement appliquées."""
    params: list[tuple[str, str]] = []
    secteur = getattr(flt, 'secteur', None)
    if secteur:
        params.append(('secteur', str(secteur)))
    for aid in getattr(flt, 'atelier_ids', None) or []:
        params.append(('atelier_ids', str(aid)))
    if getattr(flt, 'date_from', None):
        params.append(('date_from', flt.date_from.isoformat()))
    if getattr(flt, 'date_to', None):
        params.append(('date_to', flt.date_to.isoformat()))
    if getattr(flt, 'group_by', None):
        params.append(('group_by', str(flt.group_by)))
    if getattr(flt, 'periode_id', None):
        params.append(('periode_id', str(flt.periode_id)))
    for key, value in (extra or {}).items():
        if value is None:
            continue
        params.append((str(key), str(value)))
    return urlencode(params, doseq=True)


def _magatomatique_filename(prefix: str, flt) -> str:
    return f"{prefix}_{_export_period_slug(flt)}.xlsx"


def _apply_atelier_lifecycle_filters(query, flt, include_inactive: bool = False):
    """Applique le filtrage métier atelier (hors corbeille + actifs par défaut)."""
    query = query.filter(AtelierActivite.is_deleted.is_(False))
    if not include_inactive and not _has_atelier_filter(flt):
        query = query.filter(AtelierActivite.is_active.is_(True))
    return query


def _sessions_for_consumption_stats(flt):
    """Sessions filtrées pour les KPI consommation.

    On ne dépend pas de compute_volume_activity_stats(), car ce service ne renvoie
    pas les objets SessionActivite complets. On reconstruit donc une requête dédiée,
    avec les mêmes filtres stats-impact.
    """
    query = (
        db.session.query(SessionActivite)
        .join(AtelierActivite, SessionActivite.atelier_id == AtelierActivite.id)
        .filter(SessionActivite.is_deleted.is_(False))
        .filter(AtelierActivite.is_deleted.is_(False))
    )

    query = _apply_common_filters(query, flt)
    query = query.filter(func.lower(func.coalesce(SessionActivite.statut, "")) != "annulee")
    query = query.order_by(_session_date_expr().asc(), SessionActivite.id.asc())
    return query.all()


def _consumption_stats_for_filter(flt):
    try:
        return aggregate_sessions_consumption(_sessions_for_consumption_stats(flt))
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        current_app.logger.exception("Impossible de calculer les statistiques de consommation stats-impact")
        return {
            "total_kwh": 0.0,
            "total_co2": 0.0,
            "sessions_count": 0,
            "avg_kwh_per_session": 0.0,
            "top_materiels": [],
        }




def _individual_consumption_stats_for_filter(flt):
    try:
        return aggregate_individual_consumption(
            date_from=getattr(flt, 'date_from', None),
            date_to=getattr(flt, 'date_to', None),
            secteur=_resolve_secteur_scope(flt) if _resolve_secteur_scope(flt) != "__restricted__" else "__NO_ACCESS__",
            atelier_ids=getattr(flt, 'atelier_ids', None) or ([getattr(flt, 'atelier_id', None)] if getattr(flt, 'atelier_id', None) else None),
        )
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        current_app.logger.exception("Impossible de calculer les statistiques de consommation individuelle stats-impact")
        return {
            "total_kwh": 0.0,
            "total_co2": 0.0,
            "sessions_count": 0,
            "presences_count": 0,
            "participants_count": 0,
            "materiel_lines_count": 0,
            "avg_kwh_per_presence": 0.0,
            "top_materiels": [],
            "by_atelier": [],
            "by_secteur": [],
        }


def _individual_consumption_by_atelier_for_filter(flt):
    stats = _individual_consumption_stats_for_filter(flt)
    return {int(row.get('atelier_id')): row for row in (stats.get('by_atelier') or []) if row.get('atelier_id')}

def _consumption_by_atelier_for_filter(flt):
    """Agrège la consommation par atelier, en respectant les filtres Stats-Impact."""
    out = {}
    for session_obj in _sessions_for_consumption_stats(flt):
        try:
            atelier = getattr(session_obj, "atelier", None) or db.session.get(AtelierActivite, session_obj.atelier_id)
            if not atelier:
                continue
            key = _continuity_group_key(atelier)
            payload = calculate_session_consumption(session_obj, atelier=atelier)
            bucket = out.setdefault(key, {"kwh": 0.0, "co2_kg": 0.0, "sessions_count": 0})
            bucket["kwh"] += float(payload.get("total_kwh") or 0.0)
            bucket["co2_kg"] += float(payload.get("co2_kg") or 0.0)
            if payload.get("details"):
                bucket["sessions_count"] += 1
        except Exception:
            current_app.logger.exception("Impossible d'agréger la consommation pour une séance stats-impact")
    for value in out.values():
        value["kwh"] = round(value.get("kwh", 0.0), 3)
        value["co2_kg"] = round(value.get("co2_kg", 0.0), 3)
    return out


def _enrich_stats_with_consumption(stats, flt):
    """Ajoute kWh/CO₂ aux lignes table_ateliers sans casser les stats existantes."""
    if not isinstance(stats, dict):
        return stats
    by_atelier = _consumption_by_atelier_for_filter(flt)
    by_atelier_indiv = _individual_consumption_by_atelier_for_filter(flt)
    for row in stats.get("table_ateliers", []) or []:
        payload = by_atelier.get(row.get("atelier_id"), {})
        row["conso_kwh"] = round(float(payload.get("kwh") or 0.0), 3)
        row["co2_kg"] = round(float(payload.get("co2_kg") or 0.0), 3)
        row["conso_sessions"] = int(payload.get("sessions_count") or 0)
        indiv_payload = by_atelier_indiv.get(row.get("atelier_id"), {})
        row["conso_indiv_kwh"] = round(float(indiv_payload.get("kwh") or 0.0), 3)
        row["co2_indiv_kg"] = round(float(indiv_payload.get("co2_kg") or 0.0), 3)
        row["conso_indiv_presences"] = int(indiv_payload.get("presences_count") or 0)
        row["conso_indiv_participants"] = int(indiv_payload.get("participants_count") or 0)
    return stats


def _continuity_group_key(at: AtelierActivite) -> int:
    return int(getattr(at, "continuity_parent_id", None) or at.id)

