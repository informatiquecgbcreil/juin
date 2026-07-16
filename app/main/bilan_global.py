from io import BytesIO
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from flask import (
    render_template, request, redirect, url_for, current_app, send_file
)
from flask_login import login_required, current_user
from app.rbac import require_perm, can

from app.extensions import db
from app.models import (
    Subvention,
    Projet,
    SubventionProjet,
)


from app.main.common import bp

def _auto_fit_worksheet(ws, min_width: int = 10, max_width: int = 42):
    for col_cells in ws.columns:
        try:
            idx = col_cells[0].column
        except Exception:
            continue
        max_len = 0
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(idx)].width = max(min_width, min(max_len + 2, max_width))


def _write_table(ws, title: str, headers: list[str], rows: list[list], *, start_row: int = 1) -> int:
    row = start_row
    ws.cell(row=row, column=1, value=title)
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
    for idx, header in enumerate(headers, start=1):
        c = ws.cell(row=row, column=idx, value=header)
        c.font = Font(bold=True)
        c.fill = fill
    row += 1
    for data in rows:
        for idx, value in enumerate(data, start=1):
            ws.cell(row=row, column=idx, value=value)
        row += 1
    return row


def _compute_bilan_global_payload():
    has_global_scope = can("scope:all_secteurs")
    annee_raw = (request.args.get("annee") or "").strip()
    secteur_raw = (request.args.get("secteur") or "").strip()
    projet_id_raw = (request.args.get("projet_id") or "").strip()

    selected_annee = None
    if annee_raw:
        try:
            selected_annee = int(annee_raw)
        except ValueError:
            selected_annee = None

    selected_secteur = secteur_raw or None
    selected_projet_id = None
    if projet_id_raw:
        try:
            selected_projet_id = int(projet_id_raw)
        except ValueError:
            selected_projet_id = None

    if not has_global_scope:
        selected_secteur = current_user.secteur_assigne
        if selected_projet_id:
            pjt = db.session.get(Projet, selected_projet_id)
            sec_pjt = (pjt.secteur or "").strip().lower() if pjt else ""
            sec_user = (selected_secteur or "").strip().lower()
            if (not pjt) or (sec_pjt != sec_user):
                selected_projet_id = None

    q = Subvention.query.filter_by(est_archive=False)
    if selected_annee is not None:
        q = q.filter(Subvention.annee_exercice == selected_annee)
    if selected_secteur:
        q = q.filter(Subvention.secteur == selected_secteur)
    if selected_projet_id:
        q = q.join(SubventionProjet, SubventionProjet.subvention_id == Subvention.id).filter(SubventionProjet.projet_id == selected_projet_id)

    subs = q.order_by(Subvention.annee_exercice.desc(), Subvention.secteur.asc(), Subvention.nom.asc()).all()

    totals = {
        "demande": round(sum(float(s.montant_demande or 0) for s in subs), 2),
        "attribue": round(sum(float(s.montant_attribue or 0) for s in subs), 2),
        "recu": round(sum(float(s.montant_recu or 0) for s in subs), 2),
        "reel_lignes": round(sum(float(s.total_reel_lignes or 0) for s in subs), 2),
        "engage": round(sum(float(s.total_engage or 0) for s in subs), 2),
        "reste": round(sum(float(s.total_reste or 0) for s in subs), 2),
    }

    alertes = []
    for s in subs:
        recu = float(s.montant_recu or 0)
        reel_lignes = float(s.total_reel_lignes or 0)
        engage = float(s.total_engage or 0)
        if recu > 0 and reel_lignes == 0:
            alertes.append(f"{s.nom} : reçu {recu:.2f}€ mais lignes réel = 0€ (ventilation manquante).")
        if recu > 0 and reel_lignes > 0 and reel_lignes < recu:
            alertes.append(f"{s.nom} : reçu {recu:.2f}€ mais lignes réel = {reel_lignes:.2f}€ (ventilation incomplète).")
        if reel_lignes > 0 and engage > reel_lignes:
            alertes.append(f"{s.nom} : engagé {engage:.2f}€ > lignes réel {reel_lignes:.2f}€ (dépassement).")

    secteurs = current_app.config.get("SECTEURS", [])
    if not secteurs:
        secteurs = [r[0] for r in db.session.query(db.distinct(Subvention.secteur)).all() if r[0]]
    if not has_global_scope:
        secteurs = [current_user.secteur_assigne]

    projets_q = Projet.query
    if not has_global_scope:
        projets_q = projets_q.filter(Projet.secteur == current_user.secteur_assigne)
    projets = projets_q.order_by(Projet.secteur.asc(), Projet.nom.asc()).all()

    return {
        "subs": subs,
        "totals_dict": totals,
        "totals": type("Obj", (), totals),
        "alertes": alertes,
        "secteurs": secteurs,
        "projets": projets,
        "selected_annee": selected_annee,
        "selected_secteur": selected_secteur,
        "selected_projet_id": selected_projet_id,
    }


def _build_bilan_global_workbook(payload: dict):
    subs = payload["subs"]
    totals = payload["totals_dict"]
    selected_annee = payload["selected_annee"]
    selected_secteur = payload["selected_secteur"]
    selected_projet_id = payload["selected_projet_id"]
    projets = payload["projets"]
    alertes = payload["alertes"]

    projet_label = "Tous"
    if selected_projet_id:
        p = next((x for x in projets if x.id == selected_projet_id), None)
        if p:
            projet_label = f"{p.secteur} · {p.nom}"

    wb = Workbook()
    ws_params = wb.active
    ws_params.title = "Paramètres"
    _write_table(ws_params, "Périmètre exact de l'export", ["Paramètre", "Valeur"], [
        ["Année", selected_annee or "Toutes"],
        ["Secteur", selected_secteur or "Tous"],
        ["Projet", projet_label],
        ["Date d'export", date.today().isoformat()],
        ["Utilisateur", getattr(current_user, "email", None) or getattr(current_user, "username", None) or f"user#{getattr(current_user, 'id', '')}"],
        ["Subventions incluses", len(subs)],
    ], start_row=1)

    ws_syn = wb.create_sheet("Synthèse")
    _write_table(ws_syn, "Synthèse financière", ["Indicateur", "Valeur €"], [
        ["Demandé", totals.get("demande", 0)],
        ["Attribué", totals.get("attribue", 0)],
        ["Reçu", totals.get("recu", 0)],
        ["Ventilé réel", totals.get("reel_lignes", 0)],
        ["Engagé", totals.get("engage", 0)],
        ["Reste", totals.get("reste", 0)],
    ], start_row=1)

    ws_subs = wb.create_sheet("Subventions")
    _write_table(ws_subs, "Subventions incluses", ["Année", "Secteur", "Subvention", "Demandé €", "Attribué €", "Reçu €", "Ventilé réel €", "Engagé €", "Reste €"], [
        [s.annee_exercice, s.secteur, s.nom, float(s.montant_demande or 0), float(s.montant_attribue or 0), float(s.montant_recu or 0), float(s.total_reel_lignes or 0), float(s.total_engage or 0), float(s.total_reste or 0)]
        for s in subs
    ], start_row=1)

    ws_alert = wb.create_sheet("Alertes")
    _write_table(ws_alert, "Alertes", ["Message"], [[a] for a in alertes] or [["Aucune alerte"]], start_row=1)

    for ws in wb.worksheets:
        _auto_fit_worksheet(ws)
    return wb


# --------- Setup start ---------

# --------- Bilan financeurs (global) ---------

@bp.route("/bilan")
@login_required
@require_perm("bilans:view")
def bilan_global():
    payload = _compute_bilan_global_payload()
    return render_template(
        "bilan.html",
        subs=payload["subs"],
        totals=payload["totals"],
        alertes=payload["alertes"],
        secteurs=payload["secteurs"],
        projets=payload["projets"],
        selected_annee=payload["selected_annee"],
        selected_secteur=payload["selected_secteur"],
        selected_projet_id=payload["selected_projet_id"],
    )


@bp.route("/bilan/export.xlsx")
@login_required
@require_perm("bilans:view")
def bilan_global_export_xlsx():
    payload = _compute_bilan_global_payload()
    wb = _build_bilan_global_workbook(payload)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    secteur_label = (payload["selected_secteur"] or "all").replace(" ", "_")
    annee_label = payload["selected_annee"] or "all"
    return send_file(
        bio,
        as_attachment=True,
        download_name=f"bilan_global_{annee_label}_{secteur_label}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# Alias de compat (si ton layout appelle encore main.bilan)

@bp.route("/bilan-global")
@login_required
def bilan():
    return redirect(url_for("main.bilan_global"))



