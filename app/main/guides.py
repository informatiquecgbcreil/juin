"""Guides pas à pas : liste, démarrage, avancement, sortie.

Le guide actif vit en session ; le bandeau est rendu par le layout sur
toutes les pages (voir templates/_guide_bandeau.html).
"""
from flask import flash, redirect, render_template, request, url_for
from flask_login import current_user, login_required

from app.main.common import bp
from app.rbac import require_perm
from app.services.guides import (
    GUIDES,
    _user_can_any,
    avancer_guide,
    demarrer_guide,
    guide_actif_ctx,
    guides_disponibles,
    quitter_guide,
)


def _retour():
    return redirect(request.form.get("next") or request.referrer or url_for("main.guides_liste"))


@bp.route("/guides")
@login_required
@require_perm("dashboard:view")
def guides_liste():
    return render_template("guides.html", guides=guides_disponibles(current_user))


@bp.route("/guides/glossaire")
@login_required
@require_perm("dashboard:view")
def glossaire():
    """Le dico du social : glossaire de base + mots de la structure."""
    from app.aide.glossaire import glossaire_fusionne
    from app.rbac import can

    peut_editer = can("glossaire:edit")
    masques = []
    if peut_editer:
        try:
            from app.models import GlossaireTerme
            masques = [r.terme for r in GlossaireTerme.query.filter_by(masque=True).order_by(GlossaireTerme.terme.asc()).all()]
        except Exception:
            masques = []
    return render_template(
        "glossaire.html",
        glossaire=glossaire_fusionne(),
        peut_editer=peut_editer,
        masques=masques,
    )


@bp.post("/guides/glossaire/enregistrer")
@login_required
@require_perm("glossaire:edit")
def glossaire_enregistrer():
    """Créer un mot, ou modifier un mot (de base ou local)."""
    from app.aide.glossaire import _cle
    from app.extensions import db
    from app.models import GlossaireTerme

    terme = (request.form.get("terme") or "").strip()
    definition = (request.form.get("definition") or "").strip()
    if not terme or not definition:
        flash("Le mot et sa définition sont obligatoires.", "danger")
        return redirect(url_for("main.glossaire"))
    categorie = (request.form.get("categorie") or "").strip() or None
    dans_app = (request.form.get("dans_app") or "").strip() or None
    # terme_original : le nom AVANT modification (permet de renommer un mot).
    original = (request.form.get("terme_original") or terme).strip()

    row = GlossaireTerme.query.filter(
        db.func.lower(GlossaireTerme.terme) == _cle(original)
    ).first()
    if row is None:
        row = GlossaireTerme(terme=terme)
        db.session.add(row)
    row.terme = terme
    row.definition = definition
    row.categorie = categorie
    row.dans_app = dans_app
    row.masque = False
    db.session.commit()
    flash(f"« {terme} » a bien été enregistré dans le glossaire.", "success")
    return redirect(url_for("main.glossaire"))


@bp.post("/guides/glossaire/supprimer")
@login_required
@require_perm("glossaire:edit")
def glossaire_supprimer():
    """Retirer un mot : masque un terme de base, supprime un mot ajouté."""
    from app.aide.glossaire import _cle, termes_de_base
    from app.extensions import db
    from app.models import GlossaireTerme

    terme = (request.form.get("terme") or "").strip()
    if not terme:
        return redirect(url_for("main.glossaire"))
    row = GlossaireTerme.query.filter(
        db.func.lower(GlossaireTerme.terme) == _cle(terme)
    ).first()
    if _cle(terme) in termes_de_base():
        # Terme de base (éventuellement modifié) : on le masque, réversible.
        if row is None:
            row = GlossaireTerme(terme=terme, definition="")
            db.session.add(row)
        row.masque = True
        flash(f"« {terme} » est retiré de l'affichage (récupérable via « Rétablir »).", "info")
    elif row is not None:
        db.session.delete(row)
        flash(f"« {terme} » a été supprimé du glossaire.", "info")
    db.session.commit()
    return redirect(url_for("main.glossaire"))


@bp.post("/guides/glossaire/retablir")
@login_required
@require_perm("glossaire:edit")
def glossaire_retablir():
    """Rétablir la version d'origine d'un terme de base (modifié ou masqué)."""
    from app.aide.glossaire import _cle
    from app.extensions import db
    from app.models import GlossaireTerme

    terme = (request.form.get("terme") or "").strip()
    row = GlossaireTerme.query.filter(
        db.func.lower(GlossaireTerme.terme) == _cle(terme)
    ).first()
    if row is not None:
        db.session.delete(row)
        db.session.commit()
        flash(f"« {terme} » est revenu à sa version d'origine.", "success")
    return redirect(url_for("main.glossaire"))


@bp.post("/guides/glossaire/import")
@login_required
@require_perm("glossaire:edit")
def glossaire_import():
    """Import XLSX : colonnes Terme | Définition | Catégorie | Dans l'application."""
    from app.aide.glossaire import _cle
    from app.extensions import db
    from app.models import GlossaireTerme

    fichier = request.files.get("fichier")
    if not fichier or not fichier.filename:
        flash("Choisis un fichier XLSX à importer.", "danger")
        return redirect(url_for("main.glossaire"))
    try:
        from openpyxl import load_workbook
        wb = load_workbook(fichier, read_only=True, data_only=True)
        ws = wb.active
        lignes = list(ws.iter_rows(values_only=True))
    except Exception:
        flash("Fichier illisible : il faut un fichier Excel (.xlsx).", "danger")
        return redirect(url_for("main.glossaire"))

    if not lignes:
        flash("Le fichier est vide.", "warning")
        return redirect(url_for("main.glossaire"))

    # Ligne d'en-tête tolérante : on repère les colonnes par leur nom.
    entetes = [str(c or "").strip().lower() for c in lignes[0]]

    def _col(*mots):
        for i, h in enumerate(entetes):
            if any(m in h for m in mots):
                return i
        return None

    i_terme = _col("terme", "mot")
    i_def = _col("définition", "definition")
    i_cat = _col("catégorie", "categorie")
    i_app = _col("application", "dans l")
    if i_terme is None or i_def is None:
        flash("Colonnes attendues : « Terme » et « Définition » (puis « Catégorie », « Dans l'application » en option).", "danger")
        return redirect(url_for("main.glossaire"))

    def _cell(ligne, idx):
        if idx is None or idx >= len(ligne):
            return ""
        return str(ligne[idx] or "").strip()

    ajoutes = modifies = ignores = 0
    for ligne in lignes[1:]:
        terme = _cell(ligne, i_terme)
        definition = _cell(ligne, i_def)
        if not terme or not definition:
            if terme or definition:
                ignores += 1
            continue
        row = GlossaireTerme.query.filter(
            db.func.lower(GlossaireTerme.terme) == _cle(terme)
        ).first()
        if row is None:
            row = GlossaireTerme(terme=terme)
            db.session.add(row)
            ajoutes += 1
        else:
            modifies += 1
        row.terme = terme
        row.definition = definition
        row.categorie = _cell(ligne, i_cat) or None
        row.dans_app = _cell(ligne, i_app) or None
        row.masque = False
    db.session.commit()
    msg = f"Import terminé : {ajoutes} mot(s) ajouté(s), {modifies} mis à jour."
    if ignores:
        msg += f" {ignores} ligne(s) incomplète(s) ignorée(s)."
    flash(msg, "success")
    return redirect(url_for("main.glossaire"))


@bp.route("/guides/glossaire/export.xlsx")
@login_required
@require_perm("dashboard:view")
def glossaire_export():
    """Export XLSX du glossaire complet — sert aussi de modèle d'import."""
    import io
    from flask import send_file
    from openpyxl import Workbook
    from app.aide.glossaire import glossaire_fusionne

    wb = Workbook()
    ws = wb.active
    ws.title = "Glossaire"
    ws.append(["Terme", "Définition", "Catégorie", "Dans l'application"])
    for cat in glossaire_fusionne():
        for t in cat["termes"]:
            ws.append([t["terme"], t["definition"], cat["categorie"], t.get("dans_app") or ""])
    for col, largeur in (("A", 34), ("B", 90), ("C", 30), ("D", 60)):
        ws.column_dimensions[col].width = largeur
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="glossaire-dico-du-social.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.post("/guides/<key>/demarrer")
@login_required
@require_perm("dashboard:view")
def guide_demarrer(key: str):
    g = GUIDES.get(key)
    if not g or not _user_can_any(current_user, g.get("perm_any")):
        flash("Ce guide n'est pas disponible avec tes droits.", "danger")
        return redirect(url_for("main.guides_liste"))
    url = demarrer_guide(key)
    flash(f"Guide « {g['titre']} » démarré : suis le bandeau en haut de page.", "success")
    return redirect(url or url_for("main.guides_liste"))


@bp.post("/guides/suivant")
@login_required
def guide_suivant():
    url = avancer_guide(+1)
    if guide_actif_ctx() is None:
        # Dernière étape franchie : le guide est terminé.
        flash("Guide terminé, bravo ! Tu peux le relancer quand tu veux depuis la page Guides.", "success")
        return _retour()
    # Étape suivante : on y va si elle a une page cible, sinon on reste ici.
    return redirect(url) if url else _retour()


@bp.post("/guides/precedent")
@login_required
def guide_precedent():
    url = avancer_guide(-1)
    return redirect(url) if url else _retour()


@bp.post("/guides/quitter")
@login_required
def guide_quitter():
    quitter_guide()
    flash("Guide quitté. Tu peux le reprendre depuis la page Guides.", "info")
    return _retour()
