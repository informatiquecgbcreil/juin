from __future__ import annotations

import json
from collections import Counter
from datetime import date
from io import BytesIO

from flask import render_template, request, redirect, url_for, flash, abort, send_file, jsonify, current_app
from flask_login import login_required, current_user
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.extensions import db
from app.models import Partenaire, PartenaireSecteur, PartenaireIntervention, OrientationAccesDroit, Participant, Quartier
from app.rbac import require_perm, can

from . import bp


ORIENTATION_DOMAINES = {
    "logement": "Logement",
    "caf": "CAF / prestations",
    "sante": "Santé",
    "emploi": "Emploi / insertion",
    "retraite": "Retraite",
    "juridique": "Juridique",
    "administratif": "Administratif",
    "numerique": "Médiation numérique",
    "mobilite": "Mobilité",
    "aide_alimentaire": "Aide alimentaire",
    "violences": "Violences / protection",
    "handicap": "Handicap",
    "education": "Éducation / parentalité",
    "autre": "Autre",
}

ORIENTATION_STATUTS = {
    "oriente": "Orienté",
    "rdv_pris": "Rendez-vous pris",
    "accompagne": "Accompagné",
    "a_rappeler": "À rappeler",
    "resolu": "Résolu",
    "non_abouti": "Non abouti",
}

ORIENTATION_URGENCES = {
    "normale": "Normale",
    "prioritaire": "Prioritaire",
    "urgence": "Urgence",
}

PARTENAIRE_NIVEAUX_ORIENTATION = {
    "principal": "Partenaire principal",
    "relais": "Relais possible",
    "recours": "Dernier recours",
}


def _parse_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except Exception:
        return None


def _parse_int(value, default: int | None = None) -> int | None:
    try:
        if value in (None, "", "None"):
            return default
        return int(value)
    except Exception:
        return default


def _clean_choice(value: str | None, choices: dict, default: str) -> str:
    raw = (value or default).strip()
    return raw if raw in choices else default


def _selected_secteurs_from_request() -> list[str]:
    secteurs = request.values.getlist("secteur")
    cleaned = [s.strip() for s in secteurs if s and s.strip()]
    return list(dict.fromkeys(cleaned))


def _selected_orientation_domains_from_request() -> list[str]:
    raw = request.values.getlist("competence_orientation")
    return [code for code in dict.fromkeys(raw) if code in ORIENTATION_DOMAINES]


def _domain_label(code: str | None) -> str:
    return ORIENTATION_DOMAINES.get(code or "", code or "Non renseigné")


def _status_label(code: str | None) -> str:
    return ORIENTATION_STATUTS.get(code or "", code or "Non renseigné")


def _urgence_label(code: str | None) -> str:
    return ORIENTATION_URGENCES.get(code or "", code or "Non renseigné")


def _partner_competence_labels(partenaire: Partenaire) -> list[str]:
    return [_domain_label(code) for code in partenaire.competences_orientation()]


def _can_scope_all() -> bool:
    return bool(can("scope:all_secteurs") or can("participants:view_all"))


def _current_scope_secteur() -> str | None:
    return getattr(current_user, "secteur_assigne", None) or None


def _orientation_year() -> int:
    raw = request.values.get("year")
    try:
        year = int(raw)
    except Exception:
        year = date.today().year
    return max(2000, min(2100, year))


def _month_label(month: int) -> str:
    labels = {
        1: "Janvier", 2: "Février", 3: "Mars", 4: "Avril",
        5: "Mai", 6: "Juin", 7: "Juillet", 8: "Août",
        9: "Septembre", 10: "Octobre", 11: "Novembre", 12: "Décembre",
    }
    return labels.get(month, str(month))


def _participant_label(participant: Participant | None, fallback: str | None = None) -> str:
    if participant:
        return f"{participant.nom} {participant.prenom}".strip()
    return fallback or "Non rattaché"


def _quartier_label(quartier: Quartier | None) -> str:
    if not quartier:
        return "Quartier inconnu"
    suffix = " (QPV)" if quartier.is_qpv else ""
    return f"{quartier.ville} - {quartier.nom}{suffix}"


def _orientation_quartier(row: OrientationAccesDroit) -> Quartier | None:
    if row.quartier:
        return row.quartier
    participant = row.participant
    return getattr(participant, "quartier", None) if participant else None


def _orientation_quartier_label(row: OrientationAccesDroit) -> str:
    return _quartier_label(_orientation_quartier(row))


def _orientation_ville(row: OrientationAccesDroit) -> str:
    if row.ville:
        return row.ville
    quartier = _orientation_quartier(row)
    if quartier and quartier.ville:
        return quartier.ville
    participant = row.participant
    if participant and participant.ville:
        return participant.ville
    return "Ville inconnue"


def _age_bucket(participant: Participant | None) -> str:
    age = getattr(participant, "age", None) if participant else None
    if age is None:
        return "Inconnu"
    if age < 18:
        return "Moins de 18 ans"
    if age <= 25:
        return "18-25 ans"
    if age <= 59:
        return "26-59 ans"
    return "60 ans et plus"


def _qpv_bucket(participant: Participant | None, quartier: Quartier | None = None) -> str:
    quartier = quartier or (getattr(participant, "quartier", None) if participant else None)
    if not quartier:
        return "Inconnu"
    return "QPV" if quartier.is_qpv else "Hors QPV"


def _orientation_filtered_query(year: int):
    q = OrientationAccesDroit.query
    q = q.filter(OrientationAccesDroit.date_orientation >= date(year, 1, 1))
    q = q.filter(OrientationAccesDroit.date_orientation <= date(year, 12, 31))

    secteur = (request.values.get("secteur_filter") or "").strip()
    if not _can_scope_all():
        secteur = _current_scope_secteur() or secteur
    if secteur:
        q = q.filter(OrientationAccesDroit.secteur == secteur)

    domaine = (request.values.get("domaine") or "").strip()
    if domaine in ORIENTATION_DOMAINES:
        q = q.filter(OrientationAccesDroit.domaine == domaine)

    statut = (request.values.get("statut") or "").strip()
    if statut in ORIENTATION_STATUTS:
        q = q.filter(OrientationAccesDroit.statut == statut)

    partenaire_id = _parse_int(request.values.get("partenaire_id"))
    if partenaire_id:
        q = q.filter(OrientationAccesDroit.partenaire_id == partenaire_id)

    ville = (request.values.get("ville") or "").strip().lower()
    if ville:
        q = q.filter(db.func.lower(db.func.coalesce(OrientationAccesDroit.ville, "")).like(f"%{ville}%"))

    quartier_id = _parse_int(request.values.get("quartier_id"))
    if quartier_id:
        q = q.filter(OrientationAccesDroit.quartier_id == quartier_id)

    qtext = (request.values.get("q") or "").strip().lower()
    if qtext:
        like = f"%{qtext}%"
        q = (
            q.outerjoin(Participant, OrientationAccesDroit.participant_id == Participant.id)
            .outerjoin(Partenaire, OrientationAccesDroit.partenaire_id == Partenaire.id)
            .outerjoin(Quartier, OrientationAccesDroit.quartier_id == Quartier.id)
            .filter(db.or_(
                db.func.lower(OrientationAccesDroit.demande).like(like),
                db.func.lower(db.func.coalesce(OrientationAccesDroit.ville, "")).like(like),
                db.func.lower(db.func.coalesce(OrientationAccesDroit.note, "")).like(like),
                db.func.lower(db.func.coalesce(OrientationAccesDroit.demandeur_libre, "")).like(like),
                db.func.lower(db.func.coalesce(Participant.nom, "")).like(like),
                db.func.lower(db.func.coalesce(Participant.prenom, "")).like(like),
                db.func.lower(db.func.coalesce(Partenaire.nom, "")).like(like),
                db.func.lower(db.func.coalesce(Quartier.nom, "")).like(like),
            ))
        )

    return q


def _counter_rows(counter: Counter, label_fn=lambda key: key, limit: int = 10) -> list[dict]:
    return [
        {"key": key, "label": label_fn(key), "count": count}
        for key, count in counter.most_common(limit)
    ]


def _orientation_dashboard_payload(rows: list[OrientationAccesDroit]) -> dict:
    by_domaine = Counter(row.domaine for row in rows)
    by_status = Counter(row.statut for row in rows)
    by_partner = Counter(row.partenaire.nom if row.partenaire else "Sans partenaire" for row in rows)
    by_month = Counter(row.date_orientation.month for row in rows if row.date_orientation)
    by_ville = Counter(_orientation_ville(row) for row in rows)
    by_quartier = Counter(_orientation_quartier_label(row) for row in rows)
    by_qpv = Counter(_qpv_bucket(row.participant, _orientation_quartier(row)) for row in rows)
    by_genre = Counter((getattr(row.participant, "genre", None) or "Inconnu") for row in rows)
    by_age = Counter(_age_bucket(row.participant) for row in rows)
    partnerless = sum(1 for row in rows if not row.partenaire_id)
    unique_participants = {row.participant_id for row in rows if row.participant_id}

    return {
        "kpis": {
            "total": len(rows),
            "participants": len(unique_participants),
            "partenaires": len({row.partenaire_id for row in rows if row.partenaire_id}),
            "domaines": len({row.domaine for row in rows if row.domaine}),
            "a_rappeler": sum(1 for row in rows if row.statut == "a_rappeler"),
            "sans_partenaire": partnerless,
            "villes": len({label for label in by_ville if label != "Ville inconnue"}),
            "quartiers": len({label for label in by_quartier if label != "Quartier inconnu"}),
        },
        "by_domaine": _counter_rows(by_domaine, _domain_label, 12),
        "by_status": _counter_rows(by_status, _status_label, 10),
        "by_partner": _counter_rows(by_partner, lambda key: key, 12),
        "by_month": _counter_rows(by_month, _month_label, 12),
        "by_ville": _counter_rows(by_ville, lambda key: key, 12),
        "by_quartier": _counter_rows(by_quartier, lambda key: key, 12),
        "by_qpv": _counter_rows(by_qpv, lambda key: key, 10),
        "by_genre": _counter_rows(by_genre, lambda key: key, 10),
        "by_age": _counter_rows(by_age, lambda key: key, 10),
    }


def _orientation_allowed_participants() -> list[Participant]:
    q = Participant.query
    if not _can_scope_all():
        secteur = _current_scope_secteur()
        if secteur:
            q = q.filter(Participant.created_secteur == secteur)
    return q.order_by(Participant.nom.asc(), Participant.prenom.asc()).limit(500).all()


def _orientation_allowed_partners() -> list[Partenaire]:
    return Partenaire.query.order_by(Partenaire.nom.asc()).all()


def _orientation_allowed_quartiers() -> list[Quartier]:
    return Quartier.query.order_by(Quartier.ville.asc(), Quartier.nom.asc()).all()


@bp.route("/")
@login_required
@require_perm("partenaires:view")
def index():
    q = (request.args.get("q") or "").strip()
    secteurs = _selected_secteurs_from_request()
    competence = (request.args.get("competence") or "").strip()
    if competence not in ORIENTATION_DOMAINES:
        competence = ""

    base = Partenaire.query
    if q:
        like = f"%{q.lower()}%"
        base = base.filter(
            db.or_(
                db.func.lower(Partenaire.nom).like(like),
                db.func.lower(db.func.coalesce(Partenaire.contact_nom, "")).like(like),
                db.func.lower(db.func.coalesce(Partenaire.contact_prenom, "")).like(like),
                db.func.lower(db.func.coalesce(Partenaire.email_contact, "")).like(like),
                db.func.lower(db.func.coalesce(Partenaire.email_general, "")).like(like),
                db.func.lower(db.func.coalesce(Partenaire.tel_contact, "")).like(like),
                db.func.lower(db.func.coalesce(Partenaire.tel_general, "")).like(like),
                db.func.lower(db.func.coalesce(Partenaire.competences_orientation_json, "")).like(like),
                db.func.lower(db.func.coalesce(Partenaire.territoire_couvert, "")).like(like),
            )
        )

    if competence:
        base = base.filter(db.func.coalesce(Partenaire.competences_orientation_json, "").like(f"%{competence}%"))

    if secteurs:
        base = (
            base.join(PartenaireSecteur)
            .filter(PartenaireSecteur.secteur.in_(secteurs))
        )

    partenaires = base.order_by(Partenaire.nom.asc()).distinct().all()
    return render_template(
        "partenaires/index.html",
        partenaires=partenaires,
        q=q,
        secteurs=secteurs,
        competence=competence,
        orientation_domaines=ORIENTATION_DOMAINES,
        competence_labels={p.id: _partner_competence_labels(p) for p in partenaires},
    )


# --------------------------------------------------------------------------
# Carte des partenaires (un marqueur cliquable par structure)
# --------------------------------------------------------------------------
@bp.route("/carte")
@login_required
@require_perm("partenaires:view")
def carte():
    from app.services.geocodage import nombre_a_geocoder_partenaires

    secteur_choices = [
        s for (s,) in db.session.query(PartenaireSecteur.secteur)
        .distinct()
        .order_by(PartenaireSecteur.secteur.asc())
        .all() if s
    ]
    return render_template(
        "partenaires/carte.html",
        secteur_choices=secteur_choices,
        restants=nombre_a_geocoder_partenaires(),
        tile_url=current_app.config.get("CARTO_TILE_URL")
        or "https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png",
        tile_attribution=current_app.config.get("CARTO_TILE_ATTRIBUTION")
        or "© contributeurs OpenStreetMap",
    )


@bp.route("/carte/data")
@login_required
@require_perm("partenaires:view")
def carte_data():
    from app.services.cartographie import liste_partenaires

    secteur = (request.args.get("secteur") or "").strip() or None
    return jsonify(liste_partenaires(secteur=secteur))


@bp.route("/carte/geocoder", methods=["POST"])
@login_required
@require_perm("partenaires:edit")
def carte_geocoder():
    from app.services import geocodage as geo

    try:
        limit = int(current_app.config.get("GEOCODAGE_BATCH") or 50)
    except (TypeError, ValueError):
        limit = 50
    try:
        resume = geo.synchroniser_geocodages_partenaires(limit=limit)
    except geo.GeocodageError as exc:
        flash(f"Géocodage indisponible : {exc}", "warning")
        return redirect(url_for("partenaires.carte"))

    if resume["erreurs"]:
        flash(
            f"Géocodage interrompu (réseau) : {resume['localises']} partenaire(s) localisé(s), "
            f"{resume['restants']} restant(s). Réessayez plus tard.",
            "warning",
        )
    else:
        msg = f"{resume['localises']} partenaire(s) localisé(s)"
        if resume["non_localises"]:
            msg += f", {resume['non_localises']} sans coordonnées"
        if resume["restants"]:
            msg += f" · {resume['restants']} restant(s) — relancez pour continuer"
        flash(msg + ".", "success")
    return redirect(url_for("partenaires.carte"))


@bp.route("/orientations", methods=["GET", "POST"])
@login_required
@require_perm("partenaires:view")
def orientations():
    if request.method == "POST":
        if not can("partenaires:edit"):
            abort(403)

        demande = (request.form.get("demande") or "").strip()
        domaine = _clean_choice(request.form.get("domaine"), ORIENTATION_DOMAINES, "autre")
        if not demande:
            flash("La demande de la personne est obligatoire.", "danger")
            return redirect(url_for("partenaires.orientations", year=_orientation_year()))

        participant_id = _parse_int(request.form.get("participant_id"))
        partenaire_id = _parse_int(request.form.get("partenaire_id"))
        quartier_id = _parse_int(request.form.get("quartier_id"))
        participant = db.session.get(Participant, participant_id) if participant_id else None
        partenaire = db.session.get(Partenaire, partenaire_id) if partenaire_id else None
        quartier = db.session.get(Quartier, quartier_id) if quartier_id else None
        if not quartier and participant and getattr(participant, "quartier", None):
            quartier = participant.quartier
        ville = (request.form.get("ville") or "").strip()
        if not ville and quartier:
            ville = quartier.ville
        if not ville and participant and participant.ville:
            ville = participant.ville

        secteur = (request.form.get("secteur") or "").strip() or _current_scope_secteur()
        if not _can_scope_all():
            secteur = _current_scope_secteur()

        item = OrientationAccesDroit(
            date_orientation=_parse_date(request.form.get("date_orientation")) or date.today(),
            secteur=secteur or None,
            ville=ville or None,
            domaine=domaine,
            demande=demande,
            statut=_clean_choice(request.form.get("statut"), ORIENTATION_STATUTS, "oriente"),
            urgence=_clean_choice(request.form.get("urgence"), ORIENTATION_URGENCES, "normale"),
            suite_prevue=_parse_date(request.form.get("suite_prevue")),
            note=(request.form.get("note") or "").strip() or None,
            demandeur_libre=(request.form.get("demandeur_libre") or "").strip() or None,
            participant_id=participant.id if participant else None,
            partenaire_id=partenaire.id if partenaire else None,
            quartier_id=quartier.id if quartier else None,
            created_by_user_id=getattr(current_user, "id", None),
        )
        db.session.add(item)
        db.session.commit()
        flash("Orientation enregistrée.", "success")
        return redirect(url_for("partenaires.orientations", year=item.date_orientation.year, domaine=domaine))

    year = _orientation_year()
    rows = (
        _orientation_filtered_query(year)
        .order_by(OrientationAccesDroit.date_orientation.desc(), OrientationAccesDroit.created_at.desc())
        .all()
    )
    dashboard = _orientation_dashboard_payload(rows)
    partenaires = _orientation_allowed_partners()
    participants = _orientation_allowed_participants()
    quartiers = _orientation_allowed_quartiers()
    selected_domain = (request.args.get("domaine") or "").strip()
    suggested_partners = []
    if selected_domain in ORIENTATION_DOMAINES:
        suggested_partners = [
            p for p in partenaires
            if selected_domain in p.competences_orientation()
        ]

    return render_template(
        "partenaires/orientations.html",
        year=year,
        rows=rows[:120],
        dashboard=dashboard,
        partenaires=partenaires,
        participants=participants,
        quartiers=quartiers,
        suggested_partners=suggested_partners,
        orientation_domaines=ORIENTATION_DOMAINES,
        orientation_statuts=ORIENTATION_STATUTS,
        orientation_urgences=ORIENTATION_URGENCES,
        current_filters={
            "q": request.args.get("q", ""),
            "secteur_filter": request.args.get("secteur_filter", ""),
            "domaine": selected_domain,
            "statut": request.args.get("statut", ""),
            "partenaire_id": request.args.get("partenaire_id", ""),
            "ville": request.args.get("ville", ""),
            "quartier_id": request.args.get("quartier_id", ""),
        },
        can_edit_orientations=can("partenaires:edit"),
        can_scope_all=_can_scope_all(),
        current_scope_secteur=_current_scope_secteur(),
        domain_label=_domain_label,
        status_label=_status_label,
        urgence_label=_urgence_label,
        participant_label=_participant_label,
        quartier_label=_quartier_label,
        orientation_ville=_orientation_ville,
        orientation_quartier_label=_orientation_quartier_label,
    )


@bp.route("/orientations/<int:orientation_id>/status", methods=["POST"])
@login_required
@require_perm("partenaires:edit")
def update_orientation_status(orientation_id: int):
    item = db.get_or_404(OrientationAccesDroit, orientation_id)
    if not _can_scope_all() and item.secteur and item.secteur != _current_scope_secteur():
        abort(403)
    item.statut = _clean_choice(request.form.get("statut"), ORIENTATION_STATUTS, item.statut or "oriente")
    item.suite_prevue = _parse_date(request.form.get("suite_prevue"))
    extra_note = (request.form.get("note") or "").strip()
    if extra_note:
        item.note = ((item.note or "").strip() + "\n" + extra_note).strip()
    db.session.commit()
    flash("Suivi de l'orientation mis à jour.", "success")
    return redirect(url_for("partenaires.orientations", year=(item.date_orientation or date.today()).year))


@bp.route("/orientations/<int:orientation_id>/delete", methods=["POST"])
@login_required
@require_perm("partenaires:edit")
def delete_orientation(orientation_id: int):
    item = db.get_or_404(OrientationAccesDroit, orientation_id)
    if not _can_scope_all() and item.secteur and item.secteur != _current_scope_secteur():
        abort(403)
    year = (item.date_orientation or date.today()).year
    db.session.delete(item)
    db.session.commit()
    flash("Orientation supprimée.", "success")
    return redirect(url_for("partenaires.orientations", year=year))


@bp.route("/orientations/export.xlsx")
@login_required
@require_perm("partenaires:view")
def export_orientations_xlsx():
    year = _orientation_year()
    rows = (
        _orientation_filtered_query(year)
        .order_by(OrientationAccesDroit.date_orientation.asc(), OrientationAccesDroit.created_at.asc())
        .all()
    )
    dashboard = _orientation_dashboard_payload(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Synthèse"
    ws.append(["Accès aux droits - orientations", year])
    ws.append([])
    ws.append(["Indicateur", "Valeur"])
    for label, value in (
        ("Orientations", dashboard["kpis"]["total"]),
        ("Personnes rattachées", dashboard["kpis"]["participants"]),
        ("Partenaires sollicités", dashboard["kpis"]["partenaires"]),
        ("Domaines", dashboard["kpis"]["domaines"]),
        ("Villes", dashboard["kpis"]["villes"]),
        ("Quartiers", dashboard["kpis"]["quartiers"]),
        ("À rappeler", dashboard["kpis"]["a_rappeler"]),
        ("Sans partenaire", dashboard["kpis"]["sans_partenaire"]),
    ):
        ws.append([label, value])

    def write_counter(sheet_name: str, title: str, items: list[dict]):
        sheet = wb.create_sheet(sheet_name)
        sheet.append([title])
        sheet.append(["Libellé", "Nombre"])
        for item in items:
            sheet.append([item["label"], item["count"]])
        return sheet

    write_counter("Par domaine", "Demandes par domaine", dashboard["by_domaine"])
    write_counter("Par partenaire", "Orientations par partenaire", dashboard["by_partner"])
    write_counter("Par statut", "Orientations par statut", dashboard["by_status"])
    write_counter("Par ville", "Orientations par ville", dashboard["by_ville"])
    write_counter("Par quartier", "Orientations par quartier", dashboard["by_quartier"])
    write_counter("QPV genre age", "Répartition QPV", dashboard["by_qpv"])

    detail = wb.create_sheet("Détail orientations")
    detail.append([
        "Date", "Secteur", "Ville", "Quartier", "Personne", "QPV", "Genre", "Âge", "Domaine", "Demande",
        "Partenaire", "Statut", "Urgence", "Suite prévue", "Note",
    ])
    for row in rows:
        participant = row.participant
        quartier = _orientation_quartier(row)
        detail.append([
            row.date_orientation.isoformat() if row.date_orientation else "",
            row.secteur or "",
            _orientation_ville(row),
            _quartier_label(quartier) if quartier else "",
            _participant_label(participant, row.demandeur_libre),
            _qpv_bucket(participant, quartier),
            getattr(participant, "genre", None) or "",
            getattr(participant, "age", None) if participant else "",
            _domain_label(row.domaine),
            row.demande,
            row.partenaire.nom if row.partenaire else "",
            _status_label(row.statut),
            _urgence_label(row.urgence),
            row.suite_prevue.isoformat() if row.suite_prevue else "",
            row.note or "",
        ])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        if sheet.max_row >= 2:
            for cell in sheet[2]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="E8EEF9")
        for col in range(1, sheet.max_column + 1):
            width = min(42, max(12, max(len(str(sheet.cell(row=r, column=col).value or "")) for r in range(1, sheet.max_row + 1)) + 2))
            sheet.column_dimensions[get_column_letter(col)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"orientations_acces_droits_{year}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/new", methods=["GET", "POST"])
@login_required
@require_perm("partenaires:edit")
def create():
    if request.method == "POST":
        nom = (request.form.get("nom") or "").strip()
        if not nom:
            flash("Le nom du partenaire est obligatoire.", "danger")
            return redirect(url_for("partenaires.create"))

        partenaire = Partenaire(
            nom=nom,
            contact_nom=(request.form.get("contact_nom") or "").strip() or None,
            contact_prenom=(request.form.get("contact_prenom") or "").strip() or None,
            adresse=(request.form.get("adresse") or "").strip() or None,
            email_contact=(request.form.get("email_contact") or "").strip() or None,
            email_general=(request.form.get("email_general") or "").strip() or None,
            tel_contact=(request.form.get("tel_contact") or "").strip() or None,
            tel_general=(request.form.get("tel_general") or "").strip() or None,
            description=(request.form.get("description") or "").strip() or None,
            competences_orientation_json=json.dumps(_selected_orientation_domains_from_request(), ensure_ascii=False),
            territoire_couvert=(request.form.get("territoire_couvert") or "").strip() or None,
            modalites_orientation=(request.form.get("modalites_orientation") or "").strip() or None,
            niveau_orientation=_clean_choice(request.form.get("niveau_orientation"), PARTENAIRE_NIVEAUX_ORIENTATION, "relais"),
        )
        db.session.add(partenaire)
        db.session.flush()

        secteurs = _selected_secteurs_from_request()
        for secteur in secteurs:
            db.session.add(PartenaireSecteur(partenaire_id=partenaire.id, secteur=secteur))

        db.session.commit()
        flash("Partenaire créé.", "success")
        return redirect(url_for("partenaires.edit", partenaire_id=partenaire.id))

    return render_template(
        "partenaires/form.html",
        partenaire=None,
        secteurs=[],
        competences=[],
        orientation_domaines=ORIENTATION_DOMAINES,
        orientation_niveaux=PARTENAIRE_NIVEAUX_ORIENTATION,
    )


@bp.route("/<int:partenaire_id>/edit", methods=["GET", "POST"])
@login_required
@require_perm("partenaires:edit")
def edit(partenaire_id: int):
    partenaire = db.get_or_404(Partenaire, partenaire_id)

    if request.method == "POST":
        nom = (request.form.get("nom") or "").strip()
        if not nom:
            flash("Le nom du partenaire est obligatoire.", "danger")
            return redirect(url_for("partenaires.edit", partenaire_id=partenaire.id))

        partenaire.nom = nom
        partenaire.contact_nom = (request.form.get("contact_nom") or "").strip() or None
        partenaire.contact_prenom = (request.form.get("contact_prenom") or "").strip() or None
        partenaire.adresse = (request.form.get("adresse") or "").strip() or None
        partenaire.email_contact = (request.form.get("email_contact") or "").strip() or None
        partenaire.email_general = (request.form.get("email_general") or "").strip() or None
        partenaire.tel_contact = (request.form.get("tel_contact") or "").strip() or None
        partenaire.tel_general = (request.form.get("tel_general") or "").strip() or None
        partenaire.description = (request.form.get("description") or "").strip() or None
        partenaire.competences_orientation_json = json.dumps(_selected_orientation_domains_from_request(), ensure_ascii=False)
        partenaire.territoire_couvert = (request.form.get("territoire_couvert") or "").strip() or None
        partenaire.modalites_orientation = (request.form.get("modalites_orientation") or "").strip() or None
        partenaire.niveau_orientation = _clean_choice(request.form.get("niveau_orientation"), PARTENAIRE_NIVEAUX_ORIENTATION, "relais")

        secteurs = _selected_secteurs_from_request()
        PartenaireSecteur.query.filter_by(partenaire_id=partenaire.id).delete()
        for secteur in secteurs:
            db.session.add(PartenaireSecteur(partenaire_id=partenaire.id, secteur=secteur))

        db.session.commit()
        flash("Partenaire mis à jour.", "success")
        return redirect(url_for("partenaires.edit", partenaire_id=partenaire.id))

    secteurs = [s.secteur for s in partenaire.secteurs]
    return render_template(
        "partenaires/form.html",
        partenaire=partenaire,
        secteurs=secteurs,
        competences=partenaire.competences_orientation(),
        orientation_domaines=ORIENTATION_DOMAINES,
        orientation_niveaux=PARTENAIRE_NIVEAUX_ORIENTATION,
    )


@bp.route("/<int:partenaire_id>/delete", methods=["POST"])
@login_required
@require_perm("partenaires:delete")
def delete(partenaire_id: int):
    partenaire = db.get_or_404(Partenaire, partenaire_id)
    db.session.delete(partenaire)
    db.session.commit()
    flash("Partenaire supprimé.", "success")
    return redirect(url_for("partenaires.index"))


@bp.route("/<int:partenaire_id>/interventions", methods=["POST"])
@login_required
@require_perm("partenaires:edit")
def add_intervention(partenaire_id: int):
    partenaire = db.get_or_404(Partenaire, partenaire_id)
    date_value = _parse_date(request.form.get("date_intervention"))
    if not date_value:
        flash("La date d'intervention est obligatoire.", "danger")
        return redirect(url_for("partenaires.edit", partenaire_id=partenaire.id))

    intervention = PartenaireIntervention(
        partenaire_id=partenaire.id,
        secteur=(request.form.get("secteur") or "").strip() or None,
        date_intervention=date_value,
        description=(request.form.get("description") or "").strip() or None,
        created_by_user_id=getattr(current_user, "id", None),
    )
    db.session.add(intervention)
    db.session.commit()
    flash("Intervention ajoutée.", "success")
    return redirect(url_for("partenaires.edit", partenaire_id=partenaire.id))


@bp.route("/<int:partenaire_id>/interventions/<int:intervention_id>/delete", methods=["POST"])
@login_required
@require_perm("partenaires:edit")
def delete_intervention(partenaire_id: int, intervention_id: int):
    partenaire = db.get_or_404(Partenaire, partenaire_id)
    intervention = PartenaireIntervention.query.filter_by(id=intervention_id, partenaire_id=partenaire.id).first()
    if not intervention:
        abort(404)
    db.session.delete(intervention)
    db.session.commit()
    flash("Intervention supprimée.", "success")
    return redirect(url_for("partenaires.edit", partenaire_id=partenaire.id))
